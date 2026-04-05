"""
Export Service
Handles Excel and PDF export logic for schedules
"""
import io
import os
from datetime import datetime, time, timedelta
from flask import current_app
from openpyxl import Workbook
from app.models.settings import InstitutionSettings, AcademicSettings
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.drawing.image import Image as ExcelImage
from openpyxl.drawing.spreadsheet_drawing import OneCellAnchor, AnchorMarker
from openpyxl.drawing.spreadsheet_drawing import XDRPositiveSize2D
from openpyxl.utils import get_column_letter
from PIL import Image, ImageDraw, ImageFont
from reportlab.lib.pagesizes import letter, landscape, A4
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, Image as RLImage
from reportlab.lib import colors as rl_colors
from reportlab.lib.enums import TA_CENTER, TA_LEFT


# ============================================================================
# HELPER: Resolve department name / secretary from program (with fallback)
# ============================================================================

def _get_department_name(program, fallback='{Program Name}'):
    """Get the department/program display name, preferring department relationship."""
    if not program:
        return fallback
    if program.department_id and program.department:
        return program.department.department_name
    return program.program_name or fallback


def _get_secretary_name(program):
    """Get the secretary name from the program's parent department."""
    if not program:
        return ''
    if program.department_id and program.department and program.department.secretary_name:
        return program.department.secretary_name
    return ''


# ============================================================================
# STANDARDIZED HEADER FORMAT (Based on sample format)
# ============================================================================

# ----------------------------------------------------------------------------
# Centered logo helper — uses OneCellAnchor so the image is truly centered
# inside a column rather than placed at the raw top-left corner of the cell.
# ----------------------------------------------------------------------------

def add_logo_centered(ws, logo_path, col_0idx, img_width_px, img_height_px,
                      col_width_chars, header_height_pts):
    """
    Insert an image into the worksheet anchored at ``col_0idx`` (0-based) and
    row 0, with EMU offsets that visually center the logo inside the given
    column width and vertical header span.

    Args:
        ws:                Worksheet object
        logo_path:         Absolute path to the image file
        col_0idx:          0-based column index for the logo's column
        img_width_px:      Rendered image width in pixels
        img_height_px:     Rendered image height in pixels
        col_width_chars:   Excel column-width in character units
        header_height_pts: Total point-height of the rows the logo spans
    """
    try:
        if not logo_path or not os.path.exists(logo_path):
            return

        EMU_PER_PX  = 9525    # 96 DPI — 1 px = 9525 EMU
        EMU_PER_PT  = 12700   # 1 pt  = 12700 EMU
        CHAR_PX     = 7.0017  # approx px per Excel char-width unit (Calibri 11pt)

        col_emu   = col_width_chars * CHAR_PX * EMU_PER_PX
        img_w_emu = img_width_px  * EMU_PER_PX
        img_h_emu = img_height_px * EMU_PER_PX
        hdr_emu   = header_height_pts * EMU_PER_PT

        col_off = int(max(0, (col_emu  - img_w_emu) / 2))
        row_off = int(max(0, (hdr_emu  - img_h_emu) / 2))

        img = ExcelImage(logo_path)
        img.width  = img_width_px
        img.height = img_height_px

        marker   = AnchorMarker(col=col_0idx, colOff=col_off, row=0, rowOff=row_off)
        img_size = XDRPositiveSize2D(img_w_emu, img_h_emu)
        img.anchor = OneCellAnchor(_from=marker, ext=img_size)
        ws.add_image(img)
    except Exception as exc:
        print(f"[LOGO CENTERED] Could not add logo: {exc}")


def get_images_dir():
    """Get the absolute path to static/images directory"""
    base_dir = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
    return os.path.join(base_dir, 'static', 'images')


def get_institution_head_name():
    """
    Get the department president name from institution settings.
    Returns the department president name or a placeholder if not set.
    """
    try:
        settings = InstitutionSettings.query.first()
        if settings and settings.institution_head:
            return settings.institution_head.upper()
    except Exception:
        pass
    return '{Department President Name}'


def get_institution_name():
    """
    Get the institution name from institution settings.
    Returns the institution name or default 'NORZAGARAY COLLEGE' if not set.
    """
    try:
        settings = InstitutionSettings.query.first()
        if settings and settings.institution_name:
            return settings.institution_name.upper()
    except Exception:
        pass
    return 'NORZAGARAY COLLEGE'


def get_excel_header_config():
    """
    Get configurable Excel export header settings from the DB.
    Returns (line1, line2, schedule_color) with safe fallbacks.
    """
    try:
        settings = InstitutionSettings.query.first()
        if settings:
            line1 = (settings.excel_header_line1 or 'Republic of the Philippines').strip()
            line2 = (settings.excel_header_line2 or 'Municipality of Norzagaray').strip()
            color = (settings.excel_schedule_color or '').strip()
            # If set, ensure it is a valid 6-digit hex; otherwise treat as no fill
            if color and not (len(color) == 7 and color.startswith('#')):
                color = ''
            return line1, line2, color
    except Exception:
        pass
    return 'Republic of the Philippines', 'Municipality of Norzagaray', ''


def get_institution_logo_path():
    """
    Get the absolute file path to the left institution logo (seal).
    Reads from InstitutionSettings DB, falls back to default static image.
    """
    try:
        settings = InstitutionSettings.query.first()
        if settings and settings.institution_logo:
            base_dir = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
            logo_path = os.path.join(base_dir, 'static', settings.institution_logo)
            if os.path.exists(logo_path):
                return logo_path
    except Exception:
        pass
    return os.path.join(get_images_dir(), 'norzagaray-college-logo.png')


def get_institution_logo_right_path():
    """
    Get the absolute file path to the right institution logo (e.g., Bagong Pilipinas).
    Reads from InstitutionSettings DB, falls back to default static image.
    """
    try:
        settings = InstitutionSettings.query.first()
        if settings and settings.institution_logo_right:
            base_dir = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
            logo_path = os.path.join(base_dir, 'static', settings.institution_logo_right)
            if os.path.exists(logo_path):
                return logo_path
    except Exception:
        pass
    return os.path.join(get_images_dir(), 'bagong-pilipinas.png')


def format_semester_text(semester, academic_year):
    """
    Format semester text with proper ordinal suffix.
    Converts '1 Semester' to '1ST SEMESTER', '2 Semester' to '2ND SEMESTER', etc.
    
    Args:
        semester: The semester string (e.g., '1 Semester', '2 Semester')
        academic_year: The academic year string (e.g., '2025-2026')
    
    Returns:
        Formatted string like '1ST SEMESTER, AY 2025-2026'
    """
    if not semester:
        return f"AY {academic_year}" if academic_year else ""
    
    semester_upper = semester.upper().strip()
    
    # Map semester numbers to ordinals
    ordinal_map = {
        '1': '1ST', 'FIRST': '1ST', '1ST': '1ST',
        '2': '2ND', 'SECOND': '2ND', '2ND': '2ND',
        '3': '3RD', 'THIRD': '3RD', '3RD': '3RD',
        'SUMMER': 'SUMMER', 'MID-YEAR': 'MID-YEAR'
    }
    
    # Extract the number/name from semester string
    # Handle formats like "1 Semester", "First Semester", "1ST SEMESTER"
    for key, ordinal in ordinal_map.items():
        if semester_upper.startswith(key) or key in semester_upper:
            formatted_semester = f"{ordinal} SEMESTER"
            break
    else:
        # Fallback: just uppercase the semester
        formatted_semester = semester_upper
        if 'SEMESTER' not in formatted_semester:
            formatted_semester += ' SEMESTER'
    
    if academic_year:
        return f"{formatted_semester}, AY {academic_year}"
    return formatted_semester


def create_standard_pdf_header(
    report_title,
    office_name="OFFICE OF THE REGISTRAR",
    subtitle=None,
    period_text=None,
    department_logo_path=None
):
    """
    Create a standardized PDF header matching the institutional format.
    
    Args:
        report_title: Main title of the report (e.g., "CLASS SCHEDULE REPORT")
        office_name: Office name (e.g., "OFFICE OF THE COLLEGE DEAN", "OFFICE OF THE REGISTRAR")
        subtitle: Optional subtitle text
        period_text: Optional period text (e.g., "First Semester, A.Y. 2025-2026")
        department_logo_path: Optional path to program logo
    
    Returns:
        List of reportlab elements for the header
    """
    elements = []
    styles = getSampleStyleSheet()
    
    # Try to add centered logo (larger size: 1.0 inch)
    logo_path = get_institution_logo_path()
    try:
        if os.path.exists(logo_path):
            logo = RLImage(logo_path, width=1.0*inch, height=1.0*inch)
            # Center the logo using a table
            logo_table = Table([[logo]], colWidths=[8*inch])
            logo_table.setStyle(TableStyle([
                ('ALIGN', (0, 0), (0, 0), 'CENTER'),
                ('VALIGN', (0, 0), (0, 0), 'MIDDLE'),
            ]))
            elements.append(logo_table)
    except Exception as e:
        print(f"[PDF HEADER] Could not add logo: {e}")
    
    # Title styles (Times font for formal documents)
    institution_style = ParagraphStyle(
        'InstitutionName',
        parent=styles['Normal'],
        fontSize=12,
        fontName='Times-Bold',
        alignment=TA_CENTER,
        spaceAfter=2
    )
    
    address_style = ParagraphStyle(
        'Address',
        parent=styles['Normal'],
        fontSize=10,
        fontName='Times-Roman',
        alignment=TA_CENTER,
        spaceAfter=2
    )
    
    office_style = ParagraphStyle(
        'OfficeName',
        parent=styles['Normal'],
        fontSize=10,
        fontName='Times-Bold',
        alignment=TA_CENTER,
        spaceAfter=4
    )
    
    report_title_style = ParagraphStyle(
        'ReportTitle',
        parent=styles['Normal'],
        fontSize=11,
        fontName='Times-Bold',
        alignment=TA_CENTER,
        spaceAfter=3
    )
    
    period_style = ParagraphStyle(
        'PeriodText',
        parent=styles['Normal'],
        fontSize=10,
        fontName='Times-Roman',
        alignment=TA_CENTER,
        spaceAfter=12
    )
    
    # Add institution header
    elements.append(Paragraph(get_institution_name(), institution_style))
    elements.append(Paragraph('Municipal Compound, Norzagaray, Bulacan', address_style))
    elements.append(Paragraph(office_name.upper(), office_style))
    elements.append(Paragraph(report_title.upper(), report_title_style))
    
    if subtitle:
        elements.append(Paragraph(subtitle, period_style))
    
    if period_text:
        elements.append(Paragraph(period_text, period_style))
    
    elements.append(Spacer(1, 0.15*inch))
    
    return elements


def create_standard_excel_header(
    ws,
    report_title,
    office_name="OFFICE OF THE REGISTRAR",
    subtitle=None,
    period_text=None,
    merge_columns='A:F',
    start_row=1,
    department_logo_path=None
):
    """
    Create a standardized Excel header matching the institutional format.
    
    Args:
        ws: Worksheet object
        report_title: Main title of the report
        office_name: Office name (e.g., "OFFICE OF THE COLLEGE DEAN")
        subtitle: Optional subtitle text
        period_text: Optional period text (e.g., "First Semester, A.Y. 2025-2026")
        merge_columns: Column range to merge (e.g., 'A:F', 'A:H')
        start_row: Starting row for header
        department_logo_path: Optional path to program logo
    
    Returns:
        Next available row after header
    """
    
    # Parse merge columns
    start_col = merge_columns.split(':')[0]
    end_col = merge_columns.split(':')[1]
    
    # Font styles (Times New Roman for formal documents)
    title_font = Font(name='Times New Roman', size=14, bold=True)
    header_font = Font(name='Times New Roman', size=12, bold=True)
    normal_font = Font(name='Times New Roman', size=10)
    center_align = Alignment(horizontal='center', vertical='center')
    
    current_row = start_row
    
    # Try to add centered logo using proper EMU anchor
    logo_path = get_institution_logo_path()
    try:
        if os.path.exists(logo_path):
            col_count = ord(end_col) - ord(start_col) + 1
            center_col_idx = ord(start_col) + (col_count // 2) - 1
            center_col_0idx = center_col_idx - ord('A')
            center_col_letter = chr(center_col_idx) if center_col_idx >= ord('A') else start_col
            col_w = ws.column_dimensions[center_col_letter].width or 14
            ws.row_dimensions[current_row].height = 65
            ws.row_dimensions[current_row + 1].height = 5
            add_logo_centered(ws, logo_path, col_0idx=center_col_0idx,
                              img_width_px=80, img_height_px=80,
                              col_width_chars=col_w, header_height_pts=65)
            current_row += 2
    except Exception as e:
        print(f"[EXCEL HEADER] Could not add logo: {e}")
    
    # Institution name
    ws.merge_cells(f'{start_col}{current_row}:{end_col}{current_row}')
    ws[f'{start_col}{current_row}'] = get_institution_name()
    ws[f'{start_col}{current_row}'].font = title_font
    ws[f'{start_col}{current_row}'].alignment = center_align
    current_row += 1
    
    # Address
    ws.merge_cells(f'{start_col}{current_row}:{end_col}{current_row}')
    ws[f'{start_col}{current_row}'] = 'Municipal Compound, Norzagaray, Bulacan'
    ws[f'{start_col}{current_row}'].font = normal_font
    ws[f'{start_col}{current_row}'].alignment = center_align
    current_row += 1
    
    # Office name
    ws.merge_cells(f'{start_col}{current_row}:{end_col}{current_row}')
    ws[f'{start_col}{current_row}'] = office_name.upper()
    ws[f'{start_col}{current_row}'].font = header_font
    ws[f'{start_col}{current_row}'].alignment = center_align
    current_row += 1
    
    # Report title
    ws.merge_cells(f'{start_col}{current_row}:{end_col}{current_row}')
    ws[f'{start_col}{current_row}'] = report_title.upper()
    ws[f'{start_col}{current_row}'].font = Font(name='Times New Roman', size=11, bold=False)
    ws[f'{start_col}{current_row}'].alignment = center_align
    current_row += 1
    
    # Subtitle (if provided)
    if subtitle:
        ws.merge_cells(f'{start_col}{current_row}:{end_col}{current_row}')
        ws[f'{start_col}{current_row}'] = subtitle
        ws[f'{start_col}{current_row}'].font = normal_font
        ws[f'{start_col}{current_row}'].alignment = center_align
        current_row += 1
    
    # Period text (if provided)
    if period_text:
        ws.merge_cells(f'{start_col}{current_row}:{end_col}{current_row}')
        ws[f'{start_col}{current_row}'] = period_text
        ws[f'{start_col}{current_row}'].font = normal_font
        ws[f'{start_col}{current_row}'].alignment = center_align
        current_row += 1
    
    # Add empty row for spacing
    current_row += 1
    
    return current_row


def create_posting_style_excel_header(
    ws,
    report_title,
    office_name="OFFICE OF THE REGISTRAR",
    subtitle=None,
    last_col='G',
    left_col_width=14,
    right_col_width=20
):
    """
    Create Excel header matching the class schedule posting format.
    Dual logos (left + right), left-aligned institutional text, centered title.
    
    Args:
        ws: Worksheet object
        report_title: Main title (e.g., "SYSTEM ACTIVITY LOG REPORT")
        office_name: Office name (e.g., "OFFICE OF THE SYSTEM ADMINISTRATOR")
        subtitle: Optional subtitle text (e.g., generated date)
        last_col: Last column letter for right logo and merge range
    
    Returns:
        Next available row after header
    """
    images_dir = get_images_dir()
    last_col_idx = ord(last_col) - ord('A') + 1
    # Header row total height: rows 1-4 (18+15+15+15 = 63 pt)
    _HDR_HEIGHT_PT = 63
    # Left logo always full size; right logo capped to its column width so it
    # never overflows past the last column.
    _CHAR_PX = 7.0017
    left_logo_px  = 75
    right_logo_px = min(80, int(right_col_width * _CHAR_PX))

    # Left logo (Institution Seal) — column A, full size
    logo_left_path = get_institution_logo_path()
    add_logo_centered(ws, logo_left_path, col_0idx=0,
                      img_width_px=left_logo_px, img_height_px=left_logo_px,
                      col_width_chars=left_col_width,
                      header_height_pts=_HDR_HEIGHT_PT)

    # Right logo (Bagong Pilipinas) — last column, full size
    logo_right_path = get_institution_logo_right_path()
    add_logo_centered(ws, logo_right_path,
                      col_0idx=ord(last_col) - ord('A'),
                      img_width_px=right_logo_px, img_height_px=right_logo_px,
                      col_width_chars=right_col_width,
                      header_height_pts=_HDR_HEIGHT_PT)

    # Auto-calculate indent for header text at column B so it clears the logo
    # even when column A is narrower than the logo width.
    _col_a_px   = left_col_width * _CHAR_PX
    _overflow_px = max(0.0, left_logo_px - _col_a_px)
    _text_indent = int(_overflow_px / _CHAR_PX) + 1 if _overflow_px > 0 else 0
    left_align = Alignment(horizontal='left', vertical='center', indent=_text_indent)

    # Header text — left-aligned at column B (rows 1-4)
    _hdr_line1, _hdr_line2, _ = get_excel_header_config()
    ws['B1'] = _hdr_line1
    ws['B1'].font = Font(size=11)
    ws['B1'].alignment = left_align
    
    ws['B2'] = _hdr_line2
    ws['B2'].font = Font(size=11)
    ws['B2'].alignment = left_align
    
    ws['B3'] = get_institution_name()
    ws['B3'].font = Font(bold=True, size=11)
    ws['B3'].alignment = left_align
    
    ws['B4'] = office_name.upper()
    ws['B4'].font = Font(bold=True, size=11)
    ws['B4'].alignment = left_align
    
    # Row heights
    ws.row_dimensions[1].height = 18
    ws.row_dimensions[2].height = 15
    ws.row_dimensions[3].height = 15
    ws.row_dimensions[4].height = 15
    ws.row_dimensions[5].height = 12  # Spacer
    
    center_align = Alignment(horizontal='center', vertical='center')
    
    # Row 6: Report title (centered, bold, size 12)
    ws.merge_cells(start_row=6, start_column=1, end_row=6, end_column=last_col_idx)
    title_cell = ws.cell(row=6, column=1, value=report_title.upper())
    title_cell.font = Font(bold=True, size=12)
    title_cell.alignment = center_align
    
    current_row = 7
    
    # Row 7: Subtitle (if provided)
    if subtitle:
        ws.merge_cells(start_row=7, start_column=1, end_row=7, end_column=last_col_idx)
        sub_cell = ws.cell(row=7, column=1, value=subtitle)
        sub_cell.font = Font(bold=True, size=11)
        sub_cell.alignment = center_align
        current_row = 8
    
    # Spacer row
    ws.row_dimensions[current_row].height = 8
    current_row += 1
    
    return current_row


def add_pdf_info_section(
    info_items,
    col_widths=None
):
    """
    Create a PDF info section with label-value pairs.
    
    Args:
        info_items: List of dicts with 'label' and 'value' keys
        col_widths: Optional column widths [label_width, value_width]
    
    Returns:
        Table element with info items
    """
    styles = getSampleStyleSheet()
    
    header_style = ParagraphStyle(
        'InfoHeader',
        parent=styles['Normal'],
        fontSize=10,
        fontName='Times-Bold'
    )
    
    if col_widths is None:
        col_widths = [2*inch, 4*inch]
    
    table_data = []
    for item in info_items:
        label = Paragraph(f"<b>{item['label']}:</b>", header_style)
        value = Paragraph(str(item['value']), header_style)
        table_data.append([label, value])
    
    info_table = Table(table_data, colWidths=col_widths)
    info_table.setStyle(TableStyle([
        ('ALIGN', (0, 0), (0, -1), 'LEFT'),
        ('ALIGN', (1, 0), (1, -1), 'LEFT'),
        ('VALIGN', (0, 0), (-1, -1), 'TOP'),
        ('TOPPADDING', (0, 0), (-1, -1), 2),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 2),
    ]))
    
    return info_table


def add_excel_info_section(ws, info_items, start_row, start_col='A'):
    """
    Add an info section to Excel with label-value pairs.
    
    Args:
        ws: Worksheet object
        info_items: List of dicts with 'label' and 'value' keys
        start_row: Starting row
        start_col: Starting column letter
    
    Returns:
        Next available row
    """
    header_font = Font(name='Times New Roman', size=10, bold=True)
    normal_font = Font(name='Times New Roman', size=10)
    
    current_row = start_row
    for item in info_items:
        col_idx = ord(start_col) - ord('A') + 1
        ws.cell(row=current_row, column=col_idx, value=f"{item['label']}:").font = header_font
        ws.cell(row=current_row, column=col_idx + 1, value=str(item['value'])).font = normal_font
        current_row += 1
    
    return current_row + 1  # Add spacing


def add_pdf_rating_scale_legend():
    """Add a standard rating scale legend for PDF exports"""
    styles = getSampleStyleSheet()
    legend_style = ParagraphStyle(
        'Legend',
        parent=styles['Normal'],
        fontSize=8,
        fontName='Times-Roman'
    )
    return Paragraph(
        '<b>Rating Scale:</b> 4.50-5.00: Outstanding | 3.50-4.49: Highly Satisfactory | '
        '2.50-3.49: Satisfactory | 1.50-2.49: Needs Improvement | 1.00-1.49: Poor',
        legend_style
    )


# ============================================================================
# HELPER FUNCTIONS FOR EXCEL EXPORT
# ============================================================================

def add_institution_logos(ws):
    """Add institution logos to the worksheet if they exist, centered in their columns."""
    try:
        # Header rows 1-4 total height: 18+15+15+15 = 63 pt
        _HDR_HEIGHT_PT = 63
        col_a_w = ws.column_dimensions['A'].width or 14
        col_g_w = ws.column_dimensions['G'].width or 20

        # Left logo (Norzagaray Department) — column A, 120x120
        add_logo_centered(ws, get_institution_logo_path(),
                          col_0idx=0, img_width_px=120, img_height_px=120,
                          col_width_chars=col_a_w, header_height_pts=_HDR_HEIGHT_PT)

        # Right logo (Bagong Pilipinas) — column G, 120x120
        add_logo_centered(ws, get_institution_logo_right_path(),
                          col_0idx=6, img_width_px=120, img_height_px=120,
                          col_width_chars=col_g_w, header_height_pts=_HDR_HEIGHT_PT)

    except Exception as e:
        print(f"[LOGO ERROR] Error adding logos: {str(e)}")


def add_institution_header(ws, program_name):
    """Add institution header to the worksheet"""
    # Header - Institution Information (centered across columns C-E)
    ws['C1'] = 'Republic of the Philippines'
    ws['C1'].font = Font(bold=True, size=11)
    ws['C1'].alignment = Alignment(horizontal='center', vertical='center')
    ws.merge_cells('C1:E1')
    
    ws['C2'] = 'Municipality of Norzagaray'
    ws['C2'].font = Font(size=10)
    ws['C2'].alignment = Alignment(horizontal='center', vertical='center')
    ws.merge_cells('C2:E2')
    
    ws['C3'] = get_institution_name()
    ws['C3'].font = Font(bold=True, size=11)
    ws['C3'].alignment = Alignment(horizontal='center', vertical='center')
    ws.merge_cells('C3:E3')
    
    dept_name = program_name.upper() if program_name else '{Program Name}'
    ws['C4'] = f'{dept_name}'
    ws['C4'].font = Font(bold=True, size=11)
    ws['C4'].alignment = Alignment(horizontal='center', vertical='center')
    ws.merge_cells('C4:E4')


def add_schedule_title(ws, title, semester_text, section_display):
    """Add schedule title and metadata"""
    # Title
    ws['A6'] = title
    ws['A6'].font = Font(bold=True, size=12)
    ws['A6'].alignment = Alignment(horizontal='center', vertical='center')
    ws.merge_cells('A6:G6')
    
    # Semester and AY - BOLD
    ws['A7'] = semester_text
    ws['A7'].font = Font(bold=True, size=11)
    ws['A7'].alignment = Alignment(horizontal='center', vertical='center')
    ws.merge_cells('A7:G7')
    
    # Section/Faculty/Room info
    ws['A8'] = section_display
    ws['A8'].font = Font(bold=True, size=11)
    ws['A8'].alignment = Alignment(horizontal='center', vertical='center')
    ws.merge_cells('A8:G8')


def generate_time_slots(start_hour=7, end_hour=20):
    """Generate 30-minute time slots from start_hour to end_hour"""
    time_slots = []
    current_time = time(start_hour, 0)
    
    while current_time.hour < end_hour:
        next_time = (datetime.combine(datetime.today(), current_time) + timedelta(minutes=30)).time()
        
        # Format time slots (7:00-7:30, 12:00-12:30, 1:00-1:30, etc.)
        start_str = current_time.strftime('%#I:%M').lstrip('0') if current_time.hour != 12 else current_time.strftime('12:%M')
        end_str = next_time.strftime('%#I:%M').lstrip('0') if next_time.hour != 12 else next_time.strftime('12:%M')
        
        time_slots.append(f"{start_str}-{end_str}")
        current_time = next_time
    
    return time_slots


def add_column_headers(ws):
    """Add day column headers to the worksheet"""
    operation_days = AcademicSettings.get_active_operation_days()
    headers = ['TIME'] + [d.upper() for d in operation_days]
    _, _, schedule_color = get_excel_header_config()
    if schedule_color:
        hex_color = schedule_color.lstrip('#')
        header_fill = PatternFill(start_color=hex_color, end_color=hex_color, fill_type='solid')
        header_font = Font(bold=True, size=10, color='FFFFFF')
    else:
        header_fill = None
        header_font = Font(bold=True, size=10, color='000000')
    header_alignment = Alignment(horizontal='center', vertical='center')
    
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=10, column=col_idx, value=header)
        cell.font = header_font
        if header_fill:
            cell.fill = header_fill
        cell.alignment = header_alignment


def write_time_slots(ws, time_slots):
    """Write time slots to the worksheet"""
    for idx, time_slot in enumerate(time_slots, start=11):
        ws.cell(row=idx, column=1, value=time_slot)
        ws.cell(row=idx, column=1).alignment = Alignment(horizontal='left', vertical='center')
        ws.cell(row=idx, column=1).font = Font(size=10)


def get_day_column_mapping():
    """Get mapping of day names to column indices (dynamic based on operation days)"""
    operation_days = AcademicSettings.get_active_operation_days()
    return {day: idx + 2 for idx, day in enumerate(operation_days)}


def place_schedule_in_grid(ws, schedules, start_hour=7):
    """Place schedules in the weekly grid"""
    day_columns = get_day_column_mapping()
    
    for schedule in schedules:
        if schedule.day_of_week not in day_columns:
            continue
        
        col_idx = day_columns[schedule.day_of_week]
        
        # Calculate which 30-minute slot this starts in
        start_minutes = schedule.start_time.hour * 60 + schedule.start_time.minute
        end_minutes = schedule.end_time.hour * 60 + schedule.end_time.minute
        slot_start_minutes = start_hour * 60
        
        # Find start row (each row is 30 minutes, starting from row 11)
        start_row = 11 + ((start_minutes - slot_start_minutes) // 30)
        
        # Calculate how many rows to merge
        duration_minutes = end_minutes - start_minutes
        rows_to_merge = max(1, (duration_minutes + 29) // 30)
        
        # Build cell content - always show all 3 lines
        subject_code = schedule.subject.subject_code if schedule.subject else 'TBA'
        room_display = schedule.room.room_number if schedule.room else 'TBA'
        faculty_name = schedule.faculty.full_name if schedule.faculty else 'TBA'
        
        cell_content = f"{subject_code}\n{room_display}\n{faculty_name}"
        
        # Write to cell
        cell = ws.cell(row=start_row, column=col_idx, value=cell_content)
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.font = Font(size=10)
        
        # Merge cells if needed
        if rows_to_merge > 1:
            end_row = start_row + rows_to_merge - 1
            ws.merge_cells(start_row=start_row, start_column=col_idx, 
                         end_row=end_row, end_column=col_idx)


def place_faculty_schedule_in_grid(ws, schedules, start_hour=7):
    """Place faculty schedules in the weekly grid - shows subject/room/section"""
    day_columns = get_day_column_mapping()
    
    for schedule in schedules:
        if schedule.day_of_week not in day_columns:
            continue
        
        col_idx = day_columns[schedule.day_of_week]
        
        # Calculate which 30-minute slot this starts in
        start_minutes = schedule.start_time.hour * 60 + schedule.start_time.minute
        end_minutes = schedule.end_time.hour * 60 + schedule.end_time.minute
        slot_start_minutes = start_hour * 60
        
        # Find start row (each row is 30 minutes, starting from row 11)
        start_row = 11 + ((start_minutes - slot_start_minutes) // 30)
        
        # Calculate how many rows to merge
        duration_minutes = end_minutes - start_minutes
        rows_to_merge = max(1, (duration_minutes + 29) // 30)
        
        # Build cell content - for faculty: subject/room/section
        subject_code = schedule.subject.subject_code if schedule.subject else 'TBA'
        room_display = schedule.room.room_number if schedule.room else 'TBA'
        section = schedule.section
        section_display = ''
        if section:
            if hasattr(section, 'full_section_name'):
                section_display = section.full_section_name
            elif section.program:
                section_display = f"{section.program.program_code}-{section.year_level}{section.section_name}"
            else:
                section_display = f"{section.year_level}{section.section_name}"
        else:
            section_display = 'TBA'
        
        cell_content = f"{subject_code}\n{room_display}\n{section_display}"
        
        # Write to cell
        cell = ws.cell(row=start_row, column=col_idx, value=cell_content)
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.font = Font(size=10)
        
        # Merge cells if needed
        if rows_to_merge > 1:
            end_row = start_row + rows_to_merge - 1
            ws.merge_cells(start_row=start_row, start_column=col_idx, 
                         end_row=end_row, end_column=col_idx)


def place_room_schedule_in_grid(ws, schedules, start_hour=7):
    """Place room schedules in the weekly grid - shows subject/section/faculty"""
    day_columns = get_day_column_mapping()
    
    for schedule in schedules:
        if schedule.day_of_week not in day_columns:
            continue
        
        col_idx = day_columns[schedule.day_of_week]
        
        # Calculate which 30-minute slot this starts in
        start_minutes = schedule.start_time.hour * 60 + schedule.start_time.minute
        end_minutes = schedule.end_time.hour * 60 + schedule.end_time.minute
        slot_start_minutes = start_hour * 60
        
        # Find start row (each row is 30 minutes, starting from row 11)
        start_row = 11 + ((start_minutes - slot_start_minutes) // 30)
        
        # Calculate how many rows to merge
        duration_minutes = end_minutes - start_minutes
        rows_to_merge = max(1, (duration_minutes + 29) // 30)
        
        # Build cell content - for room: subject/section/faculty
        subject_code = schedule.subject.subject_code if schedule.subject else 'TBA'
        section = schedule.section
        section_display = ''
        if section:
            if hasattr(section, 'full_section_name'):
                section_display = section.full_section_name
            elif section.program:
                section_display = f"{section.program.program_code}-{section.year_level}{section.section_name}"
            else:
                section_display = f"{section.year_level}{section.section_name}"
        else:
            section_display = 'TBA'
        faculty_name = schedule.faculty.full_name if schedule.faculty else 'TBA'
        
        cell_content = f"{subject_code}\n{section_display}\n{faculty_name}"
        
        # Write to cell
        cell = ws.cell(row=start_row, column=col_idx, value=cell_content)
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.font = Font(size=10)
        
        # Merge cells if needed
        if rows_to_merge > 1:
            end_row = start_row + rows_to_merge - 1
            ws.merge_cells(start_row=start_row, start_column=col_idx, 
                         end_row=end_row, end_column=col_idx)


def apply_grid_borders(ws, last_row):
    """Apply borders to the schedule grid"""
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    for row in range(10, last_row + 1):
        for col in range(1, 8):  # A to G
            ws.cell(row=row, column=col).border = thin_border


def add_signature_section(ws, sig_start_row, program_id, dept_name, dean_name=None):
    """Add signature section with dean and department president"""
    # Prepared by section (column D)
    ws.cell(row=sig_start_row, column=4, value='Prepared by :')
    ws.cell(row=sig_start_row, column=4).font = Font(size=10)
    ws.cell(row=sig_start_row, column=4).alignment = Alignment(horizontal='left')
    
    # Noted section (column F)
    ws.cell(row=sig_start_row, column=6, value='Noted :')
    ws.cell(row=sig_start_row, column=6).font = Font(size=10)
    ws.cell(row=sig_start_row, column=6).alignment = Alignment(horizontal='left')
    
    # Dean name (column D) - Use provided dean name or default - BOLD and UPPERCASE
    dean_display_name = dean_name.upper() if dean_name else '{Dean Name}'
    ws.cell(row=sig_start_row + 2, column=4, value=dean_display_name)
    ws.cell(row=sig_start_row + 2, column=4).font = Font(bold=True, size=10)
    ws.cell(row=sig_start_row + 2, column=4).alignment = Alignment(horizontal='left')
    
    # Department President name (column F) - BOLD
    ws.cell(row=sig_start_row + 2, column=6, value=get_institution_head_name())
    ws.cell(row=sig_start_row + 2, column=6).font = Font(bold=True, size=10)
    ws.cell(row=sig_start_row + 2, column=6).alignment = Alignment(horizontal='left')
    
    # Dean title (column D) - Directly under dean name - CENTERED
    dean_title = f"Dean, {dept_name}" if dept_name else "Dean, {Program Name}"
    ws.cell(row=sig_start_row + 3, column=4, value=dean_title)
    ws.cell(row=sig_start_row + 3, column=4).font = Font(size=10)
    ws.cell(row=sig_start_row + 3, column=4).alignment = Alignment(horizontal='center')
    
    # Department President title (column F)
    ws.cell(row=sig_start_row + 3, column=6, value='Department President')
    ws.cell(row=sig_start_row + 3, column=6).font = Font(size=10)
    ws.cell(row=sig_start_row + 3, column=6).alignment = Alignment(horizontal='left')


def set_column_widths(ws):
    """Set standard column widths for schedule grid"""
    ws.column_dimensions['A'].width = 12.71
    for col in ['B', 'C', 'D', 'E', 'F', 'G']:
        ws.column_dimensions[col].width = 20.71


# ============================================================================
# PDF EXPORT HELPER FUNCTIONS (ReportLab)
# ============================================================================

def create_pdf_schedule(schedules, title, semester_text, section_display, dept_name, filename, start_hour=7, end_hour=20):
    """Create PDF schedule with exact same template as Excel export"""
    output = io.BytesIO()
    
    # Create document with landscape orientation to match Excel layout
    doc = SimpleDocTemplate(
        output,
        pagesize=landscape(letter),
        rightMargin=0.5*inch,
        leftMargin=0.5*inch,
        topMargin=0.5*inch,
        bottomMargin=0.5*inch
    )
    
    story = []
    styles = getSampleStyleSheet()
    
    # Custom styles
    header_style = ParagraphStyle(
        'CustomHeader',
        parent=styles['Normal'],
        fontSize=11,
        alignment=TA_CENTER,
        fontName='Helvetica-Bold'
    )
    
    subheader_style = ParagraphStyle(
        'CustomSubHeader',
        parent=styles['Normal'],
        fontSize=10,
        alignment=TA_CENTER,
        fontName='Helvetica'
    )
    
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Normal'],
        fontSize=12,
        alignment=TA_CENTER,
        fontName='Helvetica-Bold'
    )
    
    # Add institution header
    story.append(Paragraph('Republic of the Philippines', header_style))
    story.append(Paragraph('Municipality of Norzagaray', subheader_style))
    story.append(Paragraph(get_institution_name(), header_style))
    story.append(Paragraph(dept_name.upper() if dept_name else '{Program Name}', header_style))
    story.append(Spacer(1, 0.1*inch))
    
    # Add title and metadata
    story.append(Paragraph(title, title_style))
    story.append(Paragraph(semester_text, header_style))
    story.append(Paragraph(section_display, header_style))
    story.append(Spacer(1, 0.2*inch))
    
    # Generate time slots with dynamic range
    time_slots = generate_time_slots(start_hour=start_hour, end_hour=end_hour)
    
    # Create schedule grid data
    grid_data = []
    
    # Header row
    operation_days = AcademicSettings.get_active_operation_days()
    headers = ['TIME'] + [d.upper() for d in operation_days]
    grid_data.append(headers)
    
    # Day column mapping
    day_columns = {day: idx + 1 for idx, day in enumerate(operation_days)}
    
    # Initialize grid with empty cells
    num_days = len(operation_days)
    for time_slot in time_slots:
        row = [time_slot] + [''] * num_days
        grid_data.append(row)
    
    # Place schedules in grid
    for schedule in schedules:
        day = schedule.day_of_week
        start_time = schedule.start_time
        end_time = schedule.end_time
        
        if day not in day_columns:
            continue
        
        col_idx = day_columns[day]
        
        # Calculate which 30-minute slot this starts in (matching Excel export logic)
        start_minutes = start_time.hour * 60 + start_time.minute
        end_minutes = end_time.hour * 60 + end_time.minute
        slot_start_minutes = start_hour * 60  # Use dynamic start hour
        
        # Find start row index (each row is 30 minutes)
        start_row_idx = (start_minutes - slot_start_minutes) // 30
        
        # Calculate how many rows to span
        duration_minutes = end_minutes - start_minutes
        rows_to_span = max(1, (duration_minutes + 29) // 30)
        
        # Build schedule cell content
        subject_display = schedule.subject.course_description if schedule.subject else 'N/A'
        subject_code = schedule.subject.subject_code if schedule.subject else ''
        type_display = schedule.schedule_type.upper() if schedule.schedule_type else ''
        faculty_display = schedule.faculty.full_name if schedule.faculty else 'TBA'
        room_display = schedule.room.room_number if schedule.room else 'TBA'
        units = schedule.subject.total_units if schedule.subject else 0
        
        cell_content = f"{subject_code}\n{subject_display}\n{type_display}\n{faculty_display}\n{room_display}\n({units} units)"
        
        # Add to grid (row index + 1 because of header row)
        grid_row_idx = start_row_idx + 1
        if 0 <= start_row_idx < len(time_slots):
            if grid_data[grid_row_idx][col_idx] == '':
                grid_data[grid_row_idx][col_idx] = cell_content
            else:
                # Multiple schedules in same slot - append
                grid_data[grid_row_idx][col_idx] += f"\n\n{cell_content}"
    
    # Create table
    col_widths = [1*inch] + [1.5*inch] * num_days
    table = Table(grid_data, colWidths=col_widths, repeatRows=1)
    
    # Style table
    table.setStyle(TableStyle([
        # Header row
        ('BACKGROUND', (0, 0), (-1, 0), rl_colors.HexColor('#2563eb')),
        ('TEXTCOLOR', (0, 0), (-1, 0), rl_colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 10),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 8),
        ('TOPPADDING', (0, 0), (-1, 0), 8),
        
        # Grid
        ('GRID', (0, 0), (-1, -1), 1, rl_colors.black),
        ('ALIGN', (0, 1), (-1, -1), 'CENTER'),
        ('VALIGN', (0, 1), (-1, -1), 'MIDDLE'),
        ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 1), (-1, -1), 8),
        
        # Time column
        ('ALIGN', (0, 1), (0, -1), 'LEFT'),
        ('FONTSIZE', (0, 1), (0, -1), 9),
    ]))
    
    story.append(table)
    
    # Build PDF
    doc.build(story)
    output.seek(0)
    return output

# ============================================================================
# EXPORT SERVICE FUNCTIONS
# ============================================================================

def generate_class_schedule_excel(section, schedules, current_settings, user):
    """Generate Excel file for class schedule - grid format with posting-style header/footer"""
    # Get time range from settings (default to 7-20 if not set)
    start_hour = (current_settings.schedule_start_time.hour if current_settings and current_settings.schedule_start_time else 7)
    end_hour = (current_settings.schedule_end_time.hour if current_settings and current_settings.schedule_end_time else 20)
    
    # Create workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Schedule"
    
    # Get program info
    program = section.program
    dept_code = program.program_code if program else ''
    dept_display_name = _get_department_name(program).upper()
    
    # =========================================================================
    # LOGOS - Centered within their respective columns (A=left, G=right)
    # =========================================================================
    # Header rows 1-4: 18+15+15+15 = 63 pt total height
    _HDR_HEIGHT_PT = 63
    # Column A width = 12.71 chars; Column G width = 20.71 chars (set_column_widths)
    add_logo_centered(ws, get_institution_logo_path(),
                      col_0idx=0, img_width_px=75, img_height_px=75,
                      col_width_chars=12.71, header_height_pts=_HDR_HEIGHT_PT)
    add_logo_centered(ws, get_institution_logo_right_path(),
                      col_0idx=6, img_width_px=80, img_height_px=80,
                      col_width_chars=20.71, header_height_pts=_HDR_HEIGHT_PT)

    # =========================================================================
    # HEADER TEXT - Left aligned at column B (rows 1-4)
    # =========================================================================
    ws['B1'] = 'Republic of the Philippines'
    ws['B1'].font = Font(size=11)
    ws['B1'].alignment = Alignment(horizontal='left', vertical='center')
    
    ws['B2'] = 'Municipality of Norzagaray'
    ws['B2'].font = Font(size=11)
    ws['B2'].alignment = Alignment(horizontal='left', vertical='center')
    
    ws['B3'] = get_institution_name()
    ws['B3'].font = Font(bold=True, size=11)
    ws['B3'].alignment = Alignment(horizontal='left', vertical='center')
    
    ws['B4'] = dept_display_name
    ws['B4'].font = Font(bold=True, size=11)
    ws['B4'].alignment = Alignment(horizontal='left', vertical='center')
    
    # Set row heights for header
    ws.row_dimensions[1].height = 18
    ws.row_dimensions[2].height = 15
    ws.row_dimensions[3].height = 15
    ws.row_dimensions[4].height = 15
    ws.row_dimensions[5].height = 12  # Spacer row
    
    # =========================================================================
    # SCHEDULE TITLE - Centered across A-G (rows 6-8)
    # =========================================================================
    ws.merge_cells('A6:G6')
    ws['A6'] = 'CLASS SCHEDULE'
    ws['A6'].font = Font(bold=True, size=12)
    ws['A6'].alignment = Alignment(horizontal='center', vertical='center')
    
    semester_text = format_semester_text(current_settings.semester, current_settings.academic_year) if current_settings else "CLASS SCHEDULE"
    ws.merge_cells('A7:G7')
    ws['A7'] = semester_text
    ws['A7'].font = Font(bold=True, size=11)
    ws['A7'].alignment = Alignment(horizontal='center', vertical='center')
    
    section_display = section.full_section_name if hasattr(section, 'full_section_name') else f"{dept_code}-{section.year_level}{section.section_name}"
    ws.merge_cells('A8:G8')
    ws['A8'] = section_display
    ws['A8'].font = Font(bold=True, size=11)
    ws['A8'].alignment = Alignment(horizontal='center', vertical='center')
    
    # Row 9: Spacer
    ws.row_dimensions[9].height = 8
    
    # =========================================================================
    # GRID SCHEDULE - Time slots and day columns
    # =========================================================================
    # Add column headers
    add_column_headers(ws)
    
    # Generate and write time slots with dynamic time range
    time_slots = generate_time_slots(start_hour=start_hour, end_hour=end_hour)
    write_time_slots(ws, time_slots)
    
    # Place schedules in grid with dynamic start hour
    place_schedule_in_grid(ws, schedules, start_hour=start_hour)
    
    # Apply borders
    last_row = 11 + len(time_slots) - 1
    apply_grid_borders(ws, last_row)
    
    # =========================================================================
    # SIGNATURE SECTION - Same as posting version (merged cells B-D and E-G)
    # =========================================================================
    sig_start_row = last_row + 3
    
    # Prepared by label (column B)
    ws.cell(row=sig_start_row, column=2, value='Prepared by:')
    ws.cell(row=sig_start_row, column=2).font = Font(size=11)
    ws.cell(row=sig_start_row, column=2).alignment = Alignment(horizontal='left', vertical='center')
    
    # Noted by label (column E)
    ws.cell(row=sig_start_row, column=5, value='Noted by:')
    ws.cell(row=sig_start_row, column=5).font = Font(size=11)
    ws.cell(row=sig_start_row, column=5).alignment = Alignment(horizontal='left', vertical='center')
    
    # Blank row for signature space
    sig_start_row += 2
    
    # Dean name - BOLD and UPPERCASE (column B-D centered)
    dean_display_name = user.full_name.upper() if user and user.full_name else '{Dean Name}'
    ws.merge_cells(f'B{sig_start_row}:D{sig_start_row}')
    ws.cell(row=sig_start_row, column=2, value=dean_display_name)
    ws.cell(row=sig_start_row, column=2).font = Font(bold=True, size=11)
    ws.cell(row=sig_start_row, column=2).alignment = Alignment(horizontal='center', vertical='center')
    
    # Department President name - BOLD (column E-G centered)
    ws.merge_cells(f'E{sig_start_row}:G{sig_start_row}')
    ws.cell(row=sig_start_row, column=5, value=get_institution_head_name())
    ws.cell(row=sig_start_row, column=5).font = Font(bold=True, size=11)
    ws.cell(row=sig_start_row, column=5).alignment = Alignment(horizontal='center', vertical='center')
    
    # Dean title (column B-D centered)
    sig_start_row += 1
    dean_title = f"Dean, {dept_display_name}" if dept_display_name else "Dean, {Program Name}"
    ws.merge_cells(f'B{sig_start_row}:D{sig_start_row}')
    ws.cell(row=sig_start_row, column=2, value=dean_title)
    ws.cell(row=sig_start_row, column=2).font = Font(size=11)
    ws.cell(row=sig_start_row, column=2).alignment = Alignment(horizontal='center', vertical='center')
    
    # Department President title (column E-G centered)
    ws.merge_cells(f'E{sig_start_row}:G{sig_start_row}')
    ws.cell(row=sig_start_row, column=5, value='Department President')
    ws.cell(row=sig_start_row, column=5).font = Font(size=11)
    ws.cell(row=sig_start_row, column=5).alignment = Alignment(horizontal='center', vertical='center')
    
    # Set column widths
    set_column_widths(ws)
    
    # Page setup
    ws.page_setup.fitToPage = True
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    
    # Save and return
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    filename = f"{dept_code}_{section.year_level}{section.section_name}_Schedule.xlsx".replace(' ', '_')
    return output, filename

def generate_faculty_schedule_excel(faculty, schedules, current_settings, user):
    """Generate Excel file for faculty schedule - grid format with posting-style header/footer"""
    # Get time range from settings (default to 7-20 if not set)
    start_hour = (current_settings.schedule_start_time.hour if current_settings and current_settings.schedule_start_time else 7)
    end_hour = (current_settings.schedule_end_time.hour if current_settings and current_settings.schedule_end_time else 20)
    
    # Create workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Schedule"
    
    # Get department info from faculty
    department = faculty.department
    dept_display_name = (department.department_name if department else '{Program Name}').upper()
    
    # =========================================================================
    # LOGOS - Centered within their respective columns (A=left, G=right)
    # =========================================================================
    _HDR_HEIGHT_PT = 63
    add_logo_centered(ws, get_institution_logo_path(),
                      col_0idx=0, img_width_px=75, img_height_px=75,
                      col_width_chars=12.71, header_height_pts=_HDR_HEIGHT_PT)
    add_logo_centered(ws, get_institution_logo_right_path(),
                      col_0idx=6, img_width_px=80, img_height_px=80,
                      col_width_chars=20.71, header_height_pts=_HDR_HEIGHT_PT)

    # =========================================================================
    # HEADER TEXT - Left aligned at column B (rows 1-4)
    # =========================================================================
    _hdr_line1, _hdr_line2, _ = get_excel_header_config()
    ws['B1'] = _hdr_line1
    ws['B1'].font = Font(size=11)
    ws['B1'].alignment = Alignment(horizontal='left', vertical='center')
    
    ws['B2'] = _hdr_line2
    ws['B2'].font = Font(size=11)
    ws['B2'].alignment = Alignment(horizontal='left', vertical='center')
    
    ws['B3'] = get_institution_name()
    ws['B3'].font = Font(bold=True, size=11)
    ws['B3'].alignment = Alignment(horizontal='left', vertical='center')
    
    ws['B4'] = dept_display_name
    ws['B4'].font = Font(bold=True, size=11)
    ws['B4'].alignment = Alignment(horizontal='left', vertical='center')
    
    # Set row heights for header
    ws.row_dimensions[1].height = 18
    ws.row_dimensions[2].height = 15
    ws.row_dimensions[3].height = 15
    ws.row_dimensions[4].height = 15
    ws.row_dimensions[5].height = 12  # Spacer row
    
    # =========================================================================
    # SCHEDULE TITLE - Centered across A-G (rows 6-8)
    # =========================================================================
    ws.merge_cells('A6:G6')
    ws['A6'] = 'FACULTY SCHEDULE'
    ws['A6'].font = Font(bold=True, size=12)
    ws['A6'].alignment = Alignment(horizontal='center', vertical='center')
    
    semester_text = format_semester_text(current_settings.semester, current_settings.academic_year) if current_settings else "FACULTY SCHEDULE"
    ws.merge_cells('A7:G7')
    ws['A7'] = semester_text
    ws['A7'].font = Font(bold=True, size=11)
    ws['A7'].alignment = Alignment(horizontal='center', vertical='center')
    
    ws.merge_cells('A8:G8')
    ws['A8'] = faculty.full_name
    ws['A8'].font = Font(bold=True, size=11)
    ws['A8'].alignment = Alignment(horizontal='center', vertical='center')
    
    # Row 9: Spacer
    ws.row_dimensions[9].height = 8
    
    # =========================================================================
    # GRID SCHEDULE - Time slots and day columns
    # =========================================================================
    # Add column headers
    add_column_headers(ws)
    
    # Generate and write time slots with dynamic time range
    time_slots = generate_time_slots(start_hour=start_hour, end_hour=end_hour)
    write_time_slots(ws, time_slots)
    
    # Place schedules in grid with dynamic start hour (for faculty, use section as cell content)
    place_faculty_schedule_in_grid(ws, schedules, start_hour=start_hour)
    
    # Apply borders
    last_row = 11 + len(time_slots) - 1
    apply_grid_borders(ws, last_row)
    
    # =========================================================================
    # SIGNATURE SECTION - 3-column: Conforme / Prepared by / Noted
    # =========================================================================
    sig_start_row = last_row + 3
    
    # Conforme label (column A)
    ws.cell(row=sig_start_row, column=1, value='Conforme:')
    ws.cell(row=sig_start_row, column=1).font = Font(size=11)
    ws.cell(row=sig_start_row, column=1).alignment = Alignment(horizontal='left', vertical='center')
    
    # Prepared by label (column C)
    ws.cell(row=sig_start_row, column=3, value='Prepared by:')
    ws.cell(row=sig_start_row, column=3).font = Font(size=11)
    ws.cell(row=sig_start_row, column=3).alignment = Alignment(horizontal='left', vertical='center')
    
    # Noted label (column F)
    ws.cell(row=sig_start_row, column=6, value='Noted:')
    ws.cell(row=sig_start_row, column=6).font = Font(size=11)
    ws.cell(row=sig_start_row, column=6).alignment = Alignment(horizontal='left', vertical='center')
    
    # Blank row for signature space
    sig_start_row += 2
    
    # Faculty name - BOLD and UPPERCASE (column A-B centered)
    faculty_display_name = faculty.full_name.upper() if faculty and faculty.full_name else '{Faculty Name}'
    ws.merge_cells(f'A{sig_start_row}:B{sig_start_row}')
    ws.cell(row=sig_start_row, column=1, value=faculty_display_name)
    ws.cell(row=sig_start_row, column=1).font = Font(bold=True, size=11)
    ws.cell(row=sig_start_row, column=1).alignment = Alignment(horizontal='center', vertical='center')
    
    # Dean name - BOLD and UPPERCASE (column C-E centered)
    dean_display_name = user.full_name.upper() if user and user.full_name else '{Dean Name}'
    ws.merge_cells(f'C{sig_start_row}:E{sig_start_row}')
    ws.cell(row=sig_start_row, column=3, value=dean_display_name)
    ws.cell(row=sig_start_row, column=3).font = Font(bold=True, size=11)
    ws.cell(row=sig_start_row, column=3).alignment = Alignment(horizontal='center', vertical='center')
    
    # Department President name - BOLD (column F-G centered)
    ws.merge_cells(f'F{sig_start_row}:G{sig_start_row}')
    ws.cell(row=sig_start_row, column=6, value=get_institution_head_name())
    ws.cell(row=sig_start_row, column=6).font = Font(bold=True, size=11)
    ws.cell(row=sig_start_row, column=6).alignment = Alignment(horizontal='center', vertical='center')
    
    # Titles row
    sig_start_row += 1
    
    # Faculty title (column A-B centered)
    ws.merge_cells(f'A{sig_start_row}:B{sig_start_row}')
    ws.cell(row=sig_start_row, column=1, value='GE Instructor')
    ws.cell(row=sig_start_row, column=1).font = Font(size=11)
    ws.cell(row=sig_start_row, column=1).alignment = Alignment(horizontal='center', vertical='center')
    
    # Dean title (column C-E centered)
    dean_title = f"Dean, {dept_display_name}" if dept_display_name else "Dean, {Program Name}"
    ws.merge_cells(f'C{sig_start_row}:E{sig_start_row}')
    ws.cell(row=sig_start_row, column=3, value=dean_title)
    ws.cell(row=sig_start_row, column=3).font = Font(size=11)
    ws.cell(row=sig_start_row, column=3).alignment = Alignment(horizontal='center', vertical='center')
    
    # Department President title (column F-G centered)
    ws.merge_cells(f'F{sig_start_row}:G{sig_start_row}')
    ws.cell(row=sig_start_row, column=6, value='Department President')
    ws.cell(row=sig_start_row, column=6).font = Font(size=11)
    ws.cell(row=sig_start_row, column=6).alignment = Alignment(horizontal='center', vertical='center')
    
    # Set column widths
    set_column_widths(ws)
    
    # Page setup
    ws.page_setup.fitToPage = True
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    
    # Save and return
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    filename = f"{faculty.last_name}_{faculty.first_name}_Schedule.xlsx"
    return output, filename

def generate_room_schedule_excel(room, schedules, current_settings, user):
    """Generate Excel file for room schedule - grid format with posting-style header/footer"""
    # Get time range from settings (default to 7-20 if not set)
    start_hour = (current_settings.schedule_start_time.hour if current_settings and current_settings.schedule_start_time else 7)
    end_hour = (current_settings.schedule_end_time.hour if current_settings and current_settings.schedule_end_time else 20)
    
    # Create workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Schedule"
    
    # Get program info from the first schedule's section
    program = None
    if schedules and schedules[0].section:
        program = schedules[0].section.program
    
    dept_display_name = _get_department_name(program).upper()
    
    # =========================================================================
    # LOGOS - Centered within their respective columns (A=left, G=right)
    # =========================================================================
    _HDR_HEIGHT_PT = 63
    add_logo_centered(ws, get_institution_logo_path(),
                      col_0idx=0, img_width_px=75, img_height_px=75,
                      col_width_chars=12.71, header_height_pts=_HDR_HEIGHT_PT)
    add_logo_centered(ws, get_institution_logo_right_path(),
                      col_0idx=6, img_width_px=80, img_height_px=80,
                      col_width_chars=20.71, header_height_pts=_HDR_HEIGHT_PT)

    # =========================================================================
    # HEADER TEXT - Left aligned at column B (rows 1-4)
    # =========================================================================
    _hdr_line1, _hdr_line2, _ = get_excel_header_config()
    ws['B1'] = _hdr_line1
    ws['B1'].font = Font(size=11)
    ws['B1'].alignment = Alignment(horizontal='left', vertical='center')
    
    ws['B2'] = _hdr_line2
    ws['B2'].font = Font(size=11)
    ws['B2'].alignment = Alignment(horizontal='left', vertical='center')
    
    ws['B3'] = get_institution_name()
    ws['B3'].font = Font(bold=True, size=11)
    ws['B3'].alignment = Alignment(horizontal='left', vertical='center')
    
    ws['B4'] = dept_display_name
    ws['B4'].font = Font(bold=True, size=11)
    ws['B4'].alignment = Alignment(horizontal='left', vertical='center')
    
    # Set row heights for header
    ws.row_dimensions[1].height = 18
    ws.row_dimensions[2].height = 15
    ws.row_dimensions[3].height = 15
    ws.row_dimensions[4].height = 15
    ws.row_dimensions[5].height = 12  # Spacer row
    
    # =========================================================================
    # SCHEDULE TITLE - Centered across A-G (rows 6-8)
    # =========================================================================
    ws.merge_cells('A6:G6')
    ws['A6'] = 'ROOM SCHEDULE'
    ws['A6'].font = Font(bold=True, size=12)
    ws['A6'].alignment = Alignment(horizontal='center', vertical='center')
    
    semester_text = format_semester_text(current_settings.semester, current_settings.academic_year) if current_settings else "ROOM SCHEDULE"
    ws.merge_cells('A7:G7')
    ws['A7'] = semester_text
    ws['A7'].font = Font(bold=True, size=11)
    ws['A7'].alignment = Alignment(horizontal='center', vertical='center')
    
    ws.merge_cells('A8:G8')
    ws['A8'] = room.room_number
    ws['A8'].font = Font(bold=True, size=11)
    ws['A8'].alignment = Alignment(horizontal='center', vertical='center')
    
    # Row 9: Spacer
    ws.row_dimensions[9].height = 8
    
    # =========================================================================
    # GRID SCHEDULE - Time slots and day columns
    # =========================================================================
    # Add column headers
    add_column_headers(ws)
    
    # Generate and write time slots with dynamic time range
    time_slots = generate_time_slots(start_hour=start_hour, end_hour=end_hour)
    write_time_slots(ws, time_slots)
    
    # Place schedules in grid with dynamic start hour (for room, use section/faculty)
    place_room_schedule_in_grid(ws, schedules, start_hour=start_hour)
    
    # Apply borders
    last_row = 11 + len(time_slots) - 1
    apply_grid_borders(ws, last_row)
    
    # =========================================================================
    # SIGNATURE SECTION - Same as posting version (merged cells B-D and E-G)
    # =========================================================================
    sig_start_row = last_row + 3
    
    # Prepared by label (column B)
    ws.cell(row=sig_start_row, column=2, value='Prepared by:')
    ws.cell(row=sig_start_row, column=2).font = Font(size=11)
    ws.cell(row=sig_start_row, column=2).alignment = Alignment(horizontal='left', vertical='center')
    
    # Noted by label (column E)
    ws.cell(row=sig_start_row, column=5, value='Noted by:')
    ws.cell(row=sig_start_row, column=5).font = Font(size=11)
    ws.cell(row=sig_start_row, column=5).alignment = Alignment(horizontal='left', vertical='center')
    
    # Blank row for signature space
    sig_start_row += 2
    
    # Dean name - BOLD and UPPERCASE (column B-D centered)
    dean_display_name = user.full_name.upper() if user and user.full_name else '{Dean Name}'
    ws.merge_cells(f'B{sig_start_row}:D{sig_start_row}')
    ws.cell(row=sig_start_row, column=2, value=dean_display_name)
    ws.cell(row=sig_start_row, column=2).font = Font(bold=True, size=11)
    ws.cell(row=sig_start_row, column=2).alignment = Alignment(horizontal='center', vertical='center')
    
    # Department President name - BOLD (column E-G centered)
    ws.merge_cells(f'E{sig_start_row}:G{sig_start_row}')
    ws.cell(row=sig_start_row, column=5, value=get_institution_head_name())
    ws.cell(row=sig_start_row, column=5).font = Font(bold=True, size=11)
    ws.cell(row=sig_start_row, column=5).alignment = Alignment(horizontal='center', vertical='center')
    
    # Dean title (column B-D centered)
    sig_start_row += 1
    dean_title = f"Dean, {dept_display_name}" if dept_display_name else "Dean, {Program Name}"
    ws.merge_cells(f'B{sig_start_row}:D{sig_start_row}')
    ws.cell(row=sig_start_row, column=2, value=dean_title)
    ws.cell(row=sig_start_row, column=2).font = Font(size=11)
    ws.cell(row=sig_start_row, column=2).alignment = Alignment(horizontal='center', vertical='center')
    
    # Department President title (column E-G centered)
    ws.merge_cells(f'E{sig_start_row}:G{sig_start_row}')
    ws.cell(row=sig_start_row, column=5, value='Department President')
    ws.cell(row=sig_start_row, column=5).font = Font(size=11)
    ws.cell(row=sig_start_row, column=5).alignment = Alignment(horizontal='center', vertical='center')
    
    # Set column widths
    set_column_widths(ws)
    
    # Page setup
    ws.page_setup.fitToPage = True
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    
    # Save and return
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    filename = f"Room_{room.room_number}_Schedule.xlsx".replace(' ', '_')
    return output, filename


def generate_faculty_lineup_excel(department_or_program, faculty_schedule_data, current_settings, user):
    """
    Generate Excel file for Faculty Line-Up report (per department).
    Lists all faculty with their subjects, sections, units, and total hours.
    
    Args:
        department_or_program: Department or Program model instance
        faculty_schedule_data: List of dicts with faculty info and their schedule rows
            [{ 'faculty': Faculty, 'rows': [{'subject_code', 'section_name', 'units', 'hours'}], 'total_hours': float }]
        current_settings: AcademicSettings instance
        user: Current user (dean)
    
    Returns:
        (output, filename) tuple
    """
    from app.models.department import Department
    from app.models.program import Program
    wb = Workbook()
    ws = wb.active
    ws.title = "Faculty Line-Up"
    
    # Handle both Department and Program objects
    if isinstance(department_or_program, Department):
        dept_display_name = department_or_program.department_name.upper() if department_or_program else '{Program Name}'
    else:
        dept_display_name = _get_department_name(department_or_program, department_or_program.program_name if department_or_program else '{Program Name}').upper()
    
    # Format semester subtitle
    semester_text = ''
    if current_settings:
        semester_text = format_semester_text(current_settings.semester, current_settings.academic_year)

    # =========================================================================
    # HEADER — Posting-style with dual logos (columns A-H)
    # =========================================================================
    current_row = create_posting_style_excel_header(
        ws,
        report_title='FACULTY LINE-UP',
        office_name=dept_display_name,
        subtitle=semester_text if semester_text else None,
        last_col='H',
        left_col_width=5,
        right_col_width=10
    )
    
    # =========================================================================
    # COLUMN HEADERS — Rows current_row and current_row+1 (merged vertically)
    # =========================================================================
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    header_font = Font(name='Times New Roman', bold=True, size=10)
    header_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
    header_fill = PatternFill(start_color='FFFFFF', end_color='FFFFFF', fill_type='solid')
    
    # Define columns: A=NO, B=NAME, C=LENGTH OF SERVICE, D=SUBJECTS, E=SECTIONS, F=NO. OF UNITS, G=NO. OF HOURS, H=TOTAL HOURS
    headers = [
        ('A', 'NO.', 5),
        ('B', 'NAME', 30),
        ('C', 'LENGTH OF\nSERVICE\n\n(No. of\nYrs./Semesters)', 20),
        ('D', 'SUBJECTS', 15),
        ('E', 'SECTIONS', 18),
        ('F', 'NO. OF\nUNITS', 10),
        ('G', 'NO. OF\nHOURS', 10),
        ('H', 'TOTAL\nHOURS', 10),
    ]
    
    h_row1 = current_row
    h_row2 = current_row + 1
    
    for col_letter, title, width in headers:
        col_idx = ord(col_letter) - ord('A') + 1
        # Merge 2 rows for header
        ws.merge_cells(start_row=h_row1, start_column=col_idx, end_row=h_row2, end_column=col_idx)
        cell = ws.cell(row=h_row1, column=col_idx, value=title)
        cell.font = header_font
        cell.alignment = header_align
        cell.border = thin_border
        cell.fill = header_fill
        # Bottom cell of merge also needs border
        ws.cell(row=h_row2, column=col_idx).border = thin_border
        # Column width
        ws.column_dimensions[col_letter].width = width
    
    ws.row_dimensions[h_row1].height = 35
    ws.row_dimensions[h_row2].height = 35
    
    current_row = h_row2 + 1
    
    # =========================================================================
    # DATA ROWS — One row per section, subject code merged when same subject
    # =========================================================================
    data_font = Font(name='Times New Roman', size=10)
    data_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
    left_align = Alignment(horizontal='left', vertical='center', wrap_text=True)
    
    faculty_num = 0
    
    for entry in faculty_schedule_data:
        faculty = entry['faculty']
        rows = entry['rows']
        total_hours = entry['total_hours']
        
        if not rows:
            # Faculty with no schedules — single row
            rows = [{'subject_code': '', 'section_name': '', 'units': '', 'hours': '', '_subject_group_size': 0}]
        
        faculty_num += 1
        start_data_row = current_row
        num_rows = len(rows)
        
        for i, row_data in enumerate(rows):
            r = current_row + i
            group_size = row_data.get('_subject_group_size', 0)
            
            # D: Subject code — merge if this subject spans multiple sections
            if group_size > 1:
                ws.merge_cells(start_row=r, start_column=4, end_row=r + group_size - 1, end_column=4)
                cell_d = ws.cell(row=r, column=4, value=row_data.get('subject_code', ''))
                cell_d.font = data_font
                cell_d.alignment = data_align
                cell_d.border = thin_border
                for mr in range(r, r + group_size):
                    ws.cell(row=mr, column=4).border = thin_border
            elif group_size == 1 or (group_size == 0 and row_data.get('subject_code', '')):
                # Single section subject or standalone
                if row_data.get('subject_code', ''):
                    cell_d = ws.cell(row=r, column=4, value=row_data.get('subject_code', ''))
                    cell_d.font = data_font
                    cell_d.alignment = data_align
                cell_d = ws.cell(row=r, column=4)
                cell_d.border = thin_border
            else:
                # Continuation row of a merged subject — just set border
                ws.cell(row=r, column=4).border = thin_border
            
            # E: Section name — always one per row
            cell_e = ws.cell(row=r, column=5, value=row_data.get('section_name', ''))
            cell_e.font = data_font
            cell_e.alignment = data_align
            cell_e.border = thin_border
            
            # F: No. of Units — merge same as subject code
            if group_size > 1:
                ws.merge_cells(start_row=r, start_column=6, end_row=r + group_size - 1, end_column=6)
                cell_f = ws.cell(row=r, column=6, value=row_data.get('units', ''))
                cell_f.font = data_font
                cell_f.alignment = data_align
                cell_f.border = thin_border
                for mr in range(r, r + group_size):
                    ws.cell(row=mr, column=6).border = thin_border
            elif group_size == 1 or (group_size == 0 and row_data.get('units', '') != ''):
                if row_data.get('units', '') != '':
                    cell_f = ws.cell(row=r, column=6, value=row_data.get('units', ''))
                    cell_f.font = data_font
                    cell_f.alignment = data_align
                cell_f = ws.cell(row=r, column=6)
                cell_f.border = thin_border
            else:
                ws.cell(row=r, column=6).border = thin_border
            
            # G: No. of Hours (per section row)
            hours_val = row_data.get('hours', '')
            cell_g = ws.cell(row=r, column=7, value=hours_val if hours_val != '' else '')
            cell_g.font = data_font
            cell_g.alignment = data_align
            cell_g.border = thin_border
        
        end_data_row = start_data_row + num_rows - 1
        
        # A: NO. — merged across faculty rows
        if num_rows > 1:
            ws.merge_cells(start_row=start_data_row, start_column=1, end_row=end_data_row, end_column=1)
        cell_a = ws.cell(row=start_data_row, column=1, value=faculty_num)
        cell_a.font = data_font
        cell_a.alignment = data_align
        cell_a.border = thin_border
        for mr in range(start_data_row, end_data_row + 1):
            ws.cell(row=mr, column=1).border = thin_border
        
        # B: NAME — merged across faculty rows
        if num_rows > 1:
            ws.merge_cells(start_row=start_data_row, start_column=2, end_row=end_data_row, end_column=2)
        # Format name with salutation
        faculty_name = faculty.full_name if faculty.full_name else ''
        cell_b = ws.cell(row=start_data_row, column=2, value=faculty_name)
        cell_b.font = data_font
        cell_b.alignment = left_align
        cell_b.border = thin_border
        for mr in range(start_data_row, end_data_row + 1):
            ws.cell(row=mr, column=2).border = thin_border
        
        # C: LENGTH OF SERVICE — merged, placeholder
        if num_rows > 1:
            ws.merge_cells(start_row=start_data_row, start_column=3, end_row=end_data_row, end_column=3)
        cell_c = ws.cell(row=start_data_row, column=3, value='')
        cell_c.font = data_font
        cell_c.alignment = data_align
        cell_c.border = thin_border
        for mr in range(start_data_row, end_data_row + 1):
            ws.cell(row=mr, column=3).border = thin_border
        
        # H: TOTAL HOURS — merged across faculty rows
        if num_rows > 1:
            ws.merge_cells(start_row=start_data_row, start_column=8, end_row=end_data_row, end_column=8)
        cell_h = ws.cell(row=start_data_row, column=8, value=total_hours if total_hours else '')
        cell_h.font = Font(name='Times New Roman', bold=True, size=10)
        cell_h.alignment = data_align
        cell_h.border = thin_border
        for mr in range(start_data_row, end_data_row + 1):
            ws.cell(row=mr, column=8).border = thin_border
        
        current_row = end_data_row + 1
    
    # =========================================================================
    # SIGNATURE SECTION (stacked single-column layout)
    # =========================================================================
    sig_row = current_row + 2
    sig_font = Font(name='Times New Roman', size=11)
    sig_bold = Font(name='Times New Roman', bold=True, size=11)
    sig_align = Alignment(horizontal='left', vertical='center')
    
    # "Prepared by:" label
    ws.cell(row=sig_row, column=1, value='Prepared by:')
    ws.cell(row=sig_row, column=1).font = sig_font
    ws.cell(row=sig_row, column=1).alignment = sig_align
    
    # Blank row for signature space
    sig_row += 2
    
    # Secretary name — BOLD and UPPERCASE
    if isinstance(department_or_program, Department):
        secretary_name = (department_or_program.secretary_name or '').upper()
    else:
        secretary_name = _get_secretary_name(department_or_program).upper()
    ws.cell(row=sig_row, column=1, value=secretary_name)
    ws.cell(row=sig_row, column=1).font = sig_bold
    ws.cell(row=sig_row, column=1).alignment = sig_align
    
    # Secretary title
    sig_row += 1
    ws.cell(row=sig_row, column=1, value="Dean's Secretary")
    ws.cell(row=sig_row, column=1).font = sig_font
    ws.cell(row=sig_row, column=1).alignment = sig_align
    
    # Blank row
    sig_row += 2
    
    # "Checked by:" label
    ws.cell(row=sig_row, column=1, value='Checked by:')
    ws.cell(row=sig_row, column=1).font = sig_font
    ws.cell(row=sig_row, column=1).alignment = sig_align
    
    # Blank row for signature space
    sig_row += 2
    
    # Dean name — BOLD and UPPERCASE
    dean_display_name = user.full_name.upper() if user and user.full_name else '{Dean Name}'
    ws.cell(row=sig_row, column=1, value=dean_display_name)
    ws.cell(row=sig_row, column=1).font = sig_bold
    ws.cell(row=sig_row, column=1).alignment = sig_align
    
    # Dean title
    sig_row += 1
    dean_title = f"Dean, {dept_display_name}"
    ws.cell(row=sig_row, column=1, value=dean_title)
    ws.cell(row=sig_row, column=1).font = sig_font
    ws.cell(row=sig_row, column=1).alignment = sig_align
    
    # =========================================================================
    # PAGE SETUP
    # =========================================================================
    ws.page_setup.fitToPage = True
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.page_setup.orientation = 'landscape'
    
    # Save and return
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    # Determine code for filename
    if isinstance(department_or_program, Department):
        dept_code = department_or_program.department_code or 'DEPT'
    elif isinstance(department_or_program, Program):
        dept_code = department_or_program.program_code or 'DEPT'
    else:
        dept_code = 'DEPT'
    ay_slug = current_settings.academic_year.replace('-', '_') if current_settings else ''
    sem_slug = current_settings.semester.replace(' ', '_') if current_settings else ''
    filename = f"{dept_code}_Faculty_Lineup_{sem_slug}_{ay_slug}.xlsx"
    return output, filename


def generate_daily_faculty_schedule_excel(program, day_of_week, faculty_schedule_data, current_settings, user):
    """
    Generate Excel file for a program-wide daily faculty schedule.
    Lists all faculty with their time slots, subjects, sections, and rooms for a specific day.

    Args:
        program: Program model instance
        day_of_week: Day string (e.g. "Monday")
        faculty_schedule_data: List of dicts:
            [{ 'faculty': Faculty, 'rows': [{'time', 'subject', 'section', 'room'}] }]
        current_settings: AcademicSettings instance
        user: Current user (dean)

    Returns:
        (output, filename) tuple
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Faculty Schedule"

    dept_display_name = _get_department_name(program, program.program_name if program else '{Program Name}').upper()

    # Format semester subtitle
    semester_text = ''
    if current_settings:
        semester_text = format_semester_text(current_settings.semester, current_settings.academic_year)

    # =========================================================================
    # HEADER — Posting-style with dual logos (columns A-F)
    # =========================================================================
    current_row = create_posting_style_excel_header(
        ws,
        report_title='FACULTY SCHEDULE',
        office_name=dept_display_name,
        subtitle=semester_text if semester_text else None,
        last_col='F',
        left_col_width=5,
        right_col_width=12
    )

    # =========================================================================
    # DAY LABEL — Centered bold row (e.g. "MONDAY")
    # =========================================================================
    ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=6)
    day_cell = ws.cell(row=current_row, column=1, value=day_of_week.upper())
    day_cell.font = Font(name='Times New Roman', bold=True, size=12)
    day_cell.alignment = Alignment(horizontal='center', vertical='center')
    current_row += 2  # spacer

    # =========================================================================
    # COLUMN HEADERS — 2-row merged
    # =========================================================================
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    header_font = Font(name='Times New Roman', bold=True, size=10)
    header_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
    header_fill = PatternFill(start_color='FFFFFF', end_color='FFFFFF', fill_type='solid')

    headers = [
        ('A', 'NO.', 5),
        ('B', 'NAME OF INSTRUCTOR', 32),
        ('C', 'TIME', 22),
        ('D', 'SUBJECTS\nHANDLED', 18),
        ('E', 'YEAR/SECTION', 18),
        ('F', 'ROOM', 12),
    ]

    h_row1 = current_row
    h_row2 = current_row + 1

    for col_letter, title, width in headers:
        col_idx = ord(col_letter) - ord('A') + 1
        ws.merge_cells(start_row=h_row1, start_column=col_idx, end_row=h_row2, end_column=col_idx)
        cell = ws.cell(row=h_row1, column=col_idx, value=title)
        cell.font = header_font
        cell.alignment = header_align
        cell.border = thin_border
        cell.fill = header_fill
        ws.cell(row=h_row2, column=col_idx).border = thin_border
        ws.column_dimensions[col_letter].width = width

    ws.row_dimensions[h_row1].height = 28
    ws.row_dimensions[h_row2].height = 28

    current_row = h_row2 + 1

    # =========================================================================
    # DATA ROWS
    # =========================================================================
    data_font = Font(name='Times New Roman', size=10)
    data_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
    left_align = Alignment(horizontal='left', vertical='center', wrap_text=True)

    faculty_num = 0

    for entry in faculty_schedule_data:
        faculty = entry['faculty']
        rows = entry['rows']

        if not rows:
            continue

        faculty_num += 1
        num_rows = len(rows)
        start_data_row = current_row

        for i, row_data in enumerate(rows):
            r = current_row + i

            # C: TIME
            cell_c = ws.cell(row=r, column=3, value=row_data.get('time', ''))
            cell_c.font = data_font
            cell_c.alignment = data_align
            cell_c.border = thin_border

            # D: SUBJECTS HANDLED
            cell_d = ws.cell(row=r, column=4, value=row_data.get('subject', ''))
            cell_d.font = data_font
            cell_d.alignment = data_align
            cell_d.border = thin_border

            # E: YEAR/SECTION
            cell_e = ws.cell(row=r, column=5, value=row_data.get('section', ''))
            cell_e.font = data_font
            cell_e.alignment = data_align
            cell_e.border = thin_border

            # F: ROOM
            cell_f = ws.cell(row=r, column=6, value=row_data.get('room', ''))
            cell_f.font = data_font
            cell_f.alignment = data_align
            cell_f.border = thin_border

        end_data_row = start_data_row + num_rows - 1

        # A: NO. — merged across faculty rows
        if num_rows > 1:
            ws.merge_cells(start_row=start_data_row, start_column=1, end_row=end_data_row, end_column=1)
        cell_a = ws.cell(row=start_data_row, column=1, value=faculty_num)
        cell_a.font = data_font
        cell_a.alignment = data_align
        cell_a.border = thin_border
        for mr in range(start_data_row, end_data_row + 1):
            ws.cell(row=mr, column=1).border = thin_border

        # B: NAME OF INSTRUCTOR — merged across faculty rows, UPPERCASE
        if num_rows > 1:
            ws.merge_cells(start_row=start_data_row, start_column=2, end_row=end_data_row, end_column=2)
        faculty_name = faculty.full_name.upper() if faculty.full_name else ''
        cell_b = ws.cell(row=start_data_row, column=2, value=faculty_name)
        cell_b.font = Font(name='Times New Roman', bold=True, size=10)
        cell_b.alignment = left_align
        cell_b.border = thin_border
        for mr in range(start_data_row, end_data_row + 1):
            ws.cell(row=mr, column=2).border = thin_border

        current_row = end_data_row + 1

    # =========================================================================
    # SIGNATURE SECTION (stacked single-column layout)
    # =========================================================================
    sig_row = current_row + 2
    sig_font = Font(name='Times New Roman', size=11)
    sig_bold = Font(name='Times New Roman', bold=True, size=11)
    sig_align = Alignment(horizontal='left', vertical='center')

    # "Prepared by:" label
    ws.cell(row=sig_row, column=1, value='Prepared by:')
    ws.cell(row=sig_row, column=1).font = sig_font
    ws.cell(row=sig_row, column=1).alignment = sig_align

    sig_row += 2

    # Secretary name — BOLD and UPPERCASE
    secretary_name = _get_secretary_name(program).upper()
    ws.cell(row=sig_row, column=1, value=secretary_name)
    ws.cell(row=sig_row, column=1).font = sig_bold
    ws.cell(row=sig_row, column=1).alignment = sig_align

    sig_row += 1
    ws.cell(row=sig_row, column=1, value="Dean's Secretary")
    ws.cell(row=sig_row, column=1).font = sig_font
    ws.cell(row=sig_row, column=1).alignment = sig_align

    # =========================================================================
    # PAGE SETUP
    # =========================================================================
    ws.page_setup.fitToPage = True
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.page_setup.orientation = 'landscape'

    # Save and return
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    dept_code = program.program_code or 'DEPT'
    ay_slug = current_settings.academic_year.replace('-', '_') if current_settings else ''
    sem_slug = current_settings.semester.replace(' ', '_') if current_settings else ''
    filename = f"{dept_code}_Faculty_Schedule_{day_of_week}_{sem_slug}_{ay_slug}.xlsx"
    return output, filename


def generate_class_schedule_excel_for_posting(section, schedules, current_settings, user):
    """Generate Excel file for class schedule (posting version) - matches official format exactly"""
    # Create workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Schedule"
    
    # Define styles
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    header_font = Font(bold=True, size=11)
    header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    
    # Get program info for logo and names
    program = section.program
    dept_display_name = _get_department_name(program).upper()
    dept_code = program.program_code if program else ''
    
    # =========================================================================
    # COLUMN WIDTHS - Data starts at column A
    # =========================================================================
    ws.column_dimensions['A'].width = 14   # Subject Code
    ws.column_dimensions['B'].width = 32   # Description
    ws.column_dimensions['C'].width = 6    # Lec
    ws.column_dimensions['D'].width = 6    # Lab
    ws.column_dimensions['E'].width = 12   # Day
    ws.column_dimensions['F'].width = 18   # Time
    ws.column_dimensions['G'].width = 8    # Room
    ws.column_dimensions['H'].width = 18   # Professor
    
    # =========================================================================
    # LOGOS AND HEADER - Rows 1-4 (starting at column A)
    # =========================================================================
    # Column widths: A=14, H=18 (posting format) — header rows total 63 pt
    _HDR_HEIGHT_PT = 63
    add_logo_centered(ws, get_institution_logo_path(),
                      col_0idx=0, img_width_px=75, img_height_px=75,
                      col_width_chars=14, header_height_pts=_HDR_HEIGHT_PT)
    add_logo_centered(ws, get_institution_logo_right_path(),
                      col_0idx=7, img_width_px=80, img_height_px=80,
                      col_width_chars=18, header_height_pts=_HDR_HEIGHT_PT)
    
    # Header text - LEFT ALIGNED next to logo (column B)
    _hdr_line1, _hdr_line2, _ = get_excel_header_config()
    # Row 1
    ws['B1'] = _hdr_line1
    ws['B1'].font = Font(size=11)
    ws['B1'].alignment = Alignment(horizontal='left', vertical='center')
    
    # Row 2
    ws['B2'] = _hdr_line2
    ws['B2'].font = Font(size=11)
    ws['B2'].alignment = Alignment(horizontal='left', vertical='center')
    
    # Row 3: NORZAGARAY COLLEGE (bold)
    ws['B3'] = get_institution_name()
    ws['B3'].font = Font(bold=True, size=11)
    ws['B3'].alignment = Alignment(horizontal='left', vertical='center')
    
    # Row 4: Program name (bold)
    ws['B4'] = dept_display_name
    ws['B4'].font = Font(bold=True, size=11)
    ws['B4'].alignment = Alignment(horizontal='left', vertical='center')
    
    # Set row heights for header
    ws.row_dimensions[1].height = 18
    ws.row_dimensions[2].height = 15
    ws.row_dimensions[3].height = 15
    ws.row_dimensions[4].height = 15
    ws.row_dimensions[5].height = 12  # Spacer row
    
    # =========================================================================
    # SCHEDULE TITLE - Rows 6-8 (Centered across A-H)
    # =========================================================================
    # Row 6: CLASS SCHEDULE
    ws.merge_cells('A6:H6')
    ws['A6'] = 'CLASS SCHEDULE'
    ws['A6'].font = Font(bold=True, size=12)
    ws['A6'].alignment = Alignment(horizontal='center', vertical='center')
    
    # Row 7: Semester and Academic Year (format: "1ST SEMESTER, AY 2025-2026")
    ws.merge_cells('A7:H7')
    # Use format_semester_text for proper ordinal formatting
    semester_text = format_semester_text(current_settings.semester, current_settings.academic_year) if current_settings else ""
    ws['A7'] = semester_text
    ws['A7'].font = Font(bold=True, size=11)
    ws['A7'].alignment = Alignment(horizontal='center', vertical='center')
    
    # Row 8: Section name (e.g., "BSCS-4A" or "BSCS/ACT-1A" based on shared program settings)
    ws.merge_cells('A8:H8')
    # Use full_section_name which respects shared program settings based on year level (keep the dash)
    if hasattr(section, 'full_section_name'):
        section_display = section.full_section_name
    else:
        section_display = f"{dept_code}-{section.year_level}{section.section_name}"
    ws['A8'] = section_display
    ws['A8'].font = Font(bold=True, size=11)
    ws['A8'].alignment = Alignment(horizontal='center', vertical='center')
    
    # Row 9: Spacer
    ws.row_dimensions[9].height = 8
    
    # =========================================================================
    # TABLE HEADERS - Row 10-11 (starting at column A)
    # =========================================================================
    # Subject Code (A10:A11)
    ws.merge_cells('A10:A11')
    ws['A10'] = 'Subject Code'
    ws['A10'].font = header_font
    ws['A10'].alignment = header_alignment
    ws['A10'].border = thin_border
    ws['A11'].border = thin_border
    
    # Description (B10:B11)
    ws.merge_cells('B10:B11')
    ws['B10'] = 'Description'
    ws['B10'].font = header_font
    ws['B10'].alignment = header_alignment
    ws['B10'].border = thin_border
    ws['B11'].border = thin_border
    
    # Units header (C10:D10)
    ws.merge_cells('C10:D10')
    ws['C10'] = 'Units'
    ws['C10'].font = header_font
    ws['C10'].alignment = header_alignment
    ws['C10'].border = thin_border
    ws['D10'].border = thin_border  # Add border to merged cell
    
    # Units sub-headers (C11, D11)
    ws['C11'] = 'Lec'
    ws['C11'].font = header_font
    ws['C11'].alignment = header_alignment
    ws['C11'].border = thin_border
    
    ws['D11'] = 'Lab'
    ws['D11'].font = header_font
    ws['D11'].alignment = header_alignment
    ws['D11'].border = thin_border
    
    # Day (E10:E11)
    ws.merge_cells('E10:E11')
    ws['E10'] = 'Day'
    ws['E10'].font = header_font
    ws['E10'].alignment = header_alignment
    ws['E10'].border = thin_border
    ws['E11'].border = thin_border
    
    # Time (F10:F11)
    ws.merge_cells('F10:F11')
    ws['F10'] = 'Time'
    ws['F10'].font = header_font
    ws['F10'].alignment = header_alignment
    ws['F10'].border = thin_border
    ws['F11'].border = thin_border
    
    # Room (G10:G11)
    ws.merge_cells('G10:G11')
    ws['G10'] = 'Room'
    ws['G10'].font = header_font
    ws['G10'].alignment = header_alignment
    ws['G10'].border = thin_border
    ws['G11'].border = thin_border
    
    # Professor (H10:H11)
    ws.merge_cells('H10:H11')
    ws['H10'] = 'Professor'
    ws['H10'].font = header_font
    ws['H10'].alignment = header_alignment
    ws['H10'].border = thin_border
    ws['H11'].border = thin_border
    
    # Set header row heights
    ws.row_dimensions[10].height = 20
    ws.row_dimensions[11].height = 20
    
    # =========================================================================
    # SCHEDULE DATA - Starting Row 12, Column A
    # =========================================================================
    current_row = 12
    total_lec_units = 0
    total_lab_units = 0
    
    # Sort schedules by subject code, then by day
    day_order = {'Monday': 1, 'Tuesday': 2, 'Wednesday': 3, 'Thursday': 4, 'Friday': 5, 'Saturday': 6, 'Sunday': 7}
    sorted_schedules = sorted(schedules, key=lambda s: (
        s.subject.subject_code if s.subject else '',
        day_order.get(s.day_of_week, 99),
        s.start_time
    ))
    
    # Track which subject+type combinations we've already counted for units
    counted_subject_types = set()
    
    for schedule in sorted_schedules:
        subject = schedule.subject
        schedule_type = getattr(schedule, 'schedule_type', 'lecture') or 'lecture'
        
        # Get units based on schedule type - show only relevant units for this row
        if schedule_type.lower() in ['laboratory', 'lab']:
            # Lab schedule - show only lab units
            show_lec = ''
            show_lab = subject.lab_units if subject and subject.lab_units else ''
        else:
            # Lecture schedule - show only lecture units
            show_lec = subject.lec_units if subject and subject.lec_units else ''
            show_lab = ''
        
        # Track totals (only count once per subject+type combination)
        subject_type_key = (subject.id if subject else None, schedule_type.lower())
        if subject and subject_type_key not in counted_subject_types:
            if schedule_type.lower() in ['laboratory', 'lab'] and subject.lab_units:
                total_lab_units += float(subject.lab_units)
            elif schedule_type.lower() not in ['laboratory', 'lab'] and subject.lec_units:
                total_lec_units += float(subject.lec_units)
            counted_subject_types.add(subject_type_key)
        
        # Format time (e.g., "2:00 PM-5:00 PM")
        start_time = schedule.start_time.strftime('%I:%M %p').lstrip('0') if schedule.start_time else ''
        end_time = schedule.end_time.strftime('%I:%M %p').lstrip('0') if schedule.end_time else ''
        time_display = f"{start_time}-{end_time}" if start_time and end_time else ''
        
        # Format faculty name (Surname first: "Last, First M.")
        faculty_display = ''
        if schedule.faculty:
            faculty_display = schedule.faculty.full_name
        
        # Write row data (starting at column B = 2)
        row_data = [
            (subject.subject_code if subject else '', 'left'),
            (subject.course_description if subject else '', 'left'),
            (show_lec, 'center'),
            (show_lab, 'center'),
            (schedule.day_of_week if schedule.day_of_week else '', 'left'),
            (time_display, 'left'),
            (schedule.room.room_number if schedule.room else 'TBA', 'left'),
            (faculty_display, 'left')
        ]
        
        # Write data starting at column A (column index 1)
        for col_idx, (value, align) in enumerate(row_data, start=1):
            cell = ws.cell(row=current_row, column=col_idx, value=value)
            cell.font = Font(size=10)
            cell.alignment = Alignment(horizontal=align, vertical='center', wrap_text=True)
            cell.border = thin_border
        
        current_row += 1
    
    # =========================================================================
    # TOTAL ROW (starting at column A)
    # =========================================================================
    # Merge A-B for TOTAL label
    ws.merge_cells(f'A{current_row}:B{current_row}')
    ws.cell(row=current_row, column=1, value='TOTAL')
    ws.cell(row=current_row, column=1).font = Font(bold=False, size=10)
    ws.cell(row=current_row, column=1).alignment = Alignment(horizontal='center', vertical='center')
    ws.cell(row=current_row, column=1).border = thin_border
    ws.cell(row=current_row, column=2).border = thin_border
    
    # Total units in merged Lec+Lab columns (C-D)
    total_units = total_lec_units + total_lab_units
    total_units_display = int(total_units) if total_units == int(total_units) else total_units
    ws.merge_cells(f'C{current_row}:D{current_row}')
    ws.cell(row=current_row, column=3, value=total_units_display)
    ws.cell(row=current_row, column=3).font = Font(bold=False, size=10)
    ws.cell(row=current_row, column=3).alignment = Alignment(horizontal='center', vertical='center')
    ws.cell(row=current_row, column=3).border = thin_border
    ws.cell(row=current_row, column=4).border = thin_border
    
    # Empty cells for the rest of total row (E-H)
    for col_idx in range(5, 9):
        ws.cell(row=current_row, column=col_idx).border = thin_border
    
    current_row += 1
    
    # =========================================================================
    # SIGNATURE SECTION (centered on table A-H)
    # =========================================================================
    sig_start_row = current_row + 2
    
    # Prepared by label (column B)
    ws.cell(row=sig_start_row, column=2, value='Prepared by:')
    ws.cell(row=sig_start_row, column=2).font = Font(size=11)
    ws.cell(row=sig_start_row, column=2).alignment = Alignment(horizontal='left', vertical='center')
    
    # Noted by label (column F)
    ws.cell(row=sig_start_row, column=6, value='Noted by:')
    ws.cell(row=sig_start_row, column=6).font = Font(size=11)
    ws.cell(row=sig_start_row, column=6).alignment = Alignment(horizontal='left', vertical='center')
    
    # Blank row for signature space
    sig_start_row += 2
    
    # Dean name - BOLD and UPPERCASE (column B-D centered)
    dean_display_name = user.full_name.upper() if user and user.full_name else '{Dean Name}'
    ws.merge_cells(f'B{sig_start_row}:D{sig_start_row}')
    ws.cell(row=sig_start_row, column=2, value=dean_display_name)
    ws.cell(row=sig_start_row, column=2).font = Font(bold=True, size=11)
    ws.cell(row=sig_start_row, column=2).alignment = Alignment(horizontal='center', vertical='center')
    
    # Department President name - BOLD (column F-H centered)
    ws.merge_cells(f'F{sig_start_row}:H{sig_start_row}')
    ws.cell(row=sig_start_row, column=6, value=get_institution_head_name())
    ws.cell(row=sig_start_row, column=6).font = Font(bold=True, size=11)
    ws.cell(row=sig_start_row, column=6).alignment = Alignment(horizontal='center', vertical='center')
    
    # Dean title (column B-D centered)
    sig_start_row += 1
    dean_title = f"Dean, {dept_display_name}" if dept_display_name else "Dean, {Program Name}"
    ws.merge_cells(f'B{sig_start_row}:D{sig_start_row}')
    ws.cell(row=sig_start_row, column=2, value=dean_title)
    ws.cell(row=sig_start_row, column=2).font = Font(size=11)
    ws.cell(row=sig_start_row, column=2).alignment = Alignment(horizontal='center', vertical='center')
    
    # Department President title (column F-H centered)
    ws.merge_cells(f'F{sig_start_row}:H{sig_start_row}')
    ws.cell(row=sig_start_row, column=6, value='Department President')
    ws.cell(row=sig_start_row, column=6).font = Font(size=11)
    ws.cell(row=sig_start_row, column=6).alignment = Alignment(horizontal='center', vertical='center')
    
    # =========================================================================
    # PAGE SETUP
    # =========================================================================
    ws.page_setup.fitToPage = True
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    
    # =========================================================================
    # SAVE AND RETURN
    # =========================================================================
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    filename = f"{dept_code}_{section.year_level}{section.section_name}_Schedule_Posting.xlsx".replace(' ', '_')
    return output, filename

def generate_class_schedule_pdf(section, schedules, current_settings, user):
    """Generate PDF file for class schedule"""
    # Get time range from settings (default to 7-20 if not set)
    start_hour = (current_settings.schedule_start_time.hour if current_settings and current_settings.schedule_start_time else 7)
    end_hour = (current_settings.schedule_end_time.hour if current_settings and current_settings.schedule_end_time else 20)
    
    # Prepare metadata
    dept_name = _get_department_name(section.program)
    dept_code = section.program.program_code if section.program else ''
    semester_text = f"{current_settings.semester.upper()}, AY {current_settings.academic_year}" if current_settings else "CLASS SCHEDULE"
    section_display = section.full_section_name if hasattr(section, 'full_section_name') else f"{dept_code}-{section.year_level}{section.section_name}"
    
    # Create PDF with dynamic time range
    output = create_pdf_schedule(
        schedules=schedules,
        title='CLASS SCHEDULE',
        semester_text=semester_text,
        section_display=section_display,
        dept_name=dept_name.upper(),
        filename=f"{dept_code}_{section.year_level}{section.section_name}_Schedule.pdf",
        start_hour=start_hour,
        end_hour=end_hour
    )
    
    filename = f"{dept_code}_{section.year_level}{section.section_name}_Schedule.pdf".replace(' ', '_')
    return output, filename

def generate_faculty_schedule_pdf(faculty, schedules, current_settings, user):
    """Generate PDF file for faculty schedule"""
    # Get time range from settings (default to 7-20 if not set)
    start_hour = (current_settings.schedule_start_time.hour if current_settings and current_settings.schedule_start_time else 7)
    end_hour = (current_settings.schedule_end_time.hour if current_settings and current_settings.schedule_end_time else 20)
    
    # Prepare metadata
    dept_name = faculty.department.department_name if (hasattr(faculty, 'department') and faculty.department) else '{Program Name}'
    semester_text = f"{current_settings.semester.upper()}, AY {current_settings.academic_year}" if current_settings else "FACULTY SCHEDULE"
    
    # Create PDF with dynamic time range
    output = create_pdf_schedule(
        schedules=schedules,
        title='FACULTY SCHEDULE',
        semester_text=semester_text,
        section_display=faculty.full_name,
        dept_name=dept_name.upper(),
        filename=f"{faculty.full_name}_Schedule.pdf",
        start_hour=start_hour,
        end_hour=end_hour
    )
    
    filename = f"{faculty.last_name}_{faculty.first_name}_Schedule.pdf"
    return output, filename

def generate_room_schedule_pdf(room, schedules, current_settings, user):
    """Generate PDF file for room schedule"""
    # Get time range from settings (default to 7-20 if not set)
    start_hour = (current_settings.schedule_start_time.hour if current_settings and current_settings.schedule_start_time else 7)
    end_hour = (current_settings.schedule_end_time.hour if current_settings and current_settings.schedule_end_time else 20)
    
    # Prepare metadata
    building_name = room.building.building_name if room.building else 'Building'
    dept_name = 'DEPT'  # Room schedules are typically department-wide
    semester_text = f"{current_settings.semester.upper()}, AY {current_settings.academic_year}" if current_settings else "ROOM SCHEDULE"
    room_display = f"{building_name} - Room {room.room_number}"
    
    # Create PDF with dynamic time range
    output = create_pdf_schedule(
        schedules=schedules,
        title='ROOM SCHEDULE',
        semester_text=semester_text,
        section_display=room_display,
        dept_name=dept_name,
        filename=f"Room_{room.room_number}_Schedule.pdf",
        start_hour=start_hour,
        end_hour=end_hour
    )
    
    filename = f"Room_{room.room_number}_Schedule.pdf".replace(' ', '_')
    return output, filename


# ============================================================================
# USER EXPORT FUNCTIONS
# ============================================================================

def export_users_excel(users, include_archived=False):
    """
    Generate Excel file for user list export
    
    Args:
        users: List of User objects to export
        include_archived: Whether to include archived users
    
    Returns:
        Tuple of (BytesIO output, filename)
    """
    output = io.BytesIO()
    wb = Workbook()
    ws = wb.active
    ws.title = "Users"
    
    # Define styles
    header_fill = PatternFill(start_color="2563EB", end_color="2563EB", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, size=12)
    border = Border(
        left=Side(style='thin', color='DDDDDD'),
        right=Side(style='thin', color='DDDDDD'),
        top=Side(style='thin', color='DDDDDD'),
        bottom=Side(style='thin', color='DDDDDD')
    )
    
    # Title row
    ws.merge_cells('A1:G1')
    ws['A1'] = "iSchedWise - User List Export"
    ws['A1'].font = Font(bold=True, size=16, color="1E40AF")
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 30
    
    # Export date
    ws.merge_cells('A2:G2')
    ws['A2'] = f"Exported on: {datetime.now().strftime('%B %d, %Y at %I:%M %p')}"
    ws['A2'].font = Font(size=10, italic=True, color="666666")
    ws['A2'].alignment = Alignment(horizontal='center')
    ws.row_dimensions[2].height = 20
    
    # Empty row
    ws.row_dimensions[3].height = 10
    
    # Headers
    headers = ['Full Name', 'Username', 'Email', 'Role', 'Status', 'Departments', 'Last Login']
    if include_archived:
        headers.extend(['Archived', 'Archived By', 'Archived At', 'Archive Reason'])
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = border
    ws.row_dimensions[4].height = 25
    
    # Data rows
    for row_idx, user in enumerate(users, 5):
        # Get programs for user
        programs = []
        if hasattr(user, 'user_programs') and user.user_programs:
            for ud in user.user_programs:
                if ud.program:
                    programs.append(ud.program.program_name)
        dept_str = ', '.join(programs) if programs else 'N/A'
        
        # Format last login
        last_login = user.last_login.strftime('%Y-%m-%d %H:%M') if user.last_login else 'Never'
        
        # Status
        if user.is_archived:
            status = 'Archived'
        elif user.is_active:
            status = 'Active'
        else:
            status = 'Inactive'
        
        row_data = [
            user.full_name,
            user.username,
            user.email,
            user.role.title(),
            status,
            dept_str,
            last_login
        ]
        
        if include_archived:
            archived_by_name = ''
            if user.is_archived and user.archived_by:
                from app.models.user import User
                archiver = User.query.get(user.archived_by)
                archived_by_name = archiver.full_name if archiver else 'Unknown'
            
            row_data.extend([
                'Yes' if user.is_archived else 'No',
                archived_by_name,
                user.archived_at.strftime('%Y-%m-%d %H:%M') if user.archived_at else '',
                user.archive_reason or ''
            ])
        
        for col, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col, value=value)
            cell.border = border
            cell.alignment = Alignment(horizontal='left', vertical='center')
            
            # Alternate row colors
            if row_idx % 2 == 0:
                cell.fill = PatternFill(start_color="F8FAFC", end_color="F8FAFC", fill_type="solid")
    
    # Adjust column widths
    column_widths = [25, 20, 35, 12, 12, 40, 18]
    if include_archived:
        column_widths.extend([10, 20, 18, 30])
    
    for col, width in enumerate(column_widths, 1):
        ws.column_dimensions[get_column_letter(col)].width = width
    
    # Summary row
    summary_row = len(users) + 6
    ws.merge_cells(f'A{summary_row}:G{summary_row}')
    ws[f'A{summary_row}'] = f"Total Users: {len(users)}"
    ws[f'A{summary_row}'].font = Font(bold=True, size=11)
    ws[f'A{summary_row}'].alignment = Alignment(horizontal='right')
    
    wb.save(output)
    output.seek(0)
    
    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    filename = f"Users_Export_{timestamp}.xlsx"
    
    return output, filename


def export_users_pdf(users, include_archived=False):
    """
    Generate PDF file for user list export
    
    Args:
        users: List of User objects to export
        include_archived: Whether to include archived users
    
    Returns:
        Tuple of (BytesIO output, filename)
    """
    output = io.BytesIO()
    doc = SimpleDocTemplate(
        output,
        pagesize=landscape(letter),
        rightMargin=0.5*inch,
        leftMargin=0.5*inch,
        topMargin=0.5*inch,
        bottomMargin=0.5*inch
    )
    
    elements = []
    styles = getSampleStyleSheet()
    
    # Title
    title_style = ParagraphStyle(
        'Title',
        parent=styles['Heading1'],
        fontSize=18,
        textColor=rl_colors.HexColor('#1E40AF'),
        alignment=TA_CENTER,
        spaceAfter=12
    )
    elements.append(Paragraph("iSchedWise - User List Export", title_style))
    
    # Export date
    date_style = ParagraphStyle(
        'Date',
        parent=styles['Normal'],
        fontSize=10,
        textColor=rl_colors.gray,
        alignment=TA_CENTER,
        spaceAfter=20
    )
    elements.append(Paragraph(f"Exported on: {datetime.now().strftime('%B %d, %Y at %I:%M %p')}", date_style))
    elements.append(Spacer(1, 0.2*inch))
    
    # Table headers
    headers = ['Full Name', 'Username', 'Email', 'Role', 'Status', 'Departments']
    
    # Prepare table data
    table_data = [headers]
    
    for user in users:
        # Get programs for user
        programs = []
        if hasattr(user, 'user_programs') and user.user_programs:
            for ud in user.user_programs:
                if ud.program:
                    programs.append(ud.program.program_code or ud.program.program_name[:15])
        dept_str = ', '.join(programs) if programs else 'N/A'
        
        # Status
        if user.is_archived:
            status = 'Archived'
        elif user.is_active:
            status = 'Active'
        else:
            status = 'Inactive'
        
        table_data.append([
            user.full_name[:25],
            user.username[:15],
            user.email[:30],
            user.role.title(),
            status,
            dept_str[:20]
        ])
    
    # Create table
    col_widths = [1.8*inch, 1.2*inch, 2.2*inch, 0.8*inch, 0.8*inch, 1.8*inch]
    table = Table(table_data, colWidths=col_widths, repeatRows=1)
    
    # Table styling
    table_style = TableStyle([
        # Header styling
        ('BACKGROUND', (0, 0), (-1, 0), rl_colors.HexColor('#2563EB')),
        ('TEXTCOLOR', (0, 0), (-1, 0), rl_colors.white),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 10),
        ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
        ('VALIGN', (0, 0), (-1, 0), 'MIDDLE'),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 10),
        ('TOPPADDING', (0, 0), (-1, 0), 10),
        
        # Data rows styling
        ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 1), (-1, -1), 9),
        ('ALIGN', (0, 1), (-1, -1), 'LEFT'),
        ('VALIGN', (0, 1), (-1, -1), 'MIDDLE'),
        ('BOTTOMPADDING', (0, 1), (-1, -1), 6),
        ('TOPPADDING', (0, 1), (-1, -1), 6),
        
        # Borders
        ('GRID', (0, 0), (-1, -1), 0.5, rl_colors.HexColor('#E5E7EB')),
        ('LINEBELOW', (0, 0), (-1, 0), 2, rl_colors.HexColor('#1E40AF')),
    ])
    
    # Alternate row colors
    for i in range(1, len(table_data)):
        if i % 2 == 0:
            table_style.add('BACKGROUND', (0, i), (-1, i), rl_colors.HexColor('#F8FAFC'))
    
    table.setStyle(table_style)
    elements.append(table)
    
    # Summary
    elements.append(Spacer(1, 0.3*inch))
    summary_style = ParagraphStyle(
        'Summary',
        parent=styles['Normal'],
        fontSize=11,
        alignment=TA_LEFT,
        fontName='Helvetica-Bold'
    )
    elements.append(Paragraph(f"Total Users: {len(users)}", summary_style))
    
    doc.build(elements)
    output.seek(0)
    
    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    filename = f"Users_Export_{timestamp}.pdf"
    
    return output, filename


# ============================================================================
# ARCHIVE SCHEDULE GRID EXPORT
# ============================================================================

def place_archive_in_grid(ws, archives, start_hour=7, group_by='section'):
    """Place archived schedule records in the weekly grid (uses flat Archive fields)."""
    day_columns = get_day_column_mapping()

    for archive in archives:
        if not archive.day_of_week or archive.day_of_week not in day_columns:
            continue
        if not archive.start_time or not archive.end_time:
            continue

        col_idx = day_columns[archive.day_of_week]

        start_minutes = archive.start_time.hour * 60 + archive.start_time.minute
        end_minutes   = archive.end_time.hour   * 60 + archive.end_time.minute
        slot_start_minutes = start_hour * 60

        start_row = 11 + ((start_minutes - slot_start_minutes) // 30)
        duration_minutes = end_minutes - start_minutes
        rows_to_merge = max(1, (duration_minutes + 29) // 30)

        subject_code = archive.subject_code or 'TBA'
        room         = archive.room_number   or 'TBA'
        faculty      = archive.faculty_name  or 'TBA'
        section      = archive.section_name  or 'TBA'

        if group_by == 'faculty':
            # grouped by faculty → show subject / room / section
            cell_content = f"{subject_code}\n{room}\n{section}"
        elif group_by == 'room':
            # grouped by room → show subject / section / faculty
            cell_content = f"{subject_code}\n{section}\n{faculty}"
        else:
            # grouped by section (default) → show subject / room / faculty
            cell_content = f"{subject_code}\n{room}\n{faculty}"

        cell = ws.cell(row=start_row, column=col_idx, value=cell_content)
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.font = Font(size=10)

        if rows_to_merge > 1:
            end_row = start_row + rows_to_merge - 1
            ws.merge_cells(start_row=start_row, start_column=col_idx,
                           end_row=end_row, end_column=col_idx)


def generate_archive_schedule_excel(groups, group_by, filter_info, current_settings, user):
    """Generate a grid-format Excel workbook for archived schedules.

    One worksheet per group (section / faculty / room), using the same
    institution logos, header text, time-slot grid, and signature section as
    generate_class_schedule_excel().

    Args:
        groups      : OrderedDict[str, list[Archive]]
        group_by    : 'section' | 'faculty' | 'room'
        filter_info : dict with optional keys 'academic_year', 'semester'
        current_settings : AcademicSettings instance or None
        user        : current_user (Flask-Login)

    Returns:
        (output: io.BytesIO, filename: str)
    """
    import re

    start_hour = (current_settings.schedule_start_time.hour if current_settings and current_settings.schedule_start_time else 7)
    end_hour   = (current_settings.schedule_end_time.hour if current_settings and current_settings.schedule_end_time else 20)

    # Determine page-level title and logo column widths (same as regular export)
    _HDR_HEIGHT_PT = 63
    _LEFT_COL_WIDTH  = 12.71
    _RIGHT_COL_WIDTH = 20.71

    if group_by == 'faculty':
        main_title = 'FACULTY SCHEDULE'
    elif group_by == 'room':
        main_title = 'ROOM SCHEDULE'
    else:
        main_title = 'CLASS SCHEDULE'

    # Semester text from filter_info (may be empty when no filter was applied)
    fi_ay  = (filter_info or {}).get('academic_year', '')
    fi_sem = (filter_info or {}).get('semester', '')

    # Pre-compute semester_text from filter; per-sheet override if empty
    global_semester_text = ''
    if fi_ay and fi_sem:
        global_semester_text = format_semester_text(fi_sem, fi_ay)
    elif fi_ay:
        global_semester_text = f"A.Y. {fi_ay}"
    elif fi_sem:
        global_semester_text = fi_sem.upper()

    wb = Workbook()
    # Remove the default empty sheet
    wb.remove(wb.active)

    time_slots = generate_time_slots(start_hour=start_hour, end_hour=end_hour)
    last_row   = 11 + len(time_slots) - 1

    def _sanitize_sheet_name(name):
        """Truncate to 31 chars and strip Excel-illegal characters."""
        name = re.sub(r'[/\\?*\[\]:]', '', str(name))
        return name[:31] or 'Sheet'

    for group_name, archives in groups.items():
        ws = wb.create_sheet(title=_sanitize_sheet_name(group_name))

        # ------------------------------------------------------------------
        # Determine per-sheet values that depend on archive data
        # ------------------------------------------------------------------
        first = archives[0] if archives else None

        # Semester text: prefer global (from filter), else first archive's values
        if global_semester_text:
            semester_text = global_semester_text
        elif first and first.academic_year and first.semester:
            semester_text = format_semester_text(first.semester, first.academic_year)
        elif first and first.academic_year:
            semester_text = f"A.Y. {first.academic_year}"
        else:
            semester_text = main_title

        # Row-4 department display
        if group_by == 'section' and first and first.program_name:
            dept_display = first.program_name.upper()
        elif group_by == 'faculty':
            dept_display = 'FACULTY SCHEDULE'
        elif group_by == 'room':
            dept_display = 'ROOM SCHEDULE'
        else:
            dept_display = ''

        # ------------------------------------------------------------------
        # LOGOS  (col A left, col G right)
        # ------------------------------------------------------------------
        add_logo_centered(ws, get_institution_logo_path(),
                          col_0idx=0, img_width_px=75, img_height_px=75,
                          col_width_chars=_LEFT_COL_WIDTH,
                          header_height_pts=_HDR_HEIGHT_PT)
        add_logo_centered(ws, get_institution_logo_right_path(),
                          col_0idx=6, img_width_px=80, img_height_px=80,
                          col_width_chars=_RIGHT_COL_WIDTH,
                          header_height_pts=_HDR_HEIGHT_PT)

        # ------------------------------------------------------------------
        # HEADER TEXT  (B1–B4)
        # ------------------------------------------------------------------
        ws['B1'] = 'Republic of the Philippines'
        ws['B1'].font = Font(size=11)
        ws['B1'].alignment = Alignment(horizontal='left', vertical='center')

        ws['B2'] = 'Municipality of Norzagaray'
        ws['B2'].font = Font(size=11)
        ws['B2'].alignment = Alignment(horizontal='left', vertical='center')

        ws['B3'] = get_institution_name()
        ws['B3'].font = Font(bold=True, size=11)
        ws['B3'].alignment = Alignment(horizontal='left', vertical='center')

        ws['B4'] = dept_display
        ws['B4'].font = Font(bold=True, size=11)
        ws['B4'].alignment = Alignment(horizontal='left', vertical='center')

        ws.row_dimensions[1].height = 18
        ws.row_dimensions[2].height = 15
        ws.row_dimensions[3].height = 15
        ws.row_dimensions[4].height = 15
        ws.row_dimensions[5].height = 12  # spacer

        # ------------------------------------------------------------------
        # TITLE BLOCK  (rows 6–8, merged A:G)
        # ------------------------------------------------------------------
        ws.merge_cells('A6:G6')
        ws['A6'] = main_title
        ws['A6'].font = Font(bold=True, size=12)
        ws['A6'].alignment = Alignment(horizontal='center', vertical='center')

        ws.merge_cells('A7:G7')
        ws['A7'] = semester_text
        ws['A7'].font = Font(bold=True, size=11)
        ws['A7'].alignment = Alignment(horizontal='center', vertical='center')

        ws.merge_cells('A8:G8')
        ws['A8'] = group_name
        ws['A8'].font = Font(bold=True, size=11)
        ws['A8'].alignment = Alignment(horizontal='center', vertical='center')

        ws.row_dimensions[9].height = 8  # spacer

        # ------------------------------------------------------------------
        # SCHEDULE GRID
        # ------------------------------------------------------------------
        add_column_headers(ws)
        write_time_slots(ws, time_slots)
        place_archive_in_grid(ws, archives, start_hour=start_hour, group_by=group_by)
        apply_grid_borders(ws, last_row)

        # ------------------------------------------------------------------
        # SIGNATURE SECTION
        # ------------------------------------------------------------------
        sig_row = last_row + 3

        ws.cell(row=sig_row, column=2, value='Prepared by:')
        ws.cell(row=sig_row, column=2).font = Font(size=11)
        ws.cell(row=sig_row, column=2).alignment = Alignment(horizontal='left', vertical='center')

        ws.cell(row=sig_row, column=5, value='Noted by:')
        ws.cell(row=sig_row, column=5).font = Font(size=11)
        ws.cell(row=sig_row, column=5).alignment = Alignment(horizontal='left', vertical='center')

        sig_row += 2  # blank row for signature space

        # Dean / user name (merged B:D)
        dean_name = user.full_name.upper() if user and user.full_name else '{Dean Name}'
        ws.merge_cells(f'B{sig_row}:D{sig_row}')
        ws.cell(row=sig_row, column=2, value=dean_name)
        ws.cell(row=sig_row, column=2).font = Font(bold=True, size=11)
        ws.cell(row=sig_row, column=2).alignment = Alignment(horizontal='center', vertical='center')

        # Institution head name (merged E:G)
        ws.merge_cells(f'E{sig_row}:G{sig_row}')
        ws.cell(row=sig_row, column=5, value=get_institution_head_name())
        ws.cell(row=sig_row, column=5).font = Font(bold=True, size=11)
        ws.cell(row=sig_row, column=5).alignment = Alignment(horizontal='center', vertical='center')

        sig_row += 1

        # Dean title (merged B:D)
        dean_title = f"Dean, {dept_display}" if dept_display else 'Dean'
        ws.merge_cells(f'B{sig_row}:D{sig_row}')
        ws.cell(row=sig_row, column=2, value=dean_title)
        ws.cell(row=sig_row, column=2).font = Font(size=11)
        ws.cell(row=sig_row, column=2).alignment = Alignment(horizontal='center', vertical='center')

        # Institution head title (merged E:G)
        ws.merge_cells(f'E{sig_row}:G{sig_row}')
        ws.cell(row=sig_row, column=5, value='Department President')
        ws.cell(row=sig_row, column=5).font = Font(size=11)
        ws.cell(row=sig_row, column=5).alignment = Alignment(horizontal='center', vertical='center')

        # ------------------------------------------------------------------
        # Column widths & page setup
        # ------------------------------------------------------------------
        set_column_widths(ws)
        ws.page_setup.fitToPage  = True
        ws.page_setup.fitToWidth = 1
        ws.page_setup.fitToHeight = 0

    # Save to buffer
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    filename  = f"archived_schedules_{timestamp}.xlsx"
    return output, filename
