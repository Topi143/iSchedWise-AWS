"""
Admin Tools routes - Super Admin exclusive tools and system management
"""
import os
import platform
import sys
import uuid
from datetime import datetime, timedelta, timezone
from flask import Blueprint, abort, render_template, request, jsonify, current_app, flash, redirect, url_for
from flask_login import login_required, current_user
from werkzeug.utils import secure_filename
from app.decorators import super_admin_required, role_required
from app.extensions import db
from app.services.database_backup_service import BackupError, DatabaseBackupService
from app.models import (
    User, Program, Curriculum, Faculty, Building, Room,
    Schedule, ExamSchedule, Archive, UserActivityLog,
    LoginHistory, SystemConfig,
    FacultySubjectAssignment, Section, InstitutionSettings
)
from app.utils.security_email_templates import build_branded_mail_sender, build_smtp_test_email_payload

admin_tools_bp = Blueprint('admin_tools', __name__, url_prefix='/admin')

ALLOWED_BRANDING_EXTENSIONS = {'png', 'jpg', 'jpeg', 'gif', 'webp'}

DB_ACTION_CONFIRM_PHRASES = {
    'cleanup_old_logs': 'CLEANUP OLD LOGS',
    'reset_class_schedules': 'RESET CLASS SCHEDULES',
    'reset_exam_schedules': 'RESET EXAM SCHEDULES',
    'reset_all_schedules': 'RESET ALL SCHEDULES',
    'truncate_archives': 'TRUNCATE ARCHIVES',
    'truncate_activity_logs': 'TRUNCATE ACTIVITY LOGS',
    'truncate_login_history': 'TRUNCATE LOGIN HISTORY',
}


def _normalize_confirmation_phrase(value):
    if value is None:
        return ''
    return ' '.join(str(value).strip().upper().split())


def _get_confirmation_phrase(action_key):
    return DB_ACTION_CONFIRM_PHRASES.get(action_key, '')


def _is_confirmation_phrase_valid(action_key, provided_phrase):
    expected = _get_confirmation_phrase(action_key)
    if not expected:
        return False
    return _normalize_confirmation_phrase(provided_phrase) == _normalize_confirmation_phrase(expected)


def _confirmation_phrase_error(action_key):
    expected = _get_confirmation_phrase(action_key)
    return jsonify({
        'success': False,
        'error': f'Confirmation phrase mismatch. Type "{expected}" to proceed.',
        'required_phrase': expected,
    }), 400


def _allowed_branding_file(filename):
    """Check if uploaded branding image has an allowed extension."""
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_BRANDING_EXTENSIONS


# ============================================================
# 1. SYSTEM ADMINISTRATION PANEL
# ============================================================

@admin_tools_bp.route('/')
@login_required
@super_admin_required
def system_panel():
    """Legacy system panel page has been removed."""
    abort(404)


@admin_tools_bp.route('/dashboard')
@login_required
@super_admin_required
def superadmin_dashboard():
    """Dedicated super administrator dashboard."""
    return render_template('admin/superadmin_dashboard.html', user=current_user)


@admin_tools_bp.route('/branding/update', methods=['POST'])
@login_required
@super_admin_required
def update_branding():
    """Update global application branding (name and app logo)."""
    try:
        settings = InstitutionSettings.get_settings()

        submitted_name = request.form.get('system_name', '').strip()
        if not submitted_name:
            flash('System name is required.', 'error')
            return redirect(url_for('settings.index') + '#system')
        if len(submitted_name) > 255:
            flash('System name must be 255 characters or less.', 'error')
            return redirect(url_for('settings.index') + '#system')

        old_name = settings.system_name
        settings.system_name = submitted_name
        settings.updated_by = current_user.id

        branding_file = request.files.get('branding_logo')
        branding_logo_updated = False
        if branding_file and branding_file.filename:
            if not _allowed_branding_file(branding_file.filename):
                flash('Invalid logo format. Please upload PNG, JPG, JPEG, GIF, or WebP.', 'error')
                return redirect(url_for('settings.index') + '#system')

            upload_folder = os.path.join(current_app.root_path, 'static', 'uploads', 'branding_logos')
            os.makedirs(upload_folder, exist_ok=True)

            filename = secure_filename(branding_file.filename)
            unique_filename = f"{uuid.uuid4().hex}_{filename}"
            filepath = os.path.join(upload_folder, unique_filename)

            if settings.branding_logo:
                old_logo_path = os.path.join(current_app.root_path, 'static', settings.branding_logo)
                if os.path.exists(old_logo_path):
                    try:
                        os.remove(old_logo_path)
                    except Exception:
                        pass

            branding_file.save(filepath)
            settings.branding_logo = f"uploads/branding_logos/{unique_filename}"
            branding_logo_updated = True

        db.session.commit()

        from app.utils.activity_logger import log_activity
        log_activity(
            'updated',
            'system_branding',
            entity_name='Application Branding',
            details=(
                f"System name updated from '{old_name or 'Not set'}' to '{submitted_name}'. "
                f"Branding logo updated: {'Yes' if branding_logo_updated else 'No'}"
            )
        )

        flash('Global branding updated successfully!', 'success')
        return redirect(url_for('settings.index') + '#system')
    except Exception as e:
        db.session.rollback()
        flash(f'Error updating global branding: {str(e)}', 'error')
        return redirect(url_for('settings.index') + '#system')


@admin_tools_bp.route('/branding/remove-logo', methods=['POST'])
@login_required
@super_admin_required
def remove_branding_logo():
    """Remove the global application branding logo and fall back to export-left logo."""
    try:
        settings = InstitutionSettings.get_settings()
        if not settings.branding_logo:
            flash('No global branding logo to remove.', 'info')
            return redirect(url_for('settings.index') + '#system')

        branding_logo_path = os.path.join(current_app.root_path, 'static', settings.branding_logo)
        if os.path.exists(branding_logo_path):
            try:
                os.remove(branding_logo_path)
            except Exception:
                pass

        settings.branding_logo = None
        settings.updated_by = current_user.id
        db.session.commit()

        from app.utils.activity_logger import log_activity
        log_activity('updated', 'system_branding', entity_name='Application Branding', details='Removed global branding logo')

        flash('Global branding logo removed. UI will use the fallback logo.', 'success')
        return redirect(url_for('settings.index') + '#system')
    except Exception as e:
        db.session.rollback()
        flash(f'Error removing global branding logo: {str(e)}', 'error')
        return redirect(url_for('settings.index') + '#system')


@admin_tools_bp.route('/api/system-stats')
@login_required
@super_admin_required
def api_system_stats():
    """Get system statistics for the admin panel"""
    try:
        now = datetime.now()
        last_24h = now - timedelta(hours=24)

        # Entity counts
        total_users = User.query.filter_by(is_archived=False).count()
        active_users = User.query.filter_by(is_active=True, is_archived=False).count()
        total_faculty = Faculty.query.filter_by(is_archived=False).count()
        total_departments = Program.query.filter_by(is_archived=False).count()
        total_curricula = Curriculum.query.filter_by(is_archived=False).count()
        total_buildings = Building.query.filter_by(is_archived=False).count()
        total_rooms = Room.query.count()
        total_schedules = Schedule.query.filter_by(is_active=True).count()
        total_exam_schedules = ExamSchedule.query.filter_by(is_active=True).count()
        total_archives = Archive.query.count()
        total_sections = Section.query.count()

        # Activity stats
        logs_today = UserActivityLog.query.filter(UserActivityLog.created_at >= last_24h).count()
        total_logs = UserActivityLog.query.count()

        # Session stats
        active_sessions = LoginHistory.get_active_session_count()

        # Recent activity (last 10)
        recent_activities = UserActivityLog.query\
            .options(db.joinedload(UserActivityLog.user))\
            .order_by(UserActivityLog.created_at.desc())\
            .limit(10).all()

        recent_list = []
        for log in recent_activities:
            recent_list.append({
                'id': log.id,
                'user_name': log.user.full_name if log.user else 'System',
                'user_initial': (log.user.full_name[0] if log.user and log.user.full_name else 'S'),
                'action': log.action,
                'entity_type': log.entity_type,
                'entity_name': log.entity_name,
                'details': log.details,
                'created_at': (log.created_at.isoformat() + 'Z') if log.created_at else None,
                'ip_address': log.ip_address,
            })

        # System info
        system_info = {
            'python_version': sys.version.split()[0],
            'flask_version': current_app.import_name and __import__('flask').__version__,
            'platform': platform.system() + ' ' + platform.release(),
            'debug_mode': current_app.debug,
            'server_time': datetime.now().isoformat(),
        }

        # Maintenance mode status
        maintenance_mode = SystemConfig.get('maintenance_mode', False)
        maintenance_message = SystemConfig.get('maintenance_message', '')
        session_logout_policy = current_app.config.get('SESSION_LOGOUT_POLICY', 'browser_close')

        # Table row counts for database overview
        table_counts = {
            'users': User.query.count(),
            'programs': Program.query.count(),
            'curricula': Curriculum.query.count(),
            'faculty': Faculty.query.count(),
            'buildings': Building.query.count(),
            'rooms': Room.query.count(),
            'sections': Section.query.count(),
            'schedules': Schedule.query.count(),
            'exam_schedules': ExamSchedule.query.count(),
            'archives': Archive.query.count(),
            'activity_logs': total_logs,
            'login_history': LoginHistory.query.count(),
            'system_config': SystemConfig.query.count(),
        }

        return jsonify({
            'success': True,
            'stats': {
                'total_users': total_users,
                'active_users': active_users,
                'active_sessions': active_sessions,
                'logs_today': logs_today,
                'total_logs': total_logs,
                'total_faculty': total_faculty,
                'total_departments': total_departments,
                'total_curricula': total_curricula,
                'total_buildings': total_buildings,
                'total_rooms': total_rooms,
                'total_schedules': total_schedules,
                'total_exam_schedules': total_exam_schedules,
                'total_archives': total_archives,
                'total_sections': total_sections,
            },
            'recent_activities': recent_list,
            'system_info': system_info,
            'table_counts': table_counts,
            'maintenance_mode': maintenance_mode,
            'maintenance_message': maintenance_message or '',
            'session_logout_policy': session_logout_policy,
        })
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500


# ============================================================
# 2. ACTIVITY LOGS CONSOLE
# ============================================================

@admin_tools_bp.route('/activity')
@login_required
@role_required('admin', 'super_admin')
def activity_console():
    """Redirect legacy Activity Console URL to reports activity page."""
    return redirect(url_for('reports.activity_report'))


# ============================================================
# 3. DATABASE MANAGEMENT TOOLS
# ============================================================

@admin_tools_bp.route('/database')
@login_required
@super_admin_required
def database_tools():
    """Legacy database tools page has been removed."""
    abort(404)


@admin_tools_bp.route('/api/database/stats')
@login_required
@super_admin_required
def api_database_stats():
    """Get detailed database statistics"""
    try:
        tables = {
            'users': User.query.count(),
            'programs': Program.query.count(),
            'curricula': Curriculum.query.count(),
            'faculty': Faculty.query.count(),
            'faculty_subject_assignments': FacultySubjectAssignment.query.count(),
            'buildings': Building.query.count(),
            'rooms': Room.query.count(),
            'sections': Section.query.count(),
            'schedules': Schedule.query.count(),
            'exam_schedules': ExamSchedule.query.count(),
            'archives': Archive.query.count(),
            'activity_logs': UserActivityLog.query.count(),
            'login_history': LoginHistory.query.count(),
            'system_config': SystemConfig.query.count(),
        }

        # Archived counts
        archived = {
            'users': User.query.filter_by(is_archived=True).count(),
            'programs': Program.query.filter_by(is_archived=True).count(),
            'curricula': Curriculum.query.filter_by(is_archived=True).count(),
            'faculty': Faculty.query.filter_by(is_archived=True).count(),
            'buildings': Building.query.filter_by(is_archived=True).count(),
        }

        # Old log count (configurable retention period)
        days = request.args.get('days', 90, type=int)
        days = max(7, min(365, days))  # Clamp between 7 and 365
        cutoff_date = datetime.now() - timedelta(days=days)
        old_logs = UserActivityLog.query.filter(UserActivityLog.created_at < cutoff_date).count()

        # Current academic settings
        from app.models.settings import AcademicSettings
        active_settings = AcademicSettings.query.filter_by(is_active=True).first()
        current_semester = None
        class_schedule_count = 0
        exam_schedule_count = 0
        if active_settings:
            current_semester = {
                'academic_year': active_settings.academic_year,
                'semester': active_settings.semester,
                'exam_period': active_settings.exam_period,
                'start_date': active_settings.exam_period_start.strftime('%b %d, %Y') if active_settings.exam_period_start else None,
                'end_date': active_settings.exam_period_end.strftime('%b %d, %Y') if active_settings.exam_period_end else None,
            }

            class_schedule_count = Schedule.query.filter_by(
                semester=active_settings.semester,
                academic_year=active_settings.academic_year,
            ).count()
            exam_schedule_count = ExamSchedule.query.filter_by(
                semester=active_settings.semester,
                academic_year=active_settings.academic_year,
            ).count()

        backup_meta = {
            'latest_backup_at': None,
            'has_recent_backup': False,
        }
        try:
            backups = DatabaseBackupService.list_backups()
            if backups:
                latest_backup_at = backups[0].get('created_at')
                backup_meta['latest_backup_at'] = latest_backup_at
                if latest_backup_at:
                    parsed = datetime.fromisoformat(str(latest_backup_at).replace('Z', '+00:00'))
                    now_ref = datetime.now(parsed.tzinfo) if parsed.tzinfo else datetime.now()
                    backup_meta['has_recent_backup'] = (now_ref - parsed) <= timedelta(hours=24)
        except Exception:
            # Backup metadata should not block core stats rendering.
            pass

        action_counts = {
            'cleanup_old_logs': old_logs,
            'reset_class_schedules': class_schedule_count,
            'reset_exam_schedules': exam_schedule_count,
            'reset_all_schedules': class_schedule_count + exam_schedule_count,
            'truncate_archives': tables['archives'],
            'truncate_activity_logs': tables['activity_logs'],
            'truncate_login_history': tables['login_history'],
        }

        actions = {
            key: {
                'count': value,
                'required_phrase': _get_confirmation_phrase(key),
            }
            for key, value in action_counts.items()
        }

        return jsonify({
            'success': True,
            'tables': tables,
            'archived': archived,
            'cleanup': {
                'old_logs': old_logs,
            },
            'actions': actions,
            'backup': backup_meta,
            'current_semester': current_semester,
        })
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500


@admin_tools_bp.route('/api/database/cleanup/<cleanup_type>', methods=['POST'])
@login_required
@super_admin_required
def api_database_cleanup(cleanup_type):
    """Perform database cleanup operations"""
    try:
        data = request.get_json() or {}
        try:
            days = int(data.get('days', 90))
        except (TypeError, ValueError):
            return jsonify({'success': False, 'error': 'Retention days must be a number.'}), 400

        days = max(7, min(365, days))
        cutoff = datetime.now() - timedelta(days=days)
        dry_run = bool(data.get('dry_run', False))

        if cleanup_type != 'old_logs':
            return jsonify({'success': False, 'error': 'Invalid cleanup type'}), 400

        action_key = 'cleanup_old_logs'
        required_phrase = _get_confirmation_phrase(action_key)
        query = UserActivityLog.query.filter(UserActivityLog.created_at < cutoff)
        would_delete = query.count()

        if dry_run:
            return jsonify({
                'success': True,
                'dry_run': True,
                'action': action_key,
                'days': days,
                'would_delete': would_delete,
                'required_phrase': required_phrase,
            })

        if not _is_confirmation_phrase_valid(action_key, data.get('confirm_phrase')):
            return _confirmation_phrase_error(action_key)

        deleted = query.delete()

        db.session.commit()

        from app.utils.activity_logger import log_activity
        log_activity(action_key, 'system', details=f'Cleaned up {deleted} records (older than {days} days)')

        return jsonify({
            'success': True,
            'action': action_key,
            'days': days,
            'deleted': deleted,
            'required_phrase': required_phrase,
        })
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'error': str(e)}), 500


@admin_tools_bp.route('/api/database/health')
@login_required
@super_admin_required
def api_database_health():
    """Get database health information"""
    try:
        from sqlalchemy import text

        # MySQL server info
        version_result = db.session.execute(text("SELECT VERSION()")).scalar()
        uptime_result = db.session.execute(text("SHOW STATUS LIKE 'Uptime'")).fetchone()
        uptime_seconds = int(uptime_result[1]) if uptime_result else 0

        # Database size
        db_name = db.engine.url.database
        size_query = text("""
            SELECT 
                ROUND(SUM(data_length + index_length) / 1024 / 1024, 2) AS size_mb,
                ROUND(SUM(data_length) / 1024 / 1024, 2) AS data_mb,
                ROUND(SUM(index_length) / 1024 / 1024, 2) AS index_mb
            FROM information_schema.tables 
            WHERE table_schema = :db_name
        """)
        size_result = db.session.execute(size_query, {'db_name': db_name}).fetchone()

        # Table sizes
        table_sizes_query = text("""
            SELECT 
                table_name,
                table_rows,
                ROUND((data_length + index_length) / 1024, 2) AS size_kb,
                ROUND(data_length / 1024, 2) AS data_kb,
                ROUND(index_length / 1024, 2) AS index_kb
            FROM information_schema.tables 
            WHERE table_schema = :db_name
            ORDER BY (data_length + index_length) DESC
        """)
        table_sizes = db.session.execute(table_sizes_query, {'db_name': db_name}).fetchall()

        # Connection pool info
        pool = db.engine.pool
        pool_info = {
            'pool_size': pool.size() if hasattr(pool, 'size') else 'N/A',
            'checked_in': pool.checkedin() if hasattr(pool, 'checkedin') else 'N/A',
            'checked_out': pool.checkedout() if hasattr(pool, 'checkedout') else 'N/A',
            'overflow': pool.overflow() if hasattr(pool, 'overflow') else 'N/A',
        }

        # Format uptime
        days = uptime_seconds // 86400
        hours = (uptime_seconds % 86400) // 3600
        minutes = (uptime_seconds % 3600) // 60
        uptime_str = f"{days}d {hours}h {minutes}m"

        return jsonify({
            'success': True,
            'health': {
                'status': 'connected',
                'mysql_version': version_result,
                'uptime': uptime_str,
                'uptime_seconds': uptime_seconds,
                'database_name': db_name,
                'total_size_mb': float(size_result[0] or 0),
                'data_size_mb': float(size_result[1] or 0),
                'index_size_mb': float(size_result[2] or 0),
                'pool': pool_info,
            },
            'table_sizes': [
                {
                    'name': row[0],
                    'rows': row[1] or 0,
                    'size_kb': float(row[2] or 0),
                    'data_kb': float(row[3] or 0),
                    'index_kb': float(row[4] or 0),
                }
                for row in table_sizes
            ]
        })
    except Exception as e:
        return jsonify({
            'success': True,
            'health': {'status': 'error', 'error': str(e)},
            'table_sizes': []
        })


@admin_tools_bp.route('/api/database/backup', methods=['POST'])
@login_required
@super_admin_required
def api_database_backup():
    """Create a database backup using mysqldump"""
    try:
        backup = DatabaseBackupService.create_backup(
            source='manual',
            initiated_by_user_id=current_user.id,
        )

        return jsonify({
            'success': True,
            'backup': backup,
        })
    except BackupError as e:
        return jsonify({'success': False, 'error': str(e)}), 500
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500


@admin_tools_bp.route('/api/database/backups')
@login_required
@super_admin_required
def api_database_backups_list():
    """List available database backups"""
    try:
        backups = DatabaseBackupService.list_backups()
        return jsonify({'success': True, 'backups': backups})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500


@admin_tools_bp.route('/api/database/backups/<filename>/download')
@login_required
@super_admin_required
def api_database_backup_download(filename):
    """Download a specific backup file"""
    from flask import send_file

    try:
        filepath = DatabaseBackupService.get_backup_path(filename)

        if not filepath.exists():
            return jsonify({'success': False, 'error': 'Backup not found'}), 404

        return send_file(filepath, as_attachment=True, download_name=filename)
    except BackupError as e:
        return jsonify({'success': False, 'error': str(e)}), 400
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500


@admin_tools_bp.route('/api/database/backups/<filename>', methods=['DELETE'])
@login_required
@super_admin_required
def api_database_backup_delete(filename):
    """Delete a specific backup file"""
    try:
        DatabaseBackupService.delete_backup(
            filename=filename,
            initiated_by_user_id=current_user.id,
            source='manual',
        )

        return jsonify({'success': True})
    except BackupError as e:
        status_code = 404 if str(e) == 'Backup not found' else 400
        return jsonify({'success': False, 'error': str(e)}), status_code
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500


@admin_tools_bp.route('/api/database/auto-backup-settings', methods=['GET', 'POST'])
@login_required
@super_admin_required
def api_database_auto_backup_settings():
    """Get or update automatic backup settings."""
    if request.method == 'GET':
        try:
            settings = DatabaseBackupService.get_auto_backup_settings()
            return jsonify({'success': True, 'settings': settings})
        except Exception as e:
            return jsonify({'success': False, 'error': str(e)}), 500

    try:
        data = request.get_json() or {}
        enabled = bool(data.get('enabled', False))
        retention_count = data.get('retention_count', DatabaseBackupService.DEFAULT_RETENTION_COUNT)

        try:
            retention_count = int(retention_count)
        except (TypeError, ValueError):
            return jsonify({'success': False, 'error': 'Retention count must be a number.'}), 400

        if retention_count < 1 or retention_count > 365:
            return jsonify({'success': False, 'error': 'Retention count must be between 1 and 365.'}), 400

        settings = DatabaseBackupService.update_auto_backup_settings(
            enabled=enabled,
            retention_count=retention_count,
            user_id=current_user.id,
        )

        return jsonify({'success': True, 'settings': settings})
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'error': str(e)}), 500


@admin_tools_bp.route('/api/database/truncate/<table_name>', methods=['POST'])
@login_required
@super_admin_required
def api_database_truncate(table_name):
    """Truncate a specific safe table"""
    try:
        data = request.get_json() or {}
        dry_run = bool(data.get('dry_run', False))

        safe_tables = {
            'archives': Archive,
            'activity_logs': UserActivityLog,
            'login_history': LoginHistory,
        }

        if table_name not in safe_tables:
            return jsonify({'success': False, 'error': f'Table "{table_name}" cannot be truncated for safety reasons.'}), 400

        model = safe_tables[table_name]
        action_key = f'truncate_{table_name}'
        required_phrase = _get_confirmation_phrase(action_key)
        count = model.query.count()

        if dry_run:
            return jsonify({
                'success': True,
                'dry_run': True,
                'action': action_key,
                'table': table_name,
                'would_delete': count,
                'required_phrase': required_phrase,
            })

        if not _is_confirmation_phrase_valid(action_key, data.get('confirm_phrase')):
            return _confirmation_phrase_error(action_key)

        model.query.delete()
        db.session.commit()

        from app.utils.activity_logger import log_activity
        log_activity(action_key, 'system', details=f'Truncated table: {table_name} ({count} records)')

        return jsonify({
            'success': True,
            'action': action_key,
            'table': table_name,
            'deleted': count,
            'required_phrase': required_phrase,
        })
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'error': str(e)}), 500


@admin_tools_bp.route('/api/database/reset-schedules', methods=['POST'])
@login_required
@super_admin_required
def api_database_reset_schedules():
    """Reset all schedules for the current academic settings"""
    try:
        from app.models.settings import AcademicSettings
        settings = AcademicSettings.query.filter_by(is_active=True).first()
        if not settings:
            return jsonify({'success': False, 'error': 'No active academic settings found'}), 400

        data = request.get_json() or {}
        reset_type = str(data.get('type', 'all')).strip().lower()
        dry_run = bool(data.get('dry_run', False))

        action_by_type = {
            'class': 'reset_class_schedules',
            'exam': 'reset_exam_schedules',
            'all': 'reset_all_schedules',
        }
        if reset_type not in action_by_type:
            return jsonify({'success': False, 'error': 'Invalid reset type. Use class, exam, or all.'}), 400

        action_key = action_by_type[reset_type]
        required_phrase = _get_confirmation_phrase(action_key)

        class_query = Schedule.query.filter_by(
            semester=settings.semester,
            academic_year=settings.academic_year,
        )
        exam_query = ExamSchedule.query.filter_by(
            semester=settings.semester,
            academic_year=settings.academic_year,
        )

        would_delete_class = class_query.count() if reset_type in ('class', 'all') else 0
        would_delete_exam = exam_query.count() if reset_type in ('exam', 'all') else 0

        if dry_run:
            return jsonify({
                'success': True,
                'dry_run': True,
                'action': action_key,
                'type': reset_type,
                'term': {
                    'semester': settings.semester,
                    'academic_year': settings.academic_year,
                    'exam_period': settings.exam_period,
                },
                'would_delete_class': would_delete_class,
                'would_delete_exam': would_delete_exam,
                'total_would_delete': would_delete_class + would_delete_exam,
                'required_phrase': required_phrase,
            })

        if not _is_confirmation_phrase_valid(action_key, data.get('confirm_phrase')):
            return _confirmation_phrase_error(action_key)

        deleted_class = 0
        deleted_exam = 0

        if reset_type in ('class', 'all'):
            deleted_class = class_query.delete()

        if reset_type in ('exam', 'all'):
            deleted_exam = exam_query.delete()

        db.session.commit()

        from app.utils.activity_logger import log_activity
        log_activity(action_key, 'system',
                     details=f'Reset schedules for {settings.academic_year} {settings.semester}: '
                             f'{deleted_class} class + {deleted_exam} exam schedules deleted (type: {reset_type})')

        return jsonify({
            'success': True,
            'action': action_key,
            'type': reset_type,
            'deleted_class': deleted_class,
            'deleted_exam': deleted_exam,
            'total_deleted': deleted_class + deleted_exam,
            'semester': settings.semester,
            'academic_year': settings.academic_year,
            'required_phrase': required_phrase,
        })
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'error': str(e)}), 500


# ============================================================
# 5. SECURITY & ACCESS CONTROL
# ============================================================

@admin_tools_bp.route('/security')
@login_required
@super_admin_required
def security():
    """Security and access control page (hidden by policy)."""
    # Hide-only strategy: keep backend APIs for rollback/support,
    # but make the page itself inaccessible from UI and direct URL.
    abort(404)


@admin_tools_bp.route('/api/security/overview')
@login_required
@super_admin_required
def api_security_overview():
    """Get security overview stats for the dashboard"""
    try:
        include_events_param = request.args.get('include_events', '1').strip().lower()
        include_events = include_events_param not in {'0', 'false', 'no', 'off'}

        now = datetime.now()
        today_start = now.replace(hour=0, minute=0, second=0, microsecond=0)
        week_ago = now - timedelta(days=7)

        # Active sessions
        active_sessions = LoginHistory.query.filter_by(is_active=True).count()

        # Today's logins
        todays_logins = LoginHistory.query.filter(LoginHistory.login_at >= today_start).count()

        # User stats
        total_users = User.query.filter_by(is_archived=False).count()
        active_users = User.query.filter_by(is_active=True, is_archived=False).count()
        inactive_users = User.query.filter_by(is_active=False, is_archived=False).count()
        pending_password_reset = User.query.filter_by(needs_password_change=True, is_archived=False).count()
        unverified_users = User.query.filter_by(email_verified=False, is_archived=False).count()

        # Role breakdown
        super_admins = User.query.filter_by(role='super_admin', is_archived=False).count()
        admins = User.query.filter_by(role='admin', is_archived=False).count()
        deans = User.query.filter_by(role='dean', is_archived=False).count()

        # Last 7 days login activity (per day)
        login_trend = []
        max_count = 0
        for i in range(6, -1, -1):
            day = (now - timedelta(days=i)).replace(hour=0, minute=0, second=0, microsecond=0)
            day_end = day + timedelta(days=1)
            count = LoginHistory.query.filter(
                LoginHistory.login_at >= day,
                LoginHistory.login_at < day_end
            ).count()
            if count > max_count:
                max_count = count
            login_trend.append({
                'date': day.strftime('%b %d'),
                'day': day.strftime('%a'),
                'count': count,
            })

        recent_events_payload = []
        if include_events:
            # Recent security events (last 7 days)
            security_actions = [
                'force_logout', 'force_logout_all', 'force_password_reset',
                'password_changed', 'cleanup_old_sessions'
            ]
            recent_events = UserActivityLog.query.filter(
                UserActivityLog.action.in_(security_actions),
                UserActivityLog.created_at >= week_ago
            ).order_by(UserActivityLog.created_at.desc()).limit(15).all()
            recent_events_payload = [
                {
                    'action': e.action,
                    'actor_role': e.user.role if e.user else 'system',
                    'entity_name': e.entity_name,
                    'details': e.details,
                    'created_at': (e.created_at.isoformat() + 'Z') if e.created_at else None,
                }
                for e in recent_events
            ]

        return jsonify({
            'success': True,
            'overview': {
                'active_sessions': active_sessions,
                'session_logout_policy': 'browser_close',
                'todays_logins': todays_logins,
                'total_users': total_users,
                'active_users': active_users,
                'inactive_users': inactive_users,
                'pending_password_reset': pending_password_reset,
                'unverified_users': unverified_users,
                'roles': {
                    'super_admin': super_admins,
                    'admin': admins,
                    'dean': deans,
                },
                'login_trend': login_trend,
                'max_trend': max_count,
                'recent_events': recent_events_payload,
            }
        })
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500


@admin_tools_bp.route('/api/sessions')
@login_required
@super_admin_required
def api_sessions():
    """Get active sessions"""
    try:
        sessions = LoginHistory.get_active_sessions()
        now = datetime.utcnow()
        sessions_changed = False

        # Keep active session list accurate by ending sessions tied to inactive users.
        for session in sessions:
            user = session.user
            if not user or not getattr(user, 'is_active', True) or getattr(user, 'is_archived', False):
                session.is_active = False
                session.logout_at = session.logout_at or now
                if user:
                    user.force_logout_at = now
                sessions_changed = True

        if sessions_changed:
            db.session.commit()
            sessions = LoginHistory.get_active_sessions()

        return jsonify({
            'success': True,
            'sessions': [s.to_admin_dict() for s in sessions]
        })
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500


@admin_tools_bp.route('/api/sessions/<int:id>/force-logout', methods=['POST'])
@login_required
@super_admin_required
def api_force_logout(id):
    """Force logout a specific session"""
    try:
        session_entry = LoginHistory.query.get_or_404(id)

        if session_entry.user_id == current_user.id:
            return jsonify({
                'success': False,
                'error': 'You cannot force logout your own account from this action.'
            }), 400

        if LoginHistory.force_logout(id):
            db.session.commit()

            from app.utils.activity_logger import log_activity
            log_activity('force_logout', 'user', entity_id=session_entry.user_id,
                        entity_name=session_entry.user.full_name if session_entry.user else 'Unknown',
                        details=f'Force logged out session from {session_entry.ip_address}')

            return jsonify({'success': True})
        return jsonify({'success': False, 'error': 'Session not found or already logged out'}), 404
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'error': str(e)}), 500


@admin_tools_bp.route('/api/sessions/<int:id>/force-password-reset', methods=['POST'])
@login_required
@super_admin_required
def api_force_password_reset_by_session(id):
    """Force a user to change their password on next login via a session row."""
    try:
        session_entry = LoginHistory.query.get_or_404(id)
        user = session_entry.user
        if not user:
            return jsonify({'success': False, 'error': 'Associated user not found'}), 404

        user.needs_password_change = True
        db.session.commit()

        from app.utils.activity_logger import log_activity
        log_activity(
            'force_password_reset',
            'user',
            entity_id=user.id,
            entity_name=user.full_name,
            details='Forced password reset on next login via active session control'
        )

        return jsonify({'success': True})
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'error': str(e)}), 500


@admin_tools_bp.route('/api/sessions/force-logout-all', methods=['POST'])
@login_required
@super_admin_required
def api_force_logout_all():
    """Force logout all sessions except current user"""
    try:
        count = LoginHistory.force_logout_all(except_user_id=current_user.id)
        db.session.commit()

        from app.utils.activity_logger import log_activity
        log_activity('force_logout_all', 'system', details=f'Force logged out {count} sessions')

        return jsonify({'success': True, 'count': count})
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'error': str(e)}), 500


@admin_tools_bp.route('/api/login-history')
@login_required
@super_admin_required
def api_login_history():
    """Get paginated login history"""
    try:
        page = request.args.get('page', 1, type=int)
        per_page = request.args.get('per_page', 25, type=int)
        user_id = request.args.get('user_id', type=int)
        date_from = request.args.get('date_from', '')
        date_to = request.args.get('date_to', '')

        query = LoginHistory.query.options(db.joinedload(LoginHistory.user))

        if user_id:
            query = query.filter(LoginHistory.user_id == user_id)
        if date_from:
            try:
                query = query.filter(LoginHistory.login_at >= datetime.strptime(date_from, '%Y-%m-%d'))
            except ValueError:
                pass
        if date_to:
            try:
                query = query.filter(LoginHistory.login_at < datetime.strptime(date_to, '%Y-%m-%d') + timedelta(days=1))
            except ValueError:
                pass

        total = query.count()
        entries = query.order_by(LoginHistory.login_at.desc())\
            .offset((page - 1) * per_page).limit(per_page).all()

        return jsonify({
            'success': True,
            'entries': [e.to_admin_dict() for e in entries],
            'total': total,
            'page': page,
            'per_page': per_page,
            'total_pages': (total + per_page - 1) // per_page,
        })
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500


@admin_tools_bp.route('/api/login-history/export')
@login_required
@super_admin_required
def api_login_history_export():
    """Export login history to professional Excel (.xlsx)"""
    try:
        import io
        from flask import send_file
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
        from openpyxl.worksheet.table import Table, TableStyleInfo
        from app.services.export_service import get_institution_name, create_posting_style_excel_header

        user_id = request.args.get('user_id', type=int)
        date_from = request.args.get('date_from', '')
        date_to = request.args.get('date_to', '')

        query = LoginHistory.query.options(db.joinedload(LoginHistory.user))

        if user_id:
            query = query.filter(LoginHistory.user_id == user_id)
        if date_from:
            try:
                query = query.filter(LoginHistory.login_at >= datetime.strptime(date_from, '%Y-%m-%d'))
            except ValueError:
                pass
        if date_to:
            try:
                query = query.filter(LoginHistory.login_at < datetime.strptime(date_to, '%Y-%m-%d') + timedelta(days=1))
            except ValueError:
                pass

        entries = query.order_by(LoginHistory.login_at.desc()).limit(10000).all()

        # --- Build professional Excel workbook ---
        wb = Workbook()
        ws = wb.active
        ws.title = 'Login History'

        # Styles
        border = Border(
            left=Side(style='thin', color='DEE2E6'),
            right=Side(style='thin', color='DEE2E6'),
            top=Side(style='thin', color='DEE2E6'),
            bottom=Side(style='thin', color='DEE2E6'),
        )
        header_font = Font(name='Arial', size=9, bold=True, color='FFFFFF')
        header_fill = PatternFill(start_color='495057', end_color='495057', fill_type='solid')
        header_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
        data_font = Font(name='Arial', size=9)
        data_align = Alignment(horizontal='left', vertical='center', wrap_text=True)
        center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
        zebra_fill = PatternFill(start_color='F5F6F8', end_color='F5F6F8', fill_type='solid')

        # Status color mapping
        status_fills = {
            'active': PatternFill(start_color='D4EDDA', end_color='D4EDDA', fill_type='solid'),
            'ended': PatternFill(start_color='E2E3E5', end_color='E2E3E5', fill_type='solid'),
        }

        # Institutional header (dual-logo posting format)
        current_row = create_posting_style_excel_header(
            ws, report_title='LOGIN HISTORY REPORT',
            office_name='OFFICE OF THE SYSTEM ADMINISTRATOR',
            subtitle=f'Generated: {datetime.utcnow().strftime("%B %d, %Y  %I:%M %p")} UTC',
            last_col='F'
        )

        # Filter summary row
        filters_applied = []
        if user_id:
            filters_applied.append(f'User ID: {user_id}')
        if date_from:
            filters_applied.append(f'From: {date_from}')
        if date_to:
            filters_applied.append(f'To: {date_to}')

        if filters_applied:
            ws.merge_cells(f'A{current_row}:F{current_row}')
            fc = ws.cell(row=current_row, column=1, value=f'Filters: {" | ".join(filters_applied)}')
            fc.font = Font(name='Arial', size=8, italic=True, color='6C757D')
            fc.alignment = Alignment(horizontal='center')
            current_row += 1

        # Total records info
        ws.merge_cells(f'A{current_row}:F{current_row}')
        tc = ws.cell(row=current_row, column=1, value=f'Total Records: {len(entries)}')
        tc.font = Font(name='Arial', size=9, bold=True, color='1F4788')
        tc.alignment = Alignment(horizontal='left')
        current_row += 1

        # Section banner
        banner_fill = PatternFill(start_color='1F4788', end_color='1F4788', fill_type='solid')
        ws.merge_cells(f'A{current_row}:F{current_row}')
        bc = ws.cell(row=current_row, column=1, value='SESSION LOG ENTRIES')
        bc.font = Font(name='Arial', size=10, bold=True, color='FFFFFF')
        bc.fill = banner_fill
        bc.alignment = Alignment(horizontal='center', vertical='center')
        bc.border = border
        ws.row_dimensions[current_row].height = 24
        current_row += 1

        # Column headers
        columns = [
            ('#', 5),
            ('Role', 16),
            ('Login Time', 20),
            ('Logout Time', 20),
            ('Duration', 14),
            ('Status', 12),
        ]
        header_row = current_row  # Save for auto-filter/freeze
        for ci, (col_name, col_width) in enumerate(columns, 1):
            cell = ws.cell(row=current_row, column=ci, value=col_name)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_align
            cell.border = border
            ws.column_dimensions[get_column_letter(ci)].width = col_width
        current_row += 1

        # Data rows
        for idx, e in enumerate(entries, 1):
            duration = ''
            if e.login_at and e.logout_at:
                diff = e.logout_at - e.login_at
                hours, remainder = divmod(int(diff.total_seconds()), 3600)
                minutes, seconds = divmod(remainder, 60)
                duration = f'{hours}h {minutes}m {seconds}s'
            status = 'Active' if e.is_active else 'Ended'

            row_data = [
                idx,
                (e.user.role if e.user else 'unknown').replace('_', ' ').title(),
                e.login_at.strftime('%Y-%m-%d %H:%M:%S') if e.login_at else '',
                e.logout_at.strftime('%Y-%m-%d %H:%M:%S') if e.logout_at else '',
                duration,
                status,
            ]
            for ci, val in enumerate(row_data, 1):
                cell = ws.cell(row=current_row, column=ci, value=val)
                cell.font = data_font
                cell.alignment = center_align
                cell.border = border
                # Zebra striping
                if idx % 2 == 0:
                    cell.fill = zebra_fill
            # Color-code the status cell
            status_key = status.lower()
            if status_key in status_fills:
                ws.cell(row=current_row, column=6).fill = status_fills[status_key]
            current_row += 1

        # Excel Table with auto-filter (provides fully functional filter dropdowns)
        last_data_row = max(current_row - 1, header_row)  # Handle empty data
        table_ref = f'A{header_row}:F{last_data_row}'
        tab = Table(displayName='LoginHistory', ref=table_ref)
        tab.tableStyleInfo = TableStyleInfo(
            name='TableStyleLight1', showFirstColumn=False,
            showLastColumn=False, showRowStripes=False, showColumnStripes=False
        )
        ws.add_table(tab)

        # Freeze panes (freeze below column header row)
        ws.freeze_panes = f'A{header_row + 1}'

        # Footer
        current_row += 1
        ws.merge_cells(f'A{current_row}:F{current_row}')
        fc = ws.cell(row=current_row, column=1, value=f'Report generated by iSchedWise Admin Tools — {datetime.utcnow().strftime("%Y-%m-%d %H:%M:%S")} UTC')
        fc.font = Font(name='Arial', size=8, italic=True, color='6C757D')
        fc.alignment = Alignment(horizontal='center')

        # Save to buffer
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        filename = f'login_history_{datetime.utcnow().strftime("%Y%m%d_%H%M%S")}.xlsx'
        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=filename
        )
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500


@admin_tools_bp.route('/api/maintenance/toggle', methods=['POST'])
@login_required
@super_admin_required
def api_maintenance_toggle():
    """Toggle maintenance mode on/off"""
    try:
        data = request.get_json() or {}
        current_mode = SystemConfig.get('maintenance_mode', False)
        new_mode = not current_mode

        SystemConfig.set('maintenance_mode', new_mode, user_id=current_user.id)

        # Always save the message (empty string clears it)
        message = data.get('message', '')
        SystemConfig.set('maintenance_message', message, user_id=current_user.id)

        db.session.commit()

        from app.utils.activity_logger import log_activity
        action = 'enabled' if new_mode else 'disabled'
        log_activity(f'maintenance_{action}', 'system',
                    details=f'Maintenance mode {action}' + (f' with message: {message}' if message else ''))

        return jsonify({
            'success': True,
            'maintenance_mode': new_mode,
            'message': f'Maintenance mode {action}'
        })
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'error': str(e)}), 500


@admin_tools_bp.route('/api/users/<int:id>/force-password-reset', methods=['POST'])
@login_required
@super_admin_required
def api_force_password_reset(id):
    """Force a user to change their password on next login"""
    try:
        user = User.query.get_or_404(id)
        user.needs_password_change = True
        db.session.commit()

        from app.utils.activity_logger import log_activity
        log_activity('force_password_reset', 'user', entity_id=user.id,
                    entity_name=user.full_name, details='Forced password reset on next login')

        return jsonify({'success': True})
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'error': str(e)}), 500


# ============================================================
# 6. SYSTEM CONFIGURATION
# ============================================================

ALLOWED_CONFIG_CATEGORIES = {'general', 'security', 'email', 'ai'}


# System Configuration surface intentionally disabled (hide/disable-only mode).
# Keep handlers in place for easy future re-enable.
# @admin_tools_bp.route('/configuration')
# @login_required
# @super_admin_required
def configuration():
    """System configuration page"""
    return render_template('admin/configuration.html')


# @admin_tools_bp.route('/api/config/initialize-defaults', methods=['POST'])
# @login_required
# @super_admin_required
def api_config_initialize_defaults():
    """Seed default config entries (only inserts keys that do not already exist)."""
    try:
        defaults = [
            {'config_key': 'maintenance_mode', 'config_value': 'false', 'config_type': 'boolean', 'category': 'general', 'description': 'Enable maintenance mode to block non-admin access'},
            {'config_key': 'maintenance_message', 'config_value': '', 'config_type': 'string', 'category': 'general', 'description': 'Message displayed to users during maintenance'},
            {'config_key': 'app_name', 'config_value': 'iSchedWise', 'config_type': 'string', 'category': 'general', 'description': 'Application display name'},
            {'config_key': 'timezone', 'config_value': 'Asia/Manila', 'config_type': 'string', 'category': 'general', 'description': 'System timezone for date/time display'},
            {'config_key': 'session_logout_policy', 'config_value': 'browser_close', 'config_type': 'string', 'category': 'security', 'description': 'Session ends when the browser is closed'},
            {'config_key': 'max_login_attempts', 'config_value': '5', 'config_type': 'integer', 'category': 'security', 'description': 'Max failed login attempts before lockout'},
            {'config_key': 'password_min_length', 'config_value': '8', 'config_type': 'integer', 'category': 'security', 'description': 'Minimum password length for new passwords'},
            {'config_key': 'force_password_change_days', 'config_value': '0', 'config_type': 'integer', 'category': 'security', 'description': 'Force password change after N days (0 = disabled)'},
            {'config_key': 'smtp_server', 'config_value': '', 'config_type': 'string', 'category': 'email', 'description': 'SMTP server hostname (e.g. smtp.gmail.com)'},
            {'config_key': 'smtp_port', 'config_value': '587', 'config_type': 'integer', 'category': 'email', 'description': 'SMTP server port (587 for TLS, 465 for SSL)'},
            {'config_key': 'smtp_username', 'config_value': '', 'config_type': 'string', 'category': 'email', 'description': 'SMTP authentication username'},
            {'config_key': 'smtp_password', 'config_value': '', 'config_type': 'string', 'category': 'email', 'description': 'SMTP authentication password'},
            {'config_key': 'smtp_use_tls', 'config_value': 'true', 'config_type': 'boolean', 'category': 'email', 'description': 'Use TLS encryption for SMTP connection'},
            {'config_key': 'smtp_sender_email', 'config_value': '', 'config_type': 'string', 'category': 'email', 'description': 'Default sender email address'},
            {'config_key': 'smtp_sender_name', 'config_value': 'iSchedWise', 'config_type': 'string', 'category': 'email', 'description': 'Default sender display name'},
            {'config_key': 'ai_enabled', 'config_value': 'false', 'config_type': 'boolean', 'category': 'ai', 'description': 'Enable AI-powered scheduling suggestions'},
            {'config_key': 'gemini_api_key', 'config_value': '', 'config_type': 'string', 'category': 'ai', 'description': 'Google Gemini API key for AI features'},
            {'config_key': 'ai_model', 'config_value': 'gemini-2.0-flash', 'config_type': 'string', 'category': 'ai', 'description': 'AI model name to use for generation'},
            {'config_key': 'ai_max_suggestions', 'config_value': '5', 'config_type': 'integer', 'category': 'ai', 'description': 'Maximum AI suggestions to return per request'},
        ]

        created = 0
        existing_keys = {c.config_key for c in SystemConfig.query.all()}

        for default in defaults:
            if default['config_key'] in existing_keys:
                continue
            config = SystemConfig(
                config_key=default['config_key'],
                config_value=default['config_value'],
                config_type=default['config_type'],
                category=default['category'],
                description=default['description'],
                updated_by=current_user.id,
                updated_at=datetime.utcnow()
            )
            db.session.add(config)
            created += 1

        db.session.commit()

        if created > 0:
            from app.utils.activity_logger import log_activity
            log_activity('initialize_config_defaults', 'system_config', details=f'Initialized {created} default configuration entries')

        return jsonify({
            'success': True,
            'created': created,
            'message': f'{created} default entries created' if created > 0 else 'All defaults already exist'
        })
    except Exception:
        db.session.rollback()
        return jsonify({'success': False, 'error': 'Failed to initialize configuration defaults.'}), 500


# @admin_tools_bp.route('/api/config/<category>')
# @login_required
# @super_admin_required
def api_config_get(category):
    """Get config values for a category."""
    if category not in ALLOWED_CONFIG_CATEGORIES:
        return jsonify({'success': False, 'error': 'Invalid configuration category.'}), 400

    try:
        configs = SystemConfig.get_by_category(category)
        return jsonify({'success': True, 'configs': {k: v.to_dict() for k, v in configs.items()}})
    except Exception:
        return jsonify({'success': False, 'error': 'Failed to load configuration values.'}), 500


# @admin_tools_bp.route('/api/config', methods=['PUT'])
# @login_required
# @super_admin_required
def api_config_update():
    """Update configuration values with simple server-side guardrails."""
    try:
        data = request.get_json(silent=True) or {}
        updates = data.get('updates', {})
        if not isinstance(updates, dict) or not updates:
            return jsonify({'success': False, 'error': 'No configuration updates provided.'}), 400

        validators = {
            'max_login_attempts': lambda v: 1 <= int(v) <= 20,
            'password_min_length': lambda v: 6 <= int(v) <= 128,
            'force_password_change_days': lambda v: 0 <= int(v) <= 365,
            'smtp_port': lambda v: 1 <= int(v) <= 65535,
            'ai_max_suggestions': lambda v: 1 <= int(v) <= 20,
        }

        for key, value in updates.items():
            if key in validators:
                try:
                    if not validators[key](value):
                        return jsonify({'success': False, 'error': f'Invalid value for {key.replace("_", " ")}.'}), 400
                except (TypeError, ValueError):
                    return jsonify({'success': False, 'error': f'Invalid value for {key.replace("_", " ")}.'}), 400

            SystemConfig.set(key, value, user_id=current_user.id)

        db.session.commit()

        from app.utils.activity_logger import log_settings_change
        keys_changed = ', '.join(sorted(updates.keys()))
        log_settings_change('system_config', details=f'Updated config: {keys_changed}')

        return jsonify({'success': True})
    except Exception:
        db.session.rollback()
        return jsonify({'success': False, 'error': 'Failed to update system configuration.'}), 500


# @admin_tools_bp.route('/api/config/test-email', methods=['POST'])
# @login_required
# @super_admin_required
def api_config_test_email():
    """Send a test email to verify SMTP settings."""
    try:
        from flask_mail import Message
        from app.extensions import mail

        institution_name = 'Norzagaray College'
        app_brand_name = 'iSchedWise'
        try:
            settings = InstitutionSettings.query.first()
            if settings and getattr(settings, 'institution_name', None):
                institution_name = settings.institution_name.strip() or institution_name
            if settings and getattr(settings, 'system_name', None):
                app_brand_name = settings.system_name.strip() or app_brand_name
        except Exception:
            pass

        payload = build_smtp_test_email_payload(
            full_name=current_user.full_name,
            recipient_email=current_user.email,
            institution_name=institution_name,
            app_brand_name=app_brand_name,
            sent_by=current_user.full_name,
            sent_at_utc_label=datetime.now(timezone.utc).strftime('%Y-%m-%d %H:%M:%S UTC'),
        )

        msg = Message(
            subject=payload['subject'],
            recipients=[current_user.email],
            sender=build_branded_mail_sender(
                default_sender=current_app.config.get('MAIL_DEFAULT_SENDER'),
                app_brand_name=app_brand_name,
            ),
        )
        msg.body = payload['text_body']
        msg.html = payload['html_body']
        mail.send(msg)
        return jsonify({'success': True, 'message': f'Test email sent to {current_user.email}'})
    except Exception:
        return jsonify({'success': False, 'error': 'Could not send test email. Check SMTP settings and try again.'}), 500


# @admin_tools_bp.route('/api/config/test-ai', methods=['POST'])
# @login_required
# @super_admin_required
def api_config_test_ai():
    """Test AI (Gemini) API connection."""
    try:
        api_key = SystemConfig.get('gemini_api_key') or os.environ.get('GEMINI_API_KEY', '')
        if not api_key:
            return jsonify({'success': False, 'error': 'No Gemini API key configured.'}), 400

        import google.generativeai as genai
        model_name = (SystemConfig.get('ai_model', 'gemini-2.0-flash') or '').strip() or 'gemini-2.0-flash'
        genai.configure(api_key=api_key)
        model = genai.GenerativeModel(model_name)
        response = model.generate_content('Say "Hello from iSchedWise!" in one sentence.')

        response_text = ''
        try:
            response_text = (getattr(response, 'text', '') or '').strip()
        except Exception:
            response_text = ''

        if not response_text:
            response_text = 'AI connection successful.'

        return jsonify({'success': True, 'message': f'AI responded: {response_text[:200]}'})
    except Exception:
        return jsonify({'success': False, 'error': 'AI connection test failed. Verify API key and model settings.'}), 500
