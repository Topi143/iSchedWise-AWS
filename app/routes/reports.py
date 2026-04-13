"""
Reports Routes
Handles report generation and statistics
"""
from flask import Blueprint, render_template, request, jsonify, send_file
from flask_login import login_required, current_user
from sqlalchemy import func, and_, or_
from sqlalchemy.orm import joinedload
from datetime import datetime, time, timedelta
from app.extensions import db
from app.models.schedule import Schedule
from app.models.exam_schedule import ExamSchedule
from app.models.faculty import Faculty, FacultySubjectAssignment, FacultyAvailability
from app.models.curriculum import Subject
from app.models.building import Room, Building
from app.models.program import Program
from app.models.section import Section
from app.models.settings import AcademicSettings
from app.models.activity_log import UserActivityLog
from app.models.user import User
from app.decorators import role_required
from app.ai_scheduler import ai_scheduler
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, PieChart, Reference
from openpyxl.worksheet.table import Table as XLTable, TableStyleInfo
from reportlab.lib.pagesizes import letter, landscape, A4
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak, Image as RLImage
from reportlab.lib import colors as rl_colors
from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT
from reportlab.graphics.shapes import Drawing
from reportlab.graphics.charts.barcharts import VerticalBarChart, HorizontalBarChart
from reportlab.graphics.charts.piecharts import Pie
from reportlab.graphics import renderPDF
from reportlab.pdfgen import canvas as pdf_canvas
import io
import os

reports_bp = Blueprint('reports', __name__, url_prefix='/reports')


@reports_bp.route('/')
@login_required
def index():
    """Display reports dashboard with statistics"""
    # Get current academic settings
    current_settings = AcademicSettings.query.filter_by(is_active=True).first()
    
    if current_settings:
        academic_year = current_settings.academic_year
        semester = current_settings.semester
        exam_period = current_settings.exam_period
    else:
        academic_year = None
        semester = None
        exam_period = None
    
    # Get user's program access
    user_program_ids = current_user.get_program_ids()
    
    # Get filter parameters from request
    filter_department = request.args.get('program', type=int)
    
    # Auto-set filter for single-program users (non-admin)
    if not filter_department and user_program_ids is not None and len(user_program_ids) == 1:
        filter_department = user_program_ids[0]
    
    # Calculate statistics with filters - overview needs everything
    stats = calculate_statistics(
        academic_year, 
        semester, 
        user_program_ids,
        filter_department,
        include={'counts', 'faculty', 'rooms', 'weekly', 'completion'}
    )
    
    # Get programs for filter
    if user_program_ids is None:
        programs = Program.query.filter_by(is_active=True).order_by(Program.program_code).all()
    else:
        programs = Program.query.filter(
            Program.is_active == True,
            Program.id.in_(user_program_ids)
        ).order_by(Program.program_code).all()
    
    # Generate rule-based insight banner (no AI call)
    ai_insights_banner = generate_report_insights_banner(stats)
    
    return render_template(
        'reports/overview.html',
        stats=stats,
        academic_year=academic_year,
        semester=semester,
        exam_period=exam_period,
        programs=programs,
        filter_department=filter_department,
        ai_insights_banner=ai_insights_banner,
        operation_days=AcademicSettings.get_active_operation_days()
    )


@reports_bp.route('/faculty')
@login_required
def faculty_report():
    """Display faculty workload report"""
    current_settings = AcademicSettings.query.filter_by(is_active=True).first()
    academic_year = current_settings.academic_year if current_settings else None
    semester = current_settings.semester if current_settings else None

    user_program_ids = current_user.get_program_ids()
    filter_department = request.args.get('program', type=int)
    if not filter_department and user_program_ids is not None and len(user_program_ids) == 1:
        filter_department = user_program_ids[0]

    stats = calculate_statistics(academic_year, semester, user_program_ids, filter_department,
                                  include={'counts', 'faculty'})

    if user_program_ids is None:
        programs = Program.query.filter_by(is_active=True).order_by(Program.program_code).all()
    else:
        programs = Program.query.filter(
            Program.is_active == True,
            Program.id.in_(user_program_ids)
        ).order_by(Program.program_code).all()

    return render_template(
        'reports/faculty.html',
        stats=stats,
        academic_year=academic_year,
        semester=semester,
        programs=programs,
        filter_department=filter_department
    )


@reports_bp.route('/rooms')
@login_required
def rooms_report():
    """Display room utilization report"""
    current_settings = AcademicSettings.query.filter_by(is_active=True).first()
    academic_year = current_settings.academic_year if current_settings else None
    semester = current_settings.semester if current_settings else None

    user_program_ids = current_user.get_program_ids()
    filter_department = request.args.get('program', type=int)
    if not filter_department and user_program_ids is not None and len(user_program_ids) == 1:
        filter_department = user_program_ids[0]

    stats = calculate_statistics(academic_year, semester, user_program_ids, filter_department,
                                  include={'counts', 'rooms'})

    if user_program_ids is None:
        programs = Program.query.filter_by(is_active=True).order_by(Program.program_code).all()
    else:
        programs = Program.query.filter(
            Program.is_active == True,
            Program.id.in_(user_program_ids)
        ).order_by(Program.program_code).all()

    return render_template(
        'reports/rooms.html',
        stats=stats,
        academic_year=academic_year,
        semester=semester,
        programs=programs,
        filter_department=filter_department
    )


@reports_bp.route('/weekly')
@login_required
def weekly_report():
    """Display weekly distribution report"""
    current_settings = AcademicSettings.query.filter_by(is_active=True).first()
    academic_year = current_settings.academic_year if current_settings else None
    semester = current_settings.semester if current_settings else None

    user_program_ids = current_user.get_program_ids()
    filter_department = request.args.get('program', type=int)
    if not filter_department and user_program_ids is not None and len(user_program_ids) == 1:
        filter_department = user_program_ids[0]

    stats = calculate_statistics(academic_year, semester, user_program_ids, filter_department,
                                  include={'counts', 'weekly'})

    if user_program_ids is None:
        programs = Program.query.filter_by(is_active=True).order_by(Program.program_code).all()
    else:
        programs = Program.query.filter(
            Program.is_active == True,
            Program.id.in_(user_program_ids)
        ).order_by(Program.program_code).all()

    return render_template(
        'reports/weekly.html',
        stats=stats,
        academic_year=academic_year,
        semester=semester,
        programs=programs,
        filter_department=filter_department,
        operation_days=AcademicSettings.get_active_operation_days()
    )


@reports_bp.route('/compare')
@login_required
def compare_report():
    """Display semester comparison report"""
    current_settings = AcademicSettings.query.filter_by(is_active=True).first()
    academic_year = current_settings.academic_year if current_settings else None
    semester = current_settings.semester if current_settings else None

    user_program_ids = current_user.get_program_ids()
    filter_department = request.args.get('program', type=int)
    if not filter_department and user_program_ids is not None and len(user_program_ids) == 1:
        filter_department = user_program_ids[0]

    if user_program_ids is None:
        programs = Program.query.filter_by(is_active=True).order_by(Program.program_code).all()
    else:
        programs = Program.query.filter(
            Program.is_active == True,
            Program.id.in_(user_program_ids)
        ).order_by(Program.program_code).all()

    return render_template(
        'reports/compare.html',
        academic_year=academic_year,
        semester=semester,
        programs=programs,
        filter_department=filter_department,
        operation_days=AcademicSettings.get_active_operation_days()
    )


@reports_bp.route('/activity')
@login_required
@role_required('admin', 'super_admin')
def activity_report():
    """Display user activity logs (admin only)"""
    return render_template('reports/activity.html')


def _parse_user_activity_filters(args):
    """Parse and normalize activity-log filters from request args."""
    filter_user = args.get('user_id', type=int)
    filter_action = (args.get('action', type=str) or '').strip() or None
    filter_entity = (args.get('entity_type', type=str) or '').strip() or None
    filter_search = (args.get('search', type=str) or '').strip() or None
    filter_ip_address = (args.get('ip_address', type=str) or '').strip() or None

    date_from_raw = (args.get('date_from', type=str) or '').strip()
    date_to_raw = (args.get('date_to', type=str) or '').strip()

    date_from = None
    if date_from_raw:
        try:
            date_from = datetime.strptime(date_from_raw, '%Y-%m-%d')
        except ValueError:
            date_from = None

    date_to_exclusive = None
    if date_to_raw:
        try:
            date_to_exclusive = datetime.strptime(date_to_raw, '%Y-%m-%d') + timedelta(days=1)
        except ValueError:
            date_to_exclusive = None

    return {
        'user_id': filter_user,
        'action': filter_action,
        'entity_type': filter_entity,
        'search': filter_search,
        'ip_address': filter_ip_address,
        'date_from': date_from,
        'date_to_exclusive': date_to_exclusive,
    }


def _apply_user_activity_filters(query, filters):
    """Apply parsed activity-log filters to a SQLAlchemy query."""
    if filters.get('user_id'):
        query = query.filter(UserActivityLog.user_id == filters['user_id'])

    if filters.get('action'):
        query = query.filter(UserActivityLog.action == filters['action'])

    if filters.get('entity_type'):
        query = query.filter(UserActivityLog.entity_type == filters['entity_type'])

    if filters.get('search'):
        like_search = f"%{filters['search']}%"
        query = query.filter(
            or_(
                UserActivityLog.entity_name.ilike(like_search),
                UserActivityLog.details.ilike(like_search)
            )
        )

    if filters.get('ip_address'):
        query = query.filter(UserActivityLog.ip_address.ilike(f"%{filters['ip_address']}%"))

    if filters.get('date_from'):
        query = query.filter(UserActivityLog.created_at >= filters['date_from'])

    if filters.get('date_to_exclusive'):
        query = query.filter(UserActivityLog.created_at < filters['date_to_exclusive'])

    return query


@reports_bp.route('/api/filtered-data')
@login_required
def get_filtered_data():
    """Get filtered statistics data as JSON for AJAX updates"""
    # Get current academic settings
    current_settings = AcademicSettings.query.filter_by(is_active=True).first()
    
    if current_settings:
        academic_year = current_settings.academic_year
        semester = current_settings.semester
    else:
        academic_year = None
        semester = None
    
    # Get user's program access
    user_program_ids = current_user.get_program_ids()
    
    # Get filter parameters from request
    filter_department = request.args.get('program', type=int)
    
    # Calculate statistics with filters
    # Allow pages to request only specific sections via ?include=counts,faculty
    include_param = request.args.get('include')
    if include_param:
        include_set = set(include_param.split(','))
    else:
        include_set = None  # compute everything (backward compat)
    
    stats = calculate_statistics(
        academic_year, 
        semester, 
        user_program_ids,
        filter_department,
        include=include_set
    )
    
    return jsonify({'stats': stats})


@reports_bp.route('/api/ai-summary')
@login_required
def get_ai_summary():
    """Generate AI summary of current report statistics"""
    try:
        # Get current academic settings
        current_settings = AcademicSettings.query.filter_by(is_active=True).first()
        
        if current_settings:
            academic_year = current_settings.academic_year
            semester = current_settings.semester
        else:
            academic_year = None
            semester = None
        
        # Get user's program access
        user_program_ids = current_user.get_program_ids()
        
        # Get filter parameters from request
        filter_department = request.args.get('program', type=int)
        
        # Calculate statistics with filters
        stats = calculate_statistics(
            academic_year, 
            semester, 
            user_program_ids,
            filter_department
        )
        
        # Get program name for AI context
        program_name = None
        if filter_department:
            program = Program.query.get(filter_department)
            if program:
                program_name = f"{program.program_name} ({program.program_code})"
        
        # Generate AI summary with program context
        ai_summary = ai_scheduler.generate_report_summary(
            stats, 
            academic_year, 
            semester,
            program_name
        )
        
        # Add detailed metrics to the response for dashboard display
        ai_summary['metrics'] = {
            # Schedule Progress
            'schedule_progress': {
                'percentage': stats.get('schedule_completion_rate', 0),
                'scheduled': stats.get('sections_with_schedules', 0),
                'total': stats.get('total_sections', 0)
            },
            # Faculty Load
            'faculty_load': {
                'avg_utilization': stats.get('avg_faculty_utilization', 0),
                'total_faculty': stats.get('total_faculty', 0),
                'assigned': stats.get('faculty_with_schedules', 0),
                'overloaded': stats.get('overloaded_faculty_count', 0),
                'warning': stats.get('warning_faculty_count', 0),
                'underutilized': stats.get('underutilized_faculty_count', 0),
                'load_status': 'balanced' if stats.get('overloaded_faculty_count', 0) == 0 else 'imbalanced'
            },
            # Unassigned Faculty
            'unassigned_faculty': {
                'count': stats.get('unassigned_faculty_count', 0),
                'by_department': stats.get('unassigned_faculty_by_dept', {}),
                'preview': stats.get('unassigned_faculty', [])[:5]  # First 5 names
            },
            # Unused Rooms
            'unused_rooms': {
                'count': stats.get('unused_rooms_count', 0),
                'by_type': stats.get('unused_rooms_by_type', {}),
                'preview': stats.get('unused_rooms', [])[:5]  # First 5 rooms
            },
            # Room Utilization
            'room_utilization': {
                'avg_percentage': stats.get('avg_room_utilization', 0),
                'total_hours': stats.get('total_room_hours_used', 0),
                'total_rooms': stats.get('total_rooms', 0),
                'rooms_in_use': stats.get('rooms_in_use', 0),
                'by_building': stats.get('room_utilization_by_building', {})
            }
        }
        
        return jsonify(ai_summary)
        
    except Exception as e:
        print(f"Error generating AI summary: {str(e)}")
        return jsonify({
            'ai_enabled': False,
            'error': str(e),
            'summary': 'Unable to generate AI summary at this time.',
            'insights': [],
            'recommendations': []
        }), 500

def get_dept_workload_distribution(academic_year=None, semester=None, user_program_ids=None):
    """Faculty workload status by program (for stacked bar chart).

    Returns list of dicts: [{program, normal, warning, exceeded}, ...]
    """
    from collections import defaultdict

    dept_q = Program.query.filter_by(is_active=True)
    if user_program_ids is not None:
        dept_q = dept_q.filter(Program.id.in_(user_program_ids))
    programs = dept_q.order_by(Program.program_code).all()

    system_max_units = AcademicSettings.get_default_faculty_max_units()

    # Batch-load all schedules for the period
    sched_q = Schedule.query.filter_by(is_active=True).options(db.joinedload(Schedule.subject))
    if academic_year:
        sched_q = sched_q.filter_by(academic_year=academic_year)
    if semester:
        sched_q = sched_q.filter_by(semester=semester)
    all_scheds = sched_q.all()

    sched_by_fac = defaultdict(list)
    for s in all_scheds:
        if s.faculty_id:
            sched_by_fac[s.faculty_id].append(s)

    result = []
    seen_faculty_ids = set()
    for dept in programs:
        # Faculty belongs to Department, not Program — map via program.department_id
        if not dept.department_id:
            continue
        faculty_list = Faculty.query.filter_by(
            department_id=dept.department_id, is_active=True, is_archived=False
        ).all()
        normal = warning = exceeded = 0
        for f in faculty_list:
            if f.id in seen_faculty_ids:
                continue  # avoid double-counting if multiple programs share a department
            seen_faculty_ids.add(f.id)
            current_load = sum(
                float(s.subject.total_units) if s.subject and s.subject.total_units else 0
                for s in sched_by_fac.get(f.id, [])
            )
            max_units = f.max_units if f.max_units is not None else system_max_units
            util_pct = (current_load / max_units * 100) if max_units > 0 else 0
            if util_pct >= 100:
                exceeded += 1
            elif util_pct >= 80:
                warning += 1
            else:
                normal += 1
        result.append({
            'program': dept.program_code,
            'normal': normal,
            'warning': warning,
            'exceeded': exceeded
        })
    return result


def generate_report_insights_banner(stats):
    """Generate 2-3 key insights for the reports page banner.

    Uses rule-based logic (no AI API call needed).
    Pre-computed from calculate_statistics() data.

    Args:
        stats: dict from calculate_statistics()

    Returns:
        list[str]: up to 3 short insight strings
    """
    insights = []

    # --- Faculty workload insights ---
    overloaded = stats.get('overloaded_faculty_count', 0)
    warning = stats.get('warning_faculty_count', 0)
    if overloaded > 0:
        insights.append(
            f"⚠️ {overloaded} faculty member{'s' if overloaded != 1 else ''} "
            f"exceeded maximum workload"
        )
    elif warning > 0:
        insights.append(
            f"{warning} faculty member{'s' if warning != 1 else ''} "
            f"approaching workload limit (>80%)"
        )
    else:
        avg_util = stats.get('avg_faculty_utilization', 0)
        if avg_util > 0:
            insights.append(f"Faculty workloads are balanced (avg {avg_util:.0f}% utilization)")

    # --- Room utilization insights ---
    total_rooms = stats.get('total_rooms', 0)
    rooms_in_use = stats.get('rooms_in_use', 0)
    unused_rooms = total_rooms - rooms_in_use
    if unused_rooms > 0 and total_rooms > 0:
        pct = round((unused_rooms / total_rooms) * 100)
        insights.append(f"{unused_rooms} room{'s' if unused_rooms != 1 else ''} ({pct}%) have no schedules assigned")

    # --- Schedule completion ---
    completion = stats.get('schedule_completion_rate', None)
    if completion is not None and completion < 100:
        remaining_pct = 100 - completion
        insights.append(f"Schedule completion at {completion}% — {remaining_pct}% of sections still need scheduling")

    # --- Unassigned faculty ---
    unassigned = stats.get('unassigned_faculty_count', 0)
    if unassigned > 0:
        insights.append(f"{unassigned} faculty have no schedule assignments this semester")

    # --- Weekly distribution imbalance ---
    by_day = stats.get('schedule_by_day', {})
    if by_day:
        active_days = {d: c for d, c in by_day.items() if c > 0}
        if len(active_days) >= 2:
            max_day = max(active_days, key=active_days.get)
            min_day = min(active_days, key=active_days.get)
            if active_days[max_day] > active_days[min_day] * 2:
                insights.append(
                    f"Schedule imbalance: {max_day} has {active_days[max_day]} classes vs "
                    f"{min_day}'s {active_days[min_day]}"
                )

    return insights[:3]


def calculate_statistics(academic_year=None, semester=None, user_program_ids=None, 
                        filter_department=None, include=None, include_archived=False):
    """Calculate various statistics for the dashboard.
    
    Args:
        include: Set of sections to compute. None or 'all' = everything.
                 Valid values: 'counts', 'faculty', 'rooms', 'weekly', 'completion'
        include_archived: If True, include archived (is_active=False) schedules.
                          Used by semester comparison to show historical data.
    """
    if include is None:
        include = {'counts', 'faculty', 'rooms', 'weekly', 'completion'}
    elif include == 'all':
        include = {'counts', 'faculty', 'rooms', 'weekly', 'completion'}

    stats = {}
    
    # Build base queries - include archived schedules for historical comparison
    if include_archived:
        schedule_query = Schedule.query
        exam_query = ExamSchedule.query
    else:
        schedule_query = Schedule.query.filter_by(is_active=True)
        exam_query = ExamSchedule.query.filter_by(is_active=True)
    
    if academic_year:
        schedule_query = schedule_query.filter_by(academic_year=academic_year)
        exam_query = exam_query.filter_by(academic_year=academic_year)
    
    if semester:
        schedule_query = schedule_query.filter_by(semester=semester)
        exam_query = exam_query.filter_by(semester=semester)
    
    # Track if we've already joined Section
    schedule_has_section_join = False
    exam_has_section_join = False
    
    # Filter by user program access
    if user_program_ids is not None:
        schedule_query = schedule_query.join(Section)
        exam_query = exam_query.join(Section)
        schedule_query = schedule_query.filter(Section.program_id.in_(user_program_ids))
        exam_query = exam_query.filter(Section.program_id.in_(user_program_ids))
        schedule_has_section_join = True
        exam_has_section_join = True
    
    # Apply additional filters
    if filter_department:
        if not schedule_has_section_join:
            schedule_query = schedule_query.join(Section)
            schedule_has_section_join = True
        schedule_query = schedule_query.filter(Section.program_id == filter_department)
        
        if not exam_has_section_join:
            exam_query = exam_query.join(Section)
            exam_has_section_join = True
        exam_query = exam_query.filter(Section.program_id == filter_department)
    
    # Total schedules
    stats['total_schedules'] = schedule_query.count()
    stats['total_exam_schedules'] = exam_query.count()
    
    # Active sections - apply both user program access AND filter_department
    section_query = Section.query
    if filter_department:
        section_query = section_query.filter(Section.program_id == filter_department)
    elif user_program_ids is not None:
        section_query = section_query.filter(Section.program_id.in_(user_program_ids))
    stats['total_sections'] = section_query.count()
    
    # Active faculty - filter by department (derived from program filter) OR teaching in those programs
    faculty_query = Faculty.query.filter_by(is_active=True, is_archived=False)
    
    # Base query for finding faculty teaching in relevant programs
    teaching_fac_query = db.session.query(Schedule.faculty_id).join(Section).filter(
        Schedule.faculty_id.isnot(None),
        Schedule.is_active == True
    )
    if academic_year:
        teaching_fac_query = teaching_fac_query.filter(Schedule.academic_year == academic_year)
    if semester:
        teaching_fac_query = teaching_fac_query.filter(Schedule.semester == semester)

    if filter_department:
        teaching_fac_query = teaching_fac_query.filter(Section.program_id == filter_department)
        teaching_ids = [r[0] for r in teaching_fac_query.distinct().all()]
        
        from app.models.program import Program as _Dept
        _dept_obj = _Dept.query.get(filter_department)
        
        if _dept_obj and _dept_obj.department_id:
            if teaching_ids:
                faculty_query = faculty_query.filter(
                    or_(
                        Faculty.department_id == _dept_obj.department_id,
                        Faculty.id.in_(teaching_ids)
                    )
                )
            else:
                faculty_query = faculty_query.filter(Faculty.department_id == _dept_obj.department_id)
        elif teaching_ids:
            faculty_query = faculty_query.filter(Faculty.id.in_(teaching_ids))
            
    elif user_program_ids is not None:
        teaching_fac_query = teaching_fac_query.filter(Section.program_id.in_(user_program_ids))
        teaching_ids = [r[0] for r in teaching_fac_query.distinct().all()]
        
        from app.models.program import Program as _Dept
        _department_ids = db.session.query(_Dept.department_id).filter(
            _Dept.id.in_(user_program_ids), _Dept.department_id.isnot(None)
        ).distinct().all()
        _dept_id_list = [c[0] for c in _department_ids]
        
        if _dept_id_list:
            if teaching_ids:
                faculty_query = faculty_query.filter(
                    or_(
                        Faculty.department_id.in_(_dept_id_list),
                        Faculty.id.in_(teaching_ids)
                    )
                )
            else:
                faculty_query = faculty_query.filter(Faculty.department_id.in_(_dept_id_list))
        elif teaching_ids:
            faculty_query = faculty_query.filter(Faculty.id.in_(teaching_ids))
            
    stats['total_faculty'] = faculty_query.count()
    
    # Available rooms
    stats['total_rooms'] = Room.query.filter_by(is_available=True).count()
    
    # Faculty with schedules
    faculty_with_schedules = schedule_query.filter(Schedule.faculty_id.isnot(None))\
        .with_entities(Schedule.faculty_id).distinct().count()
    stats['faculty_with_schedules'] = faculty_with_schedules
    
    # Rooms being used
    rooms_in_use = schedule_query.filter(Schedule.room_id.isnot(None))\
        .with_entities(Schedule.room_id).distinct().count()
    stats['rooms_in_use'] = rooms_in_use
    
    # Schedule type distribution
    lecture_count = schedule_query.filter(Schedule.schedule_type == 'lecture').count()
    lab_count = schedule_query.filter(Schedule.schedule_type == 'lab').count()
    stats['lecture_count'] = lecture_count
    stats['lab_count'] = lab_count
    
    # Day distribution
    if 'weekly' in include or 'counts' in include:
        days = AcademicSettings.get_active_operation_days()
        stats['schedule_by_day'] = {}
        for day in days:
            stats['schedule_by_day'][day] = schedule_query.filter(Schedule.day_of_week == day).count()
    
    # Faculty workload details - ALL faculty with enhanced utilization data
    if 'faculty' in include:
        from collections import defaultdict
        
        # BATCH: Fetch ALL schedules for the period in ONE query (with subjects + sections eager-loaded)
        all_fac_sched_query = Schedule.query.options(
            db.joinedload(Schedule.subject),
            db.joinedload(Schedule.section)
        )
        if not include_archived:
            all_fac_sched_query = all_fac_sched_query.filter_by(is_active=True)
        
        if academic_year:
            all_fac_sched_query = all_fac_sched_query.filter_by(academic_year=academic_year)
        if semester:
            all_fac_sched_query = all_fac_sched_query.filter_by(semester=semester)
        
        all_fac_schedules = all_fac_sched_query.all()
        
        # Group schedules by faculty_id in Python (O(n) instead of N queries)
        schedules_by_faculty = defaultdict(list)
        for s in all_fac_schedules:
            if s.faculty_id:
                schedules_by_faculty[s.faculty_id].append(s)
        
        # Get system default max units ONCE
        system_max_units = AcademicSettings.get_default_faculty_max_units()
        
        # Determine program filter set for Python-side filtering
        dept_filter_set = None
        if filter_department:
            dept_filter_set = {filter_department}
        elif user_program_ids is not None:
            dept_filter_set = set(user_program_ids)
        
        # Build a department_id filter set for Faculty-level filtering
        # (dept_filter_set has program IDs; Faculty uses department_id)
        dept_id_filter_set = None
        if dept_filter_set is not None:
            from app.models.program import Program as _FiltProg
            _filt_dept_rows = db.session.query(_FiltProg.department_id).filter(
                _FiltProg.id.in_(dept_filter_set),
                _FiltProg.department_id.isnot(None)
            ).distinct().all()
            dept_id_filter_set = {r[0] for r in _filt_dept_rows}
        
        faculty_list = faculty_query.options(db.joinedload(Faculty.department)).order_by(Faculty.last_name, Faculty.first_name).all()
        relevant_fac_ids = {f.id for f in faculty_list}
        faculty_workloads = []
        
        for faculty in faculty_list:
            fac_all_schedules = schedules_by_faculty.get(faculty.id, [])
            
            # Program-filtered schedules for display columns
            if dept_filter_set is not None:
                fac_schedules = [s for s in fac_all_schedules if s.section and s.section.program_id in dept_filter_set]
            else:
                fac_schedules = fac_all_schedules
            
            # Calculate units from display schedules
            total_units = sum(float(s.subject.total_units) if s.subject and s.subject.total_units else 0 for s in fac_schedules)
            lec_units = sum(float(s.subject.lec_units) if s.subject and s.subject.lec_units else 0 for s in fac_schedules)
            lab_units = sum(float(s.subject.lab_units) if s.subject and s.subject.lab_units else 0 for s in fac_schedules)

            # Weekly teaching hours from class schedule durations (filtered display context)
            weekly_minutes = 0
            for sched in fac_schedules:
                if sched.start_time and sched.end_time:
                    start_minutes = (sched.start_time.hour * 60) + sched.start_time.minute
                    end_minutes = (sched.end_time.hour * 60) + sched.end_time.minute
                    if end_minutes > start_minutes:
                        weekly_minutes += (end_minutes - start_minutes)
            weekly_hours = round(weekly_minutes / 60, 1)
            
            # Load status from ALL schedules (not dept-filtered) — matches get_load_status behavior
            current_load = sum(float(s.subject.total_units) if s.subject and s.subject.total_units else 0 for s in fac_all_schedules)
            max_units = faculty.max_units if faculty.max_units is not None else system_max_units
            utilization_pct = (current_load / max_units * 100) if max_units > 0 else 0
            
            if utilization_pct >= 100:
                load_status = 'exceeded'
            elif utilization_pct >= 80:
                load_status = 'warning'
            else:
                load_status = 'normal'
            
            faculty_workloads.append({
                'name': faculty.full_name,
                'program': faculty.department.department_code if faculty.department else 'N/A',
                'schedules': len(fac_schedules),
                'lec_units': float(lec_units),
                'lab_units': float(lab_units),
                'total_units': float(total_units),
                'weekly_hours': float(weekly_hours),
                # Enhanced utilization data
                'current_units': float(current_load),
                'max_units': int(max_units),
                'utilization_pct': round(utilization_pct, 1),
                'load_status': load_status  # 'normal', 'warning', 'exceeded'
            })
        # Sort by total units descending - show ALL faculty (no limit)
        stats['faculty_workloads'] = sorted(faculty_workloads, key=lambda x: x['total_units'], reverse=True)
        
        # Aggregate faculty utilization stats
        if faculty_workloads:
            stats['avg_faculty_utilization'] = round(
                sum([f['utilization_pct'] for f in faculty_workloads]) / len(faculty_workloads), 1
            )
            stats['overloaded_faculty_count'] = len([f for f in faculty_workloads if f['load_status'] == 'exceeded'])
            stats['warning_faculty_count'] = len([f for f in faculty_workloads if f['load_status'] == 'warning'])
            stats['underutilized_faculty_count'] = len([f for f in faculty_workloads if f['utilization_pct'] < 50])
        else:
            stats['avg_faculty_utilization'] = 0
            stats['overloaded_faculty_count'] = 0
            stats['warning_faculty_count'] = 0
            stats['underutilized_faculty_count'] = 0
        
        # Unassigned faculty peek (faculty without schedules) - top 10 for modal
        unassigned_faculty = [f for f in faculty_workloads if f['schedules'] == 0]
        stats['unassigned_faculty'] = unassigned_faculty[:10]
        stats['unassigned_faculty_count'] = len(unassigned_faculty)
        
        # Unassigned faculty by program breakdown
        unassigned_by_dept = {}
        for f in unassigned_faculty:
            dept = f['program']
            unassigned_by_dept[dept] = unassigned_by_dept.get(dept, 0) + 1
        stats['unassigned_faculty_by_dept'] = unassigned_by_dept
        
        # Unavailable faculty RIGHT NOW peek
        from datetime import datetime as _dt
        now = _dt.now()
        current_day = now.strftime('%A')  # 'Monday', 'Tuesday', etc.
        current_time = now.time()

        # ── 1. Faculty NOT available right now ──
        # Faculty who have availability slots defined for today, but the current time
        # is outside all their windows (before first slot or between/after slots).
        today_avail_records = FacultyAvailability.query.filter(
            FacultyAvailability.is_active == True,
            FacultyAvailability.day_of_week == current_day
        ).all()

        today_slots_by_faculty = {}
        for rec in today_avail_records:
            today_slots_by_faculty.setdefault(rec.faculty_id, []).append(rec)

        unavailable_faculty_ids = set()
        unavailable_now_list = []

        for fac_id, slots in today_slots_by_faculty.items():
            valid_slots = [s for s in slots if s.start_time and s.end_time]
            if not valid_slots:
                continue
            # Faculty is "not available right now" only if current time is outside all windows
            is_in_slot = any(s.start_time <= current_time < s.end_time for s in valid_slots)
            if is_in_slot:
                continue
            fac = Faculty.query.get(fac_id)
            if not fac or fac.is_archived or not fac.is_active:
                continue
            if fac.id not in relevant_fac_ids:
                continue
            unavailable_faculty_ids.add(fac_id)
            dept_name = fac.department.department_code if fac.department else 'N/A'
            sorted_slots = sorted(valid_slots, key=lambda s: s.start_time)
            slot_text = ', '.join(
                f"{s.start_time.strftime('%I:%M %p')}–{s.end_time.strftime('%I:%M %p')}"
                for s in sorted_slots
            )
            unavailable_now_list.append({
                'name': fac.full_name,
                'program': dept_name,
                'reason': f"Window: {slot_text}" if slot_text else 'No active window',
                'start_time': sorted_slots[0].start_time.strftime('%I:%M %p'),
                'end_time': sorted_slots[-1].end_time.strftime('%I:%M %p'),
                'day': current_day,
                'type': 'not_in_window'
            })

        # ── 2. Faculty with RESTRICTED weekly availability ──
        # Show faculty who only have availability on certain days of the week
        # (not available every weekday) — regardless of what day it currently is.
        from sqlalchemy import distinct, and_, func as sa_func
        ALL_WEEKDAYS = set(AcademicSettings.get_active_operation_days())
        
        # Get each faculty's available days: { faculty_id: set(day_names) }
        avail_rows = db.session.query(
            FacultyAvailability.faculty_id,
            FacultyAvailability.day_of_week
        ).filter(
            FacultyAvailability.is_active == True,
            FacultyAvailability.day_of_week != None
        ).distinct().all()
        
        faculty_avail_days = {}
        for fac_id, day in avail_rows:
            if fac_id not in faculty_avail_days:
                faculty_avail_days[fac_id] = set()
            faculty_avail_days[fac_id].add(day)
        
        day_order = {'Monday': 1, 'Tuesday': 2, 'Wednesday': 3, 'Thursday': 4,
                     'Friday': 5, 'Saturday': 6, 'Sunday': 7}
        
        restricted_list = []
        for fac_id, avail_days_set in faculty_avail_days.items():
            # Skip if already in explicit unavailable list
            if fac_id in unavailable_faculty_ids:
                continue
            # Only flag if they DON'T cover all weekdays (Mon-Sat)
            if avail_days_set >= ALL_WEEKDAYS:
                continue
            
            fac = Faculty.query.get(fac_id)
            if not fac or fac.is_archived or not fac.is_active:
                continue
            if dept_id_filter_set is not None and fac.department_id not in dept_id_filter_set:
                continue
            
            sorted_days = sorted(avail_days_set, key=lambda x: day_order.get(x, 8))
            short_days = [d[:3] for d in sorted_days]
            reason = f"Only available on {', '.join(short_days)}"
            
            dept_name = fac.department.department_code if fac.department else 'N/A'
            is_available_today = current_day in avail_days_set
            
            restricted_list.append({
                'name': fac.full_name,
                'program': dept_name,
                'reason': reason,
                'start_time': '',
                'end_time': '',
                'day': current_day,
                'type': 'restricted',  # has limited weekly availability
                'available_today': is_available_today,
                'available_days': short_days
            })
        
        restricted_list.sort(key=lambda x: x['name'])
        
        # Sort: explicit unavailable first, restricted second
        unavailable_now_list.sort(key=lambda x: x['name'])
        
        stats['unavailable_faculty_now'] = unavailable_now_list[:10]
        stats['unavailable_faculty_now_count'] = len(unavailable_now_list)
        stats['restricted_faculty'] = restricted_list[:10]
        stats['restricted_faculty_count'] = len(restricted_list)
        stats['current_day'] = current_day
        stats['current_time_display'] = now.strftime('%I:%M %p')
        
        # Unavailable by program (combine both lists)
        unavailable_by_dept = {}
        for f in unavailable_now_list + restricted_list:
            dept = f['program']
            unavailable_by_dept[dept] = unavailable_by_dept.get(dept, 0) + 1
        stats['unavailable_faculty_by_dept'] = unavailable_by_dept
        
        # ── 3. Unavailable Faculty TODAY ──
        unavailable_today_list = []
        seen_unavail_today = set()
        
        # (a) Explicit unavailable records removed — availability_type no longer exists
        
        # (b) Restricted faculty who don't have today in their available days
        for fac_id, avail_days_set in faculty_avail_days.items():
            if fac_id in seen_unavail_today:
                continue
            if current_day in avail_days_set:
                continue  # available today
            fac = Faculty.query.get(fac_id)
            if not fac or fac.is_archived or not fac.is_active:
                continue
            if dept_id_filter_set is not None and fac.department_id not in dept_id_filter_set:
                continue
            seen_unavail_today.add(fac_id)
            dept_name = fac.department.department_code if fac.department else 'N/A'
            unavailable_today_list.append({
                'name': fac.full_name,
                'program': dept_name,
                'reason': 'Not available this day',
            })
        
        # (c) Faculty who are technically available today but have NO classes scheduled
        for fac in faculty_list:
            if fac.id in seen_unavail_today:
                continue
            if fac.id in unavailable_faculty_ids:
                continue
            # Check restricted availability
            if fac.id in faculty_avail_days:
                if current_day not in faculty_avail_days[fac.id]:
                    continue  # already handled in (b)
            # Faculty is available today — check if they have classes
            today_scheds = [s for s in schedules_by_faculty.get(fac.id, []) if s.day_of_week == current_day]
            if len(today_scheds) == 0:
                seen_unavail_today.add(fac.id)
                dept_name = fac.department.department_code if fac.department else 'N/A'
                unavailable_today_list.append({
                    'name': fac.full_name,
                    'program': dept_name,
                    'reason': 'No classes today',
                })
        
        unavailable_today_list.sort(key=lambda x: x['name'])
        stats['unavailable_faculty_today_list'] = unavailable_today_list
        stats['unavailable_faculty_today_count'] = len(unavailable_today_list)
        
        # ── 4. Faculty AVAILABLE TODAY ──
        # All active faculty who are NOT in unavailable_faculty_ids and either:
        #   (a) have no restricted availability at all, or
        #   (b) have current_day in their available_days
        available_today_list = []
        for fac in faculty_list:
            if fac.id in unavailable_faculty_ids:
                continue
            # Check if this faculty has restricted availability
            if fac.id in faculty_avail_days:
                if current_day not in faculty_avail_days[fac.id]:
                    continue  # not available today
            # Faculty is available today — only include if they have classes
            dept_name = fac.department.department_code if fac.department else 'N/A'
            today_scheds = [s for s in schedules_by_faculty.get(fac.id, []) if s.day_of_week == current_day]
            schedule_count = len(today_scheds)
            if schedule_count == 0:
                continue  # No classes today = not considered available

            valid_time_ranges = [
                (s.start_time, s.end_time)
                for s in today_scheds
                if s.start_time and s.end_time
            ]
            if valid_time_ranges:
                earliest_start = min(slot[0] for slot in valid_time_ranges)
                latest_end = max(slot[1] for slot in valid_time_ranges)
                availability_span_display = f"{earliest_start.strftime('%I:%M %p')} - {latest_end.strftime('%I:%M %p')}"
            else:
                availability_span_display = 'Time unavailable'

            available_today_list.append({
                'name': fac.full_name,
                'program': dept_name,
                'schedule_count': schedule_count,
                'availability_span_display': availability_span_display,
            })
        
        available_today_list.sort(key=lambda x: x['name'])
        stats['available_faculty_today'] = available_today_list
        stats['available_faculty_today_count'] = len(available_today_list)
        
        # Faculty utilization distribution brackets
        brackets = {'0-25': 0, '25-50': 0, '50-80': 0, '80-100': 0, '100+': 0}
        for f in faculty_workloads:
            pct = f['utilization_pct']
            if pct >= 100:
                brackets['100+'] += 1
            elif pct >= 80:
                brackets['80-100'] += 1
            elif pct >= 50:
                brackets['50-80'] += 1
            elif pct >= 25:
                brackets['25-50'] += 1
            else:
                brackets['0-25'] += 1
        stats['faculty_utilization_brackets'] = brackets
    
    # Room utilization details - ALL rooms with HOURS-BASED calculation
    if 'rooms' in include:
        from collections import defaultdict as _defaultdict
        active_settings = AcademicSettings.query.filter_by(is_active=True).first()
        operation_days = active_settings.get_operation_days_list() if active_settings else AcademicSettings.get_active_operation_days()

        if active_settings:
            start_time = active_settings.schedule_start_time or time(int(active_settings.schedule_start_hour or 7), 0)
            end_time = active_settings.schedule_end_time or time(int(active_settings.schedule_end_hour or 20), 0)
            DAY_START_MINUTES = (start_time.hour * 60) + start_time.minute
            DAY_END_MINUTES = (end_time.hour * 60) + end_time.minute
        else:
            DAY_START_MINUTES = 7 * 60
            DAY_END_MINUTES = 20 * 60

        # Guard invalid ranges from settings by falling back to a safe default window.
        if DAY_END_MINUTES <= DAY_START_MINUTES:
            DAY_START_MINUTES = 7 * 60
            DAY_END_MINUTES = 20 * 60

        max_daily_hours = (DAY_END_MINUTES - DAY_START_MINUTES) / 60
        operation_day_count = len(operation_days) if operation_days else 6
        MAX_WEEKLY_HOURS = round(max_daily_hours * operation_day_count, 1)

        def _format_minutes(total_minutes):
            hour = total_minutes // 60
            minute = total_minutes % 60
            suffix = 'AM' if hour < 12 else 'PM'
            display_hour = hour % 12
            if display_hour == 0:
                display_hour = 12
            return f"{display_hour}:{minute:02d} {suffix}"

        def _merge_intervals(intervals):
            if not intervals:
                return []
            sorted_intervals = sorted(intervals, key=lambda i: i[0])
            merged = [sorted_intervals[0]]
            for start, end in sorted_intervals[1:]:
                last_start, last_end = merged[-1]
                if start <= last_end:
                    merged[-1] = (last_start, max(last_end, end))
                else:
                    merged.append((start, end))
            return merged

        def _invert_intervals(occupied, start_limit, end_limit):
            free = []
            current = start_limit
            for start, end in occupied:
                if current < start:
                    free.append((current, start))
                current = max(current, end)
            if current < end_limit:
                free.append((current, end_limit))
            return free
        
        # BATCH: Fetch ALL room schedules in ONE query (with section eager-loaded for dept filter)
        all_room_sched_query = Schedule.query.options(
            db.joinedload(Schedule.section)
        ).filter(Schedule.room_id.isnot(None))
        if not include_archived:
            all_room_sched_query = all_room_sched_query.filter_by(is_active=True)
        
        if academic_year:
            all_room_sched_query = all_room_sched_query.filter_by(academic_year=academic_year)
        if semester:
            all_room_sched_query = all_room_sched_query.filter_by(semester=semester)
        
        all_room_schedules = all_room_sched_query.all()
        
        # BATCH: Fetch ALL room exams in ONE query
        all_room_exam_query = ExamSchedule.query.options(
            db.joinedload(ExamSchedule.section)
        ).filter(ExamSchedule.room_id.isnot(None))
        if not include_archived:
            all_room_exam_query = all_room_exam_query.filter_by(is_active=True)
        
        if academic_year:
            all_room_exam_query = all_room_exam_query.filter_by(academic_year=academic_year)
        if semester:
            all_room_exam_query = all_room_exam_query.filter_by(semester=semester)
        
        all_room_exams = all_room_exam_query.all()
        
        # Determine program filter set for Python-side filtering
        room_dept_filter = None
        if filter_department:
            room_dept_filter = {filter_department}
        elif user_program_ids is not None:
            room_dept_filter = set(user_program_ids)
        
        # Group by room_id with program filtering in Python
        schedules_by_room = _defaultdict(list)
        for s in all_room_schedules:
            if room_dept_filter is not None:
                if not s.section or s.section.program_id not in room_dept_filter:
                    continue
            schedules_by_room[s.room_id].append(s)
        
        exams_by_room = _defaultdict(list)
        for e in all_room_exams:
            if room_dept_filter is not None:
                if not e.section or e.section.program_id not in room_dept_filter:
                    continue
            exams_by_room[e.room_id].append(e)
        
        rooms = Room.query.options(db.joinedload(Room.building)).filter_by(is_available=True).all()
        room_utilizations = []
        room_availability = []
        total_hours_all_rooms = 0
        
        for room in rooms:
            room_schedules = schedules_by_room.get(room.id, [])
            room_exams = exams_by_room.get(room.id, [])

            occupied_by_day = {day: [] for day in operation_days}
            for sched in room_schedules:
                day = sched.day_of_week
                if day not in occupied_by_day or not sched.start_time or not sched.end_time:
                    continue

                start_minutes = sched.start_time.hour * 60 + sched.start_time.minute
                end_minutes = sched.end_time.hour * 60 + sched.end_time.minute
                start_minutes = max(start_minutes, DAY_START_MINUTES)
                end_minutes = min(end_minutes, DAY_END_MINUTES)
                if end_minutes > start_minutes:
                    occupied_by_day[day].append((start_minutes, end_minutes))

            availability_by_day = {}
            for day in operation_days:
                merged_occupied = _merge_intervals(occupied_by_day[day])
                free_slots = _invert_intervals(merged_occupied, DAY_START_MINUTES, DAY_END_MINUTES)
                availability_by_day[day] = [
                    f"{_format_minutes(start)} - {_format_minutes(end)}"
                    for start, end in free_slots
                ]
            
            # Calculate hours used from schedules accurately (merged intervals prevent >100% overlapping inflation)
            schedule_hours = 0
            for day in operation_days:
                merged_occupied = _merge_intervals(occupied_by_day[day])
                for start, end in merged_occupied:
                    schedule_hours += (end - start) / 60.0
            
            total_hours = schedule_hours
            total_hours_all_rooms += total_hours
            utilization_pct = round((total_hours / MAX_WEEKLY_HOURS) * 100, 1) if MAX_WEEKLY_HOURS > 0 else 0
            
            total_usage = len(room_schedules)
            # Include ALL rooms, even those with zero usage
            room_utilizations.append({
                'room': room.room_number,
                'name': room.room_number,  # Alias for modal display
                'building': room.building.building_name if room.building else 'N/A',
                'building_id': room.building_id,
                'room_type': room.room_type or 'Lecture',
                'schedules': len(room_schedules),
                'exams': len(room_exams),
                'total_usage': total_usage,
                'schedule_hours': round(schedule_hours, 1),
                'total_hours': round(total_hours, 1),
                'utilization_pct': utilization_pct,
                'max_hours': MAX_WEEKLY_HOURS,
                'is_available': room.is_available
            })

            room_availability.append({
                'room': room.room_number,
                'building': room.building.building_name if room.building else 'N/A',
                'building_id': room.building_id,
                'room_type': room.room_type or 'Lecture',
                'has_schedule': len(room_schedules) > 0,
                'no_schedule': len(room_schedules) == 0,
                'available_slots_by_day': availability_by_day
            })
        
        # Sort by total hours descending - show ALL rooms (no limit)
        stats['room_utilizations'] = sorted(room_utilizations, key=lambda x: x['total_hours'], reverse=True)
        stats['room_utilization_details'] = stats['room_utilizations']  # Alias for modal
        
        # Calculate average room utilization percentage
        if room_utilizations:
            avg_room_utilization = round(sum(r['utilization_pct'] for r in room_utilizations) / len(room_utilizations), 1)
        else:
            avg_room_utilization = 0
        stats['avg_room_utilization'] = avg_room_utilization
        stats['total_room_hours_used'] = round(total_hours_all_rooms, 1)
        stats['max_weekly_hours'] = MAX_WEEKLY_HOURS

        no_schedule_count = len([r for r in room_availability if r['no_schedule']])
        stats['room_availability'] = sorted(room_availability, key=lambda x: (x['building'], x['room']))
        stats['room_availability_count'] = len(room_availability)
        stats['room_availability_no_schedule_count'] = no_schedule_count
        stats['room_availability_with_schedule_count'] = len(room_availability) - no_schedule_count
        
        # Unused rooms peek (rooms with zero usage) - top 10 for modal
        unused_rooms = [r for r in room_utilizations if r['total_hours'] == 0]
        stats['unused_rooms'] = unused_rooms[:10]
        stats['unused_rooms_count'] = len(unused_rooms)
        
        # Unused rooms by type breakdown
        unused_by_type = {}
        for r in unused_rooms:
            rtype = r['room_type']
            unused_by_type[rtype] = unused_by_type.get(rtype, 0) + 1
        stats['unused_rooms_by_type'] = unused_by_type
        
        # Room utilization by building (hours-based)
        building_utilization = {}
        for r in room_utilizations:
            bldg = r['building']
            if bldg not in building_utilization:
                building_utilization[bldg] = {'total': 0, 'total_hours': 0, 'max_hours': 0, 'in_use': 0, 'unused': 0}
            building_utilization[bldg]['total'] += 1
            building_utilization[bldg]['total_hours'] += r['total_hours']
            building_utilization[bldg]['max_hours'] += MAX_WEEKLY_HOURS
            if r['total_hours'] > 0:
                building_utilization[bldg]['in_use'] += 1
            else:
                building_utilization[bldg]['unused'] += 1
        
        # Calculate utilization percentage for each building
        for bldg in building_utilization:
            max_hrs = building_utilization[bldg]['max_hours']
            total_hrs = building_utilization[bldg]['total_hours']
            building_utilization[bldg]['utilization_pct'] = round((total_hrs / max_hrs) * 100, 1) if max_hrs > 0 else 0
            building_utilization[bldg]['total_hours'] = round(building_utilization[bldg]['total_hours'], 1)  # Round for display
        
        stats['room_utilization_by_building'] = building_utilization
    
    # Calculate schedule completion rate (based on sections with at least one schedule)
    if 'completion' in include:
        total_sections = stats.get('total_sections', 0)
        
        # Count sections that have at least one active schedule
        sections_scheduled_query = db.session.query(func.count(func.distinct(Schedule.section_id)))\
            .join(Section, Schedule.section_id == Section.id)\
            .filter(Schedule.is_active == True)
        
        if academic_year:
            sections_scheduled_query = sections_scheduled_query.filter(Schedule.academic_year == academic_year)
        if semester:
            sections_scheduled_query = sections_scheduled_query.filter(Schedule.semester == semester)
        
        if filter_department:
            sections_scheduled_query = sections_scheduled_query.filter(Section.program_id == filter_department)
        elif user_program_ids is not None:
            sections_scheduled_query = sections_scheduled_query.filter(Section.program_id.in_(user_program_ids))
        
        sections_with_schedules = sections_scheduled_query.scalar() or 0
        
        # Cap at 100% to handle edge cases
        stats['schedule_completion_rate'] = min(100, round((sections_with_schedules / total_sections * 100) if total_sections > 0 else 0))
        stats['sections_with_schedules'] = sections_with_schedules
    
    return stats


@reports_bp.route('/api/academic-periods')
@login_required
def get_academic_periods():
    """Get all unique academic year/semester combinations for comparison.
    
    Pulls periods from schedules, exam_schedules, AND faculty_subject_assignments
    so that all known academic periods appear even if one table has limited data.
    """
    try:
        # Get user's program access
        user_program_ids = current_user.get_program_ids()
        
        # Collect periods from multiple sources and merge
        all_period_tuples = set()
        
        # 1. Schedules table (include archived so past semesters appear)
        sched_query = db.session.query(
            Schedule.academic_year,
            Schedule.semester
        )
        if user_program_ids is not None:
            sched_query = sched_query.join(Section).filter(Section.program_id.in_(user_program_ids))
        for row in sched_query.distinct().all():
            if row[0] and row[1]:
                all_period_tuples.add((row[0], row[1]))
        
        # 2. Exam schedules table (include archived so past semesters appear)
        exam_query = db.session.query(
            ExamSchedule.academic_year,
            ExamSchedule.semester
        )
        if user_program_ids is not None:
            exam_query = exam_query.join(Section).filter(Section.program_id.in_(user_program_ids))
        for row in exam_query.distinct().all():
            if row[0] and row[1]:
                all_period_tuples.add((row[0], row[1]))
        
        # 3. Faculty subject assignments table
        fsa_query = db.session.query(
            FacultySubjectAssignment.academic_year,
            FacultySubjectAssignment.semester
        )
        if user_program_ids is not None:
            fsa_query = fsa_query.join(Faculty).filter(Faculty.department_id.in_(
                db.session.query(Program.department_id).filter(
                    Program.id.in_(user_program_ids), Program.department_id.isnot(None)
                )
            ))
        for row in fsa_query.distinct().all():
            if row[0] and row[1]:
                all_period_tuples.add((row[0], row[1]))
        
        # Sort: academic_year descending, semester ascending
        sorted_periods = sorted(all_period_tuples, key=lambda x: (x[0], x[1]), reverse=True)
        
        # Format periods for frontend
        formatted_periods = []
        for academic_year, semester in sorted_periods:
            formatted_periods.append({
                'academic_year': academic_year,
                'semester': semester,
                'label': f"{academic_year} - {semester}"
            })
        
        return jsonify({
            'success': True,
            'periods': formatted_periods
        })
        
    except Exception as e:
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500


@reports_bp.route('/api/semester-comparison')
@login_required
def get_semester_comparison():
    """Compare statistics between two academic periods"""
    try:
        # Get comparison parameters
        year1 = request.args.get('year1')
        sem1 = request.args.get('sem1')
        year2 = request.args.get('year2')
        sem2 = request.args.get('sem2')
        
        if not all([year1, sem1, year2, sem2]):
            return jsonify({
                'success': False,
                'error': 'Please select both periods for comparison'
            }), 400
        
        # Get user's program access
        user_program_ids = current_user.get_program_ids()
        filter_department = request.args.get('program', type=int)
        
        # Calculate stats for period 1 (include_archived=True so past semesters show data)
        stats1 = calculate_statistics(year1, sem1, user_program_ids, filter_department, include_archived=True)
        stats1['period_label'] = f"{year1} - {sem1}"
        
        # Calculate stats for period 2 (include_archived=True so past semesters show data)
        stats2 = calculate_statistics(year2, sem2, user_program_ids, filter_department, include_archived=True)
        stats2['period_label'] = f"{year2} - {sem2}"
        
        # Calculate comparison metrics (period2 - period1)
        comparison = {
            'schedule_diff': stats2['total_schedules'] - stats1['total_schedules'],
            'schedule_pct_change': calculate_pct_change(stats1['total_schedules'], stats2['total_schedules']),
            'exam_diff': stats2['total_exam_schedules'] - stats1['total_exam_schedules'],
            'exam_pct_change': calculate_pct_change(stats1['total_exam_schedules'], stats2['total_exam_schedules']),
            'faculty_assigned_diff': stats2['faculty_with_schedules'] - stats1['faculty_with_schedules'],
            'faculty_assigned_pct_change': calculate_pct_change(stats1['faculty_with_schedules'], stats2['faculty_with_schedules']),
            'rooms_used_diff': stats2['rooms_in_use'] - stats1['rooms_in_use'],
            'rooms_used_pct_change': calculate_pct_change(stats1['rooms_in_use'], stats2['rooms_in_use']),
            'sections_diff': stats2['total_sections'] - stats1['total_sections'],
            'sections_pct_change': calculate_pct_change(stats1['total_sections'], stats2['total_sections']),
            'avg_utilization_diff': stats2['avg_faculty_utilization'] - stats1['avg_faculty_utilization'],
            'lecture_diff': stats2['lecture_count'] - stats1['lecture_count'],
            'lab_diff': stats2['lab_count'] - stats1['lab_count'],
        }
        
        return jsonify({
            'success': True,
            'period1': {
                'label': stats1['period_label'],
                'total_schedules': stats1['total_schedules'],
                'total_exam_schedules': stats1['total_exam_schedules'],
                'faculty_with_schedules': stats1['faculty_with_schedules'],
                'rooms_in_use': stats1['rooms_in_use'],
                'total_sections': stats1['total_sections'],
                'lecture_count': stats1['lecture_count'],
                'lab_count': stats1['lab_count'],
                'avg_faculty_utilization': stats1['avg_faculty_utilization'],
                'schedule_by_day': stats1['schedule_by_day']
            },
            'period2': {
                'label': stats2['period_label'],
                'total_schedules': stats2['total_schedules'],
                'total_exam_schedules': stats2['total_exam_schedules'],
                'faculty_with_schedules': stats2['faculty_with_schedules'],
                'rooms_in_use': stats2['rooms_in_use'],
                'total_sections': stats2['total_sections'],
                'lecture_count': stats2['lecture_count'],
                'lab_count': stats2['lab_count'],
                'avg_faculty_utilization': stats2['avg_faculty_utilization'],
                'schedule_by_day': stats2['schedule_by_day']
            },
            'comparison': comparison
        })
        
    except Exception as e:
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500


def calculate_pct_change(old_value, new_value):
    """Calculate percentage change between two values"""
    if old_value == 0:
        return 100.0 if new_value > 0 else 0.0
    return round(((new_value - old_value) / old_value) * 100, 1)


@reports_bp.route('/api/semester-trends')
@login_required
def get_semester_trends():
    """Get key metrics across all semesters for trend line visualization (C2)."""
    try:
        user_program_ids = current_user.get_program_ids()
        filter_department = request.args.get('program', type=int)

        # Get distinct academic_year + semester from schedules
        sem_order = {'1st Semester': 1, '2nd Semester': 2, 'Summer': 3}
        period_rows = db.session.query(
            Schedule.academic_year, Schedule.semester
        ).distinct().all()

        # Dedup and sort
        periods = sorted(
            {(r[0], r[1]) for r in period_rows if r[0] and r[1]},
            key=lambda x: (x[0], sem_order.get(x[1], 9))
        )

        trend_data = []
        for ay, sem in periods:
            # Schedules count
            sq = Schedule.query.filter_by(academic_year=ay, semester=sem, is_active=True)
            eq = ExamSchedule.query.filter_by(academic_year=ay, semester=sem, is_active=True)
            fq = db.session.query(Schedule.faculty_id).filter(
                Schedule.academic_year == ay, Schedule.semester == sem,
                Schedule.is_active == True, Schedule.faculty_id.isnot(None)
            ).distinct()

            if filter_department:
                sq = sq.join(Section).filter(Section.program_id == filter_department)
                eq = eq.join(Section).filter(Section.program_id == filter_department)
                fq = fq.join(Section).filter(Section.program_id == filter_department)
            elif user_program_ids is not None:
                sq = sq.join(Section).filter(Section.program_id.in_(user_program_ids))
                eq = eq.join(Section).filter(Section.program_id.in_(user_program_ids))
                fq = fq.join(Section).filter(Section.program_id.in_(user_program_ids))

            short_sem = sem[:3] if sem else ''
            trend_data.append({
                'label': f"{short_sem} {ay}",
                'schedules': sq.count(),
                'exams': eq.count(),
                'faculty': fq.count(),
            })

        return jsonify({'success': True, 'trends': trend_data})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500


@reports_bp.route('/api/user-activity')
@login_required
@role_required('admin', 'super_admin')
def get_user_activity():
    """Get user activity logs (admin only)"""
    try:
        page = max(1, request.args.get('page', 1, type=int))
        per_page = request.args.get('per_page', 30, type=int)
        per_page = min(max(per_page, 1), 200)

        filters = _parse_user_activity_filters(request.args)

        query = UserActivityLog.query.options(joinedload(UserActivityLog.user))
        query = _apply_user_activity_filters(query, filters)

        total = query.count()
        total_pages = (total + per_page - 1) // per_page if total else 0

        logs = query.order_by(UserActivityLog.created_at.desc())\
            .offset((page - 1) * per_page).limit(per_page).all()

        log_list = []
        for log in logs:
            log_list.append({
                'id': log.id,
                'user_id': log.user_id,
                'user_name': log.user.full_name if log.user else 'System',
                'user_role': log.user.role if log.user else 'Unknown',
                'action': log.action,
                'entity_type': log.entity_type,
                'entity_id': log.entity_id,
                'entity_name': log.entity_name,
                'details': log.details,
                'ip_address': log.ip_address,
                'user_agent': log.user_agent,
                'created_at': (log.created_at.isoformat() + 'Z') if log.created_at else None,
                'created_at_iso': (log.created_at.isoformat() + 'Z') if log.created_at else None,
            })

        # Filter options for client-side dropdowns
        all_users = User.query.filter_by(is_archived=False).order_by(User.full_name).all()
        all_actions = db.session.query(UserActivityLog.action)\
            .filter(UserActivityLog.action.isnot(None), UserActivityLog.action != '')\
            .distinct().order_by(UserActivityLog.action).all()
        all_entities = db.session.query(UserActivityLog.entity_type)\
            .filter(UserActivityLog.entity_type.isnot(None), UserActivityLog.entity_type != '')\
            .distinct().order_by(UserActivityLog.entity_type).all()

        # Action breakdown stats for current filter set
        stats = {}
        if total > 0:
            action_breakdown = _apply_user_activity_filters(
                db.session.query(UserActivityLog.action, func.count(UserActivityLog.id)),
                filters
            ).group_by(UserActivityLog.action).all()
            stats['by_action'] = {action_name: count for action_name, count in action_breakdown}

        has_prev = page > 1
        has_next = total_pages > 0 and page < total_pages

        return jsonify({
            'success': True,
            'logs': log_list,
            'total': total,
            'page': page,
            'per_page': per_page,
            'total_pages': total_pages,
            'stats': stats,
            'pagination': {
                'page': page,
                'per_page': per_page,
                'total': total,
                'pages': total_pages,
                'has_prev': has_prev,
                'has_next': has_next,
            },
            'filters': {
                'users': [{'id': u.id, 'name': u.full_name, 'role': u.role} for u in all_users],
                'actions': [a[0] for a in all_actions],
                'entities': [e[0] for e in all_entities],
            }
        })

    except Exception as e:
        return jsonify({
            'success': False,
            'message': f'Error fetching user activity: {str(e)}'
        }), 500


@reports_bp.route('/api/user-activity/stats')
@login_required
@role_required('admin', 'super_admin')
def get_user_activity_stats():
    """Get user activity statistics (admin only)"""
    try:
        filters = _parse_user_activity_filters(request.args)

        total_actions = _apply_user_activity_filters(UserActivityLog.query, filters).count()

        actions_by_type = _apply_user_activity_filters(
            db.session.query(UserActivityLog.action, func.count(UserActivityLog.id)),
            filters
        ).group_by(UserActivityLog.action).all()

        actions_by_entity = _apply_user_activity_filters(
            db.session.query(UserActivityLog.entity_type, func.count(UserActivityLog.id)),
            filters
        ).group_by(UserActivityLog.entity_type).all()

        most_active_users = _apply_user_activity_filters(
            db.session.query(
                User.full_name,
                User.role,
                func.count(UserActivityLog.id).label('action_count')
            ).join(UserActivityLog, UserActivityLog.user_id == User.id),
            filters
        ).group_by(User.id, User.full_name, User.role)\
         .order_by(func.count(UserActivityLog.id).desc())\
         .limit(10).all()

        yesterday = datetime.utcnow() - timedelta(days=1)
        recent_actions = _apply_user_activity_filters(
            UserActivityLog.query.filter(UserActivityLog.created_at >= yesterday),
            filters
        ).count()

        return jsonify({
            'success': True,
            'stats': {
                'total_actions': total_actions,
                'recent_actions_24h': recent_actions,
                'actions_by_type': [{'action': a[0], 'count': a[1]} for a in actions_by_type],
                'actions_by_entity': [{'entity': e[0], 'count': e[1]} for e in actions_by_entity],
                'most_active_users': [
                    {'name': u[0], 'role': u[1], 'actions': u[2]} 
                    for u in most_active_users
                ]
            }
        })

    except Exception as e:
        return jsonify({
            'success': False,
            'message': f'Error fetching activity stats: {str(e)}'
        }), 500


@reports_bp.route('/api/user-activity/export')
@login_required
@role_required('admin', 'super_admin')
def export_user_activity():
    """Export filtered user activity logs to Excel (.xlsx)."""
    try:
        filters = _parse_user_activity_filters(request.args)

        query = UserActivityLog.query.options(joinedload(UserActivityLog.user))
        query = _apply_user_activity_filters(query, filters)
        logs = query.order_by(UserActivityLog.created_at.desc()).limit(10000).all()

        wb = Workbook()
        ws = wb.active
        ws.title = 'Activity Logs'

        border = Border(
            left=Side(style='thin', color='DEE2E6'),
            right=Side(style='thin', color='DEE2E6'),
            top=Side(style='thin', color='DEE2E6'),
            bottom=Side(style='thin', color='DEE2E6'),
        )
        header_font = Font(name='Arial', size=9, bold=True, color='FFFFFF')
        header_fill = PatternFill(start_color='495057', end_color='495057', fill_type='solid')
        header_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
        data_font = Font(name='Arial', size=9)
        data_align = Alignment(horizontal='left', vertical='center', wrap_text=True)
        center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
        zebra_fill = PatternFill(start_color='F5F6F8', end_color='F5F6F8', fill_type='solid')

        action_fills = {
            'login': PatternFill(start_color='D4EDDA', end_color='D4EDDA', fill_type='solid'),
            'logout': PatternFill(start_color='E2E3E5', end_color='E2E3E5', fill_type='solid'),
            'created': PatternFill(start_color='CCE5FF', end_color='CCE5FF', fill_type='solid'),
            'edited': PatternFill(start_color='FFF3CD', end_color='FFF3CD', fill_type='solid'),
            'deleted': PatternFill(start_color='F8D7DA', end_color='F8D7DA', fill_type='solid'),
            'archived': PatternFill(start_color='FFE5D0', end_color='FFE5D0', fill_type='solid'),
        }

        from app.services.export_service import create_posting_style_excel_header

        current_row = create_posting_style_excel_header(
            ws,
            report_title='SYSTEM ACTIVITY LOG REPORT',
            office_name='OFFICE OF THE SYSTEM ADMINISTRATOR',
            subtitle=f'Generated: {datetime.utcnow().strftime("%B %d, %Y  %I:%M %p")} UTC',
            last_col='G'
        )

        filters_applied = []
        if filters.get('user_id'):
            filters_applied.append(f"User ID: {filters['user_id']}")
        if filters.get('action'):
            filters_applied.append(f"Action: {filters['action']}")
        if filters.get('entity_type'):
            filters_applied.append(f"Entity: {filters['entity_type']}")
        if filters.get('date_from'):
            filters_applied.append(f"From: {filters['date_from'].strftime('%Y-%m-%d')}")
        if filters.get('date_to_exclusive'):
            filters_applied.append(
                f"To: {(filters['date_to_exclusive'] - timedelta(days=1)).strftime('%Y-%m-%d')}"
            )
        if filters.get('search'):
            filters_applied.append(f"Search: {filters['search']}")
        if filters.get('ip_address'):
            filters_applied.append(f"IP: {filters['ip_address']}")

        if filters_applied:
            ws.merge_cells(f'A{current_row}:G{current_row}')
            filter_cell = ws.cell(row=current_row, column=1, value=f"Filters: {' | '.join(filters_applied)}")
            filter_cell.font = Font(name='Arial', size=8, italic=True, color='6C757D')
            filter_cell.alignment = Alignment(horizontal='center')
            current_row += 1

        ws.merge_cells(f'A{current_row}:G{current_row}')
        total_cell = ws.cell(row=current_row, column=1, value=f'Total Records: {len(logs)}')
        total_cell.font = Font(name='Arial', size=9, bold=True, color='1F4788')
        total_cell.alignment = Alignment(horizontal='left')
        current_row += 1

        banner_fill = PatternFill(start_color='1F4788', end_color='1F4788', fill_type='solid')
        ws.merge_cells(f'A{current_row}:G{current_row}')
        banner_cell = ws.cell(row=current_row, column=1, value='ACTIVITY LOG ENTRIES')
        banner_cell.font = Font(name='Arial', size=10, bold=True, color='FFFFFF')
        banner_cell.fill = banner_fill
        banner_cell.alignment = Alignment(horizontal='center', vertical='center')
        banner_cell.border = border
        ws.row_dimensions[current_row].height = 24
        current_row += 1

        columns = [
            ('#', 5),
            ('Timestamp', 20),
            ('User', 22),
            ('Action', 16),
            ('Entity Type', 16),
            ('Entity Name', 28),
            ('Details', 40),
        ]
        header_row = current_row
        for index, (column_name, column_width) in enumerate(columns, start=1):
            cell = ws.cell(row=current_row, column=index, value=column_name)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_align
            cell.border = border
            ws.column_dimensions[get_column_letter(index)].width = column_width
        current_row += 1

        for index, log in enumerate(logs, start=1):
            row_data = [
                index,
                log.created_at.strftime('%Y-%m-%d %H:%M:%S') if log.created_at else '',
                log.user.full_name if log.user else 'System',
                log.action or '',
                log.entity_type or '',
                log.entity_name or '',
                log.details or '',
            ]
            for column_index, value in enumerate(row_data, start=1):
                cell = ws.cell(row=current_row, column=column_index, value=value)
                cell.font = data_font
                cell.alignment = center_align if column_index <= 5 else data_align
                cell.border = border
                if index % 2 == 0:
                    cell.fill = zebra_fill

            action_value = (log.action or '').lower()
            if action_value in action_fills:
                ws.cell(row=current_row, column=4).fill = action_fills[action_value]

            current_row += 1

        last_data_row = max(current_row - 1, header_row)
        table_ref = f'A{header_row}:G{last_data_row}'
        table = XLTable(displayName='ActivityLogs', ref=table_ref)
        table.tableStyleInfo = TableStyleInfo(
            name='TableStyleLight1',
            showFirstColumn=False,
            showLastColumn=False,
            showRowStripes=False,
            showColumnStripes=False,
        )
        ws.add_table(table)
        ws.freeze_panes = f'A{header_row + 1}'

        current_row += 1
        ws.merge_cells(f'A{current_row}:G{current_row}')
        footer_cell = ws.cell(
            row=current_row,
            column=1,
            value=f'Report generated by iSchedWise Reports — {datetime.utcnow().strftime("%Y-%m-%d %H:%M:%S")} UTC'
        )
        footer_cell.font = Font(name='Arial', size=8, italic=True, color='6C757D')
        footer_cell.alignment = Alignment(horizontal='center')

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        filename = f"activity_logs_{datetime.utcnow().strftime('%Y%m%d_%H%M%S')}.xlsx"
        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=filename,
        )
    except Exception as e:
        return jsonify({'success': False, 'message': f'Error exporting activity logs: {str(e)}'}), 500


# ============================================================================
# DAILY FACULTY SCHEDULE EXPORT (Program-wide, per day)
# ============================================================================

@reports_bp.route('/export/faculty-schedule')
@login_required
def export_daily_faculty_schedule():
    """Export a program-wide daily faculty schedule to Excel"""
    from app.services.export_service import generate_daily_faculty_schedule_excel

    program_id = request.args.get('program_id', type=int)
    day_of_week = request.args.get('day', type=str)

    if not program_id or not day_of_week:
        return jsonify({'error': 'program_id and day are required'}), 400

    valid_days = AcademicSettings.get_active_operation_days()
    # Normalize input
    day_of_week = day_of_week.strip().title()
    if day_of_week not in valid_days:
        return jsonify({'error': f'Invalid day. Must be one of: {", ".join(valid_days)}'}), 400

    program = Program.query.get_or_404(program_id)

    # Dean access check
    user_program_ids = current_user.get_program_ids()
    if user_program_ids is not None and program_id not in user_program_ids:
        return jsonify({'error': 'You do not have access to this program.'}), 403

    # Get active academic settings
    current_settings = AcademicSettings.query.filter_by(is_active=True).first()
    if not current_settings:
        return jsonify({'error': 'No active academic settings found.'}), 400

    # Get all active faculty who have schedules in this program on the given day
    # Faculty doesn't have program_id; the link is Schedule → Section → Program
    faculties = (
        Faculty.query
        .join(Schedule, Schedule.faculty_id == Faculty.id)
        .join(Section, Schedule.section_id == Section.id)
        .filter(
            Section.program_id == program_id,
            Schedule.is_active == True,
            Schedule.academic_year == current_settings.academic_year,
            Schedule.semester == current_settings.semester,
            Schedule.day_of_week == day_of_week,
            Faculty.is_archived == False,
        )
        .distinct()
        .order_by(Faculty.last_name, Faculty.first_name)
        .all()
    )

    # Build schedule data for each faculty
    faculty_schedule_data = []

    for faculty in faculties:
        schedules = (
            Schedule.query
            .join(Section, Schedule.section_id == Section.id)
            .filter(
                Schedule.faculty_id == faculty.id,
                Schedule.is_active == True,
                Schedule.academic_year == current_settings.academic_year,
                Schedule.semester == current_settings.semester,
                Schedule.day_of_week == day_of_week,
                Section.program_id == program_id,
            )
            .options(
                db.joinedload(Schedule.subject),
                db.joinedload(Schedule.section),
                db.joinedload(Schedule.room)
            )
            .order_by(Schedule.start_time)
            .all()
        )

        if not schedules:
            continue

        # Build raw rows sorted chronologically (already ordered by start_time)
        raw_rows = []
        for sched in schedules:
            from datetime import datetime as dt
            subject_code = sched.subject.subject_code if sched.subject else ''
            sched_type = (sched.schedule_type or 'lecture').lower()
            is_lab = sched_type in ('laboratory', 'lab')
            section_str = sched.section.full_section_name if sched.section else ''
            room_str = sched.room.room_number if sched.room else ''

            raw_rows.append({
                'subject_code': subject_code,
                'is_lab': is_lab,
                'section': section_str,
                'room': room_str,
                'start_time': sched.start_time,
                'end_time': sched.end_time,
            })

        # Merge consecutive lec+lab (or lab+lec) for same subject, section, room
        merged_rows = []
        i = 0
        while i < len(raw_rows):
            curr = raw_rows[i]
            # Check if next row can be merged (same subject, section, room, different type, consecutive)
            if i + 1 < len(raw_rows):
                nxt = raw_rows[i + 1]
                same_subject = curr['subject_code'] == nxt['subject_code'] and curr['subject_code'] != ''
                same_section = curr['section'] == nxt['section']
                same_room = curr['room'] == nxt['room']
                different_type = curr['is_lab'] != nxt['is_lab']
                consecutive = curr['end_time'] == nxt['start_time'] if (curr['end_time'] and nxt['start_time']) else False

                if same_subject and same_section and same_room and different_type and consecutive:
                    # Merge into one "Lec & Lab" row with combined time
                    combined_start = curr['start_time']
                    combined_end = nxt['end_time']
                    merged_rows.append({
                        'subject_code': curr['subject_code'],
                        'suffix': 'Lec & Lab',
                        'section': curr['section'],
                        'room': curr['room'],
                        'start_time': combined_start,
                        'end_time': combined_end,
                    })
                    i += 2
                    continue

            # No merge — single row
            suffix = 'Lab' if curr['is_lab'] else 'Lec'
            merged_rows.append({
                'subject_code': curr['subject_code'],
                'suffix': suffix,
                'section': curr['section'],
                'room': curr['room'],
                'start_time': curr['start_time'],
                'end_time': curr['end_time'],
            })
            i += 1

        # Format final rows
        rows = []
        for mr in merged_rows:
            time_str = ''
            if mr['start_time'] and mr['end_time']:
                from datetime import datetime as dt
                start_fmt = dt.combine(dt.today(), mr['start_time']).strftime('%I:%M %p').lstrip('0')
                end_fmt = dt.combine(dt.today(), mr['end_time']).strftime('%I:%M %p').lstrip('0')
                time_str = f"{start_fmt}-{end_fmt}"

            subject_str = f"{mr['subject_code']} {mr['suffix']}" if mr['subject_code'] else ''

            rows.append({
                'time': time_str,
                'subject': subject_str,
                'section': mr['section'],
                'room': mr['room'],
            })

        if not rows:
            continue

        faculty_schedule_data.append({
            'faculty': faculty,
            'rows': rows,
        })

    output, filename = generate_daily_faculty_schedule_excel(
        program, day_of_week, faculty_schedule_data, current_settings, current_user
    )

    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=filename
    )


# ============================================================================
# EXCEL EXPORT WITH ISO 25010 COMPLIANCE - ENHANCED VERSION
# ============================================================================

@reports_bp.route('/export/excel')
@login_required
def export_excel():
    """Export comprehensive reports to Excel with executive summary, detailed analytics, and insights"""
    try:
        # Get current academic settings
        current_settings = AcademicSettings.query.filter_by(is_active=True).first()
        
        if current_settings:
            academic_year = current_settings.academic_year
            semester = current_settings.semester
            exam_period = current_settings.exam_period
        else:
            academic_year = "N/A"
            semester = "N/A"
            exam_period = "N/A"
        
        # Get user's program access
        user_program_ids = current_user.get_program_ids()
        
        # Get filter parameters
        filter_department = request.args.get('program', type=int)
        
        # Calculate statistics
        stats = calculate_statistics(
            academic_year if academic_year != "N/A" else None,
            semester if semester != "N/A" else None,
            user_program_ids,
            filter_department
        )
        
        # Get program name for header
        program_name = "All Departments"
        if filter_department:
            program = Program.query.get(filter_department)
            if program:
                program_name = f"{program.program_name} ({program.program_code})"
        
        # Create workbook with comprehensive sheets
        wb = Workbook()
        
        # Sheet 1: Executive Summary (Overview + Key Insights)
        ws_summary = wb.active
        ws_summary.title = "Executive Summary"
        create_executive_summary_sheet(ws_summary, stats, academic_year, semester, program_name)
        
        # Sheet 2: Faculty Workload Analysis
        ws_faculty = wb.create_sheet("Faculty Analysis")
        create_enhanced_faculty_sheet(ws_faculty, stats, academic_year, semester, program_name)
        
        # Sheet 3: Room Utilization Analysis
        ws_rooms = wb.create_sheet("Room Analysis")
        create_enhanced_room_sheet(ws_rooms, stats, academic_year, semester, program_name)
        
        # Sheet 4: Weekly Distribution
        ws_weekly = wb.create_sheet("Weekly Distribution")
        create_enhanced_weekly_sheet(ws_weekly, stats, academic_year, semester, program_name)
        
        # Save to BytesIO
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        # Generate filename
        filename = f"iSchedWise_Report_{academic_year}_{semester}"
        if filter_department:
            dept = Program.query.get(filter_department)
            if dept:
                filename += f"_{dept.program_code}"
        filename += f"_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        
        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=filename
        )
        
    except Exception as e:
        print(f"Error exporting Excel: {str(e)}")
        import traceback
        traceback.print_exc()
        return jsonify({'error': str(e)}), 500


# ============================================================================
# EXCEL EXPORT — PROFESSIONAL REPORT FORMAT
# ============================================================================

# -- Shared Excel Styles --
_XL_PRIMARY = '1F4788'
_XL_HEADER_BG = '495057'
_XL_ZEBRA = 'F5F6F8'
_XL_BORDER_CLR = 'DEE2E6'
_XL_GREEN = 'D4EDDA'
_XL_YELLOW = 'FFF3CD'
_XL_RED = 'F8D7DA'
_XL_ORANGE = 'FFE5D0'
_XL_GRAY = 'E9ECEF'
_XL_MUTED = '6C757D'

_XL_DAY_COLORS = {
    'Monday': '4A90D9', 'Tuesday': '7C4DFF', 'Wednesday': '43A047',
    'Thursday': 'FB8C00', 'Friday': 'E91E63', 'Saturday': '78909C',
}

_xl_border = Border(
    left=Side(style='thin', color=_XL_BORDER_CLR),
    right=Side(style='thin', color=_XL_BORDER_CLR),
    top=Side(style='thin', color=_XL_BORDER_CLR),
    bottom=Side(style='thin', color=_XL_BORDER_CLR),
)


def _xl_header_cell(ws, row, col, value, merge_end_col=None):
    """Write a dark-gray column header cell."""
    cell = ws.cell(row=row, column=col, value=value)
    cell.font = Font(name='Arial', size=9, bold=True, color='FFFFFF')
    cell.fill = PatternFill(start_color=_XL_HEADER_BG, end_color=_XL_HEADER_BG, fill_type='solid')
    cell.alignment = Alignment(horizontal='center', vertical='center')
    cell.border = _xl_border
    if merge_end_col and merge_end_col > col:
        ws.merge_cells(start_row=row, start_column=col, end_row=row, end_column=merge_end_col)
    return cell


def _xl_section_banner(ws, row, text, start_col=1, end_col=5):
    """Write a blue section banner spanning multiple columns."""
    cell = ws.cell(row=row, column=start_col, value=text)
    cell.font = Font(name='Arial', size=10, bold=True, color='FFFFFF')
    cell.fill = PatternFill(start_color=_XL_PRIMARY, end_color=_XL_PRIMARY, fill_type='solid')
    cell.alignment = Alignment(horizontal='center', vertical='center')
    cell.border = _xl_border
    if end_col > start_col:
        ws.merge_cells(start_row=row, start_column=start_col, end_row=row, end_column=end_col)
    ws.row_dimensions[row].height = 24
    return row + 1


def _xl_data_cell(ws, row, col, value, bold=False, align='center', fill=None, number_format=None):
    """Write a styled data cell."""
    cell = ws.cell(row=row, column=col, value=value)
    cell.font = Font(name='Arial', size=9, bold=bold)
    cell.alignment = Alignment(horizontal=align, vertical='center', wrap_text=True)
    cell.border = _xl_border
    if fill:
        cell.fill = PatternFill(start_color=fill, end_color=fill, fill_type='solid')
    if number_format:
        cell.number_format = number_format
    return cell


def _xl_zebra_row(ws, row, col_start, col_end, row_idx):
    """Apply zebra striping to a row (even rows get light gray)."""
    if row_idx % 2 == 0:
        for c in range(col_start, col_end + 1):
            ws.cell(row=row, column=c).fill = PatternFill(
                start_color=_XL_ZEBRA, end_color=_XL_ZEBRA, fill_type='solid')


def _format_export_bullet(text):
    """Return a stable ASCII-safe bullet prefix for exported insight lines."""
    return f'-  {text}'


def create_reports_excel_header(ws, report_title, program_name, semester, academic_year, last_col='E'):
    """Create an institutional header with dual logos for Excel report sheets."""
    import os
    from openpyxl.drawing.image import Image as ExcelImage
    from app.services.export_service import get_institution_name

    base_dir = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
    images_dir = os.path.join(base_dir, 'static', 'images')
    
    # Left logo
    logo_left = os.path.join(images_dir, 'norzagaray-college-logo.png')
    if os.path.exists(logo_left):
        img = ExcelImage(logo_left)
        img.width = 75
        img.height = 75
        ws.add_image(img, 'A1')

    # Right logo
    logo_right = os.path.join(images_dir, 'bagong-pilipinas.png')
    if os.path.exists(logo_right):
        img = ExcelImage(logo_right)
        img.width = 80
        img.height = 80
        ws.add_image(img, f'{last_col}1')

    # Header text (next to logo)
    ws['B1'] = 'Republic of the Philippines'
    ws['B1'].font = Font(name='Arial', size=10)
    ws['B1'].alignment = Alignment(horizontal='left', vertical='center')

    ws['B2'] = 'Municipality of Norzagaray'
    ws['B2'].font = Font(name='Arial', size=10)
    ws['B2'].alignment = Alignment(horizontal='left', vertical='center')

    ws['B3'] = get_institution_name()
    ws['B3'].font = Font(name='Arial', size=10, bold=True)
    ws['B3'].alignment = Alignment(horizontal='left', vertical='center')

    ws['B4'] = program_name
    ws['B4'].font = Font(name='Arial', size=10, bold=True)
    ws['B4'].alignment = Alignment(horizontal='left', vertical='center')

    for r in range(1, 5):
        ws.row_dimensions[r].height = 17
    ws.row_dimensions[5].height = 8  # spacer

    # Report title (centered, row 6)
    last_col_idx = ord(last_col) - ord('A') + 1
    ws.merge_cells(start_row=6, start_column=1, end_row=6, end_column=last_col_idx)
    t_cell = ws.cell(row=6, column=1, value=report_title.upper())
    t_cell.font = Font(name='Arial', size=11, bold=True)
    t_cell.alignment = Alignment(horizontal='center', vertical='center')

    # Semester / AY (row 7)
    ws.merge_cells(start_row=7, start_column=1, end_row=7, end_column=last_col_idx)
    s_cell = ws.cell(row=7, column=1, value=f'{semester}, A.Y. {academic_year}')
    s_cell.font = Font(name='Arial', size=10, bold=True)
    s_cell.alignment = Alignment(horizontal='center', vertical='center')

    ws.row_dimensions[8].height = 8  # spacer
    return 9  # next available row


# -- Sheet 1: Executive Summary ---------------------------------------------------

def create_executive_summary_sheet(ws, stats, academic_year, semester, program_name):
    """Professional executive summary: KPIs, overview, insights."""

    # Column widths
    for col, w in [('A', 26), ('B', 16), ('C', 26), ('D', 16), ('E', 18)]:
        ws.column_dimensions[col].width = w

    current_row = create_reports_excel_header(ws, "SCHEDULING ANALYTICS REPORT",
                                               program_name, semester, academic_year, 'E')

    # Timestamp
    ws.merge_cells(f'A{current_row}:E{current_row}')
    ts_cell = ws.cell(row=current_row, column=1,
                      value=f'Generated: {datetime.now().strftime("%B %d, %Y  %I:%M %p")}')
    ts_cell.font = Font(name='Arial', size=8, italic=True, color=_XL_MUTED)
    ts_cell.alignment = Alignment(horizontal='center')
    current_row += 2

    # â”€â”€ KPIs â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€
    current_row = _xl_section_banner(ws, current_row, 'KEY PERFORMANCE INDICATORS', 1, 5)

    kpi_headers = ['Metric', 'Value', 'Status', 'Target', 'Variance']
    for ci, h in enumerate(kpi_headers, 1):
        _xl_header_cell(ws, current_row, ci, h)
    current_row += 1

    total_faculty = stats.get('total_faculty', 0)
    faculty_assigned = stats.get('faculty_with_schedules', 0)
    faculty_engagement = round(faculty_assigned / total_faculty * 100, 1) if total_faculty > 0 else 0
    total_rooms = stats.get('total_rooms', 0)
    rooms_in_use = stats.get('rooms_in_use', 0)
    room_engagement = round(rooms_in_use / total_rooms * 100, 1) if total_rooms > 0 else 0
    avg_room_util = stats.get('avg_room_utilization', 0)
    schedule_completion = stats.get('schedule_completion_rate', 0)

    kpis = [
        ('Schedule Completion', schedule_completion, 100, 80, 50),
        ('Faculty Engagement', faculty_engagement, 85, 80, 60),
        ('Room Utilization (Avg)', avg_room_util, 60, 50, 30),
        ('Room Engagement', room_engagement, 80, 70, 50),
    ]

    for name, val, target, good_t, warn_t in kpis:
        variance = val - target
        if val >= good_t:
            status, bg = 'Good', _XL_GREEN
        elif val >= warn_t:
            status, bg = 'Warning', _XL_YELLOW
        else:
            status, bg = 'Needs Attention', _XL_RED

        _xl_data_cell(ws, current_row, 1, name, bold=True, align='left')
        _xl_data_cell(ws, current_row, 2, f'{val}%', fill=bg)
        _xl_data_cell(ws, current_row, 3, status, fill=bg)
        _xl_data_cell(ws, current_row, 4, f'{target}%')
        _xl_data_cell(ws, current_row, 5, f'{variance:+.1f}%')
        current_row += 1

    current_row += 1

    # â”€â”€ Scheduling & Resource Overview â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€
    current_row = _xl_section_banner(ws, current_row, 'SCHEDULING & RESOURCE OVERVIEW', 1, 4)

    # Sub-headers
    for ci, h in [(1, 'Schedule Metrics'), (3, 'Resource Metrics')]:
        c = ws.cell(row=current_row, column=ci, value=h)
        c.font = Font(name='Arial', size=9, bold=True, color='FFFFFF')
        c.fill = PatternFill(start_color=_XL_PRIMARY, end_color=_XL_PRIMARY, fill_type='solid')
        c.alignment = Alignment(horizontal='center', vertical='center')
        c.border = _xl_border
        ws.merge_cells(start_row=current_row, start_column=ci, end_row=current_row, end_column=ci + 1)
    current_row += 1

    overview_rows = [
        ('Class Schedules', stats.get('total_schedules', 0),
         'Active Faculty', total_faculty),
        ('Exam Schedules', stats.get('total_exam_schedules', 0),
         'Faculty Assigned', faculty_assigned),
        ('Lecture Classes', stats.get('lecture_count', 0),
         'Unassigned Faculty', stats.get('unassigned_faculty_count', 0)),
        ('Lab Classes', stats.get('lab_count', 0),
         'Overloaded Faculty', stats.get('overloaded_faculty_count', 0)),
        ('Active Sections', stats.get('total_sections', 0),
         'Total Rooms', total_rooms),
        ('Sections Scheduled', stats.get('sections_with_schedules', 0),
         'Rooms In Use', rooms_in_use),
        ('', '',
         'Unused Rooms', stats.get('unused_rooms_count', 0)),
    ]

    for ri, (l1, v1, l2, v2) in enumerate(overview_rows):
        _xl_data_cell(ws, current_row, 1, l1, bold=True, align='left')
        _xl_data_cell(ws, current_row, 2, v1 if v1 != '' else '')
        _xl_data_cell(ws, current_row, 3, l2, bold=True, align='left')
        _xl_data_cell(ws, current_row, 4, v2)
        _xl_zebra_row(ws, current_row, 1, 4, ri)
        current_row += 1

    current_row += 1

    # â”€â”€ Key Insights â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€
    current_row = _xl_section_banner(ws, current_row, 'KEY INSIGHTS & RECOMMENDATIONS', 1, 5)

    insights = []
    uc = stats.get('unassigned_faculty_count', 0)
    if uc > 0:
        insights.append(f'{uc} faculty member(s) have no teaching assignments — review allocation.')
    ol = stats.get('overloaded_faculty_count', 0)
    if ol > 0:
        insights.append(f'{ol} faculty member(s) exceed maximum load — redistribute workload.')
    ur = stats.get('unused_rooms_count', 0)
    if ur > 0:
        insights.append(f'{ur} room(s) are not utilized — consider consolidating or repurposing.')
    if avg_room_util < 40:
        insights.append(f'Average room utilization is low ({avg_room_util}%) — maximize space usage.')
    if schedule_completion < 80:
        remaining = stats.get('total_sections', 0) - stats.get('sections_with_schedules', 0)
        insights.append(f'{remaining} section(s) still need scheduling.')
    if schedule_completion >= 95:
        insights.append('Excellent schedule completion rate — well organized!')
    if not insights:
        insights.append('All metrics are within acceptable ranges — continue monitoring.')

    for ins in insights:
        ws.merge_cells(f'A{current_row}:E{current_row}')
        c = ws.cell(row=current_row, column=1, value=_format_export_bullet(ins))
        c.font = Font(name='Arial', size=9, color='333333')
        c.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
        ws.row_dimensions[current_row].height = 20
        current_row += 1


# -- Sheet 2: Faculty Workload Analysis -------------------------------------------

def create_enhanced_faculty_sheet(ws, stats, academic_year, semester, program_name):
    """Professional faculty workload table with status coloring."""

    col_widths = [('A', 6), ('B', 32), ('C', 12), ('D', 9), ('E', 8),
                  ('F', 8), ('G', 9), ('H', 8), ('I', 9), ('J', 11)]
    for col, w in col_widths:
        ws.column_dimensions[col].width = w

    current_row = create_reports_excel_header(ws, "FACULTY WORKLOAD ANALYSIS",
                                               program_name, semester, academic_year, 'J')

    # Summary line
    ws.merge_cells(f'A{current_row}:J{current_row}')
    s = ws.cell(row=current_row, column=1,
                value=f'Total: {stats.get("total_faculty", 0)}  |  '
                      f'Assigned: {stats.get("faculty_with_schedules", 0)}  |  '
                      f'Unassigned: {stats.get("unassigned_faculty_count", 0)}  |  '
                      f'Overloaded: {stats.get("overloaded_faculty_count", 0)}')
    s.font = Font(name='Arial', size=9, bold=True, color=_XL_PRIMARY)
    s.alignment = Alignment(horizontal='center')
    current_row += 2

    # Column headers
    headers = ['#', 'Faculty Name', 'Program', 'Classes', 'Lec', 'Lab',
               'Total', 'Max', 'Util %', 'Status']
    for ci, h in enumerate(headers, 1):
        _xl_header_cell(ws, current_row, ci, h)
    ws.row_dimensions[current_row].height = 22
    current_row += 1

    # Data rows — ALL faculty
    faculty_workloads = stats.get('faculty_workloads', [])
    for idx, fac in enumerate(faculty_workloads, start=1):
        ls = fac.get('load_status', 'normal')
        if ls == 'exceeded':
            status_text, row_bg = 'Exceeded', _XL_RED
        elif ls == 'warning':
            status_text, row_bg = 'Warning', _XL_YELLOW
        elif fac['schedules'] > 0:
            status_text, row_bg = 'OK', _XL_GREEN
        else:
            status_text, row_bg = 'Unassigned', None

        vals = [
            idx, fac['name'], fac['program'], fac['schedules'],
            fac['lec_units'], fac['lab_units'], fac['total_units'],
            fac.get('max_units', 24), f"{fac.get('utilization_pct', 0)}%", status_text,
        ]
        aligns = ['center', 'left', 'center', 'center', 'center', 'center',
                  'center', 'center', 'center', 'center']
        for ci, (v, a) in enumerate(zip(vals, aligns), 1):
            _xl_data_cell(ws, current_row, ci, v, align=a)

        # Row coloring
        if row_bg:
            for ci in range(1, 11):
                ws.cell(row=current_row, column=ci).fill = PatternFill(
                    start_color=row_bg, end_color=row_bg, fill_type='solid')
        else:
            _xl_zebra_row(ws, current_row, 1, 10, idx)

        current_row += 1

    # Legend
    current_row += 1
    ws.merge_cells(f'A{current_row}:J{current_row}')
    lg = ws.cell(row=current_row, column=1,
                 value='Legend:  Exceeded = Over max units  |  Warning = 80-99% of max  '
                       '|  OK = Normal  |  Unassigned = No schedules')
    lg.font = Font(name='Arial', size=8, italic=True, color=_XL_MUTED)
    lg.alignment = Alignment(horizontal='center')


# -- Sheet 3: Room Utilization Analysis -------------------------------------------

def create_enhanced_room_sheet(ws, stats, academic_year, semester, program_name):
    """Professional room utilization with building summary + detail table."""

    col_widths = [('A', 6), ('B', 12), ('C', 20), ('D', 10), ('E', 9),
                  ('F', 9), ('G', 10), ('H', 9), ('I', 10)]
    for col, w in col_widths:
        ws.column_dimensions[col].width = w

    current_row = create_reports_excel_header(ws, "ROOM UTILIZATION ANALYSIS",
                                               program_name, semester, academic_year, 'I')

    # Summary line
    avg_util = stats.get('avg_room_utilization', 0)
    total_hours = stats.get('total_room_hours_used', 0)
    ws.merge_cells(f'A{current_row}:I{current_row}')
    s = ws.cell(row=current_row, column=1,
                value=f'Average Utilization: {avg_util}%  |  Total Hours: {total_hours} hrs  '
                      f'|  Rooms In Use: {stats.get("rooms_in_use", 0)}/{stats.get("total_rooms", 0)}')
    s.font = Font(name='Arial', size=9, bold=True, color=_XL_PRIMARY)
    s.alignment = Alignment(horizontal='center')
    current_row += 2

    # â”€â”€ Building Summary â”€â”€
    current_row = _xl_section_banner(ws, current_row, 'BUILDING UTILIZATION SUMMARY', 1, 6)

    bldg_headers = ['Building', 'Rooms', 'In Use', 'Hours Used', 'Max Hours', 'Utilization']
    for ci, h in enumerate(bldg_headers, 1):
        _xl_header_cell(ws, current_row, ci, h)
    current_row += 1

    building_util = stats.get('room_utilization_by_building', {})
    for ri, (bldg, data) in enumerate(building_util.items()):
        _xl_data_cell(ws, current_row, 1, bldg, align='left')
        _xl_data_cell(ws, current_row, 2, data.get('total', 0))
        _xl_data_cell(ws, current_row, 3, data.get('in_use', 0))
        _xl_data_cell(ws, current_row, 4, f"{data.get('total_hours', 0)} hrs")
        _xl_data_cell(ws, current_row, 5, f"{data.get('max_hours', 0)} hrs")
        _xl_data_cell(ws, current_row, 6, f"{data.get('utilization_pct', 0)}%")
        _xl_zebra_row(ws, current_row, 1, 6, ri)
        current_row += 1

    current_row += 1

    # â”€â”€ Detailed Room Table â”€â”€
    current_row = _xl_section_banner(ws, current_row, 'DETAILED ROOM UTILIZATION', 1, 9)

    room_headers = ['#', 'Room', 'Building', 'Type', 'Classes', 'Exams',
                    'Total Hrs', 'Util %', 'Status']
    for ci, h in enumerate(room_headers, 1):
        _xl_header_cell(ws, current_row, ci, h)
    current_row += 1

    room_utilizations = stats.get('room_utilizations', [])
    for idx, room in enumerate(room_utilizations, start=1):
        util_pct = room.get('utilization_pct', 0)
        if util_pct >= 60:
            status, bg = 'High', _XL_GREEN
        elif util_pct >= 30:
            status, bg = 'Medium', _XL_YELLOW
        elif util_pct > 0:
            status, bg = 'Low', _XL_ORANGE
        else:
            status, bg = 'Unused', _XL_GRAY

        vals = [
            idx, room.get('room', room.get('name', '')), room['building'],
            room.get('room_type', 'Lecture'), room.get('schedules', 0),
            room.get('exams', 0), room.get('total_hours', 0),
            f'{util_pct}%', status,
        ]
        for ci, v in enumerate(vals, 1):
            a = 'left' if ci in (2, 3) else 'center'
            _xl_data_cell(ws, current_row, ci, v, align=a)

        # Status-colored row
        for ci in range(1, 10):
            ws.cell(row=current_row, column=ci).fill = PatternFill(
                start_color=bg, end_color=bg, fill_type='solid')

        current_row += 1

    # Legend
    current_row += 1
    ws.merge_cells(f'A{current_row}:I{current_row}')
    lg = ws.cell(row=current_row, column=1,
                 value='Legend:  High (>=60%)  |  Medium (30-59%)  '
                       '|  Low (1-29%)  |  Unused (0%)')
    lg.font = Font(name='Arial', size=8, italic=True, color=_XL_MUTED)
    lg.alignment = Alignment(horizontal='center')


# -- Sheet 4: Weekly Schedule Distribution ----------------------------------------

def create_enhanced_weekly_sheet(ws, stats, academic_year, semester, program_name):
    """Professional weekly distribution with colored day labels."""

    for col, w in [('A', 14), ('B', 12), ('C', 12), ('D', 14)]:
        ws.column_dimensions[col].width = w

    current_row = create_reports_excel_header(ws, "WEEKLY SCHEDULE DISTRIBUTION",
                                               program_name, semester, academic_year, 'D')
    current_row += 1

    # Column headers
    for ci, h in enumerate(['Day', 'Schedules', 'Percentage', 'Load Level'], 1):
        _xl_header_cell(ws, current_row, ci, h)
    ws.row_dimensions[current_row].height = 22
    current_row += 1

    schedule_by_day = stats.get('schedule_by_day', {})
    day_order = AcademicSettings.get_active_operation_days()
    total_weekly = sum(schedule_by_day.values())
    num_days = len(day_order)
    avg_per_day = total_weekly / num_days if total_weekly > 0 and num_days > 0 else 0

    for day in day_order:
        count = schedule_by_day.get(day, 0)
        pct = round(count / total_weekly * 100, 1) if total_weekly > 0 else 0

        if count == 0:
            load = 'No Classes'
        elif count < avg_per_day * 0.7:
            load = 'Light'
        elif count < avg_per_day * 1.3:
            load = 'Moderate'
        else:
            load = 'Heavy'

        # Day cell (colored)
        day_clr = _XL_DAY_COLORS.get(day, _XL_HEADER_BG)
        c = _xl_data_cell(ws, current_row, 1, day, bold=True)
        c.font = Font(name='Arial', size=9, bold=True, color='FFFFFF')
        c.fill = PatternFill(start_color=day_clr, end_color=day_clr, fill_type='solid')

        _xl_data_cell(ws, current_row, 2, count)
        _xl_data_cell(ws, current_row, 3, f'{pct}%')
        _xl_data_cell(ws, current_row, 4, load)
        ws.row_dimensions[current_row].height = 20
        current_row += 1

    # Total row
    for ci, v in enumerate(['TOTAL', total_weekly, '100%', f'Avg: {avg_per_day:.1f}/day'], 1):
        c = _xl_data_cell(ws, current_row, ci, v, bold=True)
        c.font = Font(name='Arial', size=9, bold=True, color='FFFFFF')
        c.fill = PatternFill(start_color=_XL_PRIMARY, end_color=_XL_PRIMARY, fill_type='solid')
    ws.row_dimensions[current_row].height = 22

    # Insights
    current_row += 2
    current_row = _xl_section_banner(ws, current_row, 'DISTRIBUTION INSIGHTS', 1, 4)

    sorted_days = sorted(schedule_by_day.items(), key=lambda x: x[1], reverse=True)
    busiest = sorted_days[0] if sorted_days else ('N/A', 0)
    lightest = sorted_days[-1] if sorted_days else ('N/A', 0)

    weekly_insights = [
        f'Busiest day: {busiest[0]} ({busiest[1]} schedules, '
        f'{round(busiest[1] / total_weekly * 100, 1) if total_weekly else 0}%)',
        f'Lightest day: {lightest[0]} ({lightest[1]} schedules, '
        f'{round(lightest[1] / total_weekly * 100, 1) if total_weekly else 0}%)',
        f'Average per day: {avg_per_day:.1f}   |   Total weekly: {total_weekly}',
    ]
    if busiest[1] > 0 and lightest[1] > 0:
        ratio = busiest[1] / lightest[1]
        if ratio > 2:
            weekly_insights.append(f'Distribution is uneven (ratio {ratio:.1f}:1) — consider rebalancing.')
        else:
            weekly_insights.append(f'Distribution is well-balanced (ratio {ratio:.1f}:1).')

    for ins in weekly_insights:
        ws.merge_cells(f'A{current_row}:D{current_row}')
        c = ws.cell(row=current_row, column=1, value=_format_export_bullet(ins))
        c.font = Font(name='Arial', size=9, color='333333')
        c.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
        ws.row_dimensions[current_row].height = 18
        current_row += 1


# ============================================================================
# PDF EXPORT — PROFESSIONAL REPORT FORMAT
# ============================================================================

# -- Color Palette (consistent institutional branding) --
_PDF_PRIMARY = '#1F4788'       # Institution blue — section titles
_PDF_HEADER_BG = '#495057'     # Dark gray — table column headers
_PDF_ZEBRA = '#F5F6F8'         # Light gray — alternating rows
_PDF_ACCENT_GREEN = '#D4EDDA'  # Status: good
_PDF_ACCENT_YELLOW = '#FFF3CD' # Status: warning
_PDF_ACCENT_RED = '#F8D7DA'    # Status: exceeded
_PDF_BORDER = '#DEE2E6'        # Table grid lines
_PDF_MUTED = '#6C757D'         # Muted text

DAY_COLORS = {
    'Monday': '#4A90D9', 'Tuesday': '#7C4DFF', 'Wednesday': '#43A047',
    'Thursday': '#FB8C00', 'Friday': '#E91E63', 'Saturday': '#78909C',
}


def _add_page_footer(canvas, doc):
    """Draw page number and generation date at the bottom of every page."""
    canvas.saveState()
    canvas.setFont('Helvetica', 7)
    canvas.setFillColor(rl_colors.HexColor(_PDF_MUTED))
    # Page X of Y (Y filled in at final pass by reportlab)
    canvas.drawCentredString(
        doc.pagesize[0] / 2, 0.35 * inch,
        f"Page {canvas.getPageNumber()}"
    )
    canvas.drawString(
        doc.leftMargin, 0.35 * inch,
        f"Generated: {datetime.now().strftime('%B %d, %Y %I:%M %p')}"
    )
    canvas.drawRightString(
        doc.pagesize[0] - doc.rightMargin, 0.35 * inch,
        "iSchedWise V4 — Scheduling Analytics Report"
    )
    canvas.restoreState()


def _section_title(text, available_width):
    """Return a blue section-title table + thin rule line."""
    from reportlab.platypus import HRFlowable
    title_para = Paragraph(
        text,
        ParagraphStyle(
            'PDFSectionTitle', fontSize=12, fontName='Helvetica-Bold',
            textColor=rl_colors.HexColor(_PDF_PRIMARY), spaceAfter=2,
            spaceBefore=4, alignment=TA_LEFT,
        ),
    )
    rule = HRFlowable(
        width='100%', thickness=1.2,
        color=rl_colors.HexColor(_PDF_PRIMARY),
        spaceBefore=1, spaceAfter=6,
    )
    return [title_para, rule]


def _make_table(data, col_widths, header_rows=1, row_color_fn=None):
    """Build a consistently styled table.
    
    Args:
        data: list of rows (each row is a list of cell values)
        col_widths: list of column widths
        header_rows: how many rows are headers (default 1)
        row_color_fn: optional callable(row_index, row_data) -> hex color or None
    """
    tbl = Table(data, colWidths=col_widths, repeatRows=header_rows)
    style_cmds = [
        # Header row(s)
        ('BACKGROUND', (0, 0), (-1, header_rows - 1), rl_colors.HexColor(_PDF_HEADER_BG)),
        ('TEXTCOLOR', (0, 0), (-1, header_rows - 1), rl_colors.white),
        ('FONTNAME', (0, 0), (-1, header_rows - 1), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, header_rows - 1), 8),
        ('ALIGN', (0, 0), (-1, header_rows - 1), 'CENTER'),
        ('TOPPADDING', (0, 0), (-1, header_rows - 1), 5),
        ('BOTTOMPADDING', (0, 0), (-1, header_rows - 1), 5),
        # Body rows
        ('FONTNAME', (0, header_rows), (-1, -1), 'Helvetica'),
        ('FONTSIZE', (0, header_rows), (-1, -1), 8),
        ('TOPPADDING', (0, header_rows), (-1, -1), 4),
        ('BOTTOMPADDING', (0, header_rows), (-1, -1), 4),
        ('ALIGN', (0, header_rows), (-1, -1), 'CENTER'),
        # Grid for entire table
        ('GRID', (0, 0), (-1, -1), 0.5, rl_colors.HexColor(_PDF_BORDER)),
        # Zebra striping
        ('ROWBACKGROUNDS', (0, header_rows), (-1, -1),
         [rl_colors.white, rl_colors.HexColor(_PDF_ZEBRA)]),
    ]
    # Custom row coloring (overrides zebra)
    if row_color_fn:
        for i in range(header_rows, len(data)):
            color = row_color_fn(i, data[i])
            if color:
                style_cmds.append(('BACKGROUND', (0, i), (-1, i), rl_colors.HexColor(color)))
    tbl.setStyle(TableStyle(style_cmds))
    return tbl


def _create_report_header(academic_year, semester, program_name):
    """Create institutional PDF header for report pages."""
    from app.services.export_service import create_standard_pdf_header
    period_text = f"{semester}, A.Y. {academic_year}" if academic_year and semester else None
    return create_standard_pdf_header(
        report_title="SCHEDULING ANALYTICS REPORT",
        office_name=program_name if program_name else "OFFICE OF THE REGISTRAR",
        period_text=period_text,
    )


@reports_bp.route('/export/pdf')
@login_required
def export_pdf():
    """Export comprehensive scheduling analytics report to PDF."""
    try:
        # -- Gather data -------------------------------------------------------
        current_settings = AcademicSettings.query.filter_by(is_active=True).first()
        if current_settings:
            academic_year = current_settings.academic_year
            semester = current_settings.semester
        else:
            academic_year = "N/A"
            semester = "N/A"

        user_program_ids = current_user.get_program_ids()
        filter_department = request.args.get('program', type=int)

        stats = calculate_statistics(
            academic_year if academic_year != "N/A" else None,
            semester if semester != "N/A" else None,
            user_program_ids,
            filter_department,
        )

        program_name = "All Departments"
        program_name = "OFFICE OF THE REGISTRAR"
        if filter_department:
            program = Program.query.get(filter_department)
            if program:
                program_name = f"{program.program_name} ({program.program_code})"
                program_name = program.program_name.upper()

        # -- Document setup ----------------------------------------------------
        buffer = io.BytesIO()
        doc = SimpleDocTemplate(
            buffer, pagesize=A4,
            rightMargin=0.5 * inch, leftMargin=0.5 * inch,
            topMargin=0.6 * inch, bottomMargin=0.6 * inch,
        )
        styles = getSampleStyleSheet()
        AW = A4[0] - 1.0 * inch  # available width (A4 width minus margins)

        # -- Reusable paragraph styles -----------------------------------------
        subtitle_style = ParagraphStyle(
            'PDFSubtitle', parent=styles['Normal'], fontSize=9,
            textColor=rl_colors.HexColor(_PDF_MUTED), alignment=TA_CENTER,
            fontName='Helvetica', spaceAfter=4,
        )
        insight_style = ParagraphStyle(
            'PDFInsight', parent=styles['Normal'], fontSize=9,
            textColor=rl_colors.HexColor('#333333'), spaceAfter=3,
            leftIndent=12, fontName='Helvetica',
        )
        legend_style = ParagraphStyle(
            'PDFLegend', parent=styles['Normal'], fontSize=7,
            textColor=rl_colors.HexColor(_PDF_MUTED), alignment=TA_CENTER,
            fontName='Helvetica-Oblique', spaceBefore=4, spaceAfter=2,
        )

        elements = []

        # =====================================================================
        # PAGE 1 — EXECUTIVE SUMMARY
        # =====================================================================
        elements.extend(_create_report_header(academic_year, semester, program_name))
        elements.append(Paragraph(f'Program: {program_name}', subtitle_style))
        elements.append(Spacer(1, 0.12 * inch))

        # --- KPI cards as a clean table ---
        total_faculty = stats.get('total_faculty', 0)
        faculty_assigned = stats.get('faculty_with_schedules', 0)
        faculty_engagement = round((faculty_assigned / total_faculty * 100), 1) if total_faculty > 0 else 0
        total_rooms = stats.get('total_rooms', 0)
        rooms_in_use = stats.get('rooms_in_use', 0)
        room_engagement = round((rooms_in_use / total_rooms * 100), 1) if total_rooms > 0 else 0
        avg_room_util = stats.get('avg_room_utilization', 0)
        schedule_completion = stats.get('schedule_completion_rate', 0)

        def _status_text(value, good, warn):
            if value >= good:
                return 'Good'
            elif value >= warn:
                return 'Warning'
            return 'Needs Attention'

        def _status_bg(value, good, warn):
            if value >= good:
                return _PDF_ACCENT_GREEN
            elif value >= warn:
                return _PDF_ACCENT_YELLOW
            return _PDF_ACCENT_RED

        elements.extend(_section_title('KEY PERFORMANCE INDICATORS', AW))

        kpi_data = [
            ['Metric', 'Value', 'Status', 'Target'],
            ['Schedule Completion', f'{schedule_completion}%',
             _status_text(schedule_completion, 80, 50), '100%'],
            ['Faculty Engagement', f'{faculty_engagement}%',
             _status_text(faculty_engagement, 80, 60), '85%'],
            ['Room Utilization (Avg)', f'{avg_room_util}%',
             _status_text(avg_room_util, 50, 30), '60%'],
            ['Room Engagement', f'{room_engagement}%',
             _status_text(room_engagement, 70, 50), '80%'],
        ]
        cw = AW / 4
        def _kpi_row_color(i, row):
            thresholds = {1: (80, 50), 2: (80, 60), 3: (50, 30), 4: (70, 50)}
            if i in thresholds:
                val = [schedule_completion, faculty_engagement, avg_room_util, room_engagement][i - 1]
                return _status_bg(val, *thresholds[i])
            return None

        kpi_tbl = _make_table(kpi_data, [cw * 1.4, cw * 0.8, cw * 0.9, cw * 0.9])
        # Override status column backgrounds
        extra = []
        for ri, (val, good, warn) in enumerate([
            (schedule_completion, 80, 50), (faculty_engagement, 80, 60),
            (avg_room_util, 50, 30), (room_engagement, 70, 50),
        ], start=1):
            extra.append(('BACKGROUND', (2, ri), (2, ri),
                          rl_colors.HexColor(_status_bg(val, good, warn))))
        kpi_tbl.setStyle(TableStyle(extra))
        elements.append(kpi_tbl)
        elements.append(Spacer(1, 0.15 * inch))

        # --- Scheduling Overview + Resource Summary (side-by-side as 2Ã—4 grid) ---
        elements.extend(_section_title('SCHEDULING & RESOURCE OVERVIEW', AW))

        overview_data = [
            ['Schedule Metrics', '', 'Resource Metrics', ''],
            ['Class Schedules', str(stats.get('total_schedules', 0)),
             'Active Faculty', str(total_faculty)],
            ['Exam Schedules', str(stats.get('total_exam_schedules', 0)),
             'Faculty Assigned', str(faculty_assigned)],
            ['Lecture Classes', str(stats.get('lecture_count', 0)),
             'Unassigned Faculty', str(stats.get('unassigned_faculty_count', 0))],
            ['Lab Classes', str(stats.get('lab_count', 0)),
             'Overloaded Faculty', str(stats.get('overloaded_faculty_count', 0))],
            ['Active Sections', str(stats.get('total_sections', 0)),
             'Total Rooms', str(total_rooms)],
            ['Sections Scheduled', str(stats.get('sections_with_schedules', 0)),
             'Rooms In Use', str(rooms_in_use)],
            ['', '',
             'Unused Rooms', str(stats.get('unused_rooms_count', 0))],
        ]
        overview_tbl = Table(
            overview_data,
            colWidths=[cw * 1.3, cw * 0.7, cw * 1.3, cw * 0.7],
        )
        overview_tbl.setStyle(TableStyle([
            # Sub-headers row
            ('BACKGROUND', (0, 0), (1, 0), rl_colors.HexColor(_PDF_PRIMARY)),
            ('BACKGROUND', (2, 0), (3, 0), rl_colors.HexColor(_PDF_PRIMARY)),
            ('TEXTCOLOR', (0, 0), (-1, 0), rl_colors.white),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 8),
            ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
            # Body
            ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 1), (-1, -1), 8),
            ('ALIGN', (0, 1), (0, -1), 'LEFT'),
            ('ALIGN', (1, 1), (1, -1), 'CENTER'),
            ('ALIGN', (2, 1), (2, -1), 'LEFT'),
            ('ALIGN', (3, 1), (3, -1), 'CENTER'),
            ('TOPPADDING', (0, 0), (-1, -1), 4),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
            ('GRID', (0, 0), (-1, -1), 0.5, rl_colors.HexColor(_PDF_BORDER)),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1),
             [rl_colors.white, rl_colors.HexColor(_PDF_ZEBRA)]),
            # Vertical divider between left & right halves
            ('LINEAFTER', (1, 0), (1, -1), 1.5, rl_colors.HexColor(_PDF_PRIMARY)),
        ]))
        elements.append(overview_tbl)
        elements.append(Spacer(1, 0.15 * inch))

        # --- Key Insights (simple bullet list) ---
        elements.extend(_section_title('KEY INSIGHTS & RECOMMENDATIONS', AW))

        insights = []
        unassigned_count = stats.get('unassigned_faculty_count', 0)
        if unassigned_count > 0:
            insights.append(f"{unassigned_count} faculty member(s) have no teaching assignments — review allocation.")
        overloaded = stats.get('overloaded_faculty_count', 0)
        if overloaded > 0:
            insights.append(f"{overloaded} faculty member(s) exceed maximum load — redistribute workload.")
        unused_rooms = stats.get('unused_rooms_count', 0)
        if unused_rooms > 0:
            insights.append(f"{unused_rooms} room(s) are not utilized — consider consolidating or repurposing.")
        if avg_room_util < 40:
            insights.append(f"Average room utilization is low ({avg_room_util}%) — maximize space usage.")
        if schedule_completion < 80:
            remaining = stats.get('total_sections', 0) - stats.get('sections_with_schedules', 0)
            insights.append(f"{remaining} section(s) still need scheduling.")
        if schedule_completion >= 95:
            insights.append("Excellent schedule completion rate — well organized!")
        if not insights:
            insights.append("All metrics are within acceptable ranges — continue monitoring.")

        for ins in insights:
            elements.append(Paragraph(_format_export_bullet(ins), insight_style))

        elements.append(PageBreak())

        # =====================================================================
        # PAGE 2 — FACULTY WORKLOAD ANALYSIS
        # =====================================================================
        elements.extend(_create_report_header(academic_year, semester, program_name))
        elements.extend(_section_title('FACULTY WORKLOAD ANALYSIS', AW))

        fac_summary = (
            f"Total: {total_faculty}  |  Assigned: {faculty_assigned}  "
            f"|  Unassigned: {stats.get('unassigned_faculty_count', 0)}  "
            f"|  Overloaded: {stats.get('overloaded_faculty_count', 0)}"
        )
        elements.append(Paragraph(fac_summary, subtitle_style))
        elements.append(Spacer(1, 0.08 * inch))

        faculty_workloads = stats.get('faculty_workloads', [])
        if faculty_workloads:
            fac_data = [['#', 'Faculty Name', 'Dept', 'Classes', 'Lec', 'Lab', 'Total', 'Max', 'Status']]
            for idx, fac in enumerate(faculty_workloads, start=1):
                load_status = fac.get('load_status', 'normal')
                status_text = 'Over' if load_status == 'exceeded' else ('Warning' if load_status == 'warning' else 'OK')
                fac_name = fac['name'][:28] if len(fac['name']) > 28 else fac['name']
                fac_data.append([
                    str(idx), fac_name,
                    fac['program'][:8],
                    str(fac['schedules']),
                    f"{fac['lec_units']:.0f}",
                    f"{fac['lab_units']:.0f}",
                    f"{fac['total_units']:.0f}",
                    str(fac.get('max_units', 24)),
                    status_text,
                ])

            def _fac_color(i, row):
                real_idx = i - 1  # data rows start at index 1
                if real_idx < len(faculty_workloads):
                    ls = faculty_workloads[real_idx].get('load_status', 'normal')
                    if ls == 'exceeded':
                        return _PDF_ACCENT_RED
                    elif ls == 'warning':
                        return _PDF_ACCENT_YELLOW
                return None

            fac_tbl = _make_table(
                fac_data,
                [0.3 * inch, 1.9 * inch, 0.55 * inch, 0.5 * inch,
                 0.45 * inch, 0.45 * inch, 0.5 * inch, 0.45 * inch, 0.6 * inch],
                row_color_fn=_fac_color,
            )
            # Left-align faculty name column
            fac_tbl.setStyle(TableStyle([
                ('ALIGN', (1, 1), (1, -1), 'LEFT'),
            ]))
            elements.append(fac_tbl)
        else:
            elements.append(Paragraph('No faculty workload data available.', styles['Normal']))

        elements.append(Paragraph(
            'Legend:  OK = Normal load  |  Warning = 80-99% of max  |  Over = Exceeded max units',
            legend_style,
        ))

        elements.append(PageBreak())

        # =====================================================================
        # PAGE 3 — ROOM UTILIZATION ANALYSIS
        # =====================================================================
        elements.extend(_create_report_header(academic_year, semester, program_name))
        elements.extend(_section_title('ROOM UTILIZATION ANALYSIS', AW))

        total_hours = stats.get('total_room_hours_used', 0)
        room_summary = (
            f"Average Utilization: {avg_room_util}%  |  Total Hours Used: {total_hours} hrs  "
            f"|  Rooms In Use: {rooms_in_use}/{total_rooms}"
        )
        elements.append(Paragraph(room_summary, subtitle_style))
        elements.append(Spacer(1, 0.08 * inch))

        # --- Building Summary ---
        building_util = stats.get('room_utilization_by_building', {})
        if building_util:
            elements.extend(_section_title('Building Summary', AW))
            bldg_data = [['Building', 'Rooms', 'In Use', 'Hours Used', 'Max Hours', 'Utilization']]
            for bldg, data in building_util.items():
                bldg_data.append([
                    bldg[:22],
                    str(data.get('total', 0)),
                    str(data.get('in_use', 0)),
                    f"{data.get('total_hours', 0)} hrs",
                    f"{data.get('max_hours', 0)} hrs",
                    f"{data.get('utilization_pct', 0)}%",
                ])
            elements.append(_make_table(
                bldg_data,
                [1.6 * inch, 0.65 * inch, 0.65 * inch, 1.0 * inch, 1.0 * inch, 1.0 * inch],
            ))
            elements.append(Spacer(1, 0.12 * inch))

        # --- Detailed Room List ---
        room_utilizations = stats.get('room_utilizations', [])
        if room_utilizations:
            elements.extend(_section_title('Detailed Room Utilization', AW))
            room_data = [['#', 'Room', 'Building', 'Type', 'Classes', 'Exams', 'Hours', 'Util %', 'Status']]
            for idx, room in enumerate(room_utilizations, start=1):
                util_pct = room.get('utilization_pct', 0)
                if util_pct >= 60:
                    status = 'High'
                elif util_pct >= 30:
                    status = 'Medium'
                elif util_pct > 0:
                    status = 'Low'
                else:
                    status = 'Unused'
                room_data.append([
                    str(idx),
                    room.get('room', room.get('name', '')),
                    room['building'][:14],
                    room.get('room_type', 'Lecture')[:7],
                    str(room.get('schedules', 0)),
                    str(room.get('exams', 0)),
                    str(room.get('total_hours', 0)),
                    f"{util_pct}%",
                    status,
                ])

            def _room_color(i, row):
                real_idx = i - 1
                if real_idx < len(room_utilizations):
                    pct = room_utilizations[real_idx].get('utilization_pct', 0)
                    if pct >= 60:
                        return _PDF_ACCENT_GREEN
                    elif pct >= 30:
                        return _PDF_ACCENT_YELLOW
                    elif pct > 0:
                        return '#FFE5D0'
                    else:
                        return '#E9ECEF'
                return None

            room_tbl = _make_table(
                room_data,
                [0.3 * inch, 0.75 * inch, 1.1 * inch, 0.55 * inch,
                 0.5 * inch, 0.5 * inch, 0.55 * inch, 0.6 * inch, 0.6 * inch],
                row_color_fn=_room_color,
            )
            elements.append(room_tbl)
        else:
            elements.append(Paragraph('No room utilization data available.', styles['Normal']))

        elements.append(Paragraph(
            'Legend:  High (>=60%)  |  Medium (30-59%)  |  Low (1-29%)  |  Unused (0%)',
            legend_style,
        ))

        elements.append(PageBreak())

        # =====================================================================
        # PAGE 4 — WEEKLY SCHEDULE DISTRIBUTION
        # =====================================================================
        elements.extend(_create_report_header(academic_year, semester, program_name))
        elements.extend(_section_title('WEEKLY SCHEDULE DISTRIBUTION', AW))

        schedule_by_day = stats.get('schedule_by_day', {})
        day_order = AcademicSettings.get_active_operation_days()
        total_weekly = sum(schedule_by_day.values())
        num_days = len(day_order)
        avg_per_day = total_weekly / num_days if total_weekly > 0 and num_days > 0 else 0

        weekly_data = [['Day', 'Schedules', 'Percentage', 'Load Level']]
        for day in day_order:
            count = schedule_by_day.get(day, 0)
            pct = round((count / total_weekly * 100), 1) if total_weekly > 0 else 0
            if count == 0:
                load = 'No Classes'
            elif count < avg_per_day * 0.7:
                load = 'Light'
            elif count < avg_per_day * 1.3:
                load = 'Moderate'
            else:
                load = 'Heavy'
            weekly_data.append([day, str(count), f"{pct}%", load])
        weekly_data.append(['TOTAL', str(total_weekly), '100%', f'Avg: {avg_per_day:.1f}/day'])

        weekly_tbl = Table(
            weekly_data,
            colWidths=[1.5 * inch, 1.2 * inch, 1.2 * inch, 1.5 * inch],
            repeatRows=1,
        )
        weekly_style_cmds = [
            ('BACKGROUND', (0, 0), (-1, 0), rl_colors.HexColor(_PDF_HEADER_BG)),
            ('TEXTCOLOR', (0, 0), (-1, 0), rl_colors.white),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 9),
            ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 1), (-1, -1), 9),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('TOPPADDING', (0, 0), (-1, -1), 6),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
            ('GRID', (0, 0), (-1, -1), 0.5, rl_colors.HexColor(_PDF_BORDER)),
            ('ROWBACKGROUNDS', (0, 1), (-1, -2),
             [rl_colors.white, rl_colors.HexColor(_PDF_ZEBRA)]),
            # Total row
            ('BACKGROUND', (0, len(weekly_data) - 1), (-1, len(weekly_data) - 1),
             rl_colors.HexColor(_PDF_PRIMARY)),
            ('TEXTCOLOR', (0, len(weekly_data) - 1), (-1, len(weekly_data) - 1),
             rl_colors.white),
            ('FONTNAME', (0, len(weekly_data) - 1), (-1, len(weekly_data) - 1),
             'Helvetica-Bold'),
        ]
        # Colored day labels
        for di, day in enumerate(day_order, start=1):
            c = DAY_COLORS.get(day, _PDF_HEADER_BG)
            weekly_style_cmds.append(('BACKGROUND', (0, di), (0, di), rl_colors.HexColor(c)))
            weekly_style_cmds.append(('TEXTCOLOR', (0, di), (0, di), rl_colors.white))
            weekly_style_cmds.append(('FONTNAME', (0, di), (0, di), 'Helvetica-Bold'))
        weekly_tbl.setStyle(TableStyle(weekly_style_cmds))
        elements.append(weekly_tbl)
        elements.append(Spacer(1, 0.15 * inch))

        # --- Distribution Insights ---
        elements.extend(_section_title('Distribution Insights', AW))

        sorted_days = sorted(schedule_by_day.items(), key=lambda x: x[1], reverse=True)
        busiest = sorted_days[0] if sorted_days else ('N/A', 0)
        lightest = sorted_days[-1] if sorted_days else ('N/A', 0)

        weekly_insights = [
            f"Busiest day: {busiest[0]} with {busiest[1]} schedule(s) "
            f"({round(busiest[1] / total_weekly * 100, 1) if total_weekly else 0}% of total).",
            f"Lightest day: {lightest[0]} with {lightest[1]} schedule(s) "
            f"({round(lightest[1] / total_weekly * 100, 1) if total_weekly else 0}% of total).",
            f"Average schedules per day: {avg_per_day:.1f}",
            f"Total weekly schedules: {total_weekly}",
        ]
        if busiest[1] > 0 and lightest[1] > 0:
            ratio = busiest[1] / lightest[1]
            if ratio > 2:
                weekly_insights.append(
                    f"Distribution is uneven (ratio {ratio:.1f}:1) — consider rebalancing.")
            else:
                weekly_insights.append(
                    f"Distribution is well-balanced (ratio {ratio:.1f}:1).")

        for wi in weekly_insights:
            elements.append(Paragraph(_format_export_bullet(wi), insight_style))

        # =====================================================================
        # BUILD PDF
        # =====================================================================
        doc.build(elements, onFirstPage=_add_page_footer, onLaterPages=_add_page_footer)
        buffer.seek(0)

        filename = f"iSchedWise_Report_{academic_year}_{semester}"
        if filter_department:
            dept = Program.query.get(filter_department)
            if dept:
                filename += f"_{dept.program_code}"
        filename += f"_{datetime.now().strftime('%Y%m%d_%H%M%S')}.pdf"

        return send_file(
            buffer, mimetype='application/pdf',
            as_attachment=True, download_name=filename,
        )

    except Exception as e:
        print(f"Error exporting PDF: {str(e)}")
        import traceback
        traceback.print_exc()
        return jsonify({'error': str(e)}), 500
