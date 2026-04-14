"""
Archive Routes - Manage archived schedules, curricula, and programs
"""
from flask import Blueprint, render_template, request, jsonify, flash, redirect, url_for
from flask_login import login_required, current_user
from app.extensions import db
from app.models.archive import Archive
from app.models.faculty import FacultySubjectAssignment, Faculty
from app.models.schedule import Schedule
from app.models.exam_schedule import ExamSchedule
from app.models.settings import AcademicSettings
from app.models.curriculum import Curriculum
from app.models.program import Program
from app.models.section import Section
from app.models.department import Department
from app.models.building import Building
from app.utils.activity_logger import log_unarchive
from datetime import datetime
from sqlalchemy import or_, and_

archive_bp = Blueprint('archive', __name__, url_prefix='/archive')


def user_can_manage_archive():
    """Only admins/super admins can restore or permanently delete archives."""
    return bool(getattr(current_user, 'is_admin', False))


def require_archive_manage_permission():
    """Return a standardized 403 response when archive manage permission is missing."""
    if user_can_manage_archive():
        return None
    return jsonify({
        'success': False,
        'message': 'You do not have permission to restore or delete archives'
    }), 403


@archive_bp.route('/')
@login_required
def index():
    """Display archive overview dashboard"""
    return render_template('archive/overview.html', can_manage_archive=user_can_manage_archive())


@archive_bp.route('/schedules')
@login_required
def schedules_page():
    """Display schedule archives page"""
    # Get all unique academic years from archives
    academic_years = db.session.query(Archive.academic_year).distinct().order_by(Archive.academic_year.desc()).all()
    academic_years = [ay[0] for ay in academic_years if ay[0]]
    
    # Get current academic settings
    current_settings = AcademicSettings.query.filter_by(is_active=True).first()
    
    # Get programs for program filter
    user_program_ids = current_user.get_program_ids()
    if user_program_ids is not None:
        programs = Program.query.filter(
            Program.id.in_(user_program_ids),
            Program.is_archived == False
        ).order_by(Program.program_code).all()
    else:
        programs = Program.query.filter_by(is_archived=False).order_by(Program.program_code).all()

    # Get departments for faculty filter (restricted to user's accessible programs)
    if user_program_ids is not None:
        dept_ids = db.session.query(Program.department_id)\
            .filter(Program.id.in_(user_program_ids))\
            .distinct().all()
        dept_ids = [d[0] for d in dept_ids if d[0]]
        departments = Department.query.filter(Department.id.in_(dept_ids))\
            .order_by(Department.department_name).all()
    else:
        departments = Department.query.order_by(Department.department_name).all()

    # Get buildings for rooms filter
    buildings = Building.query.filter_by(is_archived=False).order_by(Building.building_name).all()

    return render_template('archive/schedules.html',
                         academic_years=academic_years,
                         current_settings=current_settings,
                         programs=programs,
                         departments=departments,
                         buildings=buildings,
                         can_manage_archive=user_can_manage_archive())


@archive_bp.route('/curriculum')
@login_required
def curriculum_page():
    """Display curriculum archives page"""
    return render_template('archive/curriculum.html', can_manage_archive=user_can_manage_archive())


@archive_bp.route('/programs')
@login_required
def programs_page():
    """Display program/program archives page"""
    return render_template('archive/programs.html', can_manage_archive=user_can_manage_archive())


@archive_bp.route('/faculty-members')
@login_required
def faculty_page():
    """Display faculty archives page"""
    return render_template('archive/faculty.html', can_manage_archive=user_can_manage_archive())


@archive_bp.route('/buildings-page')
@login_required
def buildings_page():
    """Display building archives page"""
    return render_template('archive/buildings.html', can_manage_archive=user_can_manage_archive())


@archive_bp.route('/api/archives')
@login_required
def get_archives():
    """Get archives with filters (API endpoint)"""
    try:
        # Get filter parameters
        academic_year = request.args.get('academic_year', '')
        semester = request.args.get('semester', '')
        exam_period = request.args.get('exam_period', '')
        schedule_type = request.args.get('schedule_type', '')  # 'lecture', 'lab', 'exam'
        section_id = request.args.get('section_id', type=int)
        faculty_id = request.args.get('faculty_id', type=int)
        room_id = request.args.get('room_id', type=int)
        building_id = request.args.get('building_id', type=int)
        program_id = request.args.get('program_id', type=int)
        search = request.args.get('search', '')
        
        # Get user's program access
        user_program_ids = current_user.get_program_ids()
        
        # Build query
        query = Archive.query
        
        # Filter by user's program access (through sections)
        if user_program_ids is not None:
            accessible_section_ids = db.session.query(Section.id)\
                .filter(Section.program_id.in_(user_program_ids))\
                .all()
            accessible_section_ids = [sid[0] for sid in accessible_section_ids]
            if accessible_section_ids:
                query = query.filter(Archive.section_id.in_(accessible_section_ids))
        
        # Apply filters
        if academic_year:
            query = query.filter(Archive.academic_year == academic_year)
        
        if semester:
            query = query.filter(Archive.semester == semester)
        
        if exam_period:
            query = query.filter(Archive.exam_period == exam_period)
        
        if schedule_type:
            # Handle multiple schedule types (e.g., 'lecture,lab')
            types = [t.strip() for t in schedule_type.split(',')]
            if len(types) > 1:
                query = query.filter(Archive.schedule_type.in_(types))
            else:
                query = query.filter(Archive.schedule_type == schedule_type)
        
        if section_id:
            query = query.filter(Archive.section_id == section_id)
        
        if faculty_id:
            query = query.filter(Archive.faculty_id == faculty_id)
        
        if room_id:
            query = query.filter(Archive.room_id == room_id)
        
        if building_id:
            # Get all room IDs for this building
            from app.models.building import Room
            room_ids = db.session.query(Room.id).filter_by(building_id=building_id).all()
            room_ids = [rid[0] for rid in room_ids]
            if room_ids:
                query = query.filter(Archive.room_id.in_(room_ids))
        
        if program_id:
            # Get all section IDs for this program
            section_ids = db.session.query(Section.id).filter_by(program_id=program_id).all()
            section_ids = [sid[0] for sid in section_ids]
            if section_ids:
                query = query.filter(Archive.section_id.in_(section_ids))
        
        if search:
            search_filter = or_(
                Archive.section_name.ilike(f'%{search}%'),
                Archive.subject_code.ilike(f'%{search}%'),
                Archive.course_description.ilike(f'%{search}%'),
                Archive.faculty_name.ilike(f'%{search}%'),
                Archive.room_number.ilike(f'%{search}%'),
                Archive.building_name.ilike(f'%{search}%'),
                Archive.program_name.ilike(f'%{search}%')
            )
            query = query.filter(search_filter)
        
        # Order by archived date (most recent first), eager load relationships used by to_dict()
        archives = query.options(
            db.joinedload(Archive.section).joinedload(Section.program),
            db.joinedload(Archive.faculty).joinedload(Faculty.department),
            db.joinedload(Archive.user)
        ).order_by(Archive.archived_at.desc()).all()
        
        # Convert to dictionary
        archives_data = [archive.to_dict() for archive in archives]
        
        return jsonify({
            'success': True,
            'archives': archives_data,
            'total': len(archives_data)
        })
    
    except Exception as e:
        return jsonify({
            'success': False,
            'message': f'Error fetching archives: {str(e)}'
        }), 500


@archive_bp.route('/api/archive/schedule/<int:schedule_id>', methods=['POST'])
@login_required
def archive_schedule(schedule_id):
    """Archive a class schedule"""
    try:
        schedule = Schedule.query.get_or_404(schedule_id)
        
        # Get archive reason from request
        data = request.get_json() or {}
        archive_reason = data.get('reason', 'Manual archive')
        
        # Normalize schedule_type: 'laboratory' -> 'lab'
        normalized_type = schedule.schedule_type
        if normalized_type and normalized_type.lower() in ['laboratory', 'lab']:
            normalized_type = 'lab'
        elif normalized_type and normalized_type.lower() == 'lecture':
            normalized_type = 'lecture'
        else:
            normalized_type = 'lecture'  # default
        
        # Create archive record
        archive = Archive(
            section_id=schedule.section_id,
            subject_id=schedule.subject_id,
            faculty_id=schedule.faculty_id,
            room_id=schedule.room_id,
            section_name=schedule.section.full_section_name if schedule.section else None,
            subject_code=schedule.subject.subject_code if schedule.subject else None,
            course_description=schedule.subject.course_description if schedule.subject else None,
            faculty_name=schedule.faculty.full_name if schedule.faculty else None,
            room_number=schedule.room.room_number if schedule.room else None,
            building_name=schedule.room.building.building_name if schedule.room and schedule.room.building else None,
            program_name=schedule.section.program.program_name if schedule.section and schedule.section.program else None,
            day_of_week=schedule.day_of_week,
            start_time=schedule.start_time,
            end_time=schedule.end_time,
            semester=schedule.semester,
            academic_year=schedule.academic_year,
            schedule_type=normalized_type,
            original_schedule_id=schedule.id,
            archived_by=current_user.id,
            archive_reason=archive_reason,
            archived_at=datetime.utcnow()
        )
        
        db.session.add(archive)
        
        # Delete the original schedule
        db.session.delete(schedule)
        
        db.session.commit()
        
        return jsonify({
            'success': True,
            'message': 'Schedule archived successfully',
            'archive_id': archive.id
        })
    
    except Exception as e:
        db.session.rollback()
        return jsonify({
            'success': False,
            'message': f'Error archiving schedule: {str(e)}'
        }), 500


@archive_bp.route('/api/archive/exam-schedule/<int:exam_schedule_id>', methods=['POST'])
@login_required
def archive_exam_schedule(exam_schedule_id):
    """Archive an exam schedule"""
    try:
        exam_schedule = ExamSchedule.query.get_or_404(exam_schedule_id)
        
        # Get archive reason from request
        data = request.get_json() or {}
        archive_reason = data.get('reason', 'Manual archive')
        
        # Create archive record
        archive = Archive(
            section_id=exam_schedule.section_id,
            subject_id=exam_schedule.subject_id,
            faculty_id=exam_schedule.faculty_id,
            room_id=exam_schedule.room_id,
            section_name=exam_schedule.section.full_section_name if exam_schedule.section else None,
            subject_code=exam_schedule.subject.subject_code if exam_schedule.subject else None,
            course_description=exam_schedule.subject.course_description if exam_schedule.subject else None,
            faculty_name=exam_schedule.faculty.full_name if exam_schedule.faculty else None,
            room_number=exam_schedule.room.room_number if exam_schedule.room else None,
            building_name=exam_schedule.room.building.building_name if exam_schedule.room and exam_schedule.room.building else None,
            program_name=exam_schedule.section.program.program_name if exam_schedule.section and exam_schedule.section.program else None,
            exam_date=exam_schedule.exam_date,
            start_time=exam_schedule.start_time,
            end_time=exam_schedule.end_time,
            semester=exam_schedule.semester,
            academic_year=exam_schedule.academic_year,
            schedule_type='exam',
            exam_period=exam_schedule.exam_period,
            original_schedule_id=exam_schedule.id,
            archived_by=current_user.id,
            archive_reason=archive_reason,
            archived_at=datetime.utcnow()
        )
        
        db.session.add(archive)
        
        # Delete the original exam schedule
        db.session.delete(exam_schedule)
        
        db.session.commit()
        
        return jsonify({
            'success': True,
            'message': 'Exam schedule archived successfully',
            'archive_id': archive.id
        })
    
    except Exception as e:
        db.session.rollback()
        return jsonify({
            'success': False,
            'message': f'Error archiving exam schedule: {str(e)}'
        }), 500


@archive_bp.route('/api/archive/bulk', methods=['POST'])
@login_required
def bulk_archive():
    """Bulk archive schedules by academic year and semester"""
    try:
        data = request.get_json()
        academic_year = data.get('academic_year')
        semester = data.get('semester')
        exam_period = data.get('exam_period')
        archive_reason = data.get('reason', 'Bulk archive')
        
        if not academic_year or not semester:
            return jsonify({
                'success': False,
                'message': 'Academic year and semester are required'
            }), 400
        
        archived_count = 0
        
        # Archive class schedules
        schedules = Schedule.query.filter_by(
            academic_year=academic_year,
            semester=semester,
            is_active=True
        ).all()
        
        for schedule in schedules:
            # Normalize schedule_type: 'laboratory' -> 'lab'
            normalized_type = schedule.schedule_type
            if normalized_type and normalized_type.lower() in ['laboratory', 'lab']:
                normalized_type = 'lab'
            elif normalized_type and normalized_type.lower() == 'lecture':
                normalized_type = 'lecture'
            else:
                normalized_type = 'lecture'  # default
            
            archive = Archive(
                section_id=schedule.section_id,
                subject_id=schedule.subject_id,
                faculty_id=schedule.faculty_id,
                room_id=schedule.room_id,
                section_name=schedule.section.full_section_name if schedule.section else None,
                subject_code=schedule.subject.subject_code if schedule.subject else None,
                course_description=schedule.subject.course_description if schedule.subject else None,
                faculty_name=schedule.faculty.full_name if schedule.faculty else None,
                room_number=schedule.room.room_number if schedule.room else None,
                building_name=schedule.room.building.building_name if schedule.room and schedule.room.building else None,
                program_name=schedule.section.program.program_name if schedule.section and schedule.section.program else None,
                day_of_week=schedule.day_of_week,
                start_time=schedule.start_time,
                end_time=schedule.end_time,
                semester=schedule.semester,
                academic_year=schedule.academic_year,
                schedule_type=normalized_type,
                original_schedule_id=schedule.id,
                archived_by=current_user.id,
                archive_reason=archive_reason,
                archived_at=datetime.utcnow()
            )
            db.session.add(archive)
            db.session.delete(schedule)
            archived_count += 1
        
        # Archive exam schedules
        exam_query = ExamSchedule.query.filter_by(
            academic_year=academic_year,
            semester=semester,
            is_active=True
        )
        
        if exam_period:
            exam_query = exam_query.filter_by(exam_period=exam_period)
        
        exam_schedules = exam_query.all()
        
        for exam_schedule in exam_schedules:
            archive = Archive(
                section_id=exam_schedule.section_id,
                subject_id=exam_schedule.subject_id,
                faculty_id=exam_schedule.faculty_id,
                room_id=exam_schedule.room_id,
                section_name=exam_schedule.section.full_section_name if exam_schedule.section else None,
                subject_code=exam_schedule.subject.subject_code if exam_schedule.subject else None,
                course_description=exam_schedule.subject.course_description if exam_schedule.subject else None,
                faculty_name=exam_schedule.faculty.full_name if exam_schedule.faculty else None,
                room_number=exam_schedule.room.room_number if exam_schedule.room else None,
                building_name=exam_schedule.room.building.building_name if exam_schedule.room and exam_schedule.room.building else None,
                program_name=exam_schedule.section.program.program_name if exam_schedule.section and exam_schedule.section.program else None,
                exam_date=exam_schedule.exam_date,
                start_time=exam_schedule.start_time,
                end_time=exam_schedule.end_time,
                semester=exam_schedule.semester,
                academic_year=exam_schedule.academic_year,
                schedule_type='exam',
                exam_period=exam_schedule.exam_period,
                original_schedule_id=exam_schedule.id,
                archived_by=current_user.id,
                archive_reason=archive_reason,
                archived_at=datetime.utcnow()
            )
            db.session.add(archive)
            db.session.delete(exam_schedule)
            archived_count += 1
        
        db.session.commit()
        
        return jsonify({
            'success': True,
            'message': f'Successfully archived {archived_count} schedules',
            'archived_count': archived_count
        })
    
    except Exception as e:
        db.session.rollback()
        return jsonify({
            'success': False,
            'message': f'Error bulk archiving: {str(e)}'
        }), 500


@archive_bp.route('/api/archive/<int:archive_id>', methods=['DELETE'])
@login_required
def delete_archive(archive_id):
    """Permanently delete an archive record"""
    try:
        permission_error = require_archive_manage_permission()
        if permission_error:
            return permission_error

        archive = Archive.query.get_or_404(archive_id)
        
        db.session.delete(archive)
        db.session.commit()
        
        return jsonify({
            'success': True,
            'message': 'Archive deleted permanently'
        })
    
    except Exception as e:
        db.session.rollback()
        return jsonify({
            'success': False,
            'message': f'Error deleting archive: {str(e)}'
        }), 500


@archive_bp.route('/api/archive/stats')
@login_required
def get_archive_stats():
    """Get archive statistics"""
    try:
        # Total archives
        total = Archive.query.count()
        
        # Archives by academic year
        by_year = db.session.query(
            Archive.academic_year,
            db.func.count(Archive.id)
        ).group_by(Archive.academic_year).all()
        
        # Archives by semester
        by_semester = db.session.query(
            Archive.semester,
            db.func.count(Archive.id)
        ).group_by(Archive.semester).all()
        
        # Archives by type
        by_type = db.session.query(
            Archive.schedule_type,
            db.func.count(Archive.id)
        ).group_by(Archive.schedule_type).all()
        
        # Debug: Get raw archive types
        all_archives = Archive.query.all()
        raw_types = [a.schedule_type for a in all_archives]
        
        print(f"DEBUG Archive Stats:")
        print(f"  Total archives: {total}")
        print(f"  Raw schedule types: {raw_types}")
        print(f"  Grouped by_type: {by_type}")
        
        return jsonify({
            'success': True,
            'stats': {
                'total': total,
                'by_year': {year: count for year, count in by_year if year},
                'by_semester': {sem: count for sem, count in by_semester if sem},
                'by_type': {type: count for type, count in by_type if type}
            }
        })
    
    except Exception as e:
        print(f"ERROR in get_archive_stats: {str(e)}")
        return jsonify({
            'success': False,
            'message': f'Error fetching stats: {str(e)}'
        }), 500


@archive_bp.route('/api/semester-counts')
@login_required
def get_semester_counts():
    """Return archive counts grouped by semester for a given academic year.

    Query params:
        academic_year (str): e.g. '2025-2026'
        schedule_type (str): 'lecture,lab' or 'exam' (optional)
    """
    try:
        academic_year = request.args.get('academic_year', '')
        schedule_type = request.args.get('schedule_type', '')
        program_id    = request.args.get('program_id', type=int)

        q = db.session.query(Archive.semester, db.func.count(Archive.id))

        # Respect user program access
        user_program_ids = current_user.get_program_ids()
        if user_program_ids is not None:
            accessible_depts = Program.query.filter(Program.id.in_(user_program_ids)).all()
            dept_names = [d.program_name for d in accessible_depts]
            q = q.filter(Archive.program_name.in_(dept_names))

        # Apply explicit program filter (mirrors the client-side program dropdown)
        if program_id:
            from app.models.section import Section
            section_ids = db.session.query(Section.id).filter_by(program_id=program_id).all()
            section_ids = [sid[0] for sid in section_ids]
            if section_ids:
                q = q.filter(Archive.section_id.in_(section_ids))
            else:
                # Program exists but has no sections — return empty counts
                return jsonify({'success': True, 'counts': {}})

        if academic_year:
            q = q.filter(Archive.academic_year == academic_year)

        if schedule_type:
            types = [t.strip() for t in schedule_type.split(',')]
            if len(types) > 1:
                q = q.filter(Archive.schedule_type.in_(types))
            else:
                q = q.filter(Archive.schedule_type == schedule_type)

        rows = q.group_by(Archive.semester).all()

        return jsonify({
            'success': True,
            'counts': {sem: count for sem, count in rows if sem}
        })

    except Exception as e:
        return jsonify({
            'success': False,
            'message': f'Error fetching semester counts: {str(e)}'
        }), 500


@archive_bp.route('/api/faculty-assignments')
@login_required
def get_faculty_assignment_archives():
    """Get archived faculty subject assignments with filters (API endpoint)"""
    try:
        # Get filter parameters
        academic_year = request.args.get('academic_year', '')
        semester = request.args.get('semester', '')
        faculty_id = request.args.get('faculty_id', type=int)
        program = request.args.get('program', '')
        search = request.args.get('search', '')
        
        # Build query against flag-based faculty_subject_assignments
        query = FacultySubjectAssignment.query.filter_by(is_archived=True)
        
        # Apply filters
        if academic_year:
            query = query.filter(FacultySubjectAssignment.academic_year == academic_year)

        if semester:
            query = query.filter(FacultySubjectAssignment.semester == semester)

        if faculty_id:
            query = query.filter(FacultySubjectAssignment.faculty_id == faculty_id)

        # Order by archived date (most recent first)
        assignments = query.order_by(FacultySubjectAssignment.archived_at.desc()).all()
        
        # Convert minimal data to dictionary
        archives_data = []
        for a in assignments:
            archives_data.append({
                'id': a.id,
                'faculty_id': a.faculty_id,
                'subject_id': a.subject_id,
                'faculty_name': a.faculty.full_name if a.faculty else None,
                'subject_code': a.subject.subject_code if a.subject else None,
                'course_description': a.subject.course_description if a.subject else None,
                'academic_year': a.academic_year,
                'semester': a.semester,
                'archive_reason': a.archive_reason,
                'archived_at': a.archived_at.strftime('%Y-%m-%d %H:%M:%S') if a.archived_at else None,
                'archived_by': a.archived_by
            })
        
        return jsonify({
            'success': True,
            'archives': archives_data,
            'total': len(archives_data)
        })
    
    except Exception as e:
        return jsonify({
            'success': False,
            'message': f'Error fetching archived faculty assignments: {str(e)}'
        }), 500


@archive_bp.route('/api/faculty-assignment/<int:archive_id>', methods=['DELETE'])
@login_required
def delete_faculty_assignment_archive(archive_id):
    """Permanently delete an archived faculty assignment record"""
    try:
        permission_error = require_archive_manage_permission()
        if permission_error:
            return permission_error

        assignment = FacultySubjectAssignment.query.get_or_404(archive_id)
        # permanently delete the assignment row
        db.session.delete(assignment)
        db.session.commit()
        
        return jsonify({
            'success': True,
            'message': 'Archived faculty assignment deleted successfully'
        })
    
    except Exception as e:
        db.session.rollback()
        return jsonify({
            'success': False,
            'message': f'Error deleting archived faculty assignment: {str(e)}'
        }), 500


@archive_bp.route('/api/faculty-assignment-stats')
@login_required
def get_faculty_assignment_archive_stats():
    """Get faculty assignment archive statistics"""
    try:
        # Total archived assignments
        total = FacultySubjectAssignment.query.filter_by(is_archived=True).count()
        
        # By academic year
        by_year = db.session.query(
            FacultySubjectAssignment.academic_year,
            db.func.count(FacultySubjectAssignment.id)
        ).filter(FacultySubjectAssignment.is_archived==True)\
         .group_by(FacultySubjectAssignment.academic_year)\
         .order_by(FacultySubjectAssignment.academic_year.desc()).all()
        
        # By semester
        by_semester = db.session.query(
            FacultySubjectAssignment.semester,
            db.func.count(FacultySubjectAssignment.id)
        ).filter(FacultySubjectAssignment.is_archived==True)\
         .group_by(FacultySubjectAssignment.semester).all()
        
        # By faculty
        by_faculty = db.session.query(
            FacultySubjectAssignment.faculty_id,
            db.func.count(FacultySubjectAssignment.id)
        ).filter(FacultySubjectAssignment.is_archived==True)\
         .group_by(FacultySubjectAssignment.faculty_id)\
         .order_by(db.func.count(FacultySubjectAssignment.id).desc())\
         .limit(10).all()
        
        return jsonify({
            'success': True,
            'stats': {
                'total': total,
                'by_year': {year: count for year, count in by_year if year},
                'by_semester': {sem: count for sem, count in by_semester if sem},
                'top_faculty': [{'id': fac_id, 'count': count} for fac_id, count in by_faculty if fac_id]
            }
        })
    
    except Exception as e:
        return jsonify({
            'success': False,
            'message': f'Error fetching faculty assignment archive stats: {str(e)}'
        }), 500


@archive_bp.route('/api/curricula')
@login_required
def get_archived_curricula():
    """Get archived curricula with filters (API endpoint)"""
    try:
        # Get filter parameters
        program_id = request.args.get('program_id', type=int)
        search = request.args.get('search', '')
        
        # Get user's program access
        user_program_ids = current_user.get_program_ids()
        
        # Build query for archived curricula
        query = Curriculum.query.filter_by(is_archived=True)
        
        # Filter by user's program access
        if user_program_ids is not None:
            query = query.filter(Curriculum.program_id.in_(user_program_ids))
        
        # Apply program filter
        if program_id:
            query = query.filter_by(program_id=program_id)
        
        # Apply search filter
        if search:
            search_term = f'%{search}%'
            query = query.filter(
                or_(
                    Curriculum.curriculum_code.ilike(search_term),
                    Curriculum.degree_program.ilike(search_term)
                )
            )
        
        # Order by archived date (most recent first)
        curricula = query.order_by(Curriculum.archived_at.desc()).all()
        
        # Convert to dictionary
        curricula_data = [curriculum.to_dict() for curriculum in curricula]
        
        return jsonify({
            'success': True,
            'curricula': curricula_data,
            'total': len(curricula_data)
        })
    
    except Exception as e:
        return jsonify({
            'success': False,
            'message': f'Error fetching archived curricula: {str(e)}'
        }), 500


@archive_bp.route('/api/curriculum/<int:curriculum_id>/unarchive', methods=['POST'])
@login_required
def unarchive_curriculum(curriculum_id):
    """Unarchive a curriculum"""
    try:
        permission_error = require_archive_manage_permission()
        if permission_error:
            return permission_error

        curriculum = Curriculum.query.get_or_404(curriculum_id)
        
        # Log current state for debugging
        print(f"\n=== UNARCHIVE CURRICULUM DEBUG ===")
        print(f"Curriculum ID: {curriculum_id}")
        print(f"Curriculum Code: {curriculum.curriculum_code}")
        print(f"is_archived: {curriculum.is_archived}")
        print(f"is_active: {curriculum.is_active}")
        print(f"archived_by: {curriculum.archived_by}")
        print(f"archived_at: {curriculum.archived_at}")
        print(f"===================================\n")
        
        if not curriculum.is_archived:
            return jsonify({
                'success': False,
                'message': f'Curriculum "{curriculum.curriculum_code}" is not archived (is_archived={curriculum.is_archived}, is_active={curriculum.is_active})'
            }), 400
        
        # Unarchive curriculum using helper method
        curriculum.unarchive()
        
        # Log activity
        log_unarchive('curriculum', curriculum.id, curriculum.curriculum_code)
        
        db.session.commit()
        
        flash(f'Curriculum "{curriculum.curriculum_code}" has been restored successfully', 'success')
        
        return jsonify({
            'success': True,
            'message': f'Curriculum "{curriculum.curriculum_code}" has been restored successfully'
        })
    
    except Exception as e:
        db.session.rollback()
        print(f"ERROR in unarchive_curriculum: {str(e)}")
        import traceback
        traceback.print_exc()
        return jsonify({
            'success': False,
            'message': f'Error restoring curriculum: {str(e)}'
        }), 500


@archive_bp.route('/api/curriculum/<int:curriculum_id>', methods=['DELETE'])
@login_required
def delete_curriculum_permanently(curriculum_id):
    """Permanently delete an archived curriculum"""
    try:
        permission_error = require_archive_manage_permission()
        if permission_error:
            return permission_error

        curriculum = Curriculum.query.get_or_404(curriculum_id)
        
        if not curriculum.is_archived:
            return jsonify({
                'success': False,
                'message': 'Only archived curricula can be permanently deleted'
            }), 400
        
        curriculum_code = curriculum.curriculum_code
        
        # Log activity before deletion
        from app.utils.activity_logger import log_delete
        log_delete('curriculum', curriculum.id, curriculum_code, {
            'program': curriculum.program.program_name if curriculum.program else 'N/A'
        })
        
        # Delete curriculum (cascade will delete year levels, semesters, and subjects)
        db.session.delete(curriculum)
        db.session.commit()
        
        return jsonify({
            'success': True,
            'message': f'Curriculum "{curriculum_code}" has been permanently deleted'
        })
    
    except Exception as e:
        db.session.rollback()
        return jsonify({
            'success': False,
            'message': f'Error deleting curriculum: {str(e)}'
        }), 500


@archive_bp.route('/api/curriculum-stats')
@login_required
def get_curriculum_archive_stats():
    """Get curriculum archive statistics"""
    try:
        # Get user's program access
        user_program_ids = current_user.get_program_ids()
        
        # Build base query
        query = Curriculum.query.filter_by(is_archived=True)
        
        if user_program_ids is not None:
            query = query.filter(Curriculum.program_id.in_(user_program_ids))
        
        # Total archived curricula
        total = query.count()
        
        # By program
        from app.models.program import Program
        by_department = db.session.query(
            Program.program_name, 
            db.func.count(Curriculum.id)
        ).join(Curriculum).filter(
            Curriculum.is_archived == True
        )
        
        if user_program_ids is not None:
            by_department = by_department.filter(Curriculum.program_id.in_(user_program_ids))
        
        by_department = by_department.group_by(Program.program_name).all()
        
        return jsonify({
            'success': True,
            'stats': {
                'total': total,
                'by_department': {dept: count for dept, count in by_department}
            }
        })
    
    except Exception as e:
        return jsonify({
            'success': False,
            'message': f'Error fetching curriculum archive stats: {str(e)}'
        }), 500


@archive_bp.route('/api/programs')
@login_required
def get_archived_programs():
    """Get archived programs with filters (API endpoint)"""
    try:
        # Get filter parameters
        search = request.args.get('search', '')
        
        # Get user's program access
        user_program_ids = current_user.get_program_ids()
        
        # Build query for archived programs
        query = Program.query.filter_by(is_archived=True)
        
        # Filter by user's program access
        if user_program_ids is not None:
            query = query.filter(Program.id.in_(user_program_ids))
        
        # Apply search filter
        if search:
            search_term = f'%{search}%'
            query = query.filter(
                or_(
                    Program.program_code.ilike(search_term),
                    Program.program_name.ilike(search_term)
                )
            )
        
        # Order by archived date (most recent first)
        programs = query.order_by(Program.archived_at.desc()).all()
        
        # Convert to dictionary
        programs_data = [program.to_dict() for program in programs]
        
        return jsonify({
            'success': True,
            'programs': programs_data,
            'total': len(programs_data)
        })
    
    except Exception as e:
        return jsonify({
            'success': False,
            'message': f'Error fetching archived programs: {str(e)}'
        }), 500


@archive_bp.route('/api/program/<int:program_id>/unarchive', methods=['POST'])
@login_required
def unarchive_department(program_id):
    """Unarchive a program"""
    try:
        permission_error = require_archive_manage_permission()
        if permission_error:
            return permission_error

        program = Program.query.get_or_404(program_id)
        
        # Log current state for debugging
        print(f"\n=== UNARCHIVE DEPARTMENT DEBUG ===")
        print(f"Program ID: {program_id}")
        print(f"Program Code: {program.program_code}")
        print(f"is_archived: {program.is_archived}")
        print(f"is_active: {program.is_active}")
        print(f"archived_by: {program.archived_by}")
        print(f"archived_at: {program.archived_at}")
        print(f"===================================\n")
        
        if not program.is_archived:
            return jsonify({
                'success': False,
                'message': f'Program "{program.program_code}" is not archived (is_archived={program.is_archived}, is_active={program.is_active})'
            }), 400
        
        # Unarchive program using helper method
        program.unarchive()
        
        # Log activity
        log_unarchive('program', program.id, program.program_code)
        
        db.session.commit()
        
        flash(f'Program "{program.program_code}" has been restored successfully', 'success')
        
        return jsonify({
            'success': True,
            'message': f'Program "{program.program_code}" has been restored successfully'
        })
    
    except Exception as e:
        db.session.rollback()
        print(f"ERROR in unarchive_department: {str(e)}")
        import traceback
        traceback.print_exc()
        return jsonify({
            'success': False,
            'message': f'Error restoring program: {str(e)}'
        }), 500


@archive_bp.route('/api/program/<int:program_id>', methods=['DELETE'])
@login_required
def delete_department_permanently(program_id):
    """Permanently delete an archived program"""
    try:
        permission_error = require_archive_manage_permission()
        if permission_error:
            return permission_error

        program = Program.query.get_or_404(program_id)
        
        if not program.is_archived:
            return jsonify({
                'success': False,
                'message': 'Only archived programs can be permanently deleted'
            }), 400
        
        program_code = program.program_code
        program_name = program.program_name
        
        # Log activity before deletion
        from app.utils.activity_logger import log_delete
        log_delete('program', program.id, program_code, {
            'name': program_name
        })
        
        # Delete program (cascade will delete sections, curricula, etc.)
        db.session.delete(program)
        db.session.commit()
        
        return jsonify({
            'success': True,
            'message': f'Program "{program_code}" has been permanently deleted'
        })
    
    except Exception as e:
        db.session.rollback()
        return jsonify({
            'success': False,
            'message': f'Error deleting program: {str(e)}'
        }), 500


@archive_bp.route('/api/program-stats')
@login_required
def get_department_archive_stats():
    """Get program archive statistics"""
    try:
        # Get user's program access
        user_program_ids = current_user.get_program_ids()
        
        # Build base query
        query = Program.query.filter_by(is_archived=True)
        
        if user_program_ids is not None:
            query = query.filter(Program.id.in_(user_program_ids))
        
        # Total archived programs
        total = query.count()
        
        return jsonify({
            'success': True,
            'stats': {
                'total': total
            }
        })
    
    except Exception as e:
        return jsonify({
            'success': False,
            'message': f'Error fetching program archive stats: {str(e)}'
        }), 500


# ========================================
# FACULTY ARCHIVE ROUTES
# ========================================

@archive_bp.route('/api/faculty')
@login_required
def get_archived_faculty():
    """Get all archived faculty members"""
    try:
        # Get archived faculty (no program filtering - faculty visible to all)
        query = Faculty.query.filter_by(is_archived=True)
        
        faculty_list = query.order_by(Faculty.last_name, Faculty.first_name).all()
        
        return jsonify({
            'success': True,
            'faculty': [f.to_dict() for f in faculty_list]
        })
    
    except Exception as e:
        return jsonify({
            'success': False,
            'message': f'Error fetching archived faculty: {str(e)}'
        }), 500


@archive_bp.route('/api/faculty/<int:faculty_id>/unarchive', methods=['POST'])
@login_required
def unarchive_faculty(faculty_id):
    """Unarchive a faculty member"""
    try:
        permission_error = require_archive_manage_permission()
        if permission_error:
            return permission_error

        faculty = Faculty.query.get_or_404(faculty_id)
        
        if not faculty.is_archived:
            return jsonify({
                'success': False,
                'message': 'Faculty is not archived'
            }), 400
        
        faculty_name = faculty.full_name
        
        # Unarchive faculty using helper method
        faculty.unarchive()
        
        # Log activity
        log_unarchive('faculty', faculty.id, faculty_name)
        
        db.session.commit()
        
        return jsonify({
            'success': True,
            'message': f'Faculty member {faculty_name} has been restored successfully!'
        })
    
    except Exception as e:
        db.session.rollback()
        return jsonify({
            'success': False,
            'message': f'Error unarchiving faculty: {str(e)}'
        }), 500


@archive_bp.route('/api/faculty/<int:faculty_id>', methods=['DELETE'])
@login_required
def delete_archived_faculty(faculty_id):
    """Permanently delete an archived faculty member"""
    try:
        permission_error = require_archive_manage_permission()
        if permission_error:
            return permission_error

        faculty = Faculty.query.get_or_404(faculty_id)
        
        if not faculty.is_archived:
            return jsonify({
                'success': False,
                'message': 'Only archived faculty can be permanently deleted'
            }), 400
        
        faculty_name = faculty.full_name
        
        db.session.delete(faculty)
        db.session.commit()
        
        return jsonify({
            'success': True,
            'message': f'Faculty member {faculty_name} has been permanently deleted!'
        })
    
    except Exception as e:
        db.session.rollback()
        return jsonify({
            'success': False,
            'message': f'Error deleting faculty: {str(e)}'
        }), 500


@archive_bp.route('/api/faculty/stats')
@login_required
def get_faculty_archive_stats():
    """Get statistics about archived faculty"""
    try:
        # Get archived faculty (no program filtering - faculty visible to all)
        query = Faculty.query.filter_by(is_archived=True)
        
        # Total archived faculty
        total = query.count()
        
        return jsonify({
            'success': True,
            'stats': {
                'total': total
            }
        })
    
    except Exception as e:
        return jsonify({
            'success': False,
            'message': f'Error fetching faculty archive stats: {str(e)}'
        }), 500


# ==================== SCHEDULE ARCHIVE DETAIL & BULK OPERATIONS ====================

@archive_bp.route('/api/archives/<int:archive_id>')
@login_required
def get_archive_detail(archive_id):
    """Get detailed information about a single archived schedule"""
    try:
        archive = Archive.query.get_or_404(archive_id)
        
        # Get user's program access for permission check
        user_program_ids = current_user.get_program_ids()
        if user_program_ids is not None:
            # For deans, check if they have access to this archive's program
            if archive.program_name:
                program = Program.query.filter_by(program_name=archive.program_name).first()
                if program and program.id not in user_program_ids:
                    return jsonify({
                        'success': False,
                        'message': 'You do not have permission to view this archive'
                    }), 403
        
        return jsonify({
            'success': True,
            'archive': archive.to_dict()
        })
    
    except Exception as e:
        return jsonify({
            'success': False,
            'message': f'Error fetching archive detail: {str(e)}'
        }), 500


@archive_bp.route('/api/archives/bulk-delete', methods=['POST'])
@login_required
def bulk_delete_archives():
    """Permanently delete multiple archive records"""
    try:
        permission_error = require_archive_manage_permission()
        if permission_error:
            return permission_error

        data = request.get_json()
        archive_ids = data.get('archive_ids', [])
        
        if not archive_ids:
            return jsonify({
                'success': False,
                'message': 'No archives selected'
            }), 400
        
        # Get user's program access for permission check
        user_program_ids = current_user.get_program_ids()
        
        deleted_count = 0
        errors = []
        
        for archive_id in archive_ids:
            try:
                archive = Archive.query.get(archive_id)
                if not archive:
                    errors.append(f'Archive {archive_id} not found')
                    continue
                
                # Check permission for deans
                if user_program_ids is not None:
                    if archive.program_name:
                        program = Program.query.filter_by(program_name=archive.program_name).first()
                        if program and program.id not in user_program_ids:
                            errors.append(f'No permission to delete archive {archive_id}')
                            continue
                
                db.session.delete(archive)
                deleted_count += 1
            except Exception as e:
                errors.append(f'Error deleting archive {archive_id}: {str(e)}')
        
        db.session.commit()
        
        message = f'Successfully deleted {deleted_count} archive(s)'
        if errors:
            message += f'. Errors: {len(errors)}'
        
        return jsonify({
            'success': True,
            'message': message,
            'deleted_count': deleted_count,
            'errors': errors
        })
    
    except Exception as e:
        db.session.rollback()
        return jsonify({
            'success': False,
            'message': f'Error deleting archives: {str(e)}'
        }), 500


@archive_bp.route('/export/excel')
@login_required
def export_archives_excel():
    """Export filtered archives to Excel"""
    import io
    from flask import send_file
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    
    try:
        # Get filter parameters
        academic_year = request.args.get('academic_year', '')
        semester = request.args.get('semester', '')
        schedule_type = request.args.get('schedule_type', '')
        exam_period = request.args.get('exam_period', '')
        section_id = request.args.get('section_id', type=int)
        faculty_id = request.args.get('faculty_id', type=int)
        room_id = request.args.get('room_id', type=int)
        building_id = request.args.get('building_id', type=int)
        building_name = request.args.get('building_name', '')
        section_name = request.args.get('section_name', '')
        faculty_name = request.args.get('faculty_name', '')
        room_number = request.args.get('room_number', '')
        program_id = request.args.get('program_id', type=int)
        program = request.args.get('program', '')
        group_by = request.args.get('group_by', 'section')  # 'section', 'faculty', or 'room'
        export_variant = request.args.get('export_variant', 'standard')
        
        # Get user's program access
        user_program_ids = current_user.get_program_ids()
        
        # Build query
        query = Archive.query
        
        # Apply filters
        if academic_year:
            query = query.filter(Archive.academic_year == academic_year)
        if semester:
            query = query.filter(Archive.semester == semester)
        if schedule_type:
            # Handle multiple schedule types (e.g., 'lecture,lab')
            types = [t.strip() for t in schedule_type.split(',')]
            if len(types) > 1:
                query = query.filter(Archive.schedule_type.in_(types))
            else:
                query = query.filter(Archive.schedule_type == schedule_type)
        if exam_period:
            query = query.filter(Archive.exam_period == exam_period)

        # For exam batch exports, allow exporting all sections under active filters.
        apply_section_filter = not (schedule_type == 'exam' and export_variant == 'exam_batch')

        # Prefer ID-based filters for consistent behavior with live schedules export.
        if apply_section_filter:
            if section_id:
                query = query.filter(Archive.section_id == section_id)
            elif section_name:
                query = query.filter(Archive.section_name.ilike(f'%{section_name}%'))

        if faculty_id:
            query = query.filter(Archive.faculty_id == faculty_id)
        elif faculty_name:
            query = query.filter(Archive.faculty_name.ilike(f'%{faculty_name}%'))

        if room_id:
            query = query.filter(Archive.room_id == room_id)
        elif room_number:
            query = query.filter(Archive.room_number.ilike(f'%{room_number}%'))

        if building_id:
            from app.models.building import Room
            room_ids = db.session.query(Room.id).filter(Room.building_id == building_id).all()
            room_ids = [rid[0] for rid in room_ids]
            if room_ids:
                query = query.filter(Archive.room_id.in_(room_ids))
            else:
                query = query.filter(Archive.id == -1)

        if building_name:
            query = query.filter(Archive.building_name.ilike(f'%{building_name}%'))

        if program_id:
            section_ids = db.session.query(Section.id).filter(Section.program_id == program_id).all()
            section_ids = [sid[0] for sid in section_ids]
            if section_ids:
                query = query.filter(Archive.section_id.in_(section_ids))
            else:
                query = query.filter(Archive.id == -1)
        # Apply text-based program filter only when no explicit program_id is provided.
        if program and not program_id:
            query = query.filter(Archive.program_name.ilike(f'%{program}%'))
        
        # Filter by program access
        if user_program_ids is not None:
            accessible_section_ids = db.session.query(Section.id)\
                .filter(Section.program_id.in_(user_program_ids))\
                .all()
            accessible_section_ids = [sid[0] for sid in accessible_section_ids]
            if accessible_section_ids:
                query = query.filter(Archive.section_id.in_(accessible_section_ids))
        
        # Determine grouping field and order
        day_order = db.case(
            {'Monday': 1, 'Tuesday': 2, 'Wednesday': 3, 'Thursday': 4, 'Friday': 5, 'Saturday': 6, 'Sunday': 7},
            value=Archive.day_of_week,
            else_=8
        )
        
        if group_by == 'faculty':
            order_col = Archive.faculty_name
            group_label = 'faculty'
        elif group_by == 'room':
            order_col = Archive.room_number
            group_label = 'room'
        else:
            order_col = Archive.section_name
            group_label = 'section'
        
        if schedule_type == 'exam':
            archives = query.order_by(Archive.section_name, Archive.exam_date, Archive.start_time).all()
        else:
            archives = query.order_by(order_col, day_order, Archive.start_time).all()

        from app.models.settings import AcademicSettings
        current_settings = AcademicSettings.query.order_by(AcademicSettings.id.desc()).first()
        filter_info = {'academic_year': academic_year, 'semester': semester}

        if not archives:
            return jsonify({
                'success': False,
                'message': 'No archives found for the selected filters'
            }), 404

        if schedule_type == 'exam' and export_variant in ('exam_individual', 'exam_batch'):
            from types import SimpleNamespace
            from collections import defaultdict
            from datetime import timedelta
            from openpyxl.drawing.image import Image as ExcelImage
            from openpyxl.worksheet.page import PageMargins
            from app.routes.exam_schedule import _write_exam_export_time_rows
            import os

            def _make_program_stub(program_name):
                display_name = program_name or 'DEPT'
                return SimpleNamespace(
                    program_code='',
                    program_name=display_name,
                    department=SimpleNamespace(department_name=display_name)
                )

            def _resolve_program_from_archive(archive):
                if archive.section and getattr(archive.section, 'program', None):
                    return archive.section.program
                if program_id:
                    program_obj = Program.query.get(program_id)
                    if program_obj:
                        return program_obj
                if archive.program_name:
                    program_obj = Program.query.filter_by(program_name=archive.program_name).first()
                    if program_obj:
                        return program_obj
                return _make_program_stub(archive.program_name)

            def _resolve_section_from_archive(archive, fallback_program):
                if archive.section:
                    return archive.section
                if archive.section_id:
                    section_obj = Section.query.get(archive.section_id)
                    if section_obj:
                        return section_obj
                return SimpleNamespace(
                    id=archive.section_id,
                    full_section_name=archive.section_name or 'SECTION',
                    section_name=archive.section_name or 'SECTION',
                    year_level='',
                    program=fallback_program
                )

            def _resolve_faculty_from_archive(archive):
                if archive.faculty:
                    return archive.faculty

                raw_name = (archive.faculty_name or '').strip()
                if not raw_name:
                    return None

                first_name = ''
                last_name = ''
                if ',' in raw_name:
                    parts = [p.strip() for p in raw_name.split(',', 1)]
                    last_name = parts[0] if parts else ''
                    first_name = parts[1].split()[0] if len(parts) > 1 and parts[1] else ''
                else:
                    tokens = raw_name.split()
                    if len(tokens) == 1:
                        first_name = tokens[0]
                        last_name = tokens[0]
                    elif len(tokens) > 1:
                        first_name = tokens[0]
                        last_name = tokens[-1]

                if not last_name:
                    last_name = raw_name

                return SimpleNamespace(
                    first_name=first_name,
                    last_name=last_name,
                    gender='Male',
                    full_name=raw_name
                )

            def _build_exam_like_record(archive, section_obj, seq):
                subject_obj = archive.subject if archive.subject else SimpleNamespace(
                    id=archive.subject_id or seq,
                    subject_code=archive.subject_code or ''
                )
                room_obj = archive.room if archive.room else SimpleNamespace(
                    room_number=archive.room_number or ''
                )
                faculty_obj = _resolve_faculty_from_archive(archive)

                return SimpleNamespace(
                    exam_date=archive.exam_date,
                    start_time=archive.start_time,
                    end_time=archive.end_time,
                    subject=subject_obj,
                    room=room_obj,
                    faculty=faculty_obj,
                    section=section_obj
                )

            valid_archives = [
                a for a in archives
                if a.exam_date and a.start_time and a.end_time
            ]

            if not valid_archives:
                return jsonify({
                    'success': False,
                    'message': 'No valid archived exam schedules found for export.'
                }), 404

            # -----------------------------------------------------------------
            # Individual exam export (same format as Exam page Excel export)
            # -----------------------------------------------------------------
            if export_variant == 'exam_individual':
                selected_archives = valid_archives
                if section_id:
                    selected_archives = [a for a in selected_archives if a.section_id == section_id]
                elif section_name:
                    selected_archives = [
                        a for a in selected_archives
                        if (a.section_name or '').strip().lower() == section_name.strip().lower()
                    ]

                if not selected_archives:
                    return jsonify({
                        'success': False,
                        'message': 'No archived exam schedules found for the selected section.'
                    }), 404

                unique_sections = {
                    (a.section_id, (a.section_name or '').strip().lower())
                    for a in selected_archives
                }
                if len(unique_sections) > 1:
                    return jsonify({
                        'success': False,
                        'message': 'Select a single section before exporting individual exam schedules.'
                    }), 400

                selected_archives = sorted(
                    selected_archives,
                    key=lambda a: (a.exam_date, a.start_time)
                )

                ref_archive = selected_archives[0]
                section_obj = None
                if section_id:
                    section_obj = Section.query.get(section_id)
                elif ref_archive.section_id:
                    section_obj = Section.query.get(ref_archive.section_id)
                if not section_obj:
                    fallback_program = _resolve_program_from_archive(ref_archive)
                    section_obj = _resolve_section_from_archive(ref_archive, fallback_program)

                program_obj = getattr(section_obj, 'program', None) or _resolve_program_from_archive(ref_archive)
                dept_code = getattr(program_obj, 'program_code', '') or 'ARCHIVED'
                dept_display_name = (
                    getattr(getattr(program_obj, 'department', None), 'department_name', None)
                    or getattr(program_obj, 'program_name', None)
                    or 'DEPT'
                )

                exam_schedules = [
                    _build_exam_like_record(archive, section_obj, idx)
                    for idx, archive in enumerate(selected_archives, start=1)
                ]

                wb = Workbook()
                ws = wb.active
                ws.title = 'Exam Schedule'

                ws.page_setup.paperSize = ws.PAPERSIZE_LEGAL
                ws.page_setup.orientation = 'landscape'
                ws.page_setup.fitToPage = True
                ws.page_setup.fitToWidth = 1
                ws.page_setup.fitToHeight = 0
                ws.page_margins = PageMargins(left=0.25, right=0.25, top=0.5, bottom=0.5)
                ws.print_options.horizontalCentered = True

                base_dir = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
                images_dir = os.path.join(base_dir, 'static', 'images')

                ws.column_dimensions['A'].width = 12

                logo_left_path = os.path.join(images_dir, 'norzagaray-college-logo.png')
                if os.path.exists(logo_left_path):
                    img_left = ExcelImage(logo_left_path)
                    img_left.width = 75
                    img_left.height = 75
                    ws.add_image(img_left, 'A1')

                logo_right_path = os.path.join(images_dir, 'bagong-pilipinas.png')
                if os.path.exists(logo_right_path):
                    img_right = ExcelImage(logo_right_path)
                    img_right.width = 80
                    img_right.height = 80
                    ws.add_image(img_right, 'M1')

                ws['B1'] = 'Republic of the Philippines'
                ws['B1'].font = Font(size=11)
                ws['B1'].alignment = Alignment(horizontal='left', vertical='center')

                ws['B2'] = 'Municipality of Norzagaray'
                ws['B2'].font = Font(size=11)
                ws['B2'].alignment = Alignment(horizontal='left', vertical='center')

                ws['B3'] = 'NORZAGARAY COLLEGE'
                ws['B3'].font = Font(bold=True, size=11)
                ws['B3'].alignment = Alignment(horizontal='left', vertical='center')

                ws['B4'] = dept_display_name
                ws['B4'].font = Font(bold=True, size=11)
                ws['B4'].alignment = Alignment(horizontal='left', vertical='center')

                ws.row_dimensions[1].height = 18
                ws.row_dimensions[2].height = 15
                ws.row_dimensions[3].height = 15
                ws.row_dimensions[4].height = 15
                ws.row_dimensions[5].height = 8

                exam_period_value = exam_period or (current_settings.exam_period if current_settings and current_settings.exam_period else '')
                exam_period_title = f"{exam_period_value.upper()} EXAMINATION" if exam_period_value else 'EXAMINATION'
                ws.merge_cells('A6:M6')
                ws.cell(row=6, column=1, value=exam_period_title)
                ws.cell(row=6, column=1).font = Font(bold=True, size=12)
                ws.cell(row=6, column=1).alignment = Alignment(horizontal='center', vertical='center')

                thin_border = Border(
                    left=Side(style='thin'),
                    right=Side(style='thin'),
                    top=Side(style='thin'),
                    bottom=Side(style='thin')
                )
                yellow_fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
                header_fill = PatternFill(start_color='F5F5F5', end_color='F5F5F5', fill_type='solid')

                ws.column_dimensions['B'].width = 10
                ws.column_dimensions['C'].width = 10
                ws.column_dimensions['D'].width = 10
                ws.column_dimensions['E'].width = 10
                ws.column_dimensions['F'].width = 10
                ws.column_dimensions['G'].width = 10
                ws.column_dimensions['H'].width = 10
                ws.column_dimensions['I'].width = 10
                ws.column_dimensions['J'].width = 10
                ws.column_dimensions['K'].width = 10
                ws.column_dimensions['L'].width = 10
                ws.column_dimensions['M'].width = 12

                days_order = AcademicSettings.get_active_operation_days()
                day_start_col = {day: 2 + idx * 2 for idx, day in enumerate(days_order)}

                exam_period_start = getattr(current_settings, 'exam_period_start', None) if current_settings else None
                exam_period_end = getattr(current_settings, 'exam_period_end', None) if current_settings else None

                if exam_period_start and exam_period_end:
                    date_range_str = f"{exam_period_start.strftime('%B %d')}-{exam_period_end.strftime('%d, %Y')}"
                    date_by_day = {}
                    current_date = exam_period_start
                    while current_date <= exam_period_end:
                        day_name = current_date.strftime('%A')
                        if day_name != 'Sunday':
                            date_by_day[day_name] = current_date
                        current_date += timedelta(days=1)
                else:
                    all_dates = sorted({exam.exam_date for exam in exam_schedules if exam.exam_date})
                    date_by_day = {}
                    for date_value in all_dates:
                        date_by_day[date_value.strftime('%A')] = date_value
                    if all_dates:
                        date_range_str = f"{all_dates[0].strftime('%B %d')}-{all_dates[-1].strftime('%d, %Y')}"
                    else:
                        date_range_str = ''

                ws.merge_cells('A7:M7')
                ws.cell(row=7, column=1, value=date_range_str)
                ws.cell(row=7, column=1).font = Font(size=11)
                ws.cell(row=7, column=1).alignment = Alignment(horizontal='center', vertical='center')

                section_display = (
                    getattr(section_obj, 'full_section_name', None)
                    or f"{dept_code}-{getattr(section_obj, 'year_level', '')}{getattr(section_obj, 'section_name', '')}"
                )

                current_row = 8

                ws.merge_cells(f'A{current_row}:M{current_row}')
                ws.cell(row=current_row, column=1, value=section_display)
                ws.cell(row=current_row, column=1).font = Font(bold=True, size=11)
                ws.cell(row=current_row, column=1).alignment = Alignment(horizontal='left', vertical='center')
                ws.cell(row=current_row, column=1).fill = yellow_fill
                for col in range(1, 14):
                    ws.cell(row=current_row, column=col).border = thin_border
                current_row += 1

                ws.cell(row=current_row, column=1, value='Day')
                ws.cell(row=current_row, column=1).font = Font(bold=True, size=9)
                ws.cell(row=current_row, column=1).alignment = Alignment(horizontal='center', vertical='center')
                ws.cell(row=current_row, column=1).border = thin_border
                ws.cell(row=current_row, column=1).fill = header_fill

                for day in days_order:
                    col = day_start_col[day]
                    ws.merge_cells(start_row=current_row, start_column=col, end_row=current_row, end_column=col + 1)
                    ws.cell(row=current_row, column=col, value=day)
                    ws.cell(row=current_row, column=col).font = Font(bold=True, size=9)
                    ws.cell(row=current_row, column=col).alignment = Alignment(horizontal='center', vertical='center')
                    ws.cell(row=current_row, column=col).border = thin_border
                    ws.cell(row=current_row, column=col + 1).border = thin_border
                current_row += 1

                ws.cell(row=current_row, column=1, value='Date')
                ws.cell(row=current_row, column=1).font = Font(bold=True, size=9)
                ws.cell(row=current_row, column=1).alignment = Alignment(horizontal='center', vertical='center')
                ws.cell(row=current_row, column=1).border = thin_border
                ws.cell(row=current_row, column=1).fill = header_fill

                for day in days_order:
                    col = day_start_col[day]
                    ws.merge_cells(start_row=current_row, start_column=col, end_row=current_row, end_column=col + 1)
                    date_value = date_by_day.get(day)
                    ws.cell(row=current_row, column=col, value=date_value.strftime('%B %d') if date_value else '')
                    ws.cell(row=current_row, column=col).font = Font(size=9)
                    ws.cell(row=current_row, column=col).alignment = Alignment(horizontal='center', vertical='center')
                    ws.cell(row=current_row, column=col).border = thin_border
                    ws.cell(row=current_row, column=col + 1).border = thin_border
                current_row += 1

                ws.cell(row=current_row, column=1, value='Room')
                ws.cell(row=current_row, column=1).font = Font(bold=True, size=9)
                ws.cell(row=current_row, column=1).alignment = Alignment(horizontal='center', vertical='center')
                ws.cell(row=current_row, column=1).border = thin_border
                ws.cell(row=current_row, column=1).fill = header_fill

                rooms_by_day = defaultdict(set)
                for exam in exam_schedules:
                    day_name = exam.exam_date.strftime('%A')
                    room_name = exam.room.room_number if exam.room else ''
                    if room_name:
                        rooms_by_day[day_name].add(room_name)

                for day in days_order:
                    col = day_start_col[day]
                    ws.merge_cells(start_row=current_row, start_column=col, end_row=current_row, end_column=col + 1)
                    ws.cell(row=current_row, column=col, value=', '.join(sorted(rooms_by_day[day])) if rooms_by_day[day] else '')
                    ws.cell(row=current_row, column=col).font = Font(size=9)
                    ws.cell(row=current_row, column=col).alignment = Alignment(horizontal='center', vertical='center')
                    ws.cell(row=current_row, column=col).border = thin_border
                    ws.cell(row=current_row, column=col + 1).border = thin_border
                current_row += 1

                ws.cell(row=current_row, column=1, value='Time')
                ws.cell(row=current_row, column=1).font = Font(bold=True, size=9)
                ws.cell(row=current_row, column=1).alignment = Alignment(horizontal='center', vertical='center')
                ws.cell(row=current_row, column=1).border = thin_border
                ws.cell(row=current_row, column=1).fill = header_fill

                for day in days_order:
                    col = day_start_col[day]
                    ws.cell(row=current_row, column=col, value='Subject')
                    ws.cell(row=current_row, column=col).font = Font(bold=True, size=8)
                    ws.cell(row=current_row, column=col).alignment = Alignment(horizontal='center', vertical='center')
                    ws.cell(row=current_row, column=col).border = thin_border
                    ws.cell(row=current_row, column=col).fill = header_fill
                    ws.cell(row=current_row, column=col + 1, value='Proctor')
                    ws.cell(row=current_row, column=col + 1).font = Font(bold=True, size=8)
                    ws.cell(row=current_row, column=col + 1).alignment = Alignment(horizontal='center', vertical='center')
                    ws.cell(row=current_row, column=col + 1).border = thin_border
                    ws.cell(row=current_row, column=col + 1).fill = header_fill
                current_row += 1

                current_row = _write_exam_export_time_rows(
                    ws=ws,
                    exam_schedules=exam_schedules,
                    days_order=days_order,
                    day_start_col=day_start_col,
                    current_row=current_row,
                    thin_border=thin_border,
                    current_settings=current_settings
                )

                sig_start_row = current_row + 1
                ws.cell(row=sig_start_row, column=1, value='Prepared by:')
                ws.cell(row=sig_start_row, column=1).font = Font(size=10)

                dean_name = (current_user.full_name if current_user else 'Name of the Dean').upper()
                ws.merge_cells(f'A{sig_start_row + 2}:D{sig_start_row + 2}')
                ws.cell(row=sig_start_row + 2, column=1, value=dean_name)
                ws.cell(row=sig_start_row + 2, column=1).font = Font(bold=True, size=10, underline='single')
                ws.cell(row=sig_start_row + 2, column=1).alignment = Alignment(horizontal='left')

                ws.merge_cells(f'A{sig_start_row + 3}:D{sig_start_row + 3}')
                ws.cell(row=sig_start_row + 3, column=1, value=f'Dean, {dept_display_name}')
                ws.cell(row=sig_start_row + 3, column=1).font = Font(size=10)
                ws.cell(row=sig_start_row + 3, column=1).alignment = Alignment(horizontal='left')

                output = io.BytesIO()
                wb.save(output)
                output.seek(0)

                year_level = str(getattr(section_obj, 'year_level', '') or '')
                section_slug = (getattr(section_obj, 'section_name', None) or ref_archive.section_name or 'section').replace(' ', '_')
                filename = f"{dept_code}_{year_level}{section_slug}_Exam_Schedule.xlsx".replace(' ', '_')

                return send_file(
                    output,
                    mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                    as_attachment=True,
                    download_name=filename
                )

            # -----------------------------------------------------------------
            # Batch exam export (same format as Exam page batch export)
            # -----------------------------------------------------------------
            section_groups = {}
            for archive in valid_archives:
                key = archive.section_id if archive.section_id else f"name::{(archive.section_name or '').strip().lower()}"
                if key not in section_groups:
                    fallback_program = _resolve_program_from_archive(archive)
                    section_obj = _resolve_section_from_archive(archive, fallback_program)
                    section_groups[key] = {
                        'section': section_obj,
                        'archives': []
                    }
                section_groups[key]['archives'].append(archive)

            if not section_groups:
                return jsonify({
                    'success': False,
                    'message': 'No archived exam schedules found for batch export.'
                }), 404

            section_entries = list(section_groups.values())
            section_entries.sort(
                key=lambda entry: (
                    int(getattr(entry['section'], 'year_level', 0) or 0) if str(getattr(entry['section'], 'year_level', '') or '').isdigit() else 0,
                    str(getattr(entry['section'], 'section_name', '') or getattr(entry['section'], 'full_section_name', '')).lower()
                )
            )

            first_section = section_entries[0]['section']
            program_obj = getattr(first_section, 'program', None)
            if not program_obj:
                if program_id:
                    program_obj = Program.query.get(program_id)
                if not program_obj and section_entries[0]['archives']:
                    program_obj = _resolve_program_from_archive(section_entries[0]['archives'][0])
            if not program_obj:
                program_obj = _make_program_stub(section_entries[0]['archives'][0].program_name if section_entries[0]['archives'] else 'DEPT')

            dept_code = getattr(program_obj, 'program_code', '') or 'ARCHIVED'
            dept_display_name = (
                getattr(getattr(program_obj, 'department', None), 'department_name', None)
                or getattr(program_obj, 'program_name', None)
                or 'DEPT'
            )

            wb = Workbook()
            ws = wb.active
            ws.title = 'Exam Schedule'

            ws.page_setup.paperSize = ws.PAPERSIZE_LEGAL
            ws.page_setup.orientation = 'landscape'
            ws.page_setup.fitToPage = True
            ws.page_setup.fitToWidth = 1
            ws.page_setup.fitToHeight = 0
            ws.page_margins = PageMargins(left=0.25, right=0.25, top=0.5, bottom=0.5)
            ws.print_options.horizontalCentered = True

            base_dir = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
            images_dir = os.path.join(base_dir, 'static', 'images')

            ws.column_dimensions['A'].width = 12

            logo_left_path = os.path.join(images_dir, 'norzagaray-college-logo.png')
            if os.path.exists(logo_left_path):
                img_left = ExcelImage(logo_left_path)
                img_left.width = 75
                img_left.height = 75
                ws.add_image(img_left, 'A1')

            logo_right_path = os.path.join(images_dir, 'bagong-pilipinas.png')
            if os.path.exists(logo_right_path):
                img_right = ExcelImage(logo_right_path)
                img_right.width = 80
                img_right.height = 80
                ws.add_image(img_right, 'M1')

            ws['B1'] = 'Republic of the Philippines'
            ws['B1'].font = Font(size=11)
            ws['B1'].alignment = Alignment(horizontal='left', vertical='center')

            ws['B2'] = 'Municipality of Norzagaray'
            ws['B2'].font = Font(size=11)
            ws['B2'].alignment = Alignment(horizontal='left', vertical='center')

            ws['B3'] = 'NORZAGARAY COLLEGE'
            ws['B3'].font = Font(bold=True, size=11)
            ws['B3'].alignment = Alignment(horizontal='left', vertical='center')

            ws['B4'] = dept_display_name
            ws['B4'].font = Font(bold=True, size=11)
            ws['B4'].alignment = Alignment(horizontal='left', vertical='center')

            ws.row_dimensions[1].height = 18
            ws.row_dimensions[2].height = 15
            ws.row_dimensions[3].height = 15
            ws.row_dimensions[4].height = 15
            ws.row_dimensions[5].height = 8

            exam_period_value = exam_period or (current_settings.exam_period if current_settings and current_settings.exam_period else '')
            exam_period_title = f"{exam_period_value.upper()} EXAMINATION" if exam_period_value else 'EXAMINATION'
            ws.merge_cells('A6:M6')
            ws.cell(row=6, column=1, value=exam_period_title)
            ws.cell(row=6, column=1).font = Font(bold=True, size=12)
            ws.cell(row=6, column=1).alignment = Alignment(horizontal='center', vertical='center')

            thin_border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            yellow_fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
            header_fill = PatternFill(start_color='F5F5F5', end_color='F5F5F5', fill_type='solid')

            ws.column_dimensions['B'].width = 10
            ws.column_dimensions['C'].width = 10
            ws.column_dimensions['D'].width = 10
            ws.column_dimensions['E'].width = 10
            ws.column_dimensions['F'].width = 10
            ws.column_dimensions['G'].width = 10
            ws.column_dimensions['H'].width = 10
            ws.column_dimensions['I'].width = 10
            ws.column_dimensions['J'].width = 10
            ws.column_dimensions['K'].width = 10
            ws.column_dimensions['L'].width = 10
            ws.column_dimensions['M'].width = 12

            days_order = AcademicSettings.get_active_operation_days()
            day_start_col = {day: 2 + idx * 2 for idx, day in enumerate(days_order)}

            exam_period_start = getattr(current_settings, 'exam_period_start', None) if current_settings else None
            exam_period_end = getattr(current_settings, 'exam_period_end', None) if current_settings else None

            if exam_period_start and exam_period_end:
                date_range_str = f"{exam_period_start.strftime('%B %d')}-{exam_period_end.strftime('%d, %Y')}"
                configured_date_by_day = {}
                current_date = exam_period_start
                while current_date <= exam_period_end:
                    day_name = current_date.strftime('%A')
                    if day_name != 'Sunday':
                        configured_date_by_day[day_name] = current_date
                    current_date += timedelta(days=1)
            else:
                all_dates = sorted({a.exam_date for a in valid_archives if a.exam_date})
                if all_dates:
                    date_range_str = f"{all_dates[0].strftime('%B %d')}-{all_dates[-1].strftime('%d, %Y')}"
                else:
                    date_range_str = ''
                configured_date_by_day = {}

            current_row = 8
            for entry in section_entries:
                section_obj = entry['section']
                section_archives = sorted(
                    entry['archives'],
                    key=lambda a: (a.exam_date, a.start_time)
                )

                exam_schedules = [
                    _build_exam_like_record(archive, section_obj, idx)
                    for idx, archive in enumerate(section_archives, start=1)
                ]
                if not exam_schedules:
                    continue

                if configured_date_by_day:
                    date_by_day = configured_date_by_day.copy()
                else:
                    date_by_day = {}
                    for exam in exam_schedules:
                        date_by_day[exam.exam_date.strftime('%A')] = exam.exam_date

                section_display = (
                    getattr(section_obj, 'full_section_name', None)
                    or f"{dept_code}-{getattr(section_obj, 'year_level', '')}{getattr(section_obj, 'section_name', '')}"
                )

                if current_row == 8:
                    ws.merge_cells('A7:M7')
                    ws.cell(row=7, column=1, value=date_range_str)
                    ws.cell(row=7, column=1).font = Font(size=11)
                    ws.cell(row=7, column=1).alignment = Alignment(horizontal='center', vertical='center')

                ws.merge_cells(f'A{current_row}:M{current_row}')
                ws.cell(row=current_row, column=1, value=section_display)
                ws.cell(row=current_row, column=1).font = Font(bold=True, size=11)
                ws.cell(row=current_row, column=1).alignment = Alignment(horizontal='left', vertical='center')
                ws.cell(row=current_row, column=1).fill = yellow_fill
                for col in range(1, 14):
                    ws.cell(row=current_row, column=col).border = thin_border
                current_row += 1

                ws.cell(row=current_row, column=1, value='Day')
                ws.cell(row=current_row, column=1).font = Font(bold=True, size=9)
                ws.cell(row=current_row, column=1).alignment = Alignment(horizontal='center', vertical='center')
                ws.cell(row=current_row, column=1).border = thin_border
                ws.cell(row=current_row, column=1).fill = header_fill

                for day in days_order:
                    col = day_start_col[day]
                    ws.merge_cells(start_row=current_row, start_column=col, end_row=current_row, end_column=col + 1)
                    ws.cell(row=current_row, column=col, value=day)
                    ws.cell(row=current_row, column=col).font = Font(bold=True, size=9)
                    ws.cell(row=current_row, column=col).alignment = Alignment(horizontal='center', vertical='center')
                    ws.cell(row=current_row, column=col).border = thin_border
                    ws.cell(row=current_row, column=col + 1).border = thin_border
                current_row += 1

                ws.cell(row=current_row, column=1, value='Date')
                ws.cell(row=current_row, column=1).font = Font(bold=True, size=9)
                ws.cell(row=current_row, column=1).alignment = Alignment(horizontal='center', vertical='center')
                ws.cell(row=current_row, column=1).border = thin_border
                ws.cell(row=current_row, column=1).fill = header_fill

                for day in days_order:
                    col = day_start_col[day]
                    ws.merge_cells(start_row=current_row, start_column=col, end_row=current_row, end_column=col + 1)
                    date_value = date_by_day.get(day)
                    ws.cell(row=current_row, column=col, value=date_value.strftime('%B %d') if date_value else '')
                    ws.cell(row=current_row, column=col).font = Font(size=9)
                    ws.cell(row=current_row, column=col).alignment = Alignment(horizontal='center', vertical='center')
                    ws.cell(row=current_row, column=col).border = thin_border
                    ws.cell(row=current_row, column=col + 1).border = thin_border
                current_row += 1

                ws.cell(row=current_row, column=1, value='Room')
                ws.cell(row=current_row, column=1).font = Font(bold=True, size=9)
                ws.cell(row=current_row, column=1).alignment = Alignment(horizontal='center', vertical='center')
                ws.cell(row=current_row, column=1).border = thin_border
                ws.cell(row=current_row, column=1).fill = header_fill

                rooms_by_day = defaultdict(set)
                for exam in exam_schedules:
                    day_name = exam.exam_date.strftime('%A')
                    room_name = exam.room.room_number if exam.room else ''
                    if room_name:
                        rooms_by_day[day_name].add(room_name)

                for day in days_order:
                    col = day_start_col[day]
                    ws.merge_cells(start_row=current_row, start_column=col, end_row=current_row, end_column=col + 1)
                    ws.cell(row=current_row, column=col, value=', '.join(sorted(rooms_by_day[day])) if rooms_by_day[day] else '')
                    ws.cell(row=current_row, column=col).font = Font(size=9)
                    ws.cell(row=current_row, column=col).alignment = Alignment(horizontal='center', vertical='center')
                    ws.cell(row=current_row, column=col).border = thin_border
                    ws.cell(row=current_row, column=col + 1).border = thin_border
                current_row += 1

                ws.cell(row=current_row, column=1, value='Time')
                ws.cell(row=current_row, column=1).font = Font(bold=True, size=9)
                ws.cell(row=current_row, column=1).alignment = Alignment(horizontal='center', vertical='center')
                ws.cell(row=current_row, column=1).border = thin_border
                ws.cell(row=current_row, column=1).fill = header_fill

                for day in days_order:
                    col = day_start_col[day]
                    ws.cell(row=current_row, column=col, value='Subject')
                    ws.cell(row=current_row, column=col).font = Font(bold=True, size=8)
                    ws.cell(row=current_row, column=col).alignment = Alignment(horizontal='center', vertical='center')
                    ws.cell(row=current_row, column=col).border = thin_border
                    ws.cell(row=current_row, column=col).fill = header_fill
                    ws.cell(row=current_row, column=col + 1, value='Proctor')
                    ws.cell(row=current_row, column=col + 1).font = Font(bold=True, size=8)
                    ws.cell(row=current_row, column=col + 1).alignment = Alignment(horizontal='center', vertical='center')
                    ws.cell(row=current_row, column=col + 1).border = thin_border
                    ws.cell(row=current_row, column=col + 1).fill = header_fill
                current_row += 1

                current_row = _write_exam_export_time_rows(
                    ws=ws,
                    exam_schedules=exam_schedules,
                    days_order=days_order,
                    day_start_col=day_start_col,
                    current_row=current_row,
                    thin_border=thin_border,
                    current_settings=current_settings
                )
                current_row += 1

            sig_start_row = current_row
            ws.cell(row=sig_start_row, column=1, value='Prepared by:')
            ws.cell(row=sig_start_row, column=1).font = Font(size=10)

            dean_name = (current_user.full_name if current_user else 'Name of the Dean').upper()
            ws.merge_cells(f'A{sig_start_row + 2}:D{sig_start_row + 2}')
            ws.cell(row=sig_start_row + 2, column=1, value=dean_name)
            ws.cell(row=sig_start_row + 2, column=1).font = Font(bold=True, size=10, underline='single')
            ws.cell(row=sig_start_row + 2, column=1).alignment = Alignment(horizontal='left')

            ws.merge_cells(f'A{sig_start_row + 3}:D{sig_start_row + 3}')
            ws.cell(row=sig_start_row + 3, column=1, value=f'Dean, {dept_display_name}')
            ws.cell(row=sig_start_row + 3, column=1).font = Font(size=10)
            ws.cell(row=sig_start_row + 3, column=1).alignment = Alignment(horizontal='left')

            output = io.BytesIO()
            wb.save(output)
            output.seek(0)

            filename = f"{dept_code}_Exam_Schedule_Batch_Export.xlsx".replace(' ', '_')

            return send_file(
                output,
                mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                as_attachment=True,
                download_name=filename
            )

        if export_variant == 'for_posting':
            from types import SimpleNamespace
            from app.services.export_service import generate_class_schedule_excel_for_posting

            if group_by != 'section':
                return jsonify({
                    'success': False,
                    'message': 'For Posting template is only available for class section exports.'
                }), 400

            selected_archives = archives
            if section_id:
                selected_archives = [a for a in archives if a.section_id == section_id]
            elif section_name:
                selected_archives = [
                    a for a in archives
                    if (a.section_name or '').strip().lower() == section_name.strip().lower()
                ]

            if not selected_archives:
                return jsonify({
                    'success': False,
                    'message': 'No section archives found for the selected For Posting export.'
                }), 404

            unique_sections = {(a.section_id, (a.section_name or '').strip().lower()) for a in selected_archives}
            if len(unique_sections) > 1:
                return jsonify({
                    'success': False,
                    'message': 'Select a single section before exporting the For Posting template.'
                }), 400

            ref_archive = selected_archives[0]
            section_obj = None
            if section_id:
                section_obj = Section.query.get(section_id)
            elif ref_archive.section_id:
                section_obj = Section.query.get(ref_archive.section_id)

            if not section_obj:
                fallback_program = None
                if program_id:
                    fallback_program = Program.query.get(program_id)
                if not fallback_program and ref_archive.program_name:
                    fallback_program = Program.query.filter(
                        Program.program_name == ref_archive.program_name
                    ).first()
                if not fallback_program:
                    fallback_program = SimpleNamespace(
                        program_code='',
                        program_name=ref_archive.program_name or 'Program',
                        department_id=None,
                        department=None
                    )

                section_obj = SimpleNamespace(
                    program=fallback_program,
                    full_section_name=ref_archive.section_name or 'Section',
                    year_level='',
                    section_name=ref_archive.section_name or 'Section'
                )

            day_rank = {
                'Monday': 1,
                'Tuesday': 2,
                'Wednesday': 3,
                'Thursday': 4,
                'Friday': 5,
                'Saturday': 6,
                'Sunday': 7
            }

            schedule_rows = []
            sorted_archives = sorted(
                selected_archives,
                key=lambda a: (
                    day_rank.get(a.day_of_week, 99),
                    a.start_time or datetime.min.time(),
                    a.subject_code or ''
                )
            )

            for idx, archive in enumerate(sorted_archives, start=1):
                subject_obj = archive.subject if archive.subject else SimpleNamespace(
                    id=archive.subject_id or idx,
                    subject_code=archive.subject_code or '',
                    course_description=archive.course_description or '',
                    lec_units='',
                    lab_units=''
                )
                room_obj = archive.room if archive.room else SimpleNamespace(room_number=archive.room_number or 'TBA')
                faculty_obj = archive.faculty if archive.faculty else SimpleNamespace(full_name=archive.faculty_name or 'TBA')

                schedule_rows.append(SimpleNamespace(
                    subject=subject_obj,
                    schedule_type=archive.schedule_type or 'lecture',
                    day_of_week=archive.day_of_week or '',
                    start_time=archive.start_time,
                    end_time=archive.end_time,
                    room=room_obj,
                    faculty=faculty_obj
                ))

            export_settings = current_settings
            if not export_settings:
                export_settings = SimpleNamespace(
                    academic_year=academic_year or '',
                    semester=semester or ''
                )

            output, filename = generate_class_schedule_excel_for_posting(
                section_obj,
                schedule_rows,
                export_settings,
                current_user
            )

            return send_file(
                output,
                mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                as_attachment=True,
                download_name=filename
            )

        if export_variant in ('faculty_template', 'faculty_excel'):
            from types import SimpleNamespace
            from app.services.export_service import generate_faculty_schedule_excel

            if group_by != 'faculty':
                return jsonify({
                    'success': False,
                    'message': 'Faculty export is only available on the faculty archives tab.'
                }), 400

            selected_archives = archives
            if faculty_id:
                selected_archives = [a for a in archives if a.faculty_id == faculty_id]
            elif faculty_name:
                selected_archives = [
                    a for a in archives
                    if (a.faculty_name or '').strip().lower() == faculty_name.strip().lower()
                ]

            if not selected_archives:
                return jsonify({
                    'success': False,
                    'message': 'No faculty archives found for the selected export.'
                }), 404

            unique_faculties = {(a.faculty_id, (a.faculty_name or '').strip().lower()) for a in selected_archives}
            if len(unique_faculties) > 1:
                return jsonify({
                    'success': False,
                    'message': 'Select a single faculty before exporting this format.'
                }), 400

            ref_archive = selected_archives[0]
            faculty_obj = None
            if faculty_id:
                faculty_obj = Faculty.query.get(faculty_id)
            elif ref_archive.faculty_id:
                faculty_obj = Faculty.query.get(ref_archive.faculty_id)

            if not faculty_obj:
                fallback_full_name = (ref_archive.faculty_name or 'Archived Faculty').strip()
                if ',' in fallback_full_name:
                    parts = [p.strip() for p in fallback_full_name.split(',', 1)]
                    last_name = parts[0] if parts and parts[0] else 'Archived'
                    first_name = parts[1].split()[0] if len(parts) > 1 and parts[1] else 'Faculty'
                else:
                    name_parts = fallback_full_name.split()
                    first_name = name_parts[0] if name_parts else 'Faculty'
                    last_name = name_parts[-1] if len(name_parts) > 1 else 'Archived'

                fallback_department = SimpleNamespace(
                    department_name=ref_archive.program_name or 'Department'
                )
                faculty_obj = SimpleNamespace(
                    full_name=fallback_full_name,
                    first_name=first_name,
                    last_name=last_name,
                    department=fallback_department
                )

            day_rank = {
                'Monday': 1,
                'Tuesday': 2,
                'Wednesday': 3,
                'Thursday': 4,
                'Friday': 5,
                'Saturday': 6,
                'Sunday': 7
            }

            schedule_rows = []
            for idx, archive in enumerate(sorted(
                selected_archives,
                key=lambda a: (
                    day_rank.get(a.day_of_week, 99),
                    a.start_time or datetime.min.time(),
                    a.subject_code or ''
                )
            ), start=1):
                if not archive.day_of_week or not archive.start_time or not archive.end_time:
                    continue

                subject_obj = archive.subject if archive.subject else SimpleNamespace(
                    id=archive.subject_id or idx,
                    subject_code=archive.subject_code or 'TBA'
                )
                room_obj = archive.room if archive.room else SimpleNamespace(room_number=archive.room_number or 'TBA')

                if archive.section:
                    section_obj = archive.section
                else:
                    fallback_program = SimpleNamespace(
                        program_code='',
                        program_name=archive.program_name or 'Program',
                        department_id=None,
                        department=None
                    )
                    section_obj = SimpleNamespace(
                        full_section_name=archive.section_name or 'TBA',
                        program=fallback_program,
                        year_level='',
                        section_name=archive.section_name or 'TBA'
                    )

                schedule_rows.append(SimpleNamespace(
                    day_of_week=archive.day_of_week,
                    start_time=archive.start_time,
                    end_time=archive.end_time,
                    subject=subject_obj,
                    room=room_obj,
                    section=section_obj,
                    faculty=faculty_obj
                ))

            if not schedule_rows:
                return jsonify({
                    'success': False,
                    'message': 'No valid faculty schedule rows found for export.'
                }), 404

            export_settings = current_settings or SimpleNamespace(
                academic_year=academic_year or '',
                semester=semester or '',
                schedule_start_time=None,
                schedule_end_time=None
            )

            output, filename = generate_faculty_schedule_excel(
                faculty_obj,
                schedule_rows,
                export_settings,
                current_user
            )

            return send_file(
                output,
                mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                as_attachment=True,
                download_name=filename
            )

        if export_variant in ('room_template', 'room_excel'):
            from types import SimpleNamespace
            from app.services.export_service import generate_room_schedule_excel

            if group_by != 'room':
                return jsonify({
                    'success': False,
                    'message': 'Room export is only available on the rooms archives tab.'
                }), 400

            selected_archives = archives
            if room_id:
                selected_archives = [a for a in archives if a.room_id == room_id]
            elif room_number:
                selected_archives = [
                    a for a in archives
                    if (a.room_number or '').strip().lower() == room_number.strip().lower()
                ]
                if building_name:
                    selected_archives = [
                        a for a in selected_archives
                        if (a.building_name or '').strip().lower() == building_name.strip().lower()
                    ]

            if not selected_archives:
                return jsonify({
                    'success': False,
                    'message': 'No room archives found for the selected export.'
                }), 404

            unique_rooms = {
                (
                    a.room_id,
                    (a.room_number or '').strip().lower(),
                    (a.building_name or '').strip().lower()
                )
                for a in selected_archives
            }
            if len(unique_rooms) > 1:
                return jsonify({
                    'success': False,
                    'message': 'Select a single room before exporting this format.'
                }), 400

            ref_archive = selected_archives[0]
            room_obj = None
            if room_id:
                from app.models.building import Room
                room_obj = Room.query.get(room_id)
            elif ref_archive.room_id:
                from app.models.building import Room
                room_obj = Room.query.get(ref_archive.room_id)

            if not room_obj:
                room_obj = SimpleNamespace(room_number=ref_archive.room_number or 'TBA')

            day_rank = {
                'Monday': 1,
                'Tuesday': 2,
                'Wednesday': 3,
                'Thursday': 4,
                'Friday': 5,
                'Saturday': 6,
                'Sunday': 7
            }

            schedule_rows = []
            for idx, archive in enumerate(sorted(
                selected_archives,
                key=lambda a: (
                    day_rank.get(a.day_of_week, 99),
                    a.start_time or datetime.min.time(),
                    a.subject_code or ''
                )
            ), start=1):
                if not archive.day_of_week or not archive.start_time or not archive.end_time:
                    continue

                subject_obj = archive.subject if archive.subject else SimpleNamespace(
                    id=archive.subject_id or idx,
                    subject_code=archive.subject_code or 'TBA'
                )

                if archive.section:
                    section_obj = archive.section
                else:
                    fallback_program = SimpleNamespace(
                        program_code='',
                        program_name=archive.program_name or 'Program',
                        department_id=None,
                        department=None
                    )
                    section_obj = SimpleNamespace(
                        full_section_name=archive.section_name or 'TBA',
                        program=fallback_program,
                        year_level='',
                        section_name=archive.section_name or 'TBA'
                    )

                faculty_obj = archive.faculty if archive.faculty else SimpleNamespace(
                    full_name=archive.faculty_name or 'TBA'
                )

                schedule_rows.append(SimpleNamespace(
                    day_of_week=archive.day_of_week,
                    start_time=archive.start_time,
                    end_time=archive.end_time,
                    subject=subject_obj,
                    section=section_obj,
                    faculty=faculty_obj,
                    room=room_obj
                ))

            if not schedule_rows:
                return jsonify({
                    'success': False,
                    'message': 'No valid room schedule rows found for export.'
                }), 404

            export_settings = current_settings or SimpleNamespace(
                academic_year=academic_year or '',
                semester=semester or '',
                schedule_start_time=None,
                schedule_end_time=None
            )

            output, filename = generate_room_schedule_excel(
                room_obj,
                schedule_rows,
                export_settings,
                current_user
            )

            return send_file(
                output,
                mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                as_attachment=True,
                download_name=filename
            )
        
        # Group archives dynamically
        from collections import OrderedDict
        groups = OrderedDict()
        for archive in archives:
            if group_by == 'faculty':
                key = archive.faculty_name or 'No Faculty'
            elif group_by == 'room':
                key = f"{archive.room_number or 'No Room'}{(' - ' + archive.building_name) if archive.building_name else ''}"
            else:
                key = archive.section_name or 'No Section'
            if key not in groups:
                groups[key] = []
            groups[key].append(archive)
        
        # Generate grid-format Excel (same layout as regular schedule export)
        from app.services.export_service import generate_archive_schedule_excel

        output, filename = generate_archive_schedule_excel(
            groups=groups,
            group_by=group_by,
            filter_info=filter_info,
            current_settings=current_settings,
            user=current_user
        )

        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=filename
        )

    except Exception as e:
        return jsonify({
            'success': False,
            'message': f'Error exporting archives: {str(e)}'
        }), 500


@archive_bp.route('/export/pdf')
@login_required
def export_archives_pdf():
    """Export filtered archives to PDF"""
    import io
    from flask import send_file
    from reportlab.lib.pagesizes import letter, landscape
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import inch
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib import colors
    from reportlab.lib.enums import TA_CENTER
    
    try:
        # Get filter parameters
        academic_year = request.args.get('academic_year', '')
        semester = request.args.get('semester', '')
        schedule_type = request.args.get('schedule_type', '')
        exam_period = request.args.get('exam_period', '')
        section_id = request.args.get('section_id', type=int)
        faculty_id = request.args.get('faculty_id', type=int)
        room_id = request.args.get('room_id', type=int)
        building_id = request.args.get('building_id', type=int)
        building_name = request.args.get('building_name', '')
        section_name = request.args.get('section_name', '')
        faculty_name = request.args.get('faculty_name', '')
        room_number = request.args.get('room_number', '')
        program_id = request.args.get('program_id', type=int)
        program = request.args.get('program', '')
        group_by = request.args.get('group_by', 'section')  # 'section', 'faculty', or 'room'
        
        # Get user's program access
        user_program_ids = current_user.get_program_ids()
        
        # Build query
        query = Archive.query
        
        # Apply filters
        if academic_year:
            query = query.filter(Archive.academic_year == academic_year)
        if semester:
            query = query.filter(Archive.semester == semester)
        if schedule_type:
            types = [t.strip() for t in schedule_type.split(',')]
            if len(types) > 1:
                query = query.filter(Archive.schedule_type.in_(types))
            else:
                query = query.filter(Archive.schedule_type == schedule_type)
        if exam_period:
            query = query.filter(Archive.exam_period == exam_period)

        # Prefer ID-based filters for consistent behavior with live schedules export.
        if section_id:
            query = query.filter(Archive.section_id == section_id)
        elif section_name:
            query = query.filter(Archive.section_name.ilike(f'%{section_name}%'))

        if faculty_id:
            query = query.filter(Archive.faculty_id == faculty_id)
        elif faculty_name:
            query = query.filter(Archive.faculty_name.ilike(f'%{faculty_name}%'))

        if room_id:
            query = query.filter(Archive.room_id == room_id)
        elif room_number:
            query = query.filter(Archive.room_number.ilike(f'%{room_number}%'))

        if building_id:
            from app.models.building import Room
            room_ids = db.session.query(Room.id).filter(Room.building_id == building_id).all()
            room_ids = [rid[0] for rid in room_ids]
            if room_ids:
                query = query.filter(Archive.room_id.in_(room_ids))
            else:
                query = query.filter(Archive.id == -1)

        if building_name:
            query = query.filter(Archive.building_name.ilike(f'%{building_name}%'))

        if program_id:
            section_ids = db.session.query(Section.id).filter(Section.program_id == program_id).all()
            section_ids = [sid[0] for sid in section_ids]
            if section_ids:
                query = query.filter(Archive.section_id.in_(section_ids))
            else:
                query = query.filter(Archive.id == -1)
        # Apply text-based program filter only when no explicit program_id is provided.
        if program and not program_id:
            query = query.filter(Archive.program_name.ilike(f'%{program}%'))
        
        # Filter by program access
        if user_program_ids is not None:
            accessible_section_ids = db.session.query(Section.id)\
                .filter(Section.program_id.in_(user_program_ids))\
                .all()
            accessible_section_ids = [sid[0] for sid in accessible_section_ids]
            if accessible_section_ids:
                query = query.filter(Archive.section_id.in_(accessible_section_ids))
        
        # Determine grouping and ordering
        day_order = db.case(
            {'Monday': 1, 'Tuesday': 2, 'Wednesday': 3, 'Thursday': 4, 'Friday': 5, 'Saturday': 6, 'Sunday': 7},
            value=Archive.day_of_week,
            else_=8
        )
        
        if group_by == 'faculty':
            order_col = Archive.faculty_name
            group_label = 'faculty'
        elif group_by == 'room':
            order_col = Archive.room_number
            group_label = 'room'
        else:
            order_col = Archive.section_name
            group_label = 'section'
        
        archives = query.order_by(order_col, day_order, Archive.start_time).all()

        if not archives:
            return jsonify({
                'success': False,
                'message': 'No archives found for the selected filters'
            }), 404
        
        # Group archives dynamically
        from collections import OrderedDict
        groups = OrderedDict()
        for archive in archives:
            if group_by == 'faculty':
                key = archive.faculty_name or 'No Faculty'
            elif group_by == 'room':
                key = f"{archive.room_number or 'No Room'}{(' - ' + archive.building_name) if archive.building_name else ''}"
            else:
                key = archive.section_name or 'No Section'
            if key not in groups:
                groups[key] = []
            groups[key].append(archive)
        
        # PDF column config based on group_by
        if group_by == 'faculty':
            col_headers = ['Subject', 'Section', 'Room', 'Day/Date', 'Time', 'Type', 'AY', 'Sem']
            col_widths_pdf = [1.1*inch, 1*inch, 0.9*inch, 0.9*inch, 1.3*inch, 0.7*inch, 1*inch, 0.5*inch]
        elif group_by == 'room':
            col_headers = ['Subject', 'Section', 'Faculty', 'Day/Date', 'Time', 'Type', 'AY', 'Sem']
            col_widths_pdf = [1.1*inch, 1*inch, 1.5*inch, 0.9*inch, 1.3*inch, 0.7*inch, 0.8*inch, 0.5*inch]
        else:
            col_headers = ['Subject', 'Faculty', 'Room', 'Day/Date', 'Time', 'Type', 'AY', 'Sem']
            col_widths_pdf = [1.1*inch, 1.8*inch, 0.9*inch, 0.9*inch, 1.3*inch, 0.7*inch, 1*inch, 0.5*inch]
        
        # Create PDF
        output = io.BytesIO()
        doc = SimpleDocTemplate(
            output,
            pagesize=landscape(letter),
            rightMargin=0.5*inch,
            leftMargin=0.5*inch,
            topMargin=0.5*inch,
            bottomMargin=0.5*inch
        )
        
        # Styles
        styles = getSampleStyleSheet()
        title_style = ParagraphStyle(
            'Title',
            parent=styles['Heading1'],
            fontSize=16,
            alignment=TA_CENTER,
            spaceAfter=12
        )
        subtitle_style = ParagraphStyle(
            'Subtitle',
            parent=styles['Normal'],
            fontSize=10,
            alignment=TA_CENTER,
            spaceAfter=6
        )
        group_header_style = ParagraphStyle(
            'GroupHeader',
            parent=styles['Heading2'],
            fontSize=11,
            spaceAfter=4,
            spaceBefore=12,
            textColor=colors.HexColor('#1E40AF')
        )
        
        elements = []
        
        # Title
        elements.append(Paragraph(f"Archived Schedules Report (by {group_label.title()})", title_style))
        
        # Filters info
        filter_text = []
        if academic_year:
            filter_text.append(f"Academic Year: {academic_year}")
        if semester:
            filter_text.append(f"Semester: {semester}")
        if schedule_type:
            filter_text.append(f"Type: {schedule_type}")
        if exam_period:
            filter_text.append(f"Exam Period: {exam_period}")
        
        elements.append(Paragraph(' | '.join(filter_text) if filter_text else "All Archives", subtitle_style))
        elements.append(Paragraph(f"Exported on: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}", subtitle_style))
        elements.append(Spacer(1, 0.25*inch))
        
        for group_name, group_archives in groups.items():
            # Group header
            elements.append(Paragraph(f"{group_name} ({len(group_archives)} schedule{'s' if len(group_archives) != 1 else ''})", group_header_style))
            
            # Table data
            data = [col_headers]
            
            for archive in group_archives:
                time_str = f"{archive.start_time.strftime('%I:%M %p') if archive.start_time else ''} - {archive.end_time.strftime('%I:%M %p') if archive.end_time else ''}"
                day_str = archive.day_of_week or (archive.exam_date.strftime('%Y-%m-%d') if archive.exam_date else 'N/A')
                
                if group_by == 'faculty':
                    row = [
                        archive.subject_code or '',
                        archive.section_name or '',
                        archive.room_number or '',
                        day_str, time_str,
                        archive.schedule_type or '',
                        archive.academic_year or '',
                        archive.semester[:3] if archive.semester else ''
                    ]
                elif group_by == 'room':
                    row = [
                        archive.subject_code or '',
                        archive.section_name or '',
                        archive.faculty_name[:20] + '...' if archive.faculty_name and len(archive.faculty_name) > 20 else (archive.faculty_name or ''),
                        day_str, time_str,
                        archive.schedule_type or '',
                        archive.academic_year or '',
                        archive.semester[:3] if archive.semester else ''
                    ]
                else:
                    row = [
                        archive.subject_code or '',
                        archive.faculty_name[:20] + '...' if archive.faculty_name and len(archive.faculty_name) > 20 else (archive.faculty_name or ''),
                        archive.room_number or '',
                        day_str, time_str,
                        archive.schedule_type or '',
                        archive.academic_year or '',
                        archive.semester[:3] if archive.semester else ''
                    ]
                data.append(row)
            
            table = Table(data, colWidths=col_widths_pdf)
            
            table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#2563EB')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 9),
                ('FONTSIZE', (0, 1), (-1, -1), 8),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 8),
                ('TOPPADDING', (0, 0), (-1, 0), 8),
                ('BOTTOMPADDING', (0, 1), (-1, -1), 4),
                ('TOPPADDING', (0, 1), (-1, -1), 4),
                ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
                ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#F8FAFC')]),
                ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ]))
            
            elements.append(table)
            elements.append(Spacer(1, 0.15*inch))
        
        # Total record count
        elements.append(Spacer(1, 0.1*inch))
        elements.append(Paragraph(f"Total Records: {len(archives)} ({len(groups)} {group_label}{'s' if len(groups) != 1 else ''})", subtitle_style))
        
        # Build PDF
        doc.build(elements)
        output.seek(0)
        
        filename = f"archives_{datetime.now().strftime('%Y%m%d_%H%M%S')}.pdf"
        
        return send_file(
            output,
            mimetype='application/pdf',
            as_attachment=True,
            download_name=filename
        )
    
    except Exception as e:
        return jsonify({
            'success': False,
            'message': f'Error exporting archives: {str(e)}'
        }), 500


# ==================== BUILDING ARCHIVE ENDPOINTS ====================

@archive_bp.route('/api/buildings')
@login_required
def get_archived_buildings():
    """Get all archived buildings (API endpoint)"""
    try:
        # Get search query
        search = request.args.get('search', '').strip()
        
        # Base query for archived buildings
        query = Building.query.filter_by(is_archived=True)
        
        # Apply search filter if provided
        if search:
            query = query.filter(
                or_(
                    Building.building_name.ilike(f'%{search}%'),
                    Building.building_code.ilike(f'%{search}%')
                )
            )
        
        # Order by most recently archived first
        buildings = query.order_by(Building.archived_at.desc()).all()
        
        # Convert to dictionary format
        buildings_data = []
        for building in buildings:
            # Get archived_by user name
            archived_by_name = 'Unknown'
            if building.archived_by:
                from app.models.user import User
                user = User.query.get(building.archived_by)
                if user:
                    archived_by_name = user.full_name
            
            # Count rooms for this building
            from app.models.building import Room
            rooms_count = Room.query.filter_by(building_id=building.id).count()
            
            buildings_data.append({
                'id': building.id,
                'building_name': building.building_name,
                'rooms_count': rooms_count,
                'is_active': building.is_active,
                'archived_at': building.archived_at.strftime('%Y-%m-%d %H:%M') if building.archived_at else '',
                'archived_by': archived_by_name,
                'archive_reason': building.archive_reason or 'No reason provided'
            })
        
        return jsonify({
            'success': True,
            'buildings': buildings_data
        })
    
    except Exception as e:
        return jsonify({
            'success': False,
            'message': f'Error fetching archived buildings: {str(e)}'
        }), 500


@archive_bp.route('/api/buildings/<int:building_id>/unarchive', methods=['POST'])
@login_required
def unarchive_building(building_id):
    """Restore an archived building"""
    try:
        permission_error = require_archive_manage_permission()
        if permission_error:
            return permission_error

        building = Building.query.get_or_404(building_id)
        
        # Check if building is actually archived
        if not building.is_archived:
            return jsonify({
                'success': False,
                'message': 'Building is not archived'
            }), 400
        
        # Unarchive the building
        building.unarchive()
        
        # Log activity
        log_unarchive('building', building.id, building.building_name)
        
        db.session.commit()
        
        return jsonify({
            'success': True,
            'message': f'Building "{building.building_name}" restored successfully'
        })
    
    except Exception as e:
        db.session.rollback()
        return jsonify({
            'success': False,
            'message': f'Error restoring building: {str(e)}'
        }), 500


@archive_bp.route('/api/buildings/<int:building_id>', methods=['DELETE'])
@login_required
def delete_archived_building(building_id):
    """Permanently delete an archived building"""
    try:
        permission_error = require_archive_manage_permission()
        if permission_error:
            return permission_error

        building = Building.query.get_or_404(building_id)
        
        # Check if building is actually archived
        if not building.is_archived:
            return jsonify({
                'success': False,
                'message': 'Can only permanently delete archived buildings'
            }), 400
        
        building_name = building.building_name
        
        # Delete the building (cascades to rooms)
        db.session.delete(building)
        db.session.commit()
        
        return jsonify({
            'success': True,
            'message': f'Building "{building_name}" permanently deleted'
        })
    
    except Exception as e:
        db.session.rollback()
        return jsonify({
            'success': False,
            'message': f'Error deleting building: {str(e)}'
        }), 500


@archive_bp.route('/api/buildings/stats')
@login_required
def get_building_archive_stats():
    """Get building archive statistics"""
    try:
        # Base query for archived buildings
        query = Building.query.filter_by(is_archived=True)
        
        # Total archived buildings
        total = query.count()
        
        return jsonify({
            'success': True,
            'stats': {
                'total': total
            }
        })
    
    except Exception as e:
        return jsonify({
            'success': False,
            'message': f'Error fetching building archive stats: {str(e)}'
        }), 500


# ==========================================
# Bulk Operations for Entity Archives
# ==========================================

@archive_bp.route('/api/curricula/bulk-delete', methods=['POST'])
@login_required
def bulk_delete_curricula():
    """Permanently delete multiple archived curricula"""
    try:
        permission_error = require_archive_manage_permission()
        if permission_error:
            return permission_error

        data = request.get_json()
        ids = data.get('ids', [])
        
        if not ids:
            return jsonify({
                'success': False,
                'message': 'No curricula selected'
            }), 400
        
        deleted_count = 0
        errors = []
        
        for curriculum_id in ids:
            try:
                curriculum = Curriculum.query.get(curriculum_id)
                if not curriculum:
                    errors.append(f'Curriculum {curriculum_id} not found')
                    continue
                
                if not curriculum.is_archived:
                    errors.append(f'Curriculum {curriculum.curriculum_code} is not archived')
                    continue
                
                db.session.delete(curriculum)
                deleted_count += 1
            except Exception as e:
                errors.append(f'Error deleting curriculum {curriculum_id}: {str(e)}')
        
        db.session.commit()
        
        message = f'Successfully deleted {deleted_count} curriculum(s)'
        if errors:
            message += f'. Errors: {len(errors)}'
        
        return jsonify({
            'success': True,
            'message': message,
            'deleted_count': deleted_count,
            'errors': errors
        })
    
    except Exception as e:
        db.session.rollback()
        return jsonify({
            'success': False,
            'message': f'Error deleting curricula: {str(e)}'
        }), 500


@archive_bp.route('/api/curricula/bulk-restore', methods=['POST'])
@login_required
def bulk_restore_curricula():
    """Restore multiple archived curricula"""
    try:
        permission_error = require_archive_manage_permission()
        if permission_error:
            return permission_error

        data = request.get_json()
        ids = data.get('ids', [])
        
        if not ids:
            return jsonify({
                'success': False,
                'message': 'No curricula selected'
            }), 400
        
        restored_count = 0
        errors = []
        
        for curriculum_id in ids:
            try:
                curriculum = Curriculum.query.get(curriculum_id)
                if not curriculum:
                    errors.append(f'Curriculum {curriculum_id} not found')
                    continue
                
                if not curriculum.is_archived:
                    errors.append(f'Curriculum {curriculum.curriculum_code} is not archived')
                    continue
                
                curriculum.unarchive()
                log_unarchive('curriculum', curriculum.id, curriculum.curriculum_code)
                restored_count += 1
            except Exception as e:
                errors.append(f'Error restoring curriculum {curriculum_id}: {str(e)}')
        
        db.session.commit()
        
        message = f'Successfully restored {restored_count} curriculum(s)'
        if errors:
            message += f'. Errors: {len(errors)}'
        
        return jsonify({
            'success': True,
            'message': message,
            'restored_count': restored_count,
            'errors': errors
        })
    
    except Exception as e:
        db.session.rollback()
        return jsonify({
            'success': False,
            'message': f'Error restoring curricula: {str(e)}'
        }), 500


@archive_bp.route('/api/programs/bulk-delete', methods=['POST'])
@login_required
def bulk_delete_departments():
    """Permanently delete multiple archived programs (programs)"""
    try:
        permission_error = require_archive_manage_permission()
        if permission_error:
            return permission_error

        data = request.get_json()
        ids = data.get('ids', [])
        
        if not ids:
            return jsonify({
                'success': False,
                'message': 'No programs selected'
            }), 400
        
        deleted_count = 0
        errors = []
        
        for program_id in ids:
            try:
                program = Program.query.get(program_id)
                if not program:
                    errors.append(f'Program {program_id} not found')
                    continue
                
                if not program.is_archived:
                    errors.append(f'Program {program.program_code} is not archived')
                    continue
                
                db.session.delete(program)
                deleted_count += 1
            except Exception as e:
                errors.append(f'Error deleting program {program_id}: {str(e)}')
        
        db.session.commit()
        
        message = f'Successfully deleted {deleted_count} program(s)'
        if errors:
            message += f'. Errors: {len(errors)}'
        
        return jsonify({
            'success': True,
            'message': message,
            'deleted_count': deleted_count,
            'errors': errors
        })
    
    except Exception as e:
        db.session.rollback()
        return jsonify({
            'success': False,
            'message': f'Error deleting programs: {str(e)}'
        }), 500


@archive_bp.route('/api/programs/bulk-restore', methods=['POST'])
@login_required
def bulk_restore_departments():
    """Restore multiple archived programs (programs)"""
    try:
        permission_error = require_archive_manage_permission()
        if permission_error:
            return permission_error

        data = request.get_json()
        ids = data.get('ids', [])
        
        if not ids:
            return jsonify({
                'success': False,
                'message': 'No programs selected'
            }), 400
        
        restored_count = 0
        errors = []
        
        for program_id in ids:
            try:
                program = Program.query.get(program_id)
                if not program:
                    errors.append(f'Program {program_id} not found')
                    continue
                
                if not program.is_archived:
                    errors.append(f'Program {program.program_code} is not archived')
                    continue
                
                program.unarchive()
                log_unarchive('program', program.id, program.program_code)
                restored_count += 1
            except Exception as e:
                errors.append(f'Error restoring program {program_id}: {str(e)}')
        
        db.session.commit()
        
        message = f'Successfully restored {restored_count} program(s)'
        if errors:
            message += f'. Errors: {len(errors)}'
        
        return jsonify({
            'success': True,
            'message': message,
            'restored_count': restored_count,
            'errors': errors
        })
    
    except Exception as e:
        db.session.rollback()
        return jsonify({
            'success': False,
            'message': f'Error restoring programs: {str(e)}'
        }), 500


@archive_bp.route('/api/faculty/bulk-delete', methods=['POST'])
@login_required
def bulk_delete_faculty():
    """Permanently delete multiple archived faculty members"""
    try:
        permission_error = require_archive_manage_permission()
        if permission_error:
            return permission_error

        data = request.get_json()
        ids = data.get('ids', [])
        
        if not ids:
            return jsonify({
                'success': False,
                'message': 'No faculty members selected'
            }), 400
        
        deleted_count = 0
        errors = []
        
        for faculty_id in ids:
            try:
                faculty = Faculty.query.get(faculty_id)
                if not faculty:
                    errors.append(f'Faculty {faculty_id} not found')
                    continue
                
                if not faculty.is_archived:
                    errors.append(f'Faculty {faculty.full_name} is not archived')
                    continue
                
                db.session.delete(faculty)
                deleted_count += 1
            except Exception as e:
                errors.append(f'Error deleting faculty {faculty_id}: {str(e)}')
        
        db.session.commit()
        
        message = f'Successfully deleted {deleted_count} faculty member(s)'
        if errors:
            message += f'. Errors: {len(errors)}'
        
        return jsonify({
            'success': True,
            'message': message,
            'deleted_count': deleted_count,
            'errors': errors
        })
    
    except Exception as e:
        db.session.rollback()
        return jsonify({
            'success': False,
            'message': f'Error deleting faculty: {str(e)}'
        }), 500


@archive_bp.route('/api/faculty/bulk-restore', methods=['POST'])
@login_required
def bulk_restore_faculty():
    """Restore multiple archived faculty members"""
    try:
        permission_error = require_archive_manage_permission()
        if permission_error:
            return permission_error

        data = request.get_json()
        ids = data.get('ids', [])
        
        if not ids:
            return jsonify({
                'success': False,
                'message': 'No faculty members selected'
            }), 400
        
        restored_count = 0
        errors = []
        
        for faculty_id in ids:
            try:
                faculty = Faculty.query.get(faculty_id)
                if not faculty:
                    errors.append(f'Faculty {faculty_id} not found')
                    continue
                
                if not faculty.is_archived:
                    errors.append(f'Faculty {faculty.full_name} is not archived')
                    continue
                
                faculty.unarchive()
                log_unarchive('faculty', faculty.id, faculty.full_name)
                restored_count += 1
            except Exception as e:
                errors.append(f'Error restoring faculty {faculty_id}: {str(e)}')
        
        db.session.commit()
        
        message = f'Successfully restored {restored_count} faculty member(s)'
        if errors:
            message += f'. Errors: {len(errors)}'
        
        return jsonify({
            'success': True,
            'message': message,
            'restored_count': restored_count,
            'errors': errors
        })
    
    except Exception as e:
        db.session.rollback()
        return jsonify({
            'success': False,
            'message': f'Error restoring faculty: {str(e)}'
        }), 500


@archive_bp.route('/api/buildings/bulk-delete', methods=['POST'])
@login_required
def bulk_delete_buildings():
    """Permanently delete multiple archived buildings"""
    try:
        permission_error = require_archive_manage_permission()
        if permission_error:
            return permission_error

        data = request.get_json()
        ids = data.get('ids', [])
        
        if not ids:
            return jsonify({
                'success': False,
                'message': 'No buildings selected'
            }), 400
        
        deleted_count = 0
        errors = []
        
        for building_id in ids:
            try:
                building = Building.query.get(building_id)
                if not building:
                    errors.append(f'Building {building_id} not found')
                    continue
                
                if not building.is_archived:
                    errors.append(f'Building {building.building_name} is not archived')
                    continue
                
                db.session.delete(building)
                deleted_count += 1
            except Exception as e:
                errors.append(f'Error deleting building {building_id}: {str(e)}')
        
        db.session.commit()
        
        message = f'Successfully deleted {deleted_count} building(s)'
        if errors:
            message += f'. Errors: {len(errors)}'
        
        return jsonify({
            'success': True,
            'message': message,
            'deleted_count': deleted_count,
            'errors': errors
        })
    
    except Exception as e:
        db.session.rollback()
        return jsonify({
            'success': False,
            'message': f'Error deleting buildings: {str(e)}'
        }), 500


@archive_bp.route('/api/buildings/bulk-restore', methods=['POST'])
@login_required
def bulk_restore_buildings():
    """Restore multiple archived buildings"""
    try:
        permission_error = require_archive_manage_permission()
        if permission_error:
            return permission_error

        data = request.get_json()
        ids = data.get('ids', [])
        
        if not ids:
            return jsonify({
                'success': False,
                'message': 'No buildings selected'
            }), 400
        
        restored_count = 0
        errors = []
        
        for building_id in ids:
            try:
                building = Building.query.get(building_id)
                if not building:
                    errors.append(f'Building {building_id} not found')
                    continue
                
                if not building.is_archived:
                    errors.append(f'Building {building.building_name} is not archived')
                    continue
                
                building.unarchive()
                log_unarchive('building', building.id, building.building_name)
                restored_count += 1
            except Exception as e:
                errors.append(f'Error restoring building {building_id}: {str(e)}')
        
        db.session.commit()
        
        message = f'Successfully restored {restored_count} building(s)'
        if errors:
            message += f'. Errors: {len(errors)}'
        
        return jsonify({
            'success': True,
            'message': message,
            'restored_count': restored_count,
            'errors': errors
        })
    
    except Exception as e:
        db.session.rollback()
        return jsonify({
            'success': False,
            'message': f'Error restoring buildings: {str(e)}'
        }), 500


# =================================================================
# ENTITY ARCHIVE EXPORT ROUTES
# =================================================================

@archive_bp.route('/export/<entity_type>/<format_type>')
@login_required
def export_entity(entity_type, format_type):
    """Export archived entities to Excel or PDF"""
    import io
    from flask import send_file
    
    valid_entities = ['curriculum', 'program', 'faculty', 'building']
    valid_formats = ['excel', 'pdf']
    
    if entity_type not in valid_entities:
        return jsonify({'success': False, 'message': 'Invalid entity type'}), 400
    if format_type not in valid_formats:
        return jsonify({'success': False, 'message': 'Invalid format type'}), 400
    
    try:
        if format_type == 'excel':
            return _export_entity_excel(entity_type)
        else:
            return _export_entity_pdf(entity_type)
    except Exception as e:
        return jsonify({'success': False, 'message': f'Export error: {str(e)}'}), 500


def _export_entity_excel(entity_type):
    """Export archived entity data to Excel"""
    import io
    from flask import send_file
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    
    wb = Workbook()
    ws = wb.active
    
    # Style definitions
    header_font = Font(bold=True, color='FFFFFF', size=11)
    header_fill = PatternFill(start_color='D97706', end_color='D97706', fill_type='solid')
    header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    if entity_type == 'curriculum':
        ws.title = 'Archived Curriculum'
        headers = ['#', 'Curriculum Code', 'Program', 'Year Levels', 'Status', 'Archived At', 'Archive Reason']
        items = Curriculum.query.filter_by(is_archived=True).all()
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border
        
        for idx, item in enumerate(items, 1):
            dept = Program.query.get(item.program_id) if item.program_id else None
            row_data = [
                idx,
                item.curriculum_code or '',
                dept.program_name if dept else 'N/A',
                len(item.year_levels) if hasattr(item, 'year_levels') else 0,
                'Active' if item.is_active else 'Inactive',
                item.archived_at.strftime('%Y-%m-%d %H:%M') if item.archived_at else 'N/A',
                item.archive_reason or ''
            ]
            for col, val in enumerate(row_data, 1):
                cell = ws.cell(row=idx + 1, column=col, value=val)
                cell.border = thin_border
    
    elif entity_type == 'program':
        ws.title = 'Archived Programs'
        headers = ['#', 'Code', 'Program Name', 'Status', 'Archived At', 'Archive Reason']
        items = Program.query.filter_by(is_archived=True).all()
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border
        
        for idx, item in enumerate(items, 1):
            row_data = [
                idx,
                item.program_code or '',
                item.program_name or '',
                'Active' if item.is_active else 'Inactive',
                item.archived_at.strftime('%Y-%m-%d %H:%M') if item.archived_at else 'N/A',
                item.archive_reason or ''
            ]
            for col, val in enumerate(row_data, 1):
                cell = ws.cell(row=idx + 1, column=col, value=val)
                cell.border = thin_border
    
    elif entity_type == 'faculty':
        ws.title = 'Archived Faculty'
        headers = ['#', 'Full Name', 'Department', 'Status', 'Archived At', 'Archive Reason']
        items = Faculty.query.filter_by(is_archived=True).all()
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border
        
        for idx, item in enumerate(items, 1):
            dept = Department.query.get(item.department_id) if item.department_id else None
            row_data = [
                idx,
                item.full_name or '',
                dept.department_name if dept else 'N/A',
                'Active' if item.is_active else 'Inactive',
                item.archived_at.strftime('%Y-%m-%d %H:%M') if item.archived_at else 'N/A',
                item.archive_reason or ''
            ]
            for col, val in enumerate(row_data, 1):
                cell = ws.cell(row=idx + 1, column=col, value=val)
                cell.border = thin_border
    
    elif entity_type == 'building':
        ws.title = 'Archived Buildings'
        headers = ['#', 'Building Name', 'Rooms Count', 'Status', 'Archived At', 'Archive Reason']
        items = Building.query.filter_by(is_archived=True).all()
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border
        
        for idx, item in enumerate(items, 1):
            rooms_count = item.room_count if hasattr(item, 'room_count') else 0
            row_data = [
                idx,
                item.building_name or '',
                rooms_count,
                'Active' if item.is_active else 'Inactive',
                item.archived_at.strftime('%Y-%m-%d %H:%M') if item.archived_at else 'N/A',
                item.archive_reason or ''
            ]
            for col, val in enumerate(row_data, 1):
                cell = ws.cell(row=idx + 1, column=col, value=val)
                cell.border = thin_border
    
    # Auto-width columns
    for col in range(1, ws.max_column + 1):
        max_length = 0
        column_letter = get_column_letter(col)
        for row in ws.iter_rows(min_col=col, max_col=col):
            for cell in row:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
        adjusted_width = min(max_length + 4, 50)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    filename = f'archived_{entity_type}s_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=filename
    )


def _export_entity_pdf(entity_type):
    """Export archived entity data to PDF"""
    import io
    from flask import send_file
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import letter, landscape
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import inch
    
    output = io.BytesIO()
    doc = SimpleDocTemplate(output, pagesize=landscape(letter), 
                           topMargin=0.5*inch, bottomMargin=0.5*inch,
                           leftMargin=0.5*inch, rightMargin=0.5*inch)
    styles = getSampleStyleSheet()
    elements = []
    
    # Title
    title_map = {
        'curriculum': 'Archived Curriculum',
        'program': 'Archived Programs',
        'faculty': 'Archived Faculty',
        'building': 'Archived Buildings'
    }
    title_style = ParagraphStyle('CustomTitle', parent=styles['Heading1'], 
                                  fontSize=16, spaceAfter=20, textColor=colors.HexColor('#D97706'))
    elements.append(Paragraph(title_map.get(entity_type, 'Archives'), title_style))
    elements.append(Spacer(1, 12))
    
    cell_style = ParagraphStyle('CellStyle', parent=styles['Normal'], fontSize=8, leading=10)
    
    if entity_type == 'curriculum':
        headers = ['#', 'Code', 'Program', 'Year Levels', 'Status', 'Archived At', 'Reason']
        items = Curriculum.query.filter_by(is_archived=True).all()
        data = [headers]
        for idx, item in enumerate(items, 1):
            dept = Program.query.get(item.program_id) if item.program_id else None
            data.append([
                str(idx),
                item.curriculum_code or '',
                Paragraph(dept.program_name if dept else 'N/A', cell_style),
                str(len(item.year_levels)) if hasattr(item, 'year_levels') else '0',
                'Active' if item.is_active else 'Inactive',
                item.archived_at.strftime('%Y-%m-%d') if item.archived_at else 'N/A',
                Paragraph(item.archive_reason or '', cell_style)
            ])
    
    elif entity_type == 'program':
        headers = ['#', 'Code', 'Program Name', 'Status', 'Archived At', 'Reason']
        items = Program.query.filter_by(is_archived=True).all()
        data = [headers]
        for idx, item in enumerate(items, 1):
            data.append([
                str(idx),
                item.program_code or '',
                Paragraph(item.program_name or '', cell_style),
                'Active' if item.is_active else 'Inactive',
                item.archived_at.strftime('%Y-%m-%d') if item.archived_at else 'N/A',
                Paragraph(item.archive_reason or '', cell_style)
            ])
    
    elif entity_type == 'faculty':
        headers = ['#', 'Full Name', 'Department', 'Status', 'Archived At', 'Reason']
        items = Faculty.query.filter_by(is_archived=True).all()
        data = [headers]
        for idx, item in enumerate(items, 1):
            dept = Department.query.get(item.department_id) if item.department_id else None
            data.append([
                str(idx),
                item.full_name or '',
                Paragraph(dept.department_name if dept else 'N/A', cell_style),
                'Active' if item.is_active else 'Inactive',
                item.archived_at.strftime('%Y-%m-%d') if item.archived_at else 'N/A',
                Paragraph(item.archive_reason or '', cell_style)
            ])
    
    elif entity_type == 'building':
        headers = ['#', 'Building Name', 'Rooms', 'Status', 'Archived At', 'Reason']
        items = Building.query.filter_by(is_archived=True).all()
        data = [headers]
        for idx, item in enumerate(items, 1):
            rooms_count = item.room_count if hasattr(item, 'room_count') else 0
            data.append([
                str(idx),
                item.building_name or '',
                str(rooms_count),
                'Active' if item.is_active else 'Inactive',
                item.archived_at.strftime('%Y-%m-%d') if item.archived_at else 'N/A',
                Paragraph(item.archive_reason or '', cell_style)
            ])
    
    if len(data) > 1:
        table = Table(data, repeatRows=1)
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#D97706')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 9),
            ('FONTSIZE', (0, 1), (-1, -1), 8),
            ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#FEF3C7')]),
            ('TOPPADDING', (0, 0), (-1, -1), 4),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
        ]))
        elements.append(table)
    else:
        elements.append(Paragraph('No archived records found.', styles['Normal']))
    
    doc.build(elements)
    output.seek(0)
    
    filename = f'archived_{entity_type}s_{datetime.now().strftime("%Y%m%d_%H%M%S")}.pdf'
    return send_file(
        output,
        mimetype='application/pdf',
        as_attachment=True,
        download_name=filename
    )
