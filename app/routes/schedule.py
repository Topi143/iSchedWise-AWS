"""
Schedule routes for managing class schedules
"""
from flask import Blueprint, render_template, request, redirect, url_for, flash, jsonify, send_file
from flask_login import login_required, current_user
from sqlalchemy import and_, or_
from datetime import datetime, time, timedelta
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import io
from reportlab.lib.pagesizes import letter, landscape
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, Image as RLImage
from reportlab.lib import colors as rl_colors
from reportlab.lib.enums import TA_CENTER, TA_LEFT
import os

from app.extensions import db, csrf
from app.models.schedule import Schedule
from app.models.exam_schedule import ExamSchedule
from app.models.department import Department, Section
from app.models.curriculum import Subject
from app.models.faculty import Faculty, FacultySubjectAssignment
from app.models.building import Room
from app.models.settings import AcademicSettings
from app.models.user import User
from app.models.activity_log import UserActivityLog
from app.decorators import role_required

schedule_bp = Blueprint('schedule', __name__, url_prefix='/schedule')


# ============================================================================
# HELPER FUNCTIONS FOR EXCEL EXPORT
# ============================================================================

def add_institution_logos(ws):
    """Add institution logos to the worksheet if they exist"""
    try:
        from openpyxl.drawing.image import Image as ExcelImage
        import os
        from flask import current_app
        from PIL import Image, ImageDraw, ImageFont
        
        # Get absolute path to static/images directory
        base_dir = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
        images_dir = os.path.join(base_dir, 'static', 'images')
        
        # Path to logos in static/images (use correct filenames)
        logo_left_path = os.path.join(images_dir, 'norzagaray-college-logo.png')
        logo_right_path = os.path.join(images_dir, 'bagong-pilipinas.png')
        
        # Debug: Print paths to verify
        print(f"[LOGO DEBUG] Left logo path: {logo_left_path}")
        print(f"[LOGO DEBUG] Left logo exists: {os.path.exists(logo_left_path)}")
        print(f"[LOGO DEBUG] Right logo path: {logo_right_path}")
        print(f"[LOGO DEBUG] Right logo exists: {os.path.exists(logo_right_path)}")
        
        # Add left logo (Norzagaray College) in cell A1
        if os.path.exists(logo_left_path):
            img_left = ExcelImage(logo_left_path)
            img_left.width = 100
            img_left.height = 100
            ws.add_image(img_left, 'A1')
            print(f"[LOGO DEBUG] Left logo added successfully")
        else:
            print(f"[LOGO DEBUG] Left logo file not found at: {logo_left_path}")
        
        # Add right logo (Bagong Pilipinas) in cell G1
        if os.path.exists(logo_right_path):
            img_right = ExcelImage(logo_right_path)
            img_right.width = 100
            img_right.height = 100
            ws.add_image(img_right, 'G1')
            print(f"[LOGO DEBUG] Right logo added successfully")
        else:
            # Create placeholder image for Bagong Pilipinas
            print(f"[LOGO DEBUG] Creating placeholder for Bagong Pilipinas logo")
            try:
                import io as img_io
                # Create a 100x100 placeholder image
                placeholder = Image.new('RGB', (100, 100), color='white')
                draw = ImageDraw.Draw(placeholder)
                
                # Draw a border
                draw.rectangle([(0, 0), (99, 99)], outline='#1e40af', width=2)
                
                # Add text
                try:
                    # Try to use a default font, fallback to basic if not available
                    font = ImageFont.truetype("arial.ttf", 10)
                except:
                    font = ImageFont.load_default()
                
                text = "Bagong\nPilipinas"
                draw.text((50, 50), text, fill='#1e40af', font=font, anchor='mm', align='center')
                
                # Save to BytesIO
                img_buffer = img_io.BytesIO()
                placeholder.save(img_buffer, format='PNG')
                img_buffer.seek(0)
                
                # Add to worksheet
                img_right = ExcelImage(img_buffer)
                img_right.width = 100
                img_right.height = 100
                ws.add_image(img_right, 'G1')
                print(f"[LOGO DEBUG] Placeholder image created and added successfully")
            except Exception as placeholder_error:
                print(f"[LOGO DEBUG] Could not create placeholder: {str(placeholder_error)}")
            
    except Exception as e:
        # Log error for debugging
        print(f"[LOGO ERROR] Error adding logos: {str(e)}")
        import traceback
        traceback.print_exc()
        pass  # Continue without logos if not found


def add_institution_logos_for_posting(ws, department_logo_path=None):
    """Add institution logos for posting exports - centered with table layout"""
    try:
        from openpyxl.drawing.image import Image as ExcelImage
        import os
        from PIL import Image, ImageDraw, ImageFont
        
        # Get absolute path to static/images directory
        base_dir = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
        images_dir = os.path.join(base_dir, 'static', 'images')
        
        # Path to left logo
        logo_left_path = os.path.join(images_dir, 'norzagaray-college-logo.png')
        
        # Add left logo (Norzagaray College) in cell B1 - centered with table
        if os.path.exists(logo_left_path):
            img_left = ExcelImage(logo_left_path)
            img_left.width = 100
            img_left.height = 100
            ws.add_image(img_left, 'B1')
            print(f"[LOGO DEBUG] Left logo added for posting export (centered)")
        
        # Add department logo on right side (column G)
        if department_logo_path:
            # Construct full path to department logo
            # Remove leading slash and 'static/' if present in path
            clean_path = department_logo_path.lstrip('/')
            if clean_path.startswith('static/'):
                clean_path = clean_path[7:]  # Remove 'static/' prefix
            
            logo_right_path = os.path.join(base_dir, 'static', clean_path)
            
            print(f"[LOGO DEBUG] Original path: {department_logo_path}")
            print(f"[LOGO DEBUG] Cleaned path: {clean_path}")
            print(f"[LOGO DEBUG] Full path: {logo_right_path}")
            print(f"[LOGO DEBUG] File exists: {os.path.exists(logo_right_path)}")
            
            if os.path.exists(logo_right_path):
                img_right = ExcelImage(logo_right_path)
                img_right.width = 100
                img_right.height = 100
                ws.add_image(img_right, 'G1')
                print(f"[LOGO DEBUG] Department logo added for posting export: {logo_right_path}")
            else:
                print(f"[LOGO DEBUG] Department logo not found at: {logo_right_path}")
                # Create placeholder if logo file doesn't exist
                try:
                    import io as img_io
                    placeholder = Image.new('RGB', (100, 100), color='white')
                    draw = ImageDraw.Draw(placeholder)
                    draw.rectangle([(0, 0), (99, 99)], outline='#d1d5db', width=2)
                    try:
                        font = ImageFont.truetype("arial.ttf", 9)
                    except:
                        font = ImageFont.load_default()
                    text = "Dept\nLogo"
                    draw.text((50, 50), text, fill='#9ca3af', font=font, anchor='mm', align='center')
                    img_buffer = img_io.BytesIO()
                    placeholder.save(img_buffer, format='PNG')
                    img_buffer.seek(0)
                    img_right = ExcelImage(img_buffer)
                    img_right.width = 100
                    img_right.height = 100
                    ws.add_image(img_right, 'G1')
                except Exception as placeholder_error:
                    print(f"[LOGO DEBUG] Could not create placeholder: {str(placeholder_error)}")
        else:
            # No department logo provided - create placeholder
            try:
                import io as img_io
                placeholder = Image.new('RGB', (100, 100), color='white')
                draw = ImageDraw.Draw(placeholder)
                draw.rectangle([(0, 0), (99, 99)], outline='#d1d5db', width=2)
                try:
                    font = ImageFont.truetype("arial.ttf", 9)
                except:
                    font = ImageFont.load_default()
                text = "Dept\nLogo"
                draw.text((50, 50), text, fill='#9ca3af', font=font, anchor='mm', align='center')
                img_buffer = img_io.BytesIO()
                placeholder.save(img_buffer, format='PNG')
                img_buffer.seek(0)
                img_right = ExcelImage(img_buffer)
                img_right.width = 100
                img_right.height = 100
                ws.add_image(img_right, 'G1')
                print(f"[LOGO DEBUG] Placeholder added for posting export (no dept logo)")
            except Exception as placeholder_error:
                print(f"[LOGO DEBUG] Could not create placeholder: {str(placeholder_error)}")
            
    except Exception as e:
        print(f"[LOGO ERROR] Error adding logos for posting: {str(e)}")
        import traceback
        traceback.print_exc()
        pass


def add_institution_header(ws, department_name):
    """Add institution header to the worksheet"""
    # Header - Institution Information (centered across columns C-E)
    ws['C1'] = 'Republic of the Philippines'
    ws['C1'].font = Font(bold=True, size=11)
    ws['C1'].alignment = Alignment(horizontal='center', vertical='center')
    ws.merge_cells('C1:E1')
    
    ws['C2'] = 'Municipality of Norzagaray'
    ws['C2'].font = Font(size=10)
    ws['C2'].alignment = Alignment(horizontal='center', vertical='center')
    ws.merge_cells('C2:E2')
    
    ws['C3'] = 'NORZAGARAY COLLEGE'
    ws['C3'].font = Font(bold=True, size=11)
    ws['C3'].alignment = Alignment(horizontal='center', vertical='center')
    ws.merge_cells('C3:E3')
    
    dept_name = department_name.upper() if department_name else 'COLLEGE'
    ws['C4'] = f'{dept_name}'
    ws['C4'].font = Font(bold=True, size=11)
    ws['C4'].alignment = Alignment(horizontal='center', vertical='center')
    ws.merge_cells('C4:E4')


def add_institution_header_for_posting(ws, department_name):
    """Add institution header for posting exports - centered across full table width (A-H)"""
    # Header - Institution Information (centered across columns A-H)
    ws['A1'] = 'Republic of the Philippines'
    ws['A1'].font = Font(bold=True, size=11)
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
    ws.merge_cells('A1:H1')
    
    ws['A2'] = 'Municipality of Norzagaray'
    ws['A2'].font = Font(size=10)
    ws['A2'].alignment = Alignment(horizontal='center', vertical='center')
    ws.merge_cells('A2:H2')
    
    ws['A3'] = 'NORZAGARAY COLLEGE'
    ws['A3'].font = Font(bold=True, size=11)
    ws['A3'].alignment = Alignment(horizontal='center', vertical='center')
    ws.merge_cells('A3:H3')
    
    # Use department name as provided (don't force uppercase unless it's 'COLLEGE')
    dept_name = department_name if department_name else 'COLLEGE'
    ws['A4'] = f'{dept_name}'
    ws['A4'].font = Font(bold=True, size=11)
    ws['A4'].alignment = Alignment(horizontal='center', vertical='center')
    ws.merge_cells('A4:H4')


def add_schedule_title(ws, title, semester_text, section_display):
    """Add schedule title and metadata"""
    # Title
    ws['A6'] = title
    ws['A6'].font = Font(bold=True, size=12)
    ws['A6'].alignment = Alignment(horizontal='center', vertical='center')
    ws.merge_cells('A6:G6')
    
    # Semester and AY - BOLD
    ws['A7'] = semester_text
    ws['A7'].font = Font(bold=True, size=11)
    ws['A7'].alignment = Alignment(horizontal='center', vertical='center')
    ws.merge_cells('A7:G7')
    
    # Section/Faculty/Room info
    ws['A8'] = section_display
    ws['A8'].font = Font(bold=True, size=11)
    ws['A8'].alignment = Alignment(horizontal='center', vertical='center')
    ws.merge_cells('A8:G8')


def add_schedule_title_for_posting(ws, title, semester_text, section_display):
    """Add schedule title for posting exports - centered across full table width (A-H)"""
    # Title
    ws['A6'] = title
    ws['A6'].font = Font(bold=True, size=12)
    ws['A6'].alignment = Alignment(horizontal='center', vertical='center')
    ws.merge_cells('A6:H6')
    
    # Semester and AY - BOLD
    ws['A7'] = semester_text
    ws['A7'].font = Font(bold=True, size=11)
    ws['A7'].alignment = Alignment(horizontal='center', vertical='center')
    ws.merge_cells('A7:H7')
    
    # Section/Faculty/Room info
    ws['A8'] = section_display
    ws['A8'].font = Font(bold=True, size=11)
    ws['A8'].alignment = Alignment(horizontal='center', vertical='center')
    ws.merge_cells('A8:H8')


def generate_time_slots(start_hour=7, end_hour=20):
    """Generate 30-minute time slots from start_hour to end_hour"""
    time_slots = []
    current_time = time(start_hour, 0)
    
    while current_time.hour < end_hour:
        next_time = (datetime.combine(datetime.today(), current_time) + timedelta(minutes=30)).time()
        
        # Format time slots (7:00-7:30, 12:00-12:30, 1:00-1:30, etc.)
        start_str = current_time.strftime('%#I:%M').lstrip('0') if current_time.hour != 12 else current_time.strftime('12:%M')
        end_str = next_time.strftime('%#I:%M').lstrip('0') if next_time.hour != 12 else next_time.strftime('12:%M')
        
        time_slots.append(f"{start_str}-{end_str}")
        current_time = next_time
    
    return time_slots


def add_column_headers(ws):
    """Add day column headers to the worksheet"""
    headers = ['TIME', 'MONDAY', 'TUESDAY', 'WEDNESDAY', 'THURSDAY', 'FRIDAY', 'SATURDAY']
    header_font = Font(bold=False, size=10)
    header_alignment = Alignment(horizontal='center', vertical='center')
    
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=10, column=col_idx, value=header)
        cell.font = header_font
        cell.alignment = header_alignment


def write_time_slots(ws, time_slots):
    """Write time slots to the worksheet"""
    for idx, time_slot in enumerate(time_slots, start=11):
        ws.cell(row=idx, column=1, value=time_slot)
        ws.cell(row=idx, column=1).alignment = Alignment(horizontal='left', vertical='center')
        ws.cell(row=idx, column=1).font = Font(size=10)


def get_day_column_mapping():
    """Get mapping of day names to column indices"""
    return {
        'Monday': 2,
        'Tuesday': 3,
        'Wednesday': 4,
        'Thursday': 5,
        'Friday': 6,
        'Saturday': 7
    }


def place_schedule_in_grid(ws, schedules, start_hour=7):
    """Place schedules in the weekly grid"""
    day_columns = get_day_column_mapping()
    
    for schedule in schedules:
        if schedule.day_of_week not in day_columns:
            continue
        
        col_idx = day_columns[schedule.day_of_week]
        
        # Calculate which 30-minute slot this starts in
        start_minutes = schedule.start_time.hour * 60 + schedule.start_time.minute
        end_minutes = schedule.end_time.hour * 60 + schedule.end_time.minute
        slot_start_minutes = start_hour * 60
        
        # Find start row (each row is 30 minutes, starting from row 11)
        start_row = 11 + ((start_minutes - slot_start_minutes) // 30)
        
        # Calculate how many rows to merge
        duration_minutes = end_minutes - start_minutes
        rows_to_merge = max(1, (duration_minutes + 29) // 30)
        
        # Build cell content
        subject_code = schedule.subject.subject_code if schedule.subject else 'TBA'
        room_display = schedule.room.room_number if schedule.room else 'TBA'
        faculty_name = schedule.faculty.full_name if schedule.faculty else 'TBA'
        
        cell_content = f"{subject_code}\n{room_display}\n{faculty_name}"
        
        # Write to cell
        cell = ws.cell(row=start_row, column=col_idx, value=cell_content)
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.font = Font(size=10)
        
        # Merge cells if needed
        if rows_to_merge > 1:
            end_row = start_row + rows_to_merge - 1
            ws.merge_cells(start_row=start_row, start_column=col_idx, 
                         end_row=end_row, end_column=col_idx)


def apply_grid_borders(ws, last_row):
    """Apply borders to the schedule grid"""
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    for row in range(10, last_row + 1):
        for col in range(1, 8):  # A to G
            ws.cell(row=row, column=col).border = thin_border


def add_signature_section(ws, sig_start_row, department_id, dept_name, dean_name=None):
    """Add signature section with dean and college president"""
    # Prepared by section (column D)
    ws.cell(row=sig_start_row, column=4, value='Prepared by :')
    ws.cell(row=sig_start_row, column=4).font = Font(size=10)
    ws.cell(row=sig_start_row, column=4).alignment = Alignment(horizontal='left')
    
    # Noted section (column F)
    ws.cell(row=sig_start_row, column=6, value='Noted :')
    ws.cell(row=sig_start_row, column=6).font = Font(size=10)
    ws.cell(row=sig_start_row, column=6).alignment = Alignment(horizontal='left')
    
    # Dean name (column D) - Use provided dean name or default - BOLD and UPPERCASE
    dean_display_name = dean_name.upper() if dean_name else 'NAME OF THE DEAN'
    ws.cell(row=sig_start_row + 2, column=4, value=dean_display_name)
    ws.cell(row=sig_start_row + 2, column=4).font = Font(bold=True, size=10)
    ws.cell(row=sig_start_row + 2, column=4).alignment = Alignment(horizontal='left')
    
    # College President name (column F) - BOLD
    ws.cell(row=sig_start_row + 2, column=6, value='Ma. Liberty DG. Pascual, Ph.D')
    ws.cell(row=sig_start_row + 2, column=6).font = Font(bold=True, size=10)
    ws.cell(row=sig_start_row + 2, column=6).alignment = Alignment(horizontal='left')
    
    # Dean title (column D) - Directly under dean name - CENTERED
    dean_title = f"Dean, {dept_name}" if dept_name else "Dean"
    ws.cell(row=sig_start_row + 3, column=4, value=dean_title)
    ws.cell(row=sig_start_row + 3, column=4).font = Font(size=10)
    ws.cell(row=sig_start_row + 3, column=4).alignment = Alignment(horizontal='center')
    
    # College President title (column F)
    ws.cell(row=sig_start_row + 3, column=6, value='College President')
    ws.cell(row=sig_start_row + 3, column=6).font = Font(size=10)
    ws.cell(row=sig_start_row + 3, column=6).alignment = Alignment(horizontal='left')


def set_column_widths(ws):
    """Set standard column widths for schedule grid"""
    ws.column_dimensions['A'].width = 12.71
    for col in ['B', 'C', 'D', 'E', 'F', 'G']:
        ws.column_dimensions[col].width = 20.71


# ============================================================================
# PDF EXPORT HELPER FUNCTIONS (ReportLab)
# ============================================================================

def create_pdf_schedule(schedules, title, semester_text, section_display, dept_name, filename, start_hour=7, end_hour=20):
    """Create PDF schedule with exact same template as Excel export"""
    output = io.BytesIO()
    
    # Create document with landscape orientation to match Excel layout
    doc = SimpleDocTemplate(
        output,
        pagesize=landscape(letter),
        rightMargin=0.5*inch,
        leftMargin=0.5*inch,
        topMargin=0.5*inch,
        bottomMargin=0.5*inch
    )
    
    story = []
    styles = getSampleStyleSheet()
    
    # Custom styles
    header_style = ParagraphStyle(
        'CustomHeader',
        parent=styles['Normal'],
        fontSize=11,
        alignment=TA_CENTER,
        fontName='Helvetica-Bold'
    )
    
    subheader_style = ParagraphStyle(
        'CustomSubHeader',
        parent=styles['Normal'],
        fontSize=10,
        alignment=TA_CENTER,
        fontName='Helvetica'
    )
    
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Normal'],
        fontSize=12,
        alignment=TA_CENTER,
        fontName='Helvetica-Bold'
    )
    
    # Add institution header
    story.append(Paragraph('Republic of the Philippines', header_style))
    story.append(Paragraph('Municipality of Norzagaray', subheader_style))
    story.append(Paragraph('NORZAGARAY COLLEGE', header_style))
    story.append(Paragraph(dept_name.upper() if dept_name else 'COLLEGE', header_style))
    story.append(Spacer(1, 0.1*inch))
    
    # Add title and metadata
    story.append(Paragraph(title, title_style))
    story.append(Paragraph(semester_text, header_style))
    story.append(Paragraph(section_display, header_style))
    story.append(Spacer(1, 0.2*inch))
    
    # Generate time slots with dynamic range
    time_slots = generate_time_slots(start_hour=start_hour, end_hour=end_hour)
    
    # Create schedule grid data
    grid_data = []
    
    # Header row
    headers = ['TIME', 'MONDAY', 'TUESDAY', 'WEDNESDAY', 'THURSDAY', 'FRIDAY', 'SATURDAY']
    grid_data.append(headers)
    
    # Day column mapping
    day_columns = {
        'Monday': 1,
        'Tuesday': 2,
        'Wednesday': 3,
        'Thursday': 4,
        'Friday': 5,
        'Saturday': 6
    }
    
    # Initialize grid with empty cells
    for time_slot in time_slots:
        row = [time_slot] + [''] * 6
        grid_data.append(row)
    
    # Place schedules in grid
    for schedule in schedules:
        day = schedule.day_of_week
        start_time = schedule.start_time
        end_time = schedule.end_time
        
        if day not in day_columns:
            continue
        
        col_idx = day_columns[day]
        
        # Calculate which 30-minute slot this starts in (matching Excel export logic)
        start_minutes = start_time.hour * 60 + start_time.minute
        end_minutes = end_time.hour * 60 + end_time.minute
        slot_start_minutes = start_hour * 60  # Use dynamic start hour
        
        # Find start row index (each row is 30 minutes)
        start_row_idx = (start_minutes - slot_start_minutes) // 30
        
        # Calculate how many rows to span
        duration_minutes = end_minutes - start_minutes
        rows_to_span = max(1, (duration_minutes + 29) // 30)
        
        # Build schedule cell content
        subject_display = schedule.subject.course_description if schedule.subject else 'N/A'
        subject_code = schedule.subject.subject_code if schedule.subject else ''
        type_display = schedule.schedule_type.upper() if schedule.schedule_type else ''
        faculty_display = schedule.faculty.full_name if schedule.faculty else 'TBA'
        room_display = schedule.room.room_number if schedule.room else 'TBA'
        units = schedule.subject.total_units if schedule.subject else 0
        
        cell_content = f"{subject_code}\n{subject_display}\n{type_display}\n{faculty_display}\n{room_display}\n({units} units)"
        
        # Add to grid (row index + 1 because of header row)
        grid_row_idx = start_row_idx + 1
        if 0 <= start_row_idx < len(time_slots):
            if grid_data[grid_row_idx][col_idx] == '':
                grid_data[grid_row_idx][col_idx] = cell_content
            else:
                # Multiple schedules in same slot - append
                grid_data[grid_row_idx][col_idx] += f"\n\n{cell_content}"
    
    # Create table
    col_widths = [1*inch] + [1.5*inch] * 6
    table = Table(grid_data, colWidths=col_widths, repeatRows=1)
    
    # Style table
    table.setStyle(TableStyle([
        # Header row
        ('BACKGROUND', (0, 0), (-1, 0), rl_colors.HexColor('#2563eb')),
        ('TEXTCOLOR', (0, 0), (-1, 0), rl_colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 10),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 8),
        ('TOPPADDING', (0, 0), (-1, 0), 8),
        
        # Time column
        ('ALIGN', (0, 1), (0, -1), 'CENTER'),
        ('FONTNAME', (0, 1), (0, -1), 'Helvetica'),
        ('FONTSIZE', (0, 1), (0, -1), 8),
        
        # Data cells
        ('ALIGN', (1, 1), (-1, -1), 'CENTER'),
        ('VALIGN', (1, 1), (-1, -1), 'MIDDLE'),
        ('FONTNAME', (1, 1), (-1, -1), 'Helvetica'),
        ('FONTSIZE', (1, 1), (-1, -1), 7),
        
        # Grid
        ('GRID', (0, 0), (-1, -1), 0.5, rl_colors.grey),
        ('BOX', (0, 0), (-1, -1), 1, rl_colors.black),
        
        # Alternating row colors for better readability
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [rl_colors.white, rl_colors.HexColor('#f9fafb')]),
    ]))
    
    story.append(table)
    story.append(Spacer(1, 0.3*inch))
    
    # Add signature section
    sig_data = [
        ['', '', 'Prepared by :', '', 'Noted :'],
        ['', '', '', '', ''],
        ['', '', 'Name of the Dean', '', 'Ma. Liberty DG. Pascual, Ph.D'],
        ['', '', f'Dean, {dept_name}' if dept_name else 'Dean', '', 'College President']
    ]
    
    sig_table = Table(sig_data, colWidths=[1.5*inch, 1.5*inch, 2*inch, 1.5*inch, 2.5*inch])
    sig_table.setStyle(TableStyle([
        ('FONTNAME', (2, 0), (2, 0), 'Helvetica'),
        ('FONTNAME', (4, 0), (4, 0), 'Helvetica'),
        ('FONTNAME', (2, 2), (4, 3), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 10),
        ('ALIGN', (2, 0), (2, 0), 'LEFT'),
        ('ALIGN', (4, 0), (4, 0), 'LEFT'),
        ('ALIGN', (2, 2), (2, 3), 'LEFT'),
        ('ALIGN', (4, 2), (4, 3), 'LEFT'),
        ('TOPPADDING', (0, 0), (-1, -1), 2),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 2),
    ]))
    
    story.append(sig_table)
    
    # Build PDF
    doc.build(story)
    output.seek(0)
    
    return output


@schedule_bp.route('/')
@login_required
def index():
    """Schedule management page with two-panel layout"""
    # Get current academic settings
    current_settings = AcademicSettings.query.filter_by(is_active=True).first()
    
    # Get user's department access
    user_department_ids = current_user.get_department_ids()
    
    # Get departments based on user access
    if user_department_ids is None:
        # Admin - see all departments
        departments = Department.query.filter_by(is_active=True).order_by(Department.department_code).all()
    else:
        # Dean - only assigned departments
        departments = Department.query.filter(
            Department.is_active == True,
            Department.id.in_(user_department_ids)
        ).order_by(Department.department_code).all()
    
    # Get department filter from query params
    department_filter = request.args.get('department_id', type=int)
    
    # Auto-apply filter if user has only 1 department and no filter specified
    if department_filter is None and len(departments) == 1:
        department_filter = departments[0].id
    
    # Get sections based on filter
    sections_query = Section.query.filter_by(is_active=True)
    
    # Filter by user's department access
    if user_department_ids is not None:
        sections_query = sections_query.filter(Section.department_id.in_(user_department_ids))
    
    # Apply additional department filter if specified
    if department_filter:
        sections_query = sections_query.filter_by(department_id=department_filter)
    
    sections = sections_query.order_by(Section.section_name).all()
    
    # Calculate schedule counts for current academic period
    # Initialize empty dict to prevent Jinja2 undefined errors
    section_schedule_counts = {}
    
    if current_settings:
        for section in sections:
            count = Schedule.query.filter_by(
                section_id=section.id,
                is_active=True,
                academic_year=current_settings.academic_year,
                semester=current_settings.semester
            ).count()
            section_schedule_counts[section.id] = count
    else:
        # No active settings, count all active schedules
        for section in sections:
            count = Schedule.query.filter_by(
                section_id=section.id,
                is_active=True
            ).count()
            section_schedule_counts[section.id] = count
    
    # Get selected section for class tab
    selected_section_id = request.args.get('section_id', type=int)
    selected_section = None
    schedules = []
    
    if selected_section_id:
        selected_section = Section.query.get(selected_section_id)
        if selected_section:
            # Get schedules for selected section
            schedules_query = Schedule.query.filter_by(
                section_id=selected_section_id,
                is_active=True
            )
            
            # Filter by current academic settings if available
            if current_settings:
                schedules_query = schedules_query.filter_by(
                    academic_year=current_settings.academic_year,
                    semester=current_settings.semester
                )
            
            schedules = schedules_query.order_by(
                Schedule.day_of_week,
                Schedule.start_time
            ).all()
    
    # Get faculties for faculty tab - Show any faculty with schedules in accessible departments
    # Get department filter parameter
    faculty_department_filter = request.args.get('faculty_department_id', type=int)
    
    # Build faculty query - show faculty based on their schedules, not their department assignment
    if user_department_ids is None:
        # Admin - show faculty who have schedules in any department
        if faculty_department_filter:
            # Filter by specific department - show ANY faculty with schedules in that department
            faculty_ids_with_schedules = db.session.query(Schedule.faculty_id).distinct()\
                .join(Section).filter(
                    Schedule.is_active == True,
                    Section.department_id == faculty_department_filter
                )
            if current_settings:
                faculty_ids_with_schedules = faculty_ids_with_schedules.filter(
                    Schedule.academic_year == current_settings.academic_year,
                    Schedule.semester == current_settings.semester
                )
            faculty_ids_with_schedules = [f[0] for f in faculty_ids_with_schedules.all() if f[0] is not None]
            faculties_query = Faculty.query.filter(
                Faculty.is_active == True,
                Faculty.id.in_(faculty_ids_with_schedules)
            )
        else:
            # Show all active faculty
            faculties_query = Faculty.query.filter_by(is_active=True)
    else:
        # Dean - show ANY faculty who have schedules in dean's assigned departments
        if faculty_department_filter and faculty_department_filter in user_department_ids:
            # Filter by specific department within dean's access
            # Show ANY faculty (regardless of their assigned department) with schedules in this department
            faculty_ids_with_schedules = db.session.query(Schedule.faculty_id).distinct()\
                .join(Section).filter(
                    Schedule.is_active == True,
                    Section.department_id == faculty_department_filter
                )
            if current_settings:
                faculty_ids_with_schedules = faculty_ids_with_schedules.filter(
                    Schedule.academic_year == current_settings.academic_year,
                    Schedule.semester == current_settings.semester
                )
            faculty_ids_with_schedules = [f[0] for f in faculty_ids_with_schedules.all() if f[0] is not None]
            faculties_query = Faculty.query.filter(
                Faculty.is_active == True,
                Faculty.id.in_(faculty_ids_with_schedules)
            )
        elif len(user_department_ids) == 1:
            # Single department - auto-filter to show ANY faculty with schedules in that department
            faculty_ids_with_schedules = db.session.query(Schedule.faculty_id).distinct()\
                .join(Section).filter(
                    Schedule.is_active == True,
                    Section.department_id.in_(user_department_ids)
                )
            if current_settings:
                faculty_ids_with_schedules = faculty_ids_with_schedules.filter(
                    Schedule.academic_year == current_settings.academic_year,
                    Schedule.semester == current_settings.semester
                )
            faculty_ids_with_schedules = [f[0] for f in faculty_ids_with_schedules.all() if f[0] is not None]
            faculties_query = Faculty.query.filter(
                Faculty.is_active == True,
                Faculty.id.in_(faculty_ids_with_schedules)
            )
        else:
            # Multiple departments - show ANY faculty with schedules in any of dean's departments
            faculty_ids_with_schedules = db.session.query(Schedule.faculty_id).distinct()\
                .join(Section).filter(
                    Schedule.is_active == True,
                    Section.department_id.in_(user_department_ids)
                )
            if current_settings:
                faculty_ids_with_schedules = faculty_ids_with_schedules.filter(
                    Schedule.academic_year == current_settings.academic_year,
                    Schedule.semester == current_settings.semester
                )
            faculty_ids_with_schedules = [f[0] for f in faculty_ids_with_schedules.all() if f[0] is not None]
            faculties_query = Faculty.query.filter(
                Faculty.is_active == True,
                Faculty.id.in_(faculty_ids_with_schedules)
            )
    
    faculties_list = faculties_query.order_by(Faculty.full_name).all()
    
    # Calculate faculty schedule counts for current academic period
    faculty_schedule_counts = {}
    if current_settings:
        for faculty in faculties_list:
            count = Schedule.query.filter_by(
                faculty_id=faculty.id,
                is_active=True,
                academic_year=current_settings.academic_year,
                semester=current_settings.semester
            ).count()
            faculty_schedule_counts[faculty.id] = count
    else:
        for faculty in faculties_list:
            count = Schedule.query.filter_by(
                faculty_id=faculty.id,
                is_active=True
            ).count()
            faculty_schedule_counts[faculty.id] = count
    
    # Get selected faculty
    selected_faculty_id = request.args.get('faculty_id', type=int)
    selected_faculty = None
    faculty_schedules = []
    
    if selected_faculty_id:
        selected_faculty = Faculty.query.get(selected_faculty_id)
        if selected_faculty:
            faculty_schedules_query = Schedule.query.filter_by(
                faculty_id=selected_faculty_id,
                is_active=True
            )
            if current_settings:
                faculty_schedules_query = faculty_schedules_query.filter_by(
                    academic_year=current_settings.academic_year,
                    semester=current_settings.semester
                )
            faculty_schedules = faculty_schedules_query.order_by(
                Schedule.day_of_week,
                Schedule.start_time
            ).all()
    
    # Get rooms for room tab - Show rooms based on their schedules in accessible departments
    # Get department filter parameter
    room_department_filter = request.args.get('room_department_id', type=int)
    
    # Build room query - show rooms based on their schedules, not building
    if user_department_ids is None:
        # Admin - show rooms that have schedules in any department
        if room_department_filter:
            # Filter by specific department - show ANY room with schedules in that department
            room_ids_with_schedules = db.session.query(Schedule.room_id).distinct()\
                .join(Section).filter(
                    Schedule.is_active == True,
                    Section.department_id == room_department_filter
                )
            if current_settings:
                room_ids_with_schedules = room_ids_with_schedules.filter(
                    Schedule.academic_year == current_settings.academic_year,
                    Schedule.semester == current_settings.semester
                )
            room_ids_with_schedules = [r[0] for r in room_ids_with_schedules.all() if r[0] is not None]
            rooms_query = Room.query.filter(
                Room.is_available == True,
                Room.id.in_(room_ids_with_schedules)
            )
        else:
            # Show all available rooms
            rooms_query = Room.query.filter_by(is_available=True)
    else:
        # Dean - show ANY room that has schedules in dean's assigned departments
        if room_department_filter and room_department_filter in user_department_ids:
            # Filter by specific department within dean's access
            # Show ANY room (regardless of building) with schedules in this department
            room_ids_with_schedules = db.session.query(Schedule.room_id).distinct()\
                .join(Section).filter(
                    Schedule.is_active == True,
                    Section.department_id == room_department_filter
                )
            if current_settings:
                room_ids_with_schedules = room_ids_with_schedules.filter(
                    Schedule.academic_year == current_settings.academic_year,
                    Schedule.semester == current_settings.semester
                )
            room_ids_with_schedules = [r[0] for r in room_ids_with_schedules.all() if r[0] is not None]
            rooms_query = Room.query.filter(
                Room.is_available == True,
                Room.id.in_(room_ids_with_schedules)
            )
        elif len(user_department_ids) == 1:
            # Single department - auto-filter to show ANY room with schedules in that department
            room_ids_with_schedules = db.session.query(Schedule.room_id).distinct()\
                .join(Section).filter(
                    Schedule.is_active == True,
                    Section.department_id.in_(user_department_ids)
                )
            if current_settings:
                room_ids_with_schedules = room_ids_with_schedules.filter(
                    Schedule.academic_year == current_settings.academic_year,
                    Schedule.semester == current_settings.semester
                )
            room_ids_with_schedules = [r[0] for r in room_ids_with_schedules.all() if r[0] is not None]
            rooms_query = Room.query.filter(
                Room.is_available == True,
                Room.id.in_(room_ids_with_schedules)
            )
        else:
            # Multiple departments - show ANY room with schedules in any of dean's departments
            room_ids_with_schedules = db.session.query(Schedule.room_id).distinct()\
                .join(Section).filter(
                    Schedule.is_active == True,
                    Section.department_id.in_(user_department_ids)
                )
            if current_settings:
                room_ids_with_schedules = room_ids_with_schedules.filter(
                    Schedule.academic_year == current_settings.academic_year,
                    Schedule.semester == current_settings.semester
                )
            room_ids_with_schedules = [r[0] for r in room_ids_with_schedules.all() if r[0] is not None]
            rooms_query = Room.query.filter(
                Room.is_available == True,
                Room.id.in_(room_ids_with_schedules)
            )
    
    rooms_list = rooms_query.order_by(Room.room_number).all()
    
    # Calculate room schedule counts for current academic period
    room_schedule_counts = {}
    if current_settings:
        for room in rooms_list:
            count = Schedule.query.filter_by(
                room_id=room.id,
                is_active=True,
                academic_year=current_settings.academic_year,
                semester=current_settings.semester
            ).count()
            room_schedule_counts[room.id] = count
    else:
        for room in rooms_list:
            count = Schedule.query.filter_by(
                room_id=room.id,
                is_active=True
            ).count()
            room_schedule_counts[room.id] = count
    
    # Get selected room
    selected_room_id = request.args.get('room_id', type=int)
    selected_room = None
    room_schedules = []
    
    if selected_room_id:
        selected_room = Room.query.get(selected_room_id)
        if selected_room:
            room_schedules_query = Schedule.query.filter_by(
                room_id=selected_room_id,
                is_active=True
            )
            if current_settings:
                room_schedules_query = room_schedules_query.filter_by(
                    academic_year=current_settings.academic_year,
                    semester=current_settings.semester
                )
            room_schedules = room_schedules_query.order_by(
                Schedule.day_of_week,
                Schedule.start_time
            ).all()
    
    # Get all subjects for modals
    subjects = Subject.query.order_by(Subject.subject_code).all()
    
    # Get exam schedule data
    # Get sections for exam tab (reuse the sections query)
    exam_sections = sections
    exam_department_filter = request.args.get('exam_department_id', type=int)
    
    # Auto-apply filter if user has only 1 department and no filter specified
    if exam_department_filter is None and len(departments) == 1:
        exam_department_filter = departments[0].id
    
    # Calculate exam schedule counts for current academic period
    exam_section_schedule_counts = {}
    if current_settings:
        for section in exam_sections:
            count = ExamSchedule.query.filter_by(
                section_id=section.id,
                is_active=True,
                academic_year=current_settings.academic_year,
                semester=current_settings.semester
            ).count()
            exam_section_schedule_counts[section.id] = count
    else:
        for section in exam_sections:
            count = ExamSchedule.query.filter_by(
                section_id=section.id,
                is_active=True
            ).count()
            exam_section_schedule_counts[section.id] = count
    
    # Get selected section for exam tab
    selected_exam_section_id = request.args.get('exam_section_id', type=int)
    selected_exam_section = None
    exam_schedules = []
    
    if selected_exam_section_id:
        selected_exam_section = Section.query.get(selected_exam_section_id)
        if selected_exam_section:
            exam_schedules_query = ExamSchedule.query.filter_by(
                section_id=selected_exam_section_id,
                is_active=True
            )
            if current_settings:
                exam_schedules_query = exam_schedules_query.filter_by(
                    academic_year=current_settings.academic_year,
                    semester=current_settings.semester
                )
            exam_schedules = exam_schedules_query.order_by(
                ExamSchedule.exam_date,
                ExamSchedule.start_time
            ).all()
    
    # Get all faculties and rooms for exam modals
    all_faculties = Faculty.query.filter_by(is_active=True).order_by(Faculty.full_name).all()
    all_rooms = Room.query.filter_by(is_available=True).order_by(Room.room_number).all()
    
    # Get all buildings for room filter
    from app.models.building import Building
    buildings = Building.query.filter_by(is_active=True).order_by(Building.building_name).all()
    
    # Get time range from settings for calendar view and time dropdowns
    schedule_start_hour = current_settings.schedule_start_hour if current_settings else 7
    schedule_end_hour = current_settings.schedule_end_hour if current_settings else 20
    
    return render_template(
        'schedule.html',
        sections=sections,
        selected_section=selected_section,
        schedules=schedules,
        departments=departments,
        department_filter=department_filter,
        subjects=subjects,
        faculties=faculties_list,
        selected_faculty=selected_faculty,
        faculty_schedules=faculty_schedules,
        faculty_department_filter=faculty_department_filter,
        rooms=rooms_list,
        selected_room=selected_room,
        room_schedules=room_schedules,
        room_department_filter=room_department_filter,
        current_settings=current_settings,
        section_schedule_counts=section_schedule_counts,
        faculty_schedule_counts=faculty_schedule_counts,
        room_schedule_counts=room_schedule_counts,
        # Exam schedule data
        exam_sections=exam_sections,
        selected_exam_section=selected_exam_section,
        exam_schedules=exam_schedules,
        exam_department_filter=exam_department_filter,
        exam_section_schedule_counts=exam_section_schedule_counts,
        all_faculties=all_faculties,
        all_rooms=all_rooms,
        buildings=buildings,
        # Schedule time range for calendar and time dropdowns
        schedule_start_hour=schedule_start_hour,
        schedule_end_hour=schedule_end_hour
    )


@schedule_bp.route('/add', methods=['POST'])
@login_required
def add():
    """Add a new schedule"""
    try:
        section_id = request.form.get('section_id', type=int)
        subject_id = request.form.get('subject_id', type=int)
        faculty_id = request.form.get('faculty_id', type=int)
        room_id = request.form.get('room_id', type=int)
        day_of_week = request.form.get('day_of_week')
        start_time_str = request.form.get('start_time')
        end_time_str = request.form.get('end_time')
        schedule_type = request.form.get('schedule_type', 'lecture')
        
        # Validation - faculty and room are now required
        if not all([section_id, subject_id, faculty_id, room_id, day_of_week, start_time_str, end_time_str]):
            flash('All required fields must be filled (including faculty and room).', 'error')
            return redirect(url_for('schedule.index', section_id=section_id))
        
        # Get current academic settings
        current_settings = AcademicSettings.query.filter_by(is_active=True).first()
        academic_year = current_settings.academic_year if current_settings else None
        semester = current_settings.semester if current_settings else None
        
        # Convert time strings to time objects
        start_time = datetime.strptime(start_time_str, '%H:%M').time()
        end_time = datetime.strptime(end_time_str, '%H:%M').time()
        
        # Validate time range
        if start_time >= end_time:
            flash('End time must be after start time.', 'error')
            return redirect(url_for('schedule.index', section_id=section_id))
        
        # Check for conflicts - same section, day, and overlapping time
        conflict_query = Schedule.query.filter(
            Schedule.section_id == section_id,
            Schedule.day_of_week == day_of_week,
            Schedule.is_active == True,
            or_(
                and_(Schedule.start_time <= start_time, Schedule.end_time > start_time),
                and_(Schedule.start_time < end_time, Schedule.end_time >= end_time),
                and_(Schedule.start_time >= start_time, Schedule.end_time <= end_time)
            )
        )
        
        if academic_year and semester:
            conflict_query = conflict_query.filter(
                Schedule.academic_year == academic_year,
                Schedule.semester == semester
            )
        
        if conflict_query.first():
            flash('Schedule conflict: This section already has a class at this time.', 'error')
            return redirect(url_for('schedule.index', section_id=section_id))
        
        # Check for faculty conflicts if faculty assigned
        if faculty_id:
            faculty_conflict = Schedule.query.filter(
                Schedule.faculty_id == faculty_id,
                Schedule.day_of_week == day_of_week,
                Schedule.is_active == True,
                or_(
                    and_(Schedule.start_time <= start_time, Schedule.end_time > start_time),
                    and_(Schedule.start_time < end_time, Schedule.end_time >= end_time),
                    and_(Schedule.start_time >= start_time, Schedule.end_time <= end_time)
                )
            )
            
            if academic_year and semester:
                faculty_conflict = faculty_conflict.filter(
                    Schedule.academic_year == academic_year,
                    Schedule.semester == semester
                )
            
            if faculty_conflict.first():
                flash('Faculty conflict: This faculty member is already assigned to another class at this time.', 'error')
                return redirect(url_for('schedule.index', section_id=section_id))
        
        # Check for room conflicts if room assigned
        if room_id:
            room_conflict = Schedule.query.filter(
                Schedule.room_id == room_id,
                Schedule.day_of_week == day_of_week,
                Schedule.is_active == True,
                or_(
                    and_(Schedule.start_time <= start_time, Schedule.end_time > start_time),
                    and_(Schedule.start_time < end_time, Schedule.end_time >= end_time),
                    and_(Schedule.start_time >= start_time, Schedule.end_time <= end_time)
                )
            )
            
            if academic_year and semester:
                room_conflict = room_conflict.filter(
                    Schedule.academic_year == academic_year,
                    Schedule.semester == semester
                )
            
            if room_conflict.first():
                flash('Room conflict: This room is already booked at this time.', 'error')
                return redirect(url_for('schedule.index', section_id=section_id))
        
        # Create new schedule
        new_schedule = Schedule(
            section_id=section_id,
            subject_id=subject_id,
            faculty_id=faculty_id,
            room_id=room_id,
            day_of_week=day_of_week,
            start_time=start_time,
            end_time=end_time,
            schedule_type=schedule_type,
            academic_year=academic_year,
            semester=semester,
            is_active=True
        )
        
        db.session.add(new_schedule)
        db.session.flush()  # Get the schedule ID
        
        # Log the action
        subject = Subject.query.get(subject_id)
        section = Section.query.get(section_id)
        entity_name = f"{subject.subject_code if subject else 'N/A'} - {section.section_name if section else 'N/A'}"
        
        UserActivityLog.log_action(
            user_id=current_user.id,
            action='created',
            entity_type='schedule',
            entity_id=new_schedule.id,
            entity_name=entity_name,
            details=f'Created schedule for {day_of_week} {start_time}-{end_time}',
            ip_address=request.remote_addr,
            user_agent=request.headers.get('User-Agent')
        )
        
        db.session.commit()
        
        flash('Schedule added successfully!', 'success')
        
    except Exception as e:
        db.session.rollback()
        flash(f'Error adding schedule: {str(e)}', 'error')
    
    return redirect(url_for('schedule.index', section_id=section_id))


@schedule_bp.route('/edit', methods=['POST'])
@login_required
def edit():
    """Edit an existing schedule"""
    try:
        schedule_id = request.form.get('schedule_id', type=int)
        subject_id = request.form.get('subject_id', type=int)
        faculty_id = request.form.get('faculty_id', type=int)
        room_id = request.form.get('room_id', type=int)
        day_of_week = request.form.get('day_of_week')
        start_time_str = request.form.get('start_time')
        end_time_str = request.form.get('end_time')
        schedule_type = request.form.get('schedule_type', 'lecture')
        
        # Get schedule
        schedule = Schedule.query.get_or_404(schedule_id)
        section_id = schedule.section_id
        
        # Validation - faculty and room are now required
        if not all([subject_id, faculty_id, room_id, day_of_week, start_time_str, end_time_str]):
            flash('All required fields must be filled (including faculty and room).', 'error')
            return redirect(url_for('schedule.index', section_id=section_id))
        
        # Convert time strings to time objects
        start_time = datetime.strptime(start_time_str, '%H:%M').time()
        end_time = datetime.strptime(end_time_str, '%H:%M').time()
        
        # Validate time range
        if start_time >= end_time:
            flash('End time must be after start time.', 'error')
            return redirect(url_for('schedule.index', section_id=section_id))
        
        # Check for conflicts (excluding current schedule)
        conflict_query = Schedule.query.filter(
            Schedule.id != schedule_id,
            Schedule.section_id == schedule.section_id,
            Schedule.day_of_week == day_of_week,
            Schedule.is_active == True,
            or_(
                and_(Schedule.start_time <= start_time, Schedule.end_time > start_time),
                and_(Schedule.start_time < end_time, Schedule.end_time >= end_time),
                and_(Schedule.start_time >= start_time, Schedule.end_time <= end_time)
            )
        )
        
        if schedule.academic_year and schedule.semester:
            conflict_query = conflict_query.filter(
                Schedule.academic_year == schedule.academic_year,
                Schedule.semester == schedule.semester
            )
        
        if conflict_query.first():
            flash('Schedule conflict: This section already has a class at this time.', 'error')
            return redirect(url_for('schedule.index', section_id=section_id))
        
        # Check for faculty conflicts if faculty assigned
        if faculty_id:
            faculty_conflict = Schedule.query.filter(
                Schedule.id != schedule_id,
                Schedule.faculty_id == faculty_id,
                Schedule.day_of_week == day_of_week,
                Schedule.is_active == True,
                or_(
                    and_(Schedule.start_time <= start_time, Schedule.end_time > start_time),
                    and_(Schedule.start_time < end_time, Schedule.end_time >= end_time),
                    and_(Schedule.start_time >= start_time, Schedule.end_time <= end_time)
                )
            )
            
            if schedule.academic_year and schedule.semester:
                faculty_conflict = faculty_conflict.filter(
                    Schedule.academic_year == schedule.academic_year,
                    Schedule.semester == schedule.semester
                )
            
            if faculty_conflict.first():
                flash('Faculty conflict: This faculty member is already assigned to another class at this time.', 'error')
                return redirect(url_for('schedule.index', section_id=section_id))
        
        # Check for room conflicts if room assigned
        if room_id:
            room_conflict = Schedule.query.filter(
                Schedule.id != schedule_id,
                Schedule.room_id == room_id,
                Schedule.day_of_week == day_of_week,
                Schedule.is_active == True,
                or_(
                    and_(Schedule.start_time <= start_time, Schedule.end_time > start_time),
                    and_(Schedule.start_time < end_time, Schedule.end_time >= end_time),
                    and_(Schedule.start_time >= start_time, Schedule.end_time <= end_time)
                )
            )
            
            if schedule.academic_year and schedule.semester:
                room_conflict = room_conflict.filter(
                    Schedule.academic_year == schedule.academic_year,
                    Schedule.semester == schedule.semester
                )
            
            if room_conflict.first():
                flash('Room conflict: This room is already booked at this time.', 'error')
                return redirect(url_for('schedule.index', section_id=section_id))
        
        # Update schedule
        schedule.subject_id = subject_id
        schedule.faculty_id = faculty_id
        schedule.room_id = room_id
        schedule.day_of_week = day_of_week
        schedule.start_time = start_time
        schedule.end_time = end_time
        schedule.schedule_type = schedule_type
        schedule.updated_at = datetime.utcnow()
        
        # Log the action
        subject = Subject.query.get(subject_id)
        section = Section.query.get(schedule.section_id)
        entity_name = f"{subject.subject_code if subject else 'N/A'} - {section.section_name if section else 'N/A'}"
        
        UserActivityLog.log_action(
            user_id=current_user.id,
            action='edited',
            entity_type='schedule',
            entity_id=schedule.id,
            entity_name=entity_name,
            details=f'Updated schedule for {day_of_week} {start_time}-{end_time}',
            ip_address=request.remote_addr,
            user_agent=request.headers.get('User-Agent')
        )
        
        db.session.commit()
        
        flash('Schedule updated successfully!', 'success')
        
    except Exception as e:
        db.session.rollback()
        flash(f'Error updating schedule: {str(e)}', 'error')
    
    return redirect(url_for('schedule.index', section_id=section_id))


@schedule_bp.route('/delete', methods=['POST'])
@login_required
def delete():
    """Delete a schedule"""
    try:
        schedule_id = request.form.get('schedule_id', type=int)
        
        schedule = Schedule.query.get_or_404(schedule_id)
        section_id = schedule.section_id
        
        # Soft delete (set is_active to False)
        schedule.is_active = False
        schedule.updated_at = datetime.utcnow()
        
        # Log the action
        entity_name = f"{schedule.subject.subject_code if schedule.subject else 'N/A'} - {schedule.section.section_name if schedule.section else 'N/A'}"
        
        UserActivityLog.log_action(
            user_id=current_user.id,
            action='deleted',
            entity_type='schedule',
            entity_id=schedule.id,
            entity_name=entity_name,
            details=f'Deleted schedule for {schedule.day_of_week} {schedule.start_time}-{schedule.end_time}',
            ip_address=request.remote_addr,
            user_agent=request.headers.get('User-Agent')
        )
        
        db.session.commit()
        
        flash('Schedule deleted successfully!', 'success')
        
    except Exception as e:
        db.session.rollback()
        flash(f'Error deleting schedule: {str(e)}', 'error')
    
    return redirect(url_for('schedule.index', section_id=section_id))


@schedule_bp.route('/get-curricula/<int:section_id>')
@login_required
def get_curricula_for_section(section_id):
    """Get available curricula for a section's department ONLY"""
    from flask import jsonify
    from app.models.curriculum import Curriculum
    
    try:
        # Get the section
        section = Section.query.get_or_404(section_id)
        
        # Ensure section has a department
        if not section.department_id:
            return jsonify({
                'curricula': [],
                'error': 'Section has no department assigned'
            }), 400
        
        # Get ONLY active curricula for THIS SPECIFIC department
        curricula = Curriculum.query.filter_by(
            department_id=section.department_id,
            is_active=True,
            is_archived=False
        ).order_by(Curriculum.curriculum_code).all()
        
        # Debug logging
        print(f"[CURRICULA] Section {section_id} ({section.section_name}) - Department ID: {section.department_id}")
        print(f"[CURRICULA] Found {len(curricula)} curricula for department {section.department_id}")
        
        # Format curricula for JSON response
        curricula_data = [
            {
                'id': curriculum.id,
                'curriculum_code': curriculum.curriculum_code,
                'degree_program': curriculum.degree_program,
                'department_id': curriculum.department_id,  # Include for verification
                'display': f"{curriculum.curriculum_code}"
            }
            for curriculum in curricula
        ]
        
        return jsonify({'curricula': curricula_data})
        
    except Exception as e:
        print(f"[CURRICULA ERROR] {str(e)}")
        return jsonify({'error': str(e)}), 500


@schedule_bp.route('/get-subjects/<int:section_id>')
@login_required
def get_subjects_for_section(section_id):
    """Get subjects for a specific section based on curriculum, year level, and semester"""
    from flask import jsonify
    from app.models.curriculum import Curriculum, YearLevel, Semester
    
    try:
        # Get the section
        section = Section.query.get_or_404(section_id)
        
        # Get curriculum_id from query parameter (optional)
        curriculum_id = request.args.get('curriculum_id', type=int)
        
        # Get current academic settings to determine the semester
        current_settings = AcademicSettings.query.filter_by(is_active=True).first()
        
        if not current_settings:
            return jsonify({'subjects': []})
        
        # Determine semester number from semester name
        semester_mapping = {
            '1st Semester': 1,
            '2nd Semester': 2
        }
        semester_number = semester_mapping.get(current_settings.semester, 1)
        
        # Find curriculum
        if curriculum_id:
            # Use specified curriculum
            curriculum = Curriculum.query.get(curriculum_id)
            if not curriculum or curriculum.department_id != section.department_id:
                return jsonify({'subjects': [], 'error': 'Invalid curriculum for this section'})
        else:
            # Default to first active curriculum for this department
            curriculum = Curriculum.query.filter_by(
                department_id=section.department_id,
                is_active=True,
                is_archived=False
            ).first()
        
        if not curriculum:
            return jsonify({'subjects': []})
        
        # Find the year level
        year_level = YearLevel.query.filter_by(
            curriculum_id=curriculum.id,
            year_number=section.year_level
        ).first()
        
        if not year_level:
            return jsonify({'subjects': []})
        
        # Find the semester within that year level
        semester = Semester.query.filter_by(
            year_level_id=year_level.id,
            semester_number=semester_number
        ).first()
        
        if not semester:
            return jsonify({'subjects': []})
        
        # Get all subjects for this semester
        subjects = Subject.query.filter_by(semester_id=semester.id).order_by(Subject.subject_code).all()
        
        # Format subjects for JSON response with unit information
        subjects_data = [
            {
                'id': subject.id,
                'subject_code': subject.subject_code,
                'course_description': subject.course_description,
                'lec_units': float(subject.lec_units),
                'lab_units': float(subject.lab_units),
                'total_units': subject.total_units,
                'display': f"{subject.subject_code} - {subject.course_description} ({subject.total_units} units)"
            }
            for subject in subjects
        ]
        
        return jsonify({'subjects': subjects_data})
        
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@schedule_bp.route('/get-subject-details/<int:subject_id>')
@login_required
def get_subject_details(subject_id):
    """Get detailed information about a subject including its curriculum"""
    from flask import jsonify
    import traceback
    
    try:
        print(f'[SUBJECT DETAILS] Fetching subject {subject_id}')
        subject = Subject.query.get(subject_id)
        
        if not subject:
            print(f'[SUBJECT DETAILS] Subject {subject_id} not found')
            return jsonify({'error': 'Subject not found'}), 404
        
        print(f'[SUBJECT DETAILS] Subject found: {subject.subject_code}')
        print(f'[SUBJECT DETAILS] Semester ID: {subject.semester_id}')
        
        # Get curriculum through relationships
        semester = subject.semester
        if not semester:
            print(f'[SUBJECT DETAILS] Subject {subject_id} has no semester')
            return jsonify({'error': 'Subject has no semester'}), 404
        
        print(f'[SUBJECT DETAILS] Semester found: {semester.semester_name}')
        print(f'[SUBJECT DETAILS] Year level ID: {semester.year_level_id}')
        
        year_level = semester.year_level
        if not year_level:
            print(f'[SUBJECT DETAILS] Semester {semester.id} has no year level')
            return jsonify({'error': 'Semester has no year level'}), 404
        
        print(f'[SUBJECT DETAILS] Year level found: {year_level.year_name}')
        print(f'[SUBJECT DETAILS] Curriculum ID: {year_level.curriculum_id}')
        
        curriculum = year_level.curriculum
        if not curriculum:
            print(f'[SUBJECT DETAILS] Year level {year_level.id} has no curriculum')
            return jsonify({'error': 'Year level has no curriculum'}), 404
        
        print(f'[SUBJECT DETAILS] Curriculum found: {curriculum.curriculum_code}')
        
        return jsonify({
            'id': subject.id,
            'subject_code': subject.subject_code,
            'course_description': subject.course_description,
            'curriculum_id': curriculum.id,
            'curriculum_code': curriculum.curriculum_code,
            'curriculum_name': curriculum.degree_program
        })
        
    except Exception as e:
        print(f'[SUBJECT DETAILS] ERROR: {str(e)}')
        print(f'[SUBJECT DETAILS] Traceback: {traceback.format_exc()}')
        return jsonify({'error': str(e)}), 500


@schedule_bp.route('/get-subjects-by-curriculum/<int:curriculum_id>')
@login_required
def get_subjects_by_curriculum(curriculum_id):
    """Get all subjects for a specific curriculum"""
    from flask import jsonify
    from app.models.curriculum import Curriculum, YearLevel, Semester
    
    try:
        print(f'[SUBJECTS BY CURRICULUM] Fetching subjects for curriculum {curriculum_id}')
        
        # Get the curriculum
        curriculum = Curriculum.query.get(curriculum_id)
        if not curriculum:
            print(f'[SUBJECTS BY CURRICULUM] Curriculum {curriculum_id} not found')
            return jsonify({'subjects': []})
        
        print(f'[SUBJECTS BY CURRICULUM] Curriculum found: {curriculum.curriculum_code}')
        
        # Get all subjects for this curriculum through year levels and semesters
        subjects = []
        for year_level in curriculum.year_levels:
            for semester in year_level.semesters:
                for subject in semester.subjects:
                    subjects.append({
                        'id': subject.id,
                        'subject_code': subject.subject_code,
                        'course_description': subject.course_description,
                        'lec_units': float(subject.lec_units),
                        'lab_units': float(subject.lab_units),
                        'total_units': subject.total_units,
                        'year_level': year_level.year_name,
                        'semester': semester.semester_name,
                        'display': f"{subject.subject_code} - {subject.course_description} ({subject.total_units} units)"
                    })
        
        print(f'[SUBJECTS BY CURRICULUM] Found {len(subjects)} subjects')
        
        # Sort by year level number, semester number, then subject code
        subjects.sort(key=lambda x: (x['year_level'], x['semester'], x['subject_code']))
        
        return jsonify({'subjects': subjects})
        
    except Exception as e:
        print(f'[SUBJECTS BY CURRICULUM] ERROR: {str(e)}')
        import traceback
        print(f'[SUBJECTS BY CURRICULUM] Traceback: {traceback.format_exc()}')
        return jsonify({'error': str(e)}), 500


@schedule_bp.route('/get-faculty/<int:subject_id>')
@login_required
def get_faculty_for_subject(subject_id):
    """Get faculty members assigned to a specific subject"""
    try:
        # Get current academic settings
        current_settings = AcademicSettings.query.filter_by(is_active=True).first()
        
        if not current_settings:
            return jsonify({'faculty': []})
        
        # Get faculty assignments for this subject and current academic period
        assignments = FacultySubjectAssignment.query.filter_by(
            subject_id=subject_id,
            academic_year=current_settings.academic_year,
            semester=current_settings.semester,
            is_active=True,
            is_archived=False
        ).all()
        
        # Get unique faculty IDs
        faculty_ids = list(set([assignment.faculty_id for assignment in assignments]))
        
        # Get faculty details
        faculties = Faculty.query.filter(
            Faculty.id.in_(faculty_ids),
            Faculty.is_active == True,
            Faculty.is_archived == False
        ).order_by(Faculty.full_name).all()
        
        # Format faculty for JSON response
        faculty_data = [
            {
                'id': faculty.id,
                'full_name': faculty.full_name,
                'department_code': faculty.department.department_code if faculty.department else '',
                'display': f"{faculty.full_name}" + (f" - {faculty.department.department_code}" if faculty.department else "")
            }
            for faculty in faculties
        ]
        
        return jsonify({'faculty': faculty_data})
        
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@schedule_bp.route('/get-all-faculty')
@login_required
def get_all_faculty():
    """Get all active faculty members for exam schedule modals"""
    try:
        # Get all active faculty
        faculties = Faculty.query.filter_by(
            is_active=True,
            is_archived=False
        ).order_by(Faculty.full_name).all()
        
        # Format faculty for JSON response
        faculty_data = [
            {
                'id': faculty.id,
                'full_name': faculty.full_name,
                'department_code': faculty.department.department_code if faculty.department else '',
                'display': f"{faculty.full_name}" + (f" - {faculty.department.department_code}" if faculty.department else "")
            }
            for faculty in faculties
        ]
        
        return jsonify({'faculty': faculty_data})
        
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@schedule_bp.route('/get-all-rooms')
@login_required
def get_all_rooms():
    """Get all available rooms for exam schedule modals"""
    try:
        # Get all available rooms
        rooms = Room.query.filter_by(is_available=True).order_by(Room.room_number).all()
        
        # Format rooms for JSON response
        room_data = [
            {
                'id': room.id,
                'room_number': room.room_number,
                'building_name': room.building.building_name if room.building else '',
                'display': f"{room.room_number}" + (f" - {room.building.building_name}" if room.building else "")
            }
            for room in rooms
        ]
        
        return jsonify({'rooms': room_data})
        
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@schedule_bp.route('/get-all-faculties')
@login_required
def get_all_faculties():
    """Get all active faculty members"""
    try:
        # Get all active faculty
        faculties = Faculty.query.filter(
            Faculty.is_active == True,
            Faculty.is_archived == False
        ).order_by(Faculty.full_name).all()
        
        # Format faculty for JSON response
        faculty_data = [
            {
                'id': faculty.id,
                'full_name': faculty.full_name,
                'department_code': faculty.department.department_code if faculty.department else '',
                'display': f"{faculty.full_name}" + (f" - {faculty.department.department_code}" if faculty.department else "")
            }
            for faculty in faculties
        ]
        
        return jsonify({'faculties': faculty_data})
        
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@schedule_bp.route('/ai-check-conflicts', methods=['POST'])
@login_required
@csrf.exempt  # Exempt CSRF for AJAX endpoints
def ai_check_conflicts():
    """AI-powered conflict detection and recommendations"""
    from app.ai_scheduler import ai_scheduler
    from datetime import datetime as dt
    
    try:
        data = request.get_json()
        
        # Debug logging
        print(f"[AI CHECK] Received data: {data}")
        
        if not data:
            return jsonify({'error': 'No data received', 'ai_enabled': False}), 400
        
        # Parse schedule data
        section_id = data.get('section_id')
        subject_id = data.get('subject_id')
        faculty_id = data.get('faculty_id')
        room_id = data.get('room_id')
        day_of_week = data.get('day_of_week')
        schedule_type = data.get('schedule_type', 'lecture')  # Default to lecture
        start_time_str = data.get('start_time')
        end_time_str = data.get('end_time')
        schedule_id = data.get('schedule_id')  # For edit mode
        
        # Debug logging
        print(f"[AI CHECK] Parsed - section:{section_id} day:{day_of_week} type:{schedule_type} time:{start_time_str}-{end_time_str}")
        
        if not all([section_id, day_of_week, start_time_str, end_time_str]):
            missing = []
            if not section_id: missing.append('section_id')
            if not day_of_week: missing.append('day_of_week')
            if not start_time_str: missing.append('start_time')
            if not end_time_str: missing.append('end_time')
            error_msg = f'Missing required fields: {", ".join(missing)}'
            print(f"[AI CHECK] Validation failed: {error_msg}")
            return jsonify({'error': error_msg, 'ai_enabled': False}), 400
        
        # Convert times
        start_time = dt.strptime(start_time_str, '%H:%M').time()
        end_time = dt.strptime(end_time_str, '%H:%M').time()
        
        # Get current academic settings
        current_settings = AcademicSettings.query.filter_by(is_active=True).first()
        
        # Get existing schedules for the same academic period
        existing_query = Schedule.query.filter_by(is_active=True)
        
        if current_settings:
            existing_query = existing_query.filter_by(
                academic_year=current_settings.academic_year,
                semester=current_settings.semester
            )
        
        # Exclude current schedule if editing
        if schedule_id:
            existing_query = existing_query.filter(Schedule.id != schedule_id)
        
        existing_schedules = existing_query.all()
        
        # Prepare schedule data for AI analysis
        schedule_data = {
            'section_id': section_id,
            'subject_id': subject_id,
            'faculty_id': faculty_id,
            'room_id': room_id,
            'day_of_week': day_of_week,
            'schedule_type': schedule_type,
            'start_time': start_time,
            'end_time': end_time
        }
        
        # Get AI analysis
        analysis = ai_scheduler.analyze_schedule_conflicts(schedule_data, existing_schedules)
        
        # Format response
        response = {
            'ai_enabled': analysis.get('ai_enabled', False),
            'has_conflicts': analysis.get('has_conflicts', False),
            'conflicts': [],
            'recommendations': analysis.get('recommendations', []),
            'ai_explanation': analysis.get('ai_explanation', '')
        }
        
        # Format conflicts for frontend
        for conflict in analysis.get('conflicts', []):
            schedule = conflict.get('schedule')
            response['conflicts'].append({
                'type': conflict['type'],
                'message': conflict['message'],
                'severity': conflict['severity'],
                'details': {
                    'subject': schedule.subject.subject_code if schedule and schedule.subject else 'Unknown',
                    'time': f"{schedule.start_time.strftime('%I:%M %p')} - {schedule.end_time.strftime('%I:%M %p')}" if schedule else '',
                    'day': schedule.day_of_week if schedule else ''
                }
            })
        
        return jsonify(response)
        
    except Exception as e:
        print(f"AI check conflicts error: {str(e)}")
        import traceback
        traceback.print_exc()
        return jsonify({'error': str(e), 'ai_enabled': False}), 500


@schedule_bp.route('/ai-suggest-schedule', methods=['POST'])
@login_required
@csrf.exempt  # Exempt CSRF for AJAX endpoints
def ai_suggest_schedule():
    """Get AI suggestions for optimal scheduling"""
    from flask import jsonify
    from app.ai_scheduler import ai_scheduler
    
    try:
        data = request.get_json()
        
        section_id = data.get('section_id')
        subject_id = data.get('subject_id')
        faculty_id = data.get('faculty_id')
        
        if not section_id or not subject_id:
            return jsonify({'error': 'Section and subject are required'}), 400
        
        # Get models
        section = Section.query.get_or_404(section_id)
        subject = Subject.query.get_or_404(subject_id)
        faculty = Faculty.query.get(faculty_id) if faculty_id else None
        
        # Get AI suggestions
        result = ai_scheduler.suggest_optimal_schedule(section, subject, faculty)
        
        return jsonify(result)
        
    except Exception as e:
        print(f"AI suggest schedule error: {str(e)}")
        import traceback
        traceback.print_exc()
        return jsonify({'error': str(e), 'ai_enabled': False}), 500


@schedule_bp.route('/export/class/<int:section_id>')
@login_required
def export_class_schedule(section_id):
    """Export class schedule to Excel - weekly grid format matching template"""
    try:
        section = Section.query.get_or_404(section_id)
        current_settings = AcademicSettings.query.filter_by(is_active=True).first()
        
        # Get time range from settings (default to 7-20 if not set)
        start_hour = current_settings.schedule_start_hour if current_settings else 7
        end_hour = current_settings.schedule_end_hour if current_settings else 20
        
        # Query schedules
        query = Schedule.query.filter_by(section_id=section_id, is_active=True)
        if current_settings:
            query = query.filter_by(
                academic_year=current_settings.academic_year,
                semester=current_settings.semester
            )
        schedules = query.order_by(Schedule.day_of_week, Schedule.start_time).all()
        
        # Create workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Schedule"
        
        # Add logos
        add_institution_logos(ws)
        
        # Add header
        dept_name = section.department.department_name if section.department else 'COLLEGE'
        add_institution_header(ws, dept_name)
        
        # Add title
        dept_code = section.department.department_code if section.department else ''
        semester_text = f"{current_settings.semester.upper()}, AY {current_settings.academic_year}" if current_settings else "CLASS SCHEDULE"
        section_display = f"{dept_code} {section.year_level}{section.section_name}"
        add_schedule_title(ws, 'CLASS SCHEDULE', semester_text, section_display)
        
        # Add column headers
        add_column_headers(ws)
        
        # Generate and write time slots with dynamic time range
        time_slots = generate_time_slots(start_hour=start_hour, end_hour=end_hour)
        write_time_slots(ws, time_slots)
        
        # Place schedules in grid with dynamic start hour
        place_schedule_in_grid(ws, schedules, start_hour=start_hour)
        
        # Apply borders
        last_row = 11 + len(time_slots) - 1
        apply_grid_borders(ws, last_row)
        
        # Add signature section with current user's name
        sig_start_row = last_row + 3
        dean_name = current_user.full_name if current_user else None
        # Use full department name for dean title (keep original case)
        dept_display = section.department.full_department_name if (section.department and section.department.full_department_name) else dept_name
        add_signature_section(ws, sig_start_row, section.department_id, dept_display, dean_name)
        
        # Set column widths
        set_column_widths(ws)
        
        # Save and return
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        filename = f"{dept_code}_{section.year_level}{section.section_name}_Schedule.xlsx".replace(' ', '_')
        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=filename
        )
        
    except Exception as e:
        flash(f'Error exporting schedule: {str(e)}', 'error')
        return redirect(url_for('schedule.index', section_id=section_id))


@schedule_bp.route('/export/faculty/<int:faculty_id>')
@login_required
def export_faculty_schedule(faculty_id):
    """Export faculty schedule to Excel - weekly grid format matching template"""
    try:
        faculty = Faculty.query.get_or_404(faculty_id)
        current_settings = AcademicSettings.query.filter_by(is_active=True).first()
        
        # Get time range from settings (default to 7-20 if not set)
        start_hour = current_settings.schedule_start_hour if current_settings else 7
        end_hour = current_settings.schedule_end_hour if current_settings else 20
        
        # Query schedules
        query = Schedule.query.filter_by(faculty_id=faculty_id, is_active=True)
        if current_settings:
            query = query.filter_by(
                academic_year=current_settings.academic_year,
                semester=current_settings.semester
            )
        schedules = query.order_by(Schedule.day_of_week, Schedule.start_time).all()
        
        # Create workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Schedule"
        
        # Add logos
        add_institution_logos(ws)
        
        # Add header
        dept_name = faculty.department.department_name if faculty.department else 'COLLEGE'
        add_institution_header(ws, dept_name)
        
        # Add title
        semester_text = f"{current_settings.semester.upper()}, AY {current_settings.academic_year}" if current_settings else "FACULTY SCHEDULE"
        add_schedule_title(ws, 'FACULTY SCHEDULE', semester_text, faculty.full_name)
        
        # Add column headers
        add_column_headers(ws)
        
        # Generate and write time slots with dynamic time range
        time_slots = generate_time_slots(start_hour=start_hour, end_hour=end_hour)
        write_time_slots(ws, time_slots)
        
        # Place schedules in grid with dynamic start hour
        place_schedule_in_grid(ws, schedules, start_hour=start_hour)
        
        # Apply borders
        last_row = 11 + len(time_slots) - 1
        apply_grid_borders(ws, last_row)
        
        # Add signature section with current user's name
        sig_start_row = last_row + 3
        dept_id = faculty.department_id if faculty.department else None
        dean_name = current_user.full_name if current_user else None
        # Use full department name for dean title (keep original case)
        dept_display = faculty.department.full_department_name if (faculty.department and faculty.department.full_department_name) else dept_name
        add_signature_section(ws, sig_start_row, dept_id, dept_display, dean_name)
        
        # Set column widths
        set_column_widths(ws)
        
        # Save and return
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        filename = f"{faculty.full_name.replace(' ', '_')}_Schedule.xlsx"
        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=filename
        )
        
    except Exception as e:
        flash(f'Error exporting schedule: {str(e)}', 'error')
        return redirect(url_for('schedule.index', faculty_id=faculty_id))


@schedule_bp.route('/export/room/<int:room_id>')
@login_required
def export_room_schedule(room_id):
    """Export room schedule to Excel - weekly grid format matching template"""
    try:
        room = Room.query.get_or_404(room_id)
        current_settings = AcademicSettings.query.filter_by(is_active=True).first()
        
        # Get time range from settings (default to 7-20 if not set)
        start_hour = current_settings.schedule_start_hour if current_settings else 7
        end_hour = current_settings.schedule_end_hour if current_settings else 20
        
        # Query schedules
        query = Schedule.query.filter_by(room_id=room_id, is_active=True)
        if current_settings:
            query = query.filter_by(
                academic_year=current_settings.academic_year,
                semester=current_settings.semester
            )
        schedules = query.order_by(Schedule.day_of_week, Schedule.start_time).all()
        
        # Create workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Schedule"
        
        # Add logos
        add_institution_logos(ws)
        
        # Add header
        building_display = room.building.building_name if room.building else 'BUILDING'
        add_institution_header(ws, building_display)
        
        # Add title
        semester_text = f"{current_settings.semester.upper()}, AY {current_settings.academic_year}" if current_settings else "ROOM SCHEDULE"
        add_schedule_title(ws, 'ROOM SCHEDULE', semester_text, room.room_number)
        
        # Add column headers
        add_column_headers(ws)
        
        # Generate and write time slots with dynamic time range
        time_slots = generate_time_slots(start_hour=start_hour, end_hour=end_hour)
        write_time_slots(ws, time_slots)
        
        # Place schedules in grid with dynamic start hour
        place_schedule_in_grid(ws, schedules, start_hour=start_hour)
        
        # Apply borders
        last_row = 11 + len(time_slots) - 1
        apply_grid_borders(ws, last_row)
        
        # Add signature section with current user's name
        sig_start_row = last_row + 3
        dean_name = current_user.full_name if current_user else None
        # For room exports, use current user's department full name for dean title
        dept_name = None
        department_id = None
        if current_user and hasattr(current_user, 'departments') and current_user.departments:
            first_dept = current_user.departments[0]
            dept_name = first_dept.full_department_name if first_dept.full_department_name else first_dept.department_name
            department_id = first_dept.id
        add_signature_section(ws, sig_start_row, department_id, dept_name, dean_name)
        
        # Set column widths
        set_column_widths(ws)
        
        # Save and return
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        building_name = room.building.building_name if room.building else 'Building'
        filename = f"{building_name}_{room.room_number}_Schedule.xlsx".replace(' ', '_')
        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=filename
        )
        
    except Exception as e:
        flash(f'Error exporting schedule: {str(e)}', 'error')
        return redirect(url_for('schedule.index', room_id=room_id))




# ============================================================================
# Export for Posting Routes (Simplified, print-friendly versions)
# ============================================================================

@schedule_bp.route('/export/class/<int:section_id>/posting')
@login_required
def export_class_schedule_for_posting(section_id):
    """Export class schedule for posting - table format with subject details"""
    from flask import send_file
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from datetime import datetime, time as dt_time, timedelta
    import io
    
    try:
        section = Section.query.get_or_404(section_id)
        current_settings = AcademicSettings.query.filter_by(is_active=True).first()
        
        # Query schedules
        query = Schedule.query.filter_by(section_id=section_id, is_active=True)
        if current_settings:
            query = query.filter_by(
                academic_year=current_settings.academic_year,
                semester=current_settings.semester
            )
        schedules = query.order_by(Schedule.day_of_week, Schedule.start_time).all()
        
        # Create workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Schedule"
        
        # Get department info for logo and names
        department = section.department
        dept_logo_path = department.department_logo if department else None
        
        # Add logos (use posting-specific function with department logo)
        add_institution_logos_for_posting(ws, dept_logo_path)
        
        # Add header (posting-specific, centered across A-H) - use full department name (preserve original case)
        dept_display_name = department.full_department_name if (department and department.full_department_name) else (department.department_name if department else 'COLLEGE')
        add_institution_header_for_posting(ws, dept_display_name)
        
        # Store full department name for signature section
        dept_name_for_signature = dept_display_name
        
        # Add title (posting-specific, centered across A-H)
        dept_code = section.department.department_code if section.department else ''
        semester_text = f"{current_settings.semester.upper()}, AY {current_settings.academic_year}" if current_settings else "CLASS SCHEDULE"
        section_display = f"{dept_code} {section.year_level}{section.section_name}"
        add_schedule_title_for_posting(ws, 'CLASS SCHEDULE', semester_text, section_display)
        
        # Define border style
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Row 10: "Units" label merged across columns C and D
        ws['C10'] = 'Units'
        ws['C10'].font = Font(bold=True, size=11)
        ws['C10'].alignment = Alignment(horizontal='center', vertical='center')
        ws['C10'].border = thin_border
        ws.merge_cells('C10:D10')
        # Apply border to merged cells
        for col in range(3, 5):  # C and D
            ws.cell(row=10, column=col).border = thin_border
        
        # Row 11: Column Headers
        headers = ['Subject Code', 'Description', 'Lec', 'Lab', 'Day', 'Time', 'Room', 'Professor']
        for col_idx, header in enumerate(headers, start=1):
            cell = ws.cell(row=11, column=col_idx, value=header)
            cell.font = Font(bold=True, size=11)
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = thin_border
        
        # Data rows starting from row 12
        row = 12
        total_lec_units = 0
        
        for schedule in schedules:
            # Subject code with type suffix
            subject_code = schedule.subject.subject_code if schedule.subject else 'TBA'
            type_suffix = '-Lec' if schedule.schedule_type == 'lecture' else '-Lab'
            cell = ws.cell(row=row, column=1, value=f"{subject_code}{type_suffix}")
            cell.border = thin_border
            
            # Description
            description = schedule.subject.course_description if schedule.subject else ''
            cell = ws.cell(row=row, column=2, value=description)
            cell.border = thin_border
            
            # Lecture units
            lec_units = float(schedule.subject.lec_units) if schedule.subject and schedule.schedule_type == 'lecture' else 0
            cell = ws.cell(row=row, column=3, value=str(int(lec_units)))
            cell.alignment = Alignment(horizontal='center')
            cell.border = thin_border
            total_lec_units += lec_units
            
            # Lab units
            lab_units = float(schedule.subject.lab_units) if schedule.subject and schedule.schedule_type == 'lab' else 0
            cell = ws.cell(row=row, column=4, value=str(int(lab_units)))
            cell.alignment = Alignment(horizontal='center')
            cell.border = thin_border
            
            # Day
            cell = ws.cell(row=row, column=5, value=schedule.day_of_week)
            cell.border = thin_border
            
            # Time
            time_str = f"{schedule.start_time.strftime('%I:%M %p')}-{schedule.end_time.strftime('%I:%M %p')}"
            cell = ws.cell(row=row, column=6, value=time_str)
            cell.border = thin_border
            
            # Room
            room_display = schedule.room.room_number if schedule.room else 'TBA'
            cell = ws.cell(row=row, column=7, value=room_display)
            cell.border = thin_border
            
            # Professor
            faculty_name = schedule.faculty.full_name if schedule.faculty else 'TBA'
            cell = ws.cell(row=row, column=8, value=faculty_name)
            cell.border = thin_border
            
            row += 1
        
        # Add TOTAL row
        cell = ws.cell(row=row, column=1, value='TOTAL')
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='left')
        cell.border = thin_border
        
        # Empty cells in TOTAL row with borders
        for col in range(2, 3):  # Column B (Description)
            cell = ws.cell(row=row, column=col, value='')
            cell.border = thin_border
        
        cell = ws.cell(row=row, column=3, value=str(int(total_lec_units)))
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center')
        cell.border = thin_border
        
        # Empty cells after total for proper borders
        for col in range(4, 9):  # Columns D through H
            cell = ws.cell(row=row, column=col, value='')
            cell.border = thin_border
        
        # Signature section (skip 3 rows after TOTAL)
        sig_start_row = row + 3
        
        # Prepared by: (column B)
        ws.cell(row=sig_start_row, column=2, value='Prepared by:')
        ws.cell(row=sig_start_row, column=2).font = Font(size=10)
        
        # Checked by: (column F)
        ws.cell(row=sig_start_row, column=6, value='Checked by:')
        ws.cell(row=sig_start_row, column=6).font = Font(size=10)
        
        # Get secretary name and dean name from department and current user
        secretary_name = department.secretary_name if (department and department.secretary_name) else 'Name of the Secretary'
        dean_name = current_user.full_name if current_user else 'Name of the Dean'
        
        # Capitalize names
        secretary_name = secretary_name.upper()
        dean_name = dean_name.upper()
        
        # Name placeholders (2 rows down)
        ws.cell(row=sig_start_row + 2, column=2, value=secretary_name)
        ws.cell(row=sig_start_row + 2, column=2).font = Font(bold=True, size=10)
        
        ws.cell(row=sig_start_row + 2, column=6, value=dean_name)
        ws.cell(row=sig_start_row + 2, column=6).font = Font(bold=True, size=10)
        
        # Titles (next row)
        ws.cell(row=sig_start_row + 3, column=2, value="Dean's Secretary")
        ws.cell(row=sig_start_row + 3, column=2).font = Font(size=10)
        
        # Use full department name for dean title
        dean_title = f"Dean, {dept_name_for_signature}"
        ws.cell(row=sig_start_row + 3, column=6, value=dean_title)
        ws.cell(row=sig_start_row + 3, column=6).font = Font(size=10)
        ws.cell(row=sig_start_row + 3, column=6).alignment = Alignment(horizontal='center')
        
        # Set column widths for table format
        ws.column_dimensions['A'].width = 15  # Subject Code
        ws.column_dimensions['B'].width = 35  # Description
        ws.column_dimensions['C'].width = 8   # Lec
        ws.column_dimensions['D'].width = 8   # Lab
        ws.column_dimensions['E'].width = 12  # Day
        ws.column_dimensions['F'].width = 18  # Time
        ws.column_dimensions['G'].width = 12  # Room
        ws.column_dimensions['H'].width = 25  # Professor
        
        # Save to BytesIO
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        filename = f"{dept_code}_{section.year_level}{section.section_name}_Schedule_Posting.xlsx".replace(' ', '_')
        
        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=filename
        )
        
    except Exception as e:
        flash(f'Error exporting schedule for posting: {str(e)}', 'error')
        return redirect(url_for('schedule.index', section_id=section_id))


@schedule_bp.route('/export/faculty/<int:faculty_id>/posting')
@login_required
def export_faculty_schedule_for_posting(faculty_id):
    """Export faculty schedule for posting - table format with subject details"""
    from flask import send_file
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from datetime import datetime, time as dt_time, timedelta
    import io
    
    try:
        faculty = Faculty.query.get_or_404(faculty_id)
        
        # Get current academic settings
        current_settings = AcademicSettings.query.filter_by(is_active=True).first()
        
        # Query schedules for this faculty
        query = Schedule.query.filter_by(faculty_id=faculty_id, is_active=True)
        if current_settings:
            query = query.filter_by(
                academic_year=current_settings.academic_year,
                semester=current_settings.semester
            )
        schedules = query.order_by(Schedule.day_of_week, Schedule.start_time).all()
        
        # Create workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Schedule"
        
        # Get department info for logo and names
        department = faculty.department
        dept_logo_path = department.department_logo if department else None
        
        # Add logos (use posting-specific function with department logo)
        add_institution_logos_for_posting(ws, dept_logo_path)
        
        # Add header (posting-specific, centered across A-H) - use full department name (preserve original case)
        dept_display_name = department.full_department_name if (department and department.full_department_name) else (department.department_name if department else 'COLLEGE')
        add_institution_header_for_posting(ws, dept_display_name)
        
        # Store full department name for signature section
        dept_name_for_signature = dept_display_name
        
        # Add title (posting-specific, centered across A-H)
        semester_text = f"{current_settings.semester.upper()}, AY {current_settings.academic_year}" if current_settings else "FACULTY SCHEDULE"
        add_schedule_title_for_posting(ws, 'FACULTY SCHEDULE', semester_text, f"Faculty: {faculty.full_name}")
        
        # Define border style
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Row 10: "Units" label merged across columns C and D
        ws['C10'] = 'Units'
        ws['C10'].font = Font(bold=True, size=11)
        ws['C10'].alignment = Alignment(horizontal='center', vertical='center')
        ws['C10'].border = thin_border
        ws.merge_cells('C10:D10')
        # Apply border to merged cells
        for col in range(3, 5):  # C and D
            ws.cell(row=10, column=col).border = thin_border
        
        # Row 11: Column Headers
        headers = ['Subject Code', 'Description', 'Lec', 'Lab', 'Day', 'Time', 'Room', 'Section']
        for col_idx, header in enumerate(headers, start=1):
            cell = ws.cell(row=11, column=col_idx, value=header)
            cell.font = Font(bold=True, size=11)
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = thin_border
        
        # Data rows starting from row 12
        row = 12
        total_lec_units = 0
        
        for schedule in schedules:
            # Subject code with type suffix
            subject_code = schedule.subject.subject_code if schedule.subject else 'TBA'
            type_suffix = '-Lec' if schedule.schedule_type == 'lecture' else '-Lab'
            cell = ws.cell(row=row, column=1, value=f"{subject_code}{type_suffix}")
            cell.border = thin_border
            
            # Description
            description = schedule.subject.course_description if schedule.subject else ''
            cell = ws.cell(row=row, column=2, value=description)
            cell.border = thin_border
            
            # Lecture units
            lec_units = float(schedule.subject.lec_units) if schedule.subject and schedule.schedule_type == 'lecture' else 0
            cell = ws.cell(row=row, column=3, value=str(int(lec_units)))
            cell.alignment = Alignment(horizontal='center')
            cell.border = thin_border
            total_lec_units += lec_units
            
            # Lab units
            lab_units = float(schedule.subject.lab_units) if schedule.subject and schedule.schedule_type == 'lab' else 0
            cell = ws.cell(row=row, column=4, value=str(int(lab_units)))
            cell.alignment = Alignment(horizontal='center')
            cell.border = thin_border
            
            # Day
            cell = ws.cell(row=row, column=5, value=schedule.day_of_week)
            cell.border = thin_border
            
            # Time
            time_str = f"{schedule.start_time.strftime('%I:%M %p')}-{schedule.end_time.strftime('%I:%M %p')}"
            cell = ws.cell(row=row, column=6, value=time_str)
            cell.border = thin_border
            
            # Room
            room_display = schedule.room.room_number if schedule.room else 'TBA'
            cell = ws.cell(row=row, column=7, value=room_display)
            cell.border = thin_border
            
            # Section
            section_name = schedule.section.section_name if schedule.section else 'TBA'
            cell = ws.cell(row=row, column=8, value=section_name)
            cell.border = thin_border
            
            row += 1
        
        # Add TOTAL row
        cell = ws.cell(row=row, column=1, value='TOTAL')
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='left')
        cell.border = thin_border
        
        # Empty cells in TOTAL row with borders
        for col in range(2, 3):  # Column B (Description)
            cell = ws.cell(row=row, column=col, value='')
            cell.border = thin_border
        
        cell = ws.cell(row=row, column=3, value=str(int(total_lec_units)))
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center')
        cell.border = thin_border
        
        # Empty cells after total for proper borders
        for col in range(4, 9):  # Columns D through H
            cell = ws.cell(row=row, column=col, value='')
            cell.border = thin_border
        
        # Signature section (skip 3 rows after TOTAL)
        sig_start_row = row + 3
        
        # Prepared by: (column B)
        ws.cell(row=sig_start_row, column=2, value='Prepared by:')
        ws.cell(row=sig_start_row, column=2).font = Font(size=10)
        
        # Checked by: (column F)
        ws.cell(row=sig_start_row, column=6, value='Checked by:')
        ws.cell(row=sig_start_row, column=6).font = Font(size=10)
        
        # Get secretary name and dean name from department and current user
        secretary_name = department.secretary_name if (department and department.secretary_name) else 'Name of the Secretary'
        dean_name = current_user.full_name if current_user else 'Name of the Dean'
        
        # Capitalize names
        secretary_name = secretary_name.upper()
        dean_name = dean_name.upper()
        
        # Name placeholders (2 rows down)
        ws.cell(row=sig_start_row + 2, column=2, value=secretary_name)
        ws.cell(row=sig_start_row + 2, column=2).font = Font(bold=True, size=10)
        
        ws.cell(row=sig_start_row + 2, column=6, value=dean_name)
        ws.cell(row=sig_start_row + 2, column=6).font = Font(bold=True, size=10)
        
        # Titles (next row)
        ws.cell(row=sig_start_row + 3, column=2, value="Dean's Secretary")
        ws.cell(row=sig_start_row + 3, column=2).font = Font(size=10)
        
        # Use full department name for dean title
        dean_title = f"Dean, {dept_name_for_signature}"
        ws.cell(row=sig_start_row + 3, column=6, value=dean_title)
        ws.cell(row=sig_start_row + 3, column=6).font = Font(size=10)
        ws.cell(row=sig_start_row + 3, column=6).alignment = Alignment(horizontal='center')
        
        # Set column widths for table format
        ws.column_dimensions['A'].width = 15  # Subject Code
        ws.column_dimensions['B'].width = 35  # Description
        ws.column_dimensions['C'].width = 8   # Lec
        ws.column_dimensions['D'].width = 8   # Lab
        ws.column_dimensions['E'].width = 12  # Day
        ws.column_dimensions['F'].width = 18  # Time
        ws.column_dimensions['G'].width = 12  # Room
        ws.column_dimensions['H'].width = 25  # Section
        
        # Save to BytesIO
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        filename = f"{faculty.full_name.replace(' ', '_')}_Schedule_Posting.xlsx"
        
        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=filename
        )
        
    except Exception as e:
        flash(f'Error exporting faculty schedule for posting: {str(e)}', 'error')
        return redirect(url_for('schedule.index', faculty_id=faculty_id))


@schedule_bp.route('/export/room/<int:room_id>/posting')
@login_required
def export_room_schedule_for_posting(room_id):
    """Export room schedule for posting - table format with subject details"""
    from flask import send_file
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from datetime import datetime, time as dt_time, timedelta
    import io
    
    try:
        room = Room.query.get_or_404(room_id)
        
        # Get current academic settings
        current_settings = AcademicSettings.query.filter_by(is_active=True).first()
        
        # Query schedules for this room
        query = Schedule.query.filter_by(room_id=room_id, is_active=True)
        if current_settings:
            query = query.filter_by(
                academic_year=current_settings.academic_year,
                semester=current_settings.semester
            )
        schedules = query.order_by(Schedule.day_of_week, Schedule.start_time).all()
        
        # Get department from the first schedule's section (if available)
        department = None
        if schedules and schedules[0].section:
            department = schedules[0].section.department
        
        # Create workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Schedule"
        
        # Get department logo
        dept_logo_path = department.department_logo if department else None
        
        # Add logos (use posting-specific function with department logo)
        add_institution_logos_for_posting(ws, dept_logo_path)
        
        # Add header (posting-specific, centered across A-H) - use full department name (preserve original case)
        if department:
            dept_display_name = department.full_department_name if department.full_department_name else department.department_name
        else:
            dept_display_name = 'COLLEGE'
        add_institution_header_for_posting(ws, dept_display_name)
        
        # Store full department name for signature section
        dept_name_for_signature = dept_display_name
        
        # Add title (posting-specific, centered across A-H)
        semester_text = f"{current_settings.semester.upper()}, AY {current_settings.academic_year}" if current_settings else "ROOM SCHEDULE"
        add_schedule_title_for_posting(ws, 'ROOM SCHEDULE', semester_text, f"Room: {room.room_number}")
        
        # Define border style
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Row 10: "Units" label merged across columns C and D
        ws['C10'] = 'Units'
        ws['C10'].font = Font(bold=True, size=11)
        ws['C10'].alignment = Alignment(horizontal='center', vertical='center')
        ws['C10'].border = thin_border
        ws.merge_cells('C10:D10')
        # Apply border to merged cells
        for col in range(3, 5):  # C and D
            ws.cell(row=10, column=col).border = thin_border
        
        # Row 11: Column Headers
        headers = ['Subject Code', 'Description', 'Lec', 'Lab', 'Day', 'Time', 'Section', 'Faculty']
        for col_idx, header in enumerate(headers, start=1):
            cell = ws.cell(row=11, column=col_idx, value=header)
            cell.font = Font(bold=True, size=11)
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = thin_border
        
        # Data rows starting from row 12
        row = 12
        total_lec_units = 0
        
        for schedule in schedules:
            # Subject code with type suffix
            subject_code = schedule.subject.subject_code if schedule.subject else 'TBA'
            type_suffix = '-Lec' if schedule.schedule_type == 'lecture' else '-Lab'
            cell = ws.cell(row=row, column=1, value=f"{subject_code}{type_suffix}")
            cell.border = thin_border
            
            # Description
            description = schedule.subject.course_description if schedule.subject else ''
            cell = ws.cell(row=row, column=2, value=description)
            cell.border = thin_border
            
            # Lecture units
            lec_units = float(schedule.subject.lec_units) if schedule.subject and schedule.schedule_type == 'lecture' else 0
            cell = ws.cell(row=row, column=3, value=str(int(lec_units)))
            cell.alignment = Alignment(horizontal='center')
            cell.border = thin_border
            total_lec_units += lec_units
            
            # Lab units
            lab_units = float(schedule.subject.lab_units) if schedule.subject and schedule.schedule_type == 'lab' else 0
            cell = ws.cell(row=row, column=4, value=str(int(lab_units)))
            cell.alignment = Alignment(horizontal='center')
            cell.border = thin_border
            
            # Day
            cell = ws.cell(row=row, column=5, value=schedule.day_of_week)
            cell.border = thin_border
            
            # Time
            time_str = f"{schedule.start_time.strftime('%I:%M %p')}-{schedule.end_time.strftime('%I:%M %p')}"
            cell = ws.cell(row=row, column=6, value=time_str)
            cell.border = thin_border
            
            # Section
            section_name = schedule.section.section_name if schedule.section else 'TBA'
            cell = ws.cell(row=row, column=7, value=section_name)
            cell.border = thin_border
            
            # Faculty
            faculty_name = schedule.faculty.full_name if schedule.faculty else 'TBA'
            cell = ws.cell(row=row, column=8, value=faculty_name)
            cell.border = thin_border
            
            row += 1
        
        # Add TOTAL row
        cell = ws.cell(row=row, column=1, value='TOTAL')
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='left')
        cell.border = thin_border
        
        # Empty cells in TOTAL row with borders
        for col in range(2, 3):  # Column B (Description)
            cell = ws.cell(row=row, column=col, value='')
            cell.border = thin_border
        
        cell = ws.cell(row=row, column=3, value=str(int(total_lec_units)))
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center')
        cell.border = thin_border
        
        # Empty cells after total for proper borders
        for col in range(4, 9):  # Columns D through H
            cell = ws.cell(row=row, column=col, value='')
            cell.border = thin_border
        
        # Signature section (skip 3 rows after TOTAL)
        sig_start_row = row + 3
        
        # Prepared by: (column B)
        ws.cell(row=sig_start_row, column=2, value='Prepared by:')
        ws.cell(row=sig_start_row, column=2).font = Font(size=10)
        
        # Checked by: (column F)
        ws.cell(row=sig_start_row, column=6, value='Checked by:')
        ws.cell(row=sig_start_row, column=6).font = Font(size=10)
        
        # Get secretary name and dean name from department and current user
        secretary_name = department.secretary_name if (department and department.secretary_name) else 'Name of the Secretary'
        dean_name = current_user.full_name if current_user else 'Name of the Dean'
        
        # Capitalize names
        secretary_name = secretary_name.upper()
        dean_name = dean_name.upper()
        
        # Name placeholders (2 rows down)
        ws.cell(row=sig_start_row + 2, column=2, value=secretary_name)
        ws.cell(row=sig_start_row + 2, column=2).font = Font(bold=True, size=10)
        
        ws.cell(row=sig_start_row + 2, column=6, value=dean_name)
        ws.cell(row=sig_start_row + 2, column=6).font = Font(bold=True, size=10)
        
        # Titles (next row)
        ws.cell(row=sig_start_row + 3, column=2, value="Dean's Secretary")
        ws.cell(row=sig_start_row + 3, column=2).font = Font(size=10)
        
        # Use full department name for dean title
        dean_title = f"Dean, {dept_name_for_signature}"
        ws.cell(row=sig_start_row + 3, column=6, value=dean_title)
        ws.cell(row=sig_start_row + 3, column=6).font = Font(size=10)
        ws.cell(row=sig_start_row + 3, column=6).alignment = Alignment(horizontal='center')
        
        # Set column widths for table format
        ws.column_dimensions['A'].width = 15  # Subject Code
        ws.column_dimensions['B'].width = 35  # Description
        ws.column_dimensions['C'].width = 8   # Lec
        ws.column_dimensions['D'].width = 8   # Lab
        ws.column_dimensions['E'].width = 12  # Day
        ws.column_dimensions['F'].width = 18  # Time
        ws.column_dimensions['G'].width = 25  # Section
        ws.column_dimensions['H'].width = 25  # Faculty
        
        # Save to BytesIO
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        building_name = room.building.building_name if room.building else 'Building'
        filename = f"{building_name}_{room.room_number}_Schedule_Posting.xlsx".replace(' ', '_')
        
        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=filename
        )
        
    except Exception as e:
        flash(f'Error exporting room schedule for posting: {str(e)}', 'error')
        return redirect(url_for('schedule.index', room_id=room_id))


# ============================================================================
# PDF EXPORT ROUTES
# ============================================================================

@schedule_bp.route('/export/class/<int:section_id>/pdf')
@login_required
def export_class_schedule_pdf(section_id):
    """Export class schedule to PDF - weekly grid format matching Excel template"""
    try:
        section = Section.query.get_or_404(section_id)
        current_settings = AcademicSettings.query.filter_by(is_active=True).first()
        
        # Get time range from settings (default to 7-20 if not set)
        start_hour = current_settings.schedule_start_hour if current_settings else 7
        end_hour = current_settings.schedule_end_hour if current_settings else 20
        
        # Query schedules
        query = Schedule.query.filter_by(section_id=section_id, is_active=True)
        if current_settings:
            query = query.filter_by(
                academic_year=current_settings.academic_year,
                semester=current_settings.semester
            )
        schedules = query.order_by(Schedule.day_of_week, Schedule.start_time).all()
        
        # Prepare metadata
        dept_name = section.department.department_name if section.department else 'COLLEGE'
        dept_code = section.department.department_code if section.department else ''
        semester_text = f"{current_settings.semester.upper()}, AY {current_settings.academic_year}" if current_settings else "CLASS SCHEDULE"
        section_display = f"{dept_code} {section.year_level}{section.section_name}"
        
        # Create PDF with dynamic time range
        output = create_pdf_schedule(
            schedules=schedules,
            title='CLASS SCHEDULE',
            semester_text=semester_text,
            section_display=section_display,
            dept_name=dept_name.upper(),
            filename=f"{dept_code}_{section.year_level}{section.section_name}_Schedule.pdf",
            start_hour=start_hour,
            end_hour=end_hour
        )
        
        filename = f"{dept_code}_{section.year_level}{section.section_name}_Schedule.pdf".replace(' ', '_')
        return send_file(
            output,
            mimetype='application/pdf',
            as_attachment=True,
            download_name=filename
        )
        
    except Exception as e:
        flash(f'Error exporting PDF schedule: {str(e)}', 'error')
        return redirect(url_for('schedule.index', section_id=section_id))


@schedule_bp.route('/export/faculty/<int:faculty_id>/pdf')
@login_required
def export_faculty_schedule_pdf(faculty_id):
    """Export faculty schedule to PDF - weekly grid format"""
    try:
        faculty = Faculty.query.get_or_404(faculty_id)
        current_settings = AcademicSettings.query.filter_by(is_active=True).first()
        
        # Get time range from settings (default to 7-20 if not set)
        start_hour = current_settings.schedule_start_hour if current_settings else 7
        end_hour = current_settings.schedule_end_hour if current_settings else 20
        
        # Query schedules
        query = Schedule.query.filter_by(faculty_id=faculty_id, is_active=True)
        if current_settings:
            query = query.filter_by(
                academic_year=current_settings.academic_year,
                semester=current_settings.semester
            )
        schedules = query.order_by(Schedule.day_of_week, Schedule.start_time).all()
        
        # Prepare metadata
        dept_name = faculty.department.department_name if faculty.department else 'COLLEGE'
        semester_text = f"{current_settings.semester.upper()}, AY {current_settings.academic_year}" if current_settings else "FACULTY SCHEDULE"
        faculty_display = f"{faculty.full_name}"
        
        # Create PDF with dynamic time range
        output = create_pdf_schedule(
            schedules=schedules,
            title='FACULTY SCHEDULE',
            semester_text=semester_text,
            section_display=faculty_display,
            dept_name=dept_name.upper(),
            filename=f"{faculty.full_name.replace(' ', '_')}_Schedule.pdf",
            start_hour=start_hour,
            end_hour=end_hour
        )
        
        filename = f"{faculty.full_name.replace(' ', '_')}_Schedule.pdf"
        return send_file(
            output,
            mimetype='application/pdf',
            as_attachment=True,
            download_name=filename
        )
        
    except Exception as e:
        flash(f'Error exporting faculty PDF schedule: {str(e)}', 'error')
        return redirect(url_for('schedule.index', faculty_id=faculty_id))


@schedule_bp.route('/export/room/<int:room_id>/pdf')
@login_required
def export_room_schedule_pdf(room_id):
    """Export room schedule to PDF - weekly grid format"""
    try:
        room = Room.query.get_or_404(room_id)
        current_settings = AcademicSettings.query.filter_by(is_active=True).first()
        
        # Get time range from settings (default to 7-20 if not set)
        start_hour = current_settings.schedule_start_hour if current_settings else 7
        end_hour = current_settings.schedule_end_hour if current_settings else 20
        
        # Query schedules
        query = Schedule.query.filter_by(room_id=room_id, is_active=True)
        if current_settings:
            query = query.filter_by(
                academic_year=current_settings.academic_year,
                semester=current_settings.semester
            )
        schedules = query.order_by(Schedule.day_of_week, Schedule.start_time).all()
        
        # Prepare metadata
        building_name = room.building.building_name if room.building else 'Building'
        dept_name = 'COLLEGE'  # Room schedules are typically college-wide
        semester_text = f"{current_settings.semester.upper()}, AY {current_settings.academic_year}" if current_settings else "ROOM SCHEDULE"
        room_display = f"{building_name} - Room {room.room_number}"
        
        # Create PDF with dynamic time range
        output = create_pdf_schedule(
            schedules=schedules,
            title='ROOM SCHEDULE',
            semester_text=semester_text,
            section_display=room_display,
            dept_name=dept_name,
            filename=f"Room_{room.room_number}_Schedule.pdf",
            start_hour=start_hour,
            end_hour=end_hour
        )
        
        filename = f"Room_{room.room_number}_Schedule.pdf".replace(' ', '_')
        return send_file(
            output,
            mimetype='application/pdf',
            as_attachment=True,
            download_name=filename
        )
        
    except Exception as e:
        flash(f'Error exporting room PDF schedule: {str(e)}', 'error')
        return redirect(url_for('schedule.index', room_id=room_id))




# ============================================================================
# END OF SCHEDULE ROUTES
# ============================================================================