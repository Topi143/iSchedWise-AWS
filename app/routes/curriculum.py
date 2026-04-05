"""
Curriculum management routes
"""
from flask import Blueprint, render_template, request, redirect, url_for, flash, session, jsonify, send_file
from flask_login import login_required, current_user
from app.decorators import role_required
from app.extensions import db
from app.models import Program, Curriculum, YearLevel, Semester, Subject
from app.models.schedule import Schedule
from app.models.exam_schedule import ExamSchedule
from app.models.settings import AcademicSettings
from app.utils.activity_logger import log_create, log_edit, log_delete, log_archive, log_unarchive
import pandas as pd
import io
from werkzeug.utils import secure_filename
import os

curriculum_bp = Blueprint('curriculum', __name__, url_prefix='/curriculum')


def get_available_semesters_from_settings():
    """Return available semester names from active academic settings."""
    active_settings = AcademicSettings.query.filter_by(is_active=True).first()
    if active_settings and active_settings.available_semesters:
        semesters = [s.strip() for s in active_settings.available_semesters.split(',') if s.strip()]
        if semesters:
            return semesters
    return ['1st Semester', '2nd Semester']


def get_default_semester_definitions():
    """Map available semester names to canonical semester number/name pairs."""
    canonical_semesters = [
        (1, '1st Semester'),
        (2, '2nd Semester'),
        (3, 'Summer')
    ]
    available_set = set(get_available_semesters_from_settings())
    defaults = [(num, name) for num, name in canonical_semesters if name in available_set]
    if defaults:
        return defaults
    return canonical_semesters[:2]


def get_filter_params():
    """Get program_id filter from request args"""
    program_id = request.args.get('program_id', type=int)
    return {'program_id': program_id} if program_id else {}


def build_redirect_params(curriculum_id=None, year_level_id=None, semester_id=None, **kwargs):
    """Build redirect parameters preserving filters"""
    params = get_filter_params()
    
    if curriculum_id:
        params['curriculum_id'] = curriculum_id
        params['open'] = curriculum_id
    if year_level_id:
        params['year'] = year_level_id
    if semester_id:
        params['semester'] = semester_id
    
    # Add any additional parameters
    params.update(kwargs)
    return params


def get_open_curriculum_id():
    """Get the curriculum ID that should remain open from request or session"""
    # Check if there's an open parameter in the form
    open_id = request.form.get('_open_curriculum_id')
    if not open_id:
        # Check session
        open_id = session.get('open_curriculum_id')
    return open_id


def redirect_with_open(curriculum_id=None):
    """Redirect to curriculum index, preserving the open accordion state and filters"""
    # Preserve program_id filter from request
    program_id = request.args.get('program_id', type=int)
    
    url_params = {}
    if program_id:
        url_params['program_id'] = program_id
    
    if curriculum_id:
        session['open_curriculum_id'] = curriculum_id
        url_params['open'] = curriculum_id
        return redirect(url_for('curriculum.index', **url_params))
    else:
        # Try to get from session or request
        open_id = get_open_curriculum_id()
        if open_id:
            url_params['open'] = open_id
            return redirect(url_for('curriculum.index', **url_params))
    return redirect(url_for('curriculum.index', **url_params))


@curriculum_bp.route('/')
@login_required
def index():
    """Curriculum management page"""
    # Get filter parameters
    program_id = request.args.get('program_id', type=int)
    selected_curriculum_id = request.args.get('curriculum_id', type=int)
    
    # Get user's program access
    user_program_ids = current_user.get_program_ids()
    
    # Filter programs by user access
    if user_program_ids is None:
        programs = Program.query.filter_by(is_active=True).order_by(Program.program_name).all()
    else:
        programs = Program.query.filter(
            Program.is_active == True,
            Program.id.in_(user_program_ids)
        ).order_by(Program.program_name).all()
    
    # Auto-select program if user has only 1 program and no filter is set
    if not program_id and user_program_ids is not None and len(programs) == 1:
        program_id = programs[0].id
    
    # Build query based on filters and user access
    query = Curriculum.query
    
    # Filter out archived curricula from main list
    query = query.filter_by(is_archived=False)
    
    # Filter by user's program access
    if user_program_ids is not None:
        query = query.filter(Curriculum.program_id.in_(user_program_ids))
    
    # Apply additional program filter if specified
    if program_id:
        query = query.filter_by(program_id=program_id)
    
    # Order by curriculum_code alphabetically so same programs are grouped together
    curricula = query.order_by(Curriculum.curriculum_code.asc()).all()
    
    # Filter programs by user access
    if user_program_ids is None:
        programs = Program.query.filter_by(is_active=True).order_by(Program.program_name).all()
    else:
        programs = Program.query.filter(
            Program.is_active == True,
            Program.id.in_(user_program_ids)
        ).order_by(Program.program_name).all()
    
    # If a curriculum is selected, find it
    selected_curriculum = None
    if selected_curriculum_id and curricula:
        selected_curriculum = Curriculum.query.get(selected_curriculum_id)
    
    # Get available semesters from active settings
    available_semesters = get_available_semesters_from_settings()
    
    return render_template('curriculum.html', 
                         user=current_user, 
                         curricula=curricula, 
                         programs=programs,
                         selected_department_id=program_id,
                         selected_curriculum_id=selected_curriculum_id,
                         selected_curriculum=selected_curriculum,
                         available_semesters=available_semesters)


@curriculum_bp.route('/add', methods=['POST'])
@login_required
@role_required('admin', 'super_admin')
def add():
    """Add a new curriculum"""
    try:
        curriculum_code = request.form.get('curriculum_code', '').strip().upper()
        program_id = request.form.get('program_id', '').strip()
        year_levels_count = request.form.get('year_levels', '').strip()
        
        if not all([curriculum_code, program_id, year_levels_count]):
            flash('Please fill in all required fields.', 'error')
            return redirect(url_for('curriculum.index'))
        
        try:
            year_levels_count = int(year_levels_count)
            if year_levels_count < 1 or year_levels_count > 10:
                flash('Number of year levels must be between 1 and 10.', 'error')
                return redirect(url_for('curriculum.index'))
        except ValueError:
            flash('Invalid number of year levels.', 'error')
            return redirect(url_for('curriculum.index'))
        
        try:
            program_id = int(program_id)
        except ValueError:
            flash('Invalid program selected.', 'error')
            return redirect(url_for('curriculum.index'))
        
        if Curriculum.query.filter_by(curriculum_code=curriculum_code).first():
            flash(f'Curriculum code "{curriculum_code}" already exists. Please use a different code.', 'error')
            return redirect(url_for('curriculum.index'))
        
        program = Program.query.get(program_id)
        if not program:
            flash('Selected program not found.', 'error')
            return redirect(url_for('curriculum.index'))
        
        # Auto-generate degree program from program name
        # Use the program's full name to avoid redundancy (e.g. "BEED - BEED 2025-2026")
        degree_program = program.program_name or program.program_code
        
        new_curriculum = Curriculum(
            curriculum_code=curriculum_code,
            curriculum_name=degree_program,
            program_id=program.id,
            degree_program=degree_program,
            is_active=True,
            created_by=current_user.id
        )
        
        db.session.add(new_curriculum)
        db.session.flush()
        
        # Log activity
        log_create('curriculum', new_curriculum.id, new_curriculum.curriculum_code, {
            'program': program.program_code,
            'year_levels': year_levels_count
        })
        
        # Auto-create year levels with semesters
        year_names = ['1st Year', '2nd Year', '3rd Year', '4th Year', '5th Year', 
                      '6th Year', '7th Year', '8th Year', '9th Year', '10th Year']
        default_semesters = get_default_semester_definitions()
        
        for i in range(year_levels_count):
            year_level = YearLevel(
                curriculum_id=new_curriculum.id,
                year_number=i + 1,
                year_name=year_names[i]
            )
            db.session.add(year_level)
            db.session.flush()  # Flush to get the year_level.id
            
            # Create default semesters based on current academic settings.
            for sem_num, sem_name in default_semesters:
                semester = Semester(
                    year_level_id=year_level.id,
                    semester_number=sem_num,
                    semester_name=sem_name
                )
                db.session.add(semester)
        
        db.session.commit()
        
        flash('Curriculum has been successfully added!', 'success')
        
        # Preserve program filter in redirect
        program_id = request.form.get('program_id', type=int)
        url_params = {'curriculum_id': new_curriculum.id, 'open': new_curriculum.id}
        if program_id:
            url_params['program_id'] = program_id
        return redirect(url_for('curriculum.index', **url_params))
        
    except Exception as e:
        db.session.rollback()
        flash(f'An error occurred while adding the curriculum: {str(e)}', 'error')
        return redirect(url_for('curriculum.index'))


@curriculum_bp.route('/edit', methods=['POST'])
@login_required
def edit():
    """Edit an existing curriculum"""
    try:
        curriculum_id = request.form.get('curriculum_id', '').strip()
        curriculum_code = request.form.get('curriculum_code', '').strip().upper()
        program_id = request.form.get('program_id', '').strip()
        year_levels_count = request.form.get('year_levels', '').strip()
        
        if not all([curriculum_id, curriculum_code, program_id, year_levels_count]):
            flash('Please fill in all required fields.', 'error')
            return redirect(url_for('curriculum.index'))
        
        try:
            year_levels_count = int(year_levels_count)
            if year_levels_count < 1 or year_levels_count > 10:
                flash('Number of year levels must be between 1 and 10.', 'error')
                return redirect(url_for('curriculum.index'))
        except ValueError:
            flash('Invalid number of year levels.', 'error')
            return redirect(url_for('curriculum.index'))
        
        curriculum = Curriculum.query.get(int(curriculum_id))
        if not curriculum:
            flash('Curriculum not found.', 'error')
            return redirect(url_for('curriculum.index'))
        
        if curriculum.curriculum_code != curriculum_code:
            if Curriculum.query.filter_by(curriculum_code=curriculum_code).first():
                flash(f'Curriculum code "{curriculum_code}" already exists. Please use a different code.', 'error')
                return redirect(url_for('curriculum.index'))
        
        program = Program.query.get(int(program_id))
        if not program:
            flash('Selected program not found.', 'error')
            return redirect(url_for('curriculum.index'))
        
        # Auto-generate degree program from program name
        # Use the program's full name to avoid redundancy (e.g. "BEED - BEED 2025-2026")
        degree_program = program.program_name or program.program_code
        
        curriculum.curriculum_code = curriculum_code
        curriculum.curriculum_name = degree_program
        curriculum.program_id = int(program_id)
        curriculum.degree_program = degree_program
        
        # Handle year levels - add or remove as needed
        current_year_levels = len(curriculum.year_levels)
        year_names = ['1st Year', '2nd Year', '3rd Year', '4th Year', '5th Year', '6th Year', 
                      '7th Year', '8th Year', '9th Year', '10th Year']
        default_semesters = get_default_semester_definitions()
        
        if year_levels_count > current_year_levels:
            # Add new year levels with semesters
            for i in range(current_year_levels, year_levels_count):
                year_level = YearLevel(
                    curriculum_id=curriculum.id,
                    year_number=i + 1,
                    year_name=year_names[i]
                )
                db.session.add(year_level)
                db.session.flush()  # Flush to get the year_level.id
                
                # Create default semesters for newly added year levels based on settings.
                for sem_num, sem_name in default_semesters:
                    semester = Semester(
                        year_level_id=year_level.id,
                        semester_number=sem_num,
                        semester_name=sem_name
                    )
                    db.session.add(semester)
        elif year_levels_count < current_year_levels:
            # Remove excess year levels (from the end)
            for i in range(year_levels_count, current_year_levels):
                year_level = curriculum.year_levels[i]
                db.session.delete(year_level)
        
        # Existing year levels keep their current semesters unchanged.
        
        # Log activity
        log_edit('curriculum', curriculum.id, curriculum.curriculum_code, {
            'program': curriculum.program.program_code
        })
        
        db.session.commit()
        
        flash('Curriculum has been successfully updated!', 'success')
        
        # Preserve program filter in redirect
        program_id = request.args.get('program_id', type=int)
        url_params = {'curriculum_id': curriculum.id, 'open': curriculum.id}
        if program_id:
            url_params['program_id'] = program_id
        return redirect(url_for('curriculum.index', **url_params))
        
    except Exception as e:
        db.session.rollback()
        flash(f'An error occurred while updating the curriculum: {str(e)}', 'error')
        return redirect(url_for('curriculum.index'))


# Year Level routes
@curriculum_bp.route('/year-level/edit', methods=['POST'])
@login_required
def edit_year_level():
    """Edit a year level"""
    try:
        year_level_id = request.form.get('year_level_id', '').strip()
        year_name = request.form.get('year_name', '').strip()
        
        if not all([year_level_id, year_name]):
            flash('Please fill in all required fields.', 'error')
            return redirect(url_for('curriculum.index'))
        
        year_level = YearLevel.query.get(int(year_level_id))
        if not year_level:
            flash('Year level not found.', 'error')
            return redirect(url_for('curriculum.index'))
        
        year_level.year_name = year_name
        db.session.commit()
        
        flash('Year level has been successfully updated!', 'success')
        params = build_redirect_params(
            curriculum_id=year_level.curriculum_id, 
            year_level_id=year_level.id
        )
        return redirect(url_for('curriculum.index', **params))
        
    except Exception as e:
        db.session.rollback()
        flash(f'An error occurred while updating the year level: {str(e)}', 'error')
        return redirect(url_for('curriculum.index'))


@curriculum_bp.route('/year-level/delete', methods=['POST'])
@login_required
def delete_year_level():
    """Delete a year level"""
    try:
        year_level_id = request.form.get('year_level_id', '').strip()
        
        if not year_level_id:
            flash('Invalid year level.', 'error')
            return redirect(url_for('curriculum.index'))
        
        year_level = YearLevel.query.get(int(year_level_id))
        if not year_level:
            flash('Year level not found.', 'error')
            return redirect(url_for('curriculum.index'))
        
        curriculum_id = year_level.curriculum_id
        db.session.delete(year_level)
        db.session.commit()
        
        flash('Year level has been successfully deleted!', 'success')
        params = build_redirect_params(curriculum_id=curriculum_id)
        return redirect(url_for('curriculum.index', **params))
        
    except Exception as e:
        db.session.rollback()
        flash(f'An error occurred while deleting the year level: {str(e)}', 'error')
        return redirect(url_for('curriculum.index'))


# Semester routes
@curriculum_bp.route('/semester/add', methods=['POST'])
@login_required
def add_semester():
    """Add semester(s) to a year level"""
    try:
        year_level_id = request.form.get('year_level_id', '').strip()
        semester_number = request.form.get('semester_number', '').strip()
        
        if not all([year_level_id, semester_number]):
            flash('Please fill in all required fields.', 'error')
            return redirect(url_for('curriculum.index'))
        
        try:
            semester_number = int(semester_number)
            if semester_number < 1 or semester_number > 3:
                flash('Semester number must be between 1 and 3.', 'error')
                return redirect(url_for('curriculum.index'))
        except ValueError:
            flash('Invalid semester number.', 'error')
            return redirect(url_for('curriculum.index'))
        
        try:
            year_level_id = int(year_level_id)
        except ValueError:
            flash('Invalid year level.', 'error')
            return redirect(url_for('curriculum.index'))
        
        year_level = YearLevel.query.get(year_level_id)
        if not year_level:
            flash('Year level not found.', 'error')
            return redirect(url_for('curriculum.index'))
        
        semester_names = {
            1: '1st Semester',
            2: '2nd Semester'
        }
        
        semesters_created = 0
        for i in range(1, semester_number + 1):
            existing_semester = Semester.query.filter_by(
                year_level_id=year_level_id,
                semester_number=i
            ).first()
            
            if not existing_semester:
                new_semester = Semester(
                    year_level_id=year_level_id,
                    semester_number=i,
                    semester_name=semester_names[i]
                )
                db.session.add(new_semester)
                semesters_created += 1
        
        if semesters_created > 0:
            db.session.commit()
            if semesters_created == 1:
                flash(f'{semester_names[semester_number]} has been successfully added!', 'success')
            else:
                flash(f'{semesters_created} semesters have been successfully added!', 'success')
        else:
            flash(f'All semesters up to {semester_names[semester_number]} already exist for this year level.', 'info')
        
        params = build_redirect_params(
            curriculum_id=year_level.curriculum_id, 
            year_level_id=year_level.id
        )
        return redirect(url_for('curriculum.index', **params))
        
    except Exception as e:
        db.session.rollback()
        flash(f'An error occurred while adding the semester: {str(e)}', 'error')
        return redirect(url_for('curriculum.index'))


@curriculum_bp.route('/semester/edit', methods=['POST'])
@login_required
def edit_semester():
    """Edit a semester"""
    try:
        semester_id = request.form.get('semester_id', '').strip()
        semester_name = request.form.get('semester_name', '').strip()
        
        if not all([semester_id, semester_name]):
            flash('Please fill in all required fields.', 'error')
            return redirect(url_for('curriculum.index'))
        
        semester = Semester.query.get(int(semester_id))
        if not semester:
            flash('Semester not found.', 'error')
            return redirect(url_for('curriculum.index'))
        
        semester.semester_name = semester_name
        db.session.commit()
        
        flash('Semester has been successfully updated!', 'success')
        params = build_redirect_params(
            curriculum_id=semester.year_level.curriculum_id,
            year_level_id=semester.year_level.id,
            semester_id=semester.id
        )
        return redirect(url_for('curriculum.index', **params))
        
    except Exception as e:
        db.session.rollback()
        flash(f'An error occurred while updating the semester: {str(e)}', 'error')
        return redirect(url_for('curriculum.index'))


@curriculum_bp.route('/semester/delete', methods=['POST'])
@login_required
def delete_semester():
    """Delete a semester"""
    try:
        semester_id = request.form.get('semester_id', '').strip()
        
        if not semester_id:
            flash('Invalid semester.', 'error')
            return redirect(url_for('curriculum.index'))
        
        semester = Semester.query.get(int(semester_id))
        if not semester:
            flash('Semester not found.', 'error')
            return redirect(url_for('curriculum.index'))
        
        curriculum_id = semester.year_level.curriculum_id
        year_level_id = semester.year_level.id
        db.session.delete(semester)
        db.session.commit()
        
        flash('Semester has been successfully deleted!', 'success')
        # Don't pass semester ID when deleting - the semester no longer exists
        params = build_redirect_params(
            curriculum_id=curriculum_id,
            year_level_id=year_level_id
        )
        return redirect(url_for('curriculum.index', **params))
        
    except Exception as e:
        db.session.rollback()
        flash(f'An error occurred while deleting the semester: {str(e)}', 'error')
        return redirect(url_for('curriculum.index'))


# Subject Template routes - REMOVED (templates no longer used)
# Templates have been removed from the system. Subjects now store all data directly.

@curriculum_bp.route('/subject/add', methods=['POST'])
@login_required
def add_subject():
    """Add a new subject"""
    try:
        semester_id = request.form.get('semester_id', '').strip()
        subject_code = request.form.get('subject_code', '').strip().upper()
        course_description = request.form.get('course_description', '').strip()
        lec_units = request.form.get('lec_units', '').strip()
        lab_units = request.form.get('lab_units', '').strip()
        prerequisite = request.form.get('prerequisite', '').strip()
        
        # Validate required fields
        if not all([semester_id, subject_code, course_description, lec_units, lab_units]):
            flash('Please fill in all required fields.', 'error')
            return redirect(url_for('curriculum.index'))
        
        try:
            semester_id = int(semester_id)
            lec_units = float(lec_units)
            lab_units = float(lab_units)
            
            if lec_units < 0 or lec_units > 10 or lab_units < 0 or lab_units > 10:
                flash('Units must be between 0 and 10.', 'error')
                return redirect(url_for('curriculum.index'))
        except ValueError:
            flash('Invalid input values.', 'error')
            return redirect(url_for('curriculum.index'))
        
        semester = Semester.query.get(semester_id)
        if not semester:
            flash('Semester not found.', 'error')
            return redirect(url_for('curriculum.index'))
        
        # Create new subject directly
        new_subject = Subject(
            semester_id=semester_id,
            subject_code=subject_code,
            course_description=course_description,
            lec_units=lec_units,
            lab_units=lab_units,
            prerequisite=prerequisite if prerequisite else None
        )
        
        db.session.add(new_subject)
        db.session.flush()
        
        # Log activity
        log_create('subject', new_subject.id, subject_code, {
            'description': course_description,
            'semester': semester.semester_name
        })
        
        db.session.commit()
        
        flash(f'Subject "{subject_code}" added successfully!', 'success')
        
        params = build_redirect_params(
            curriculum_id=semester.year_level.curriculum_id,
            year_level_id=semester.year_level.id,
            semester_id=semester.id
        )
        return redirect(url_for('curriculum.index', **params))
        
    except Exception as e:
        db.session.rollback()
        flash(f'An error occurred while adding the subject: {str(e)}', 'error')
        return redirect(url_for('curriculum.index'))




@curriculum_bp.route('/subject/edit', methods=['POST'])
@login_required
def edit_subject():
    """Edit a subject"""
    try:
        subject_id = request.form.get('subject_id', '').strip()
        subject_code = request.form.get('subject_code', '').strip().upper()
        course_description = request.form.get('course_description', '').strip()
        lec_units = request.form.get('lec_units', '').strip()
        lab_units = request.form.get('lab_units', '').strip()
        prerequisite = request.form.get('prerequisite', '').strip()
        
        if not all([subject_id, subject_code, course_description, lec_units, lab_units]):
            flash('Please fill in all required fields.', 'error')
            return redirect(url_for('curriculum.index'))
        
        try:
            lec_units = float(lec_units)
            lab_units = float(lab_units)
            if lec_units < 0 or lec_units > 10 or lab_units < 0 or lab_units > 10:
                flash('Units must be between 0 and 10.', 'error')
                return redirect(url_for('curriculum.index'))
        except ValueError:
            flash('Invalid units value.', 'error')
            return redirect(url_for('curriculum.index'))
        
        subject = Subject.query.get(int(subject_id))
        if not subject:
            flash('Subject not found.', 'error')
            return redirect(url_for('curriculum.index'))
        
        # Track changes
        changes = {}
        if subject.subject_code != subject_code:
            changes['code'] = f'{subject.subject_code} → {subject_code}'
        if subject.course_description != course_description:
            changes['description'] = f'{subject.course_description} → {course_description}'
        if subject.lec_units != lec_units:
            changes['lec_units'] = f'{subject.lec_units} → {lec_units}'
        if subject.lab_units != lab_units:
            changes['lab_units'] = f'{subject.lab_units} → {lab_units}'
        if subject.prerequisite != (prerequisite if prerequisite else None):
            old_prereq = subject.prerequisite or 'None'
            new_prereq = prerequisite if prerequisite else 'None'
            changes['prerequisite'] = f'{old_prereq} → {new_prereq}'
        
        # Update subject fields directly
        subject.subject_code = subject_code
        subject.course_description = course_description
        subject.lec_units = lec_units
        subject.lab_units = lab_units
        subject.prerequisite = prerequisite if prerequisite else None
        
        # Log activity with changes
        log_edit('subject', subject.id, subject_code, changes if changes else None)
        
        db.session.commit()
        
        flash('Subject has been successfully updated!', 'success')
        
        params = build_redirect_params(
            curriculum_id=subject.semester.year_level.curriculum_id,
            year_level_id=subject.semester.year_level.id,
            semester_id=subject.semester.id
        )
        return redirect(url_for('curriculum.index', **params))
        
    except Exception as e:
        db.session.rollback()
        flash(f'An error occurred while updating the subject: {str(e)}', 'error')
        return redirect(url_for('curriculum.index'))



@curriculum_bp.route('/subject/bulk-delete', methods=['POST'])
@login_required
def bulk_delete_subjects():
    """Bulk delete subjects"""
    try:
        data = request.get_json()
        ids = data.get('ids', [])
        if not ids:
            return jsonify({'success': False, 'message': 'No subjects selected'}), 400

        subjects = Subject.query.filter(Subject.id.in_(ids)).all()
        if not subjects:
            return jsonify({'success': False, 'message': 'No subjects found'}), 404

        count = 0
        for subject in subjects:
            # Delete associated schedules
            Schedule.query.filter_by(subject_id=subject.id).delete()
            ExamSchedule.query.filter_by(subject_id=subject.id).delete()
            log_delete('subject', subject.id, subject.subject_code, {'description': subject.course_description})
            db.session.delete(subject)
            count += 1

        db.session.commit()
        return jsonify({'success': True, 'message': f'Successfully deleted {count} subject{"s" if count != 1 else ""}', 'affected': count})
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'message': str(e)}), 500


@curriculum_bp.route('/subject/delete', methods=['POST'])
@login_required
def delete_subject():
    """Delete a subject"""
    try:
        subject_id = request.form.get('subject_id', '').strip()
        
        if not subject_id:
            flash('Invalid subject.', 'error')
            return redirect(url_for('curriculum.index'))
        
        subject = Subject.query.get(int(subject_id))
        if not subject:
            flash('Subject not found.', 'error')
            return redirect(url_for('curriculum.index'))
        
        curriculum_id = subject.semester.year_level.curriculum_id
        year_level_id = subject.semester.year_level.id
        semester_id = subject.semester.id
        subject_code = subject.subject_code
        
        # Log activity before deletion
        log_delete('subject', subject.id, subject_code, {'description': subject.course_description})
        
        db.session.delete(subject)
        db.session.commit()
        
        flash('Subject has been successfully deleted!', 'success')
        params = build_redirect_params(
            curriculum_id=curriculum_id,
            year_level_id=year_level_id,
            semester_id=semester_id
        )
        return redirect(url_for('curriculum.index', **params))
        
    except Exception as e:
        db.session.rollback()
        flash(f'An error occurred while deleting the subject: {str(e)}', 'error')
        return redirect(url_for('curriculum.index'))


@curriculum_bp.route('/archive', methods=['POST'])
@login_required
def archive():
    """Archive a curriculum and delete all related schedules"""
    try:
        curriculum_id = request.form.get('curriculum_id', '').strip()
        archive_reason = request.form.get('archive_reason', 'Manual archive by user').strip()
        
        if not curriculum_id:
            flash('Invalid curriculum.', 'error')
            return redirect(url_for('curriculum.index'))
        
        curriculum = Curriculum.query.get(int(curriculum_id))
        if not curriculum:
            flash('Curriculum not found.', 'error')
            return redirect(url_for('curriculum.index'))
        
        curriculum_code = curriculum.curriculum_code
        
        # Collect all subject IDs from the curriculum
        subject_ids = []
        for year_level in curriculum.year_levels:
            for semester in year_level.semesters:
                for subject in semester.subjects:
                    subject_ids.append(subject.id)
        
        # Delete all class schedules using subjects from this curriculum
        class_schedules_count = 0
        if subject_ids:
            class_schedules = Schedule.query.filter(
                Schedule.subject_id.in_(subject_ids),
                Schedule.is_active == True
            ).all()
            
            for schedule in class_schedules:
                log_delete('schedule', schedule.id, f"Schedule for {schedule.subject.subject_code}", {
                    'reason': f'Curriculum archived: {curriculum_code}',
                    'section': schedule.section.section_name if schedule.section else 'N/A',
                    'subject': schedule.subject.subject_code if schedule.subject else 'N/A',
                    'day': schedule.day_of_week,
                    'time': f"{schedule.start_time} - {schedule.end_time}"
                })
                db.session.delete(schedule)
                class_schedules_count += 1
        
        # Delete all exam schedules using subjects from this curriculum
        exam_schedules_count = 0
        if subject_ids:
            exam_schedules = ExamSchedule.query.filter(
                ExamSchedule.subject_id.in_(subject_ids),
                ExamSchedule.is_active == True
            ).all()
            
            for exam_schedule in exam_schedules:
                log_delete('exam_schedule', exam_schedule.id, f"Exam schedule for {exam_schedule.subject.subject_code}", {
                    'reason': f'Curriculum archived: {curriculum_code}',
                    'section': exam_schedule.section.section_name if exam_schedule.section else 'N/A',
                    'subject': exam_schedule.subject.subject_code if exam_schedule.subject else 'N/A',
                    'exam_date': str(exam_schedule.exam_date),
                    'time': f"{exam_schedule.start_time} - {exam_schedule.end_time}"
                })
                db.session.delete(exam_schedule)
                exam_schedules_count += 1
        
        # Archive curriculum using helper method
        curriculum.archive(user_id=current_user.id, reason=archive_reason)
        
        # Log activity with deletion counts
        log_archive('curriculum', curriculum.id, curriculum_code, {
            'program': curriculum.program.program_code,
            'reason': archive_reason,
            'deleted_class_schedules': class_schedules_count,
            'deleted_exam_schedules': exam_schedules_count
        })
        
        db.session.commit()
        
        flash(f'Curriculum "{curriculum_code}" has been archived successfully!', 'success')
        return redirect(url_for('curriculum.index'))
        
    except Exception as e:
        db.session.rollback()
        flash(f'An error occurred while archiving the curriculum: {str(e)}', 'error')
        return redirect(url_for('curriculum.index'))


@curriculum_bp.route('/delete', methods=['POST'])
@login_required
def delete():
    """Delete a curriculum permanently (only for archived curricula)"""
    try:
        curriculum_id = request.form.get('curriculum_id', '').strip()
        
        if not curriculum_id:
            flash('Invalid curriculum.', 'error')
            return redirect(url_for('archive.index'))
        
        curriculum = Curriculum.query.get(int(curriculum_id))
        if not curriculum:
            flash('Curriculum not found.', 'error')
            return redirect(url_for('archive.index'))
        
        if not curriculum.is_archived:
            flash('Only archived curricula can be permanently deleted.', 'error')
            return redirect(url_for('curriculum.index'))
        
        curriculum_code = curriculum.curriculum_code
        
        # Log activity before deletion
        log_delete('curriculum', curriculum.id, curriculum_code, {
            'program': curriculum.program.program_code
        })
        
        # Delete curriculum (cascade will delete year levels, semesters, and subjects)
        db.session.delete(curriculum)
        db.session.commit()
        
        flash(f'Curriculum "{curriculum_code}" has been permanently deleted!', 'success')
        return redirect(url_for('archive.index'))
        
    except Exception as e:
        db.session.rollback()
        flash(f'An error occurred while deleting the curriculum: {str(e)}', 'error')
        return redirect(url_for('archive.index'))


# Bulk Import Routes
@curriculum_bp.route('/bulk-import/template/<int:curriculum_id>')
@login_required
@role_required('admin', 'super_admin')
def download_bulk_import_template(curriculum_id):
    """Generate and download Excel template for bulk subject import"""
    try:
        curriculum = Curriculum.query.get(curriculum_id)
        if not curriculum:
            flash('Curriculum not found.', 'error')
            return redirect(url_for('curriculum.index'))

        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.worksheet.datavalidation import DataValidation
        from openpyxl.utils import get_column_letter

        wb = Workbook()
        wb.remove(wb.active)  # remove default sheet

        # ── Colour palette ──
        NAVY       = '1E3A5F'
        BLUE       = '2563EB'
        BLUE_LIGHT = 'EFF6FF'
        GRAY_50    = 'F9FAFB'
        GRAY_200   = 'E5E7EB'
        WHITE      = 'FFFFFF'
        AMBER      = 'D97706'

        thin_border = Border(
            left=Side('thin', GRAY_200), right=Side('thin', GRAY_200),
            top=Side('thin', GRAY_200), bottom=Side('thin', GRAY_200),
        )
        header_border = Border(
            left=Side('thin', WHITE), right=Side('thin', WHITE),
            top=Side('thin', WHITE), bottom=Side('thin', WHITE),
        )

        # ── Reusable styles ──
        header_fill = PatternFill('solid', fgColor=BLUE)
        header_font = Font(bold=True, color=WHITE, size=10)
        header_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
        center = Alignment(horizontal='center', vertical='center')
        left_wrap = Alignment(vertical='center', wrap_text=True)

        program_name = curriculum.program.program_name if curriculum.program else ''
        DATA_ROWS = 15
        col_widths = [18, 46, 13, 13, 22]
        headers = ['Subject Code', 'Course Description', 'Lecture Units', 'Lab Units', 'Prerequisite']

        for year_level in curriculum.year_levels:
            yn = year_level.year_name
            for semester in year_level.semesters:
                sn = semester.semester_name

                # Sheet name (Excel max 31 chars)
                sheet_name = f'{yn} - {sn}'
                if len(sheet_name) > 31:
                    sheet_name = sheet_name[:31]
                ws = wb.create_sheet(title=sheet_name)

                # Column widths
                for i, w in enumerate(col_widths, 1):
                    ws.column_dimensions[get_column_letter(i)].width = w

                # ══ ROW 1 — Curriculum title banner ══
                ws.merge_cells('A1:E1')
                title_cell = ws.cell(row=1, column=1,
                                     value=f'{curriculum.curriculum_code}  \u2014  {curriculum.curriculum_name}')
                title_cell.font = Font(bold=True, size=13, color=WHITE)
                title_cell.fill = PatternFill('solid', fgColor=NAVY)
                title_cell.alignment = Alignment(horizontal='center', vertical='center')
                title_cell.border = header_border
                for c in range(2, 6):
                    ws.cell(row=1, column=c).fill = PatternFill('solid', fgColor=NAVY)
                    ws.cell(row=1, column=c).border = header_border
                ws.row_dimensions[1].height = 32

                # ══ ROW 2 — Year level + Semester info ══
                ws.merge_cells('A2:E2')
                info_cell = ws.cell(row=2, column=1, value=f'{yn}  \u2022  {sn}')
                info_cell.font = Font(bold=True, size=11, color=NAVY)
                info_cell.fill = PatternFill('solid', fgColor=BLUE_LIGHT)
                info_cell.alignment = Alignment(horizontal='center', vertical='center')
                for c in range(2, 6):
                    ws.cell(row=2, column=c).fill = PatternFill('solid', fgColor=BLUE_LIGHT)
                ws.row_dimensions[2].height = 26

                # ══ ROW 3 — Thin separator ══
                for c in range(1, 6):
                    cell = ws.cell(row=3, column=c)
                    cell.fill = PatternFill('solid', fgColor=WHITE)
                    cell.border = Border(bottom=Side('thin', GRAY_200))
                ws.row_dimensions[3].height = 6

                # ══ ROW 4 — Column headers ══
                for col_idx, title in enumerate(headers, 1):
                    cell = ws.cell(row=4, column=col_idx, value=title)
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = header_align
                    cell.border = header_border
                ws.row_dimensions[4].height = 28

                # ══ ROWS 5–14 — Data rows ══
                last_row = 4 + DATA_ROWS
                for r in range(5, last_row + 1):
                    fill = PatternFill('solid', fgColor=GRAY_50) if r % 2 == 1 else PatternFill('solid', fgColor=WHITE)
                    for c in range(1, 6):
                        cell = ws.cell(row=r, column=c)
                        cell.fill = fill
                        cell.border = thin_border
                        cell.alignment = center if c in (3, 4) else left_wrap
                    ws.row_dimensions[r].height = 22

                # ── Units validation (0–9) ──
                for col_letter in ('C', 'D'):
                    num_dv = DataValidation(type='decimal', operator='between',
                                            formula1=0, formula2=9, allow_blank=True,
                                            showErrorMessage=True,
                                            error='Enter a value between 0 and 9',
                                            errorTitle='Invalid Units')
                    ws.add_data_validation(num_dv)
                    num_dv.add(f'{col_letter}5:{col_letter}{last_row}')

                # ── Freeze & print ──
                ws.freeze_panes = 'A5'
                ws.print_title_rows = '1:4'

                # ── Tip row ──
                tip_row = last_row + 2
                tip_text = ('Fill Subject Code, Description & Units  \u2022  '
                            'Empty rows are automatically skipped  \u2022  '
                            'Prerequisite is optional (leave blank or type "None")')
                ws.merge_cells(start_row=tip_row, start_column=1, end_row=tip_row, end_column=5)
                tip_cell = ws.cell(row=tip_row, column=1, value=tip_text)
                tip_cell.font = Font(italic=True, size=9, color=AMBER)
                tip_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                ws.row_dimensions[tip_row].height = 28

        # ── Write to buffer ──
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        filename = f"{curriculum.curriculum_code}_import_template.xlsx"
        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=filename
        )
        
    except Exception as e:
        flash(f'Error generating template: {str(e)}', 'error')
        return redirect(url_for('curriculum.index'))


@curriculum_bp.route('/bulk-import/<int:curriculum_id>', methods=['POST'])
@login_required
@role_required('admin', 'super_admin')
def bulk_import_subjects(curriculum_id):
    """Process bulk import of subjects from Excel file"""
    try:
        curriculum = Curriculum.query.get(curriculum_id)
        if not curriculum:
            flash('Curriculum not found.', 'error')
            return redirect(url_for('curriculum.index'))
        
        # Check if file was uploaded
        if 'bulk_import_file' not in request.files:
            flash('No file uploaded.', 'error')
            return redirect(url_for('curriculum.index', curriculum_id=curriculum_id, open=curriculum_id))
        
        file = request.files['bulk_import_file']
        
        if file.filename == '':
            flash('No file selected.', 'error')
            return redirect(url_for('curriculum.index', curriculum_id=curriculum_id, open=curriculum_id))
        
        # Validate file extension
        if not file.filename.endswith(('.xlsx', '.xls')):
            flash('Invalid file format. Please upload an Excel file (.xlsx or .xls)', 'error')
            return redirect(url_for('curriculum.index', curriculum_id=curriculum_id, open=curriculum_id))
        
        # Read Excel workbook (each sheet = one Year Level + Semester)
        from openpyxl import load_workbook as _load_wb
        wb = _load_wb(file, data_only=True)

        subjects_added = 0
        errors = []

        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]

            # ── Parse Year Level / Semester from row 2 info cell ──
            info_value = ws.cell(row=2, column=1).value
            if not info_value or '\u2022' not in str(info_value):
                errors.append(f"Sheet '{sheet_name}': Could not detect Year Level / Semester info — skipped")
                continue

            parts = str(info_value).split('\u2022')
            if len(parts) != 2:
                errors.append(f"Sheet '{sheet_name}': Unexpected info format — skipped")
                continue

            year_name = parts[0].strip()
            semester_name = parts[1].strip()

            # Match year level
            year_level = None
            for yl in curriculum.year_levels:
                if yl.year_name.strip().lower() == year_name.lower():
                    year_level = yl
                    break
            if not year_level:
                errors.append(f"Sheet '{sheet_name}': Year level '{year_name}' not found in curriculum")
                continue

            # Match semester
            semester = None
            for sem in year_level.semesters:
                if sem.semester_name.strip().lower() == semester_name.lower():
                    semester = sem
                    break
            if not semester:
                errors.append(f"Sheet '{sheet_name}': Semester '{semester_name}' not found in {year_level.year_name}")
                continue

            # ── Read data rows (row 5+) ──
            # Headers at row 4: Subject Code | Course Description | Lecture Units | Lab Units | Prerequisite
            for row_num in range(5, ws.max_row + 1):
                try:
                    raw_code = ws.cell(row=row_num, column=1).value
                    raw_desc = ws.cell(row=row_num, column=2).value
                    lec_val = ws.cell(row=row_num, column=3).value
                    lab_val = ws.cell(row=row_num, column=4).value
                    prereq_val = ws.cell(row=row_num, column=5).value

                    raw_code_text = str(raw_code or '').strip()
                    raw_desc_text = str(raw_desc or '').strip()

                    # Skip truly empty rows.
                    if raw_code_text == '' and raw_desc_text == '' and lec_val is None and lab_val is None and (prereq_val is None or str(prereq_val).strip() == ''):
                        continue  # skip empty rows

                    # Skip template informational/tip rows (e.g., merged footer note row).
                    # This prevents false validation errors like "Row 21 required fields" when using generated templates.
                    note_markers = (
                        'tip:',
                        'empty rows are automatically skipped',
                        'do not modify sheet names',
                        'do not remove columns',
                    )
                    code_lc = raw_code_text.lower()
                    if any(marker in code_lc for marker in note_markers):
                        continue

                    subject_code = raw_code_text.upper()
                    course_description = raw_desc_text

                    if not subject_code or not course_description:
                        errors.append(f"Sheet '{sheet_name}' Row {row_num}: Subject Code and Course Description are required")
                        continue

                    lec_units = float(lec_val) if lec_val is not None else 0.0
                    lab_units = float(lab_val) if lab_val is not None else 0.0
                    prerequisite = str(prereq_val).strip() if prereq_val and str(prereq_val).strip().lower() not in ('', 'none') else None

                    # Check duplicate
                    existing_subject = Subject.query.filter_by(
                        semester_id=semester.id,
                        subject_code=subject_code
                    ).first()

                    if existing_subject:
                        errors.append(f"Sheet '{sheet_name}' Row {row_num}: Subject '{subject_code}' already exists in {year_level.year_name} - {semester.semester_name}")
                        continue

                    new_subject = Subject(
                        semester_id=semester.id,
                        subject_code=subject_code,
                        course_description=course_description,
                        lec_units=lec_units,
                        lab_units=lab_units,
                        prerequisite=prerequisite
                    )
                    db.session.add(new_subject)
                    subjects_added += 1

                except Exception as e:
                    errors.append(f"Sheet '{sheet_name}' Row {row_num}: {str(e)}")
                    continue

        # Commit all changes
        db.session.commit()

        # Show results
        if subjects_added > 0:
            flash(f'Successfully imported {subjects_added} subject(s)!', 'success')

        if errors:
            error_msg = f'{len(errors)} error(s) occurred:<br>' + '<br>'.join(errors[:10])
            if len(errors) > 10:
                error_msg += f'<br>... and {len(errors) - 10} more errors'
            flash(error_msg, 'error')

        if subjects_added == 0 and not errors:
            flash('No subjects were imported. Please make sure your file contains valid subject data with all required fields filled in.', 'info')

        return redirect(url_for('curriculum.index', curriculum_id=curriculum_id, open=curriculum_id))

    except Exception as e:
        db.session.rollback()
        flash(f'An error occurred during bulk import: {str(e)}', 'error')
        return redirect(url_for('curriculum.index', curriculum_id=curriculum_id, open=curriculum_id))

