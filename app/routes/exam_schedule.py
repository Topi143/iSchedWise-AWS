"""
Exam Schedule routes for managing exam schedules
Supports multi-user concurrent scheduling with optimistic locking
"""
from flask import Blueprint, request, flash, redirect, url_for, jsonify, send_file
from flask_login import login_required, current_user
from sqlalchemy import and_, or_
from datetime import datetime, time as dt_time
import io
import traceback
from app.extensions import db, csrf
from app.models.exam_schedule import ExamSchedule
from app.models.program import Program
from app.models.section import Section
from app.models.curriculum import Subject
from app.models.faculty import Faculty, FacultyAvailability
from app.models.building import Room, Building
from app.models.settings import AcademicSettings
from app.decorators import role_required
from app.utils.activity_logger import log_create, log_edit, log_delete
from app.routes.socket_events import broadcast_schedule_change

exam_schedule_bp = Blueprint('exam_schedule', __name__, url_prefix='/exam-schedule')


def _coerce_setting_time(value, fallback_hour):
    if isinstance(value, dt_time):
        return value
    if isinstance(value, int):
        return dt_time(value, 0)
    if isinstance(value, str):
        try:
            parts = value.split(':')
            return dt_time(int(parts[0]), int(parts[1]) if len(parts) > 1 else 0)
        except Exception:
            pass
    return dt_time(fallback_hour, 0)


def _minutes_of(t):
    return t.hour * 60 + t.minute


def _fmt_setting_time(t):
    return t.strftime('%I:%M %p').lstrip('0')


@exam_schedule_bp.route('/add', methods=['POST'])
@login_required
def add():
    """Add a new exam schedule"""
    is_ajax = request.headers.get('X-Requested-With') == 'XMLHttpRequest'
    ajax_warnings = []
    try:
        section_id = request.form.get('section_id', type=int)
        subject_id = request.form.get('subject_id', type=int)
        faculty_id = request.form.get('faculty_id', type=int)
        room_id = request.form.get('room_id', type=int)
        exam_date_str = request.form.get('exam_date')
        start_time_str = request.form.get('start_time')
        end_time_str = request.form.get('end_time')
        schedule_type = request.form.get('schedule_type', 'lecture')  # lecture or lab
        
        # Get current academic settings (including exam_period)
        current_settings = AcademicSettings.query.filter_by(is_active=True).first()
        if not current_settings:
            msg = 'No active academic settings found. Please configure academic settings first.'
            if is_ajax:
                return jsonify({'success': False, 'error': msg}), 400
            flash(msg, 'danger')
            return redirect(url_for('schedule.create_page', type='exam', section_id=section_id))
        
        academic_year = current_settings.academic_year
        semester = current_settings.semester
        exam_period = current_settings.exam_period
        
        # Validation - faculty and room are now required
        if not all([section_id, subject_id, faculty_id, room_id, exam_date_str, start_time_str, end_time_str]):
            msg = 'All required fields must be filled (including faculty and room)'
            if is_ajax:
                return jsonify({'success': False, 'error': msg}), 400
            flash(msg, 'danger')
            return redirect(url_for('schedule.create_page', type='exam', section_id=section_id))
        
        # Convert date and time strings
        exam_date = datetime.strptime(exam_date_str, '%Y-%m-%d').date()
        start_time = datetime.strptime(start_time_str, '%H:%M').time()
        end_time = datetime.strptime(end_time_str, '%H:%M').time()
        
        # Validate exam date is not in the past
        from datetime import date as date_class
        today = date_class.today()
        if exam_date < today:
            msg = 'Cannot schedule exams in the past. Please select a future date.'
            if is_ajax:
                return jsonify({'success': False, 'error': msg}), 400
            flash(msg, 'danger')
            return redirect(url_for('schedule.create_page', type='exam', section_id=section_id))
        
        # Validate exam date is within the configured exam period range
        if current_settings:
            exam_period_start = getattr(current_settings, 'exam_period_start', None)
            exam_period_end = getattr(current_settings, 'exam_period_end', None)
            
            if exam_period_start and exam_period_end:
                if exam_date < exam_period_start or exam_date > exam_period_end:
                    start_str = exam_period_start.strftime('%B %d, %Y')
                    end_str = exam_period_end.strftime('%B %d, %Y')
                    msg = f'Exam date must be within the exam period ({start_str} to {end_str})'
                    if is_ajax:
                        return jsonify({'success': False, 'error': msg}), 400
                    flash(msg, 'danger')
                    return redirect(url_for('schedule.create_page', type='exam', section_id=section_id))
        
        # Validate time range
        if start_time >= end_time:
            msg = 'End time must be after start time'
            if is_ajax:
                return jsonify({'success': False, 'error': msg}), 400
            flash(msg, 'danger')
            return redirect(url_for('schedule.create_page', type='exam', section_id=section_id))
        
        # Validate exam times are within configured exam hours
        if current_settings:
            exam_start_cfg = _coerce_setting_time(getattr(current_settings, 'exam_start_time', None), current_settings.exam_start_hour or 7)
            exam_end_cfg = _coerce_setting_time(getattr(current_settings, 'exam_end_time', None), current_settings.exam_end_hour or 17)

            if _minutes_of(start_time) < _minutes_of(exam_start_cfg):
                msg = f'Start time must be at or after {_fmt_setting_time(exam_start_cfg)}'
                if is_ajax:
                    return jsonify({'success': False, 'error': msg}), 400
                flash(msg, 'danger')
                return redirect(url_for('schedule.create_page', type='exam', section_id=section_id))
            
            if _minutes_of(end_time) > _minutes_of(exam_end_cfg):
                msg = f'End time must be at or before {_fmt_setting_time(exam_end_cfg)}'
                if is_ajax:
                    return jsonify({'success': False, 'error': msg}), 400
                flash(msg, 'danger')
                return redirect(url_for('schedule.create_page', type='exam', section_id=section_id))
            
            # Check for lunch overlap and add warning (not blocking)
            if current_settings.exam_lunch_start and current_settings.exam_lunch_end:
                lunch_start = current_settings.exam_lunch_start
                lunch_end = current_settings.exam_lunch_end
                # Check overlap: exam starts before lunch ends AND exam ends after lunch starts
                start_mins = start_time.hour * 60 + start_time.minute
                end_mins = end_time.hour * 60 + end_time.minute
                lunch_start_mins = lunch_start.hour * 60 + lunch_start.minute
                lunch_end_mins = lunch_end.hour * 60 + lunch_end.minute
                
                if start_mins < lunch_end_mins and end_mins > lunch_start_mins:
                    lunch_start_str = lunch_start.strftime('%I:%M %p').lstrip('0')
                    lunch_end_str = lunch_end.strftime('%I:%M %p').lstrip('0')
                    warn_msg = f'Note: This exam overlaps with the lunch break ({lunch_start_str} - {lunch_end_str})'
                    if is_ajax:
                        ajax_warnings.append(warn_msg)
                    else:
                        flash(warn_msg, 'warning')
            
            # Check exam duration limit
            exam_duration_limit = current_settings.exam_duration_limit if hasattr(current_settings, 'exam_duration_limit') else 120
            exam_duration = (end_time.hour * 60 + end_time.minute) - (start_time.hour * 60 + start_time.minute)
            if exam_duration > exam_duration_limit:
                hours = exam_duration_limit // 60
                mins = exam_duration_limit % 60
                limit_text = f'{hours} hour{"s" if hours > 1 else ""}' if mins == 0 else f'{hours}h {mins}m'
                warn_msg = f'Exam duration ({exam_duration} min) exceeds the maximum allowed ({limit_text})'
                if is_ajax:
                    ajax_warnings.append(warn_msg)
                else:
                    flash(warn_msg, 'warning')
        
        # Check for conflicts - same section, date, and overlapping time
        # Use pessimistic locking to prevent race conditions when multiple users add exam schedules
        conflict_query = ExamSchedule.query.filter(
            ExamSchedule.section_id == section_id,
            ExamSchedule.exam_date == exam_date,
            ExamSchedule.is_active == True,
            or_(
                and_(ExamSchedule.start_time <= start_time, ExamSchedule.end_time > start_time),
                and_(ExamSchedule.start_time < end_time, ExamSchedule.end_time >= end_time),
                and_(ExamSchedule.start_time >= start_time, ExamSchedule.end_time <= end_time)
            )
        )
        
        if academic_year and semester:
            conflict_query = conflict_query.filter(
                ExamSchedule.academic_year == academic_year,
                ExamSchedule.semester == semester
            )
        
        # Pessimistic lock to prevent concurrent inserts for same slot
        conflict_query = conflict_query.with_for_update(nowait=True)
        
        if conflict_query.first():
            msg = 'Exam schedule conflict: This section already has an exam scheduled at this time'
            if is_ajax:
                return jsonify({'success': False, 'error': msg}), 409
            flash(msg, 'danger')
            return redirect(url_for('schedule.create_page', type='exam', section_id=section_id))
        
        # Check for faculty conflicts if faculty assigned
        if faculty_id:
            faculty_conflict = ExamSchedule.query.filter(
                ExamSchedule.faculty_id == faculty_id,
                ExamSchedule.exam_date == exam_date,
                ExamSchedule.is_active == True,
                or_(
                    and_(ExamSchedule.start_time <= start_time, ExamSchedule.end_time > start_time),
                    and_(ExamSchedule.start_time < end_time, ExamSchedule.end_time >= end_time),
                    and_(ExamSchedule.start_time >= start_time, ExamSchedule.end_time <= end_time)
                )
            )
            
            if academic_year and semester:
                faculty_conflict = faculty_conflict.filter(
                    ExamSchedule.academic_year == academic_year,
                    ExamSchedule.semester == semester
                )
            
            # Pessimistic lock to prevent concurrent faculty double-booking
            faculty_conflict = faculty_conflict.with_for_update(nowait=True)
            
            if faculty_conflict.first():
                msg = 'Faculty conflict: This faculty member is already assigned to another exam at this time'
                if is_ajax:
                    return jsonify({'success': False, 'error': msg}), 409
                flash(msg, 'danger')
                return redirect(url_for('schedule.create_page', type='exam', section_id=section_id))
        
        # Check for room conflicts if room assigned
        if room_id:
            room_conflict = ExamSchedule.query.filter(
                ExamSchedule.room_id == room_id,
                ExamSchedule.exam_date == exam_date,
                ExamSchedule.is_active == True,
                or_(
                    and_(ExamSchedule.start_time <= start_time, ExamSchedule.end_time > start_time),
                    and_(ExamSchedule.start_time < end_time, ExamSchedule.end_time >= end_time),
                    and_(ExamSchedule.start_time >= start_time, ExamSchedule.end_time <= end_time)
                )
            )
            
            if academic_year and semester:
                room_conflict = room_conflict.filter(
                    ExamSchedule.academic_year == academic_year,
                    ExamSchedule.semester == semester
                )
            
            # Pessimistic lock to prevent concurrent room double-booking
            room_conflict = room_conflict.with_for_update(nowait=True)
            
            if room_conflict.first():
                msg = 'Room conflict: This room is already assigned to another exam at this time'
                if is_ajax:
                    return jsonify({'success': False, 'error': msg}), 409
                flash(msg, 'danger')
                return redirect(url_for('schedule.create_page', type='exam', section_id=section_id))
        
        # Check proctor availability
        if faculty_id:
            availability_result = FacultyAvailability.check_faculty_available(
                faculty_id, exam_date, start_time, end_time
            )
            if availability_result['status'] == 'unavailable':
                faculty = Faculty.query.get(faculty_id)
                faculty_name = faculty.full_name if faculty else 'The selected faculty'
                reason = availability_result.get('reason')
                reason_text = f" Reason: {reason}" if reason else ""
                msg = f'Proctor unavailable: {faculty_name} has marked themselves as unavailable for this time slot.{reason_text}'
                if is_ajax:
                    return jsonify({'success': False, 'error': msg}), 409
                flash(msg, 'danger')
                return redirect(url_for('schedule.create_page', type='exam', section_id=section_id))
            elif availability_result['status'] == 'not_in_schedule':
                faculty = Faculty.query.get(faculty_id)
                faculty_name = faculty.full_name if faculty else 'The selected faculty'
                warn_msg = f'Warning: {faculty_name} is not marked as available on this date/time. Exam scheduled anyway.'
                if is_ajax:
                    ajax_warnings.append(warn_msg)
                else:
                    flash(warn_msg, 'warning')
        
        # Check for soft-deleted exam in the same slot (uk_exam_section_slot)
        existing_inactive_exam = ExamSchedule.query.filter_by(
            section_id=section_id,
            exam_date=exam_date,
            start_time=start_time,
            end_time=end_time,
            academic_year=academic_year,
            semester=semester,
            exam_period=exam_period,
            is_active=False
        ).first()

        if existing_inactive_exam:
            # Reactivate and update the soft-deleted exam schedule
            existing_inactive_exam.subject_id = subject_id
            existing_inactive_exam.faculty_id = faculty_id
            existing_inactive_exam.room_id = room_id
            existing_inactive_exam.schedule_type = schedule_type
            existing_inactive_exam.is_active = True
            existing_inactive_exam.version = (existing_inactive_exam.version or 1) + 1
            existing_inactive_exam.updated_at = datetime.utcnow()
            db.session.flush()
            exam_schedule = existing_inactive_exam
        else:
            # Create new exam schedule
            exam_schedule = ExamSchedule(
                section_id=section_id,
                subject_id=subject_id,
                faculty_id=faculty_id,
                room_id=room_id,
                exam_date=exam_date,
                start_time=start_time,
                end_time=end_time,
                semester=semester,
                academic_year=academic_year,
                exam_period=exam_period,
                schedule_type=schedule_type,
                is_active=True
            )
            db.session.add(exam_schedule)
            db.session.flush()
        
        # Log activity
        log_create('exam_schedule', exam_schedule.id, f'{exam_schedule.subject.subject_code} - {exam_schedule.section.full_section_name}', {
            'exam_date': str(exam_date),
            'exam_period': exam_period
        })
        
        db.session.commit()
        
        # Broadcast exam schedule creation to all connected users
        broadcast_schedule_change(exam_schedule, 'created', 'exam')
        
        if is_ajax:
            return jsonify({
                'success': True,
                'message': 'Exam schedule added successfully!',
                'warnings': ajax_warnings,
                'section_id': section_id,
                'exam_schedule_id': exam_schedule.id
            })
        flash('Exam schedule added successfully!', 'success')
        
    except Exception as e:
        db.session.rollback()
        if is_ajax:
            return jsonify({'success': False, 'error': f'Error adding exam schedule: {str(e)}'}), 500
        flash(f'Error adding exam schedule: {str(e)}', 'danger')
    
    return redirect(url_for('schedule.create_page', type='exam', section_id=section_id))


@exam_schedule_bp.route('/edit', methods=['POST'])
@login_required
def edit():
    """Edit an existing exam schedule with optimistic locking"""
    is_ajax = request.headers.get('X-Requested-With') == 'XMLHttpRequest'
    ajax_warnings = []
    try:
        exam_schedule_id = request.form.get('exam_schedule_id', type=int)
        submitted_version = request.form.get('version', type=int)
        
        # Get exam schedule with pessimistic lock to prevent concurrent modification
        exam_schedule = ExamSchedule.query.filter_by(id=exam_schedule_id).with_for_update().first()
        
        if not exam_schedule:
            msg = 'Exam schedule not found'
            if is_ajax:
                return jsonify({'success': False, 'error': msg}), 404
            flash(msg, 'danger')
            return redirect(url_for('schedule.create_page', type='exam'))
        
        # Optimistic locking: Check if version matches
        if submitted_version is not None and exam_schedule.version != submitted_version:
            msg = 'This exam schedule was modified by another user. Please refresh and try again.'
            if is_ajax:
                return jsonify({'success': False, 'error': msg}), 409
            flash(msg, 'danger')
            return redirect(url_for('schedule.exam_edit_page', exam_id=exam_schedule_id))
        
        # Check if locked by another user
        if exam_schedule.is_locked_by_other(current_user.id):
            lock_info = exam_schedule.get_lock_info()
            msg = f'This exam schedule is being edited by {lock_info["locked_by_name"]}. Please try again later.'
            if is_ajax:
                return jsonify({'success': False, 'error': msg}), 409
            flash(msg, 'danger')
            return redirect(url_for('schedule.exam_edit_page', exam_id=exam_schedule_id))
        
        section_id = request.form.get('section_id', type=int)
        subject_id = request.form.get('subject_id', type=int)
        faculty_id = request.form.get('faculty_id', type=int)
        room_id = request.form.get('room_id', type=int)
        exam_date_str = request.form.get('exam_date')
        start_time_str = request.form.get('start_time')
        end_time_str = request.form.get('end_time')
        schedule_type = request.form.get('schedule_type', 'lecture')  # lecture or lab
        
        # Get current academic settings
        current_settings = AcademicSettings.query.filter_by(is_active=True).first()
        if not current_settings:
            msg = 'No active academic settings found. Please configure academic settings first.'
            if is_ajax:
                return jsonify({'success': False, 'error': msg}), 400
            flash(msg, 'danger')
            return redirect(url_for('schedule.exam_edit_page', exam_id=exam_schedule_id))
        
        academic_year = current_settings.academic_year
        semester = current_settings.semester
        exam_period = current_settings.exam_period
        
        # Validation - faculty and room are now required
        if not all([section_id, subject_id, faculty_id, room_id, exam_date_str, start_time_str, end_time_str]):
            msg = 'All required fields must be filled (including faculty and room)'
            if is_ajax:
                return jsonify({'success': False, 'error': msg}), 400
            flash(msg, 'danger')
            return redirect(url_for('schedule.exam_edit_page', exam_id=exam_schedule_id))
        
        # Convert date and time strings
        exam_date = datetime.strptime(exam_date_str, '%Y-%m-%d').date()
        start_time = datetime.strptime(start_time_str, '%H:%M').time()
        end_time = datetime.strptime(end_time_str, '%H:%M').time()
        
        # Validate exam date is not in the past
        from datetime import date as date_class
        today = date_class.today()
        if exam_date < today:
            msg = 'Cannot schedule exams in the past. Please select a future date.'
            if is_ajax:
                return jsonify({'success': False, 'error': msg}), 400
            flash(msg, 'danger')
            return redirect(url_for('schedule.exam_edit_page', exam_id=exam_schedule_id))
        
        # Validate exam date is within the configured exam period range
        if current_settings:
            exam_period_start = getattr(current_settings, 'exam_period_start', None)
            exam_period_end = getattr(current_settings, 'exam_period_end', None)
            
            if exam_period_start and exam_period_end:
                if exam_date < exam_period_start or exam_date > exam_period_end:
                    start_str = exam_period_start.strftime('%B %d, %Y')
                    end_str = exam_period_end.strftime('%B %d, %Y')
                    msg = f'Exam date must be within the exam period ({start_str} to {end_str})'
                    if is_ajax:
                        return jsonify({'success': False, 'error': msg}), 400
                    flash(msg, 'danger')
                    return redirect(url_for('schedule.exam_edit_page', exam_id=exam_schedule_id))
        
        # Validate time range
        if start_time >= end_time:
            msg = 'End time must be after start time'
            if is_ajax:
                return jsonify({'success': False, 'error': msg}), 400
            flash(msg, 'danger')
            return redirect(url_for('schedule.exam_edit_page', exam_id=exam_schedule_id))

        # Validate exam times are within configured exam hours
        if current_settings:
            exam_start_cfg = _coerce_setting_time(getattr(current_settings, 'exam_start_time', None), current_settings.exam_start_hour or 7)
            exam_end_cfg = _coerce_setting_time(getattr(current_settings, 'exam_end_time', None), current_settings.exam_end_hour or 17)

            if _minutes_of(start_time) < _minutes_of(exam_start_cfg):
                msg = f'Start time must be at or after {_fmt_setting_time(exam_start_cfg)}'
                if is_ajax:
                    return jsonify({'success': False, 'error': msg}), 400
                flash(msg, 'danger')
                return redirect(url_for('schedule.exam_edit_page', exam_id=exam_schedule_id))

            if _minutes_of(end_time) > _minutes_of(exam_end_cfg):
                msg = f'End time must be at or before {_fmt_setting_time(exam_end_cfg)}'
                if is_ajax:
                    return jsonify({'success': False, 'error': msg}), 400
                flash(msg, 'danger')
                return redirect(url_for('schedule.exam_edit_page', exam_id=exam_schedule_id))
        
        # Check for conflicts - same section, date, and overlapping time (excluding current exam)
        conflict_query = ExamSchedule.query.filter(
            ExamSchedule.id != exam_schedule_id,
            ExamSchedule.section_id == section_id,
            ExamSchedule.exam_date == exam_date,
            ExamSchedule.is_active == True,
            or_(
                and_(ExamSchedule.start_time <= start_time, ExamSchedule.end_time > start_time),
                and_(ExamSchedule.start_time < end_time, ExamSchedule.end_time >= end_time),
                and_(ExamSchedule.start_time >= start_time, ExamSchedule.end_time <= end_time)
            )
        )
        
        if academic_year and semester:
            conflict_query = conflict_query.filter(
                ExamSchedule.academic_year == academic_year,
                ExamSchedule.semester == semester
            )
        
        if conflict_query.first():
            msg = 'Exam schedule conflict: This section already has an exam scheduled at this time'
            if is_ajax:
                return jsonify({'success': False, 'error': msg}), 409
            flash(msg, 'danger')
            return redirect(url_for('schedule.exam_edit_page', exam_id=exam_schedule_id))
        
        # Check for faculty conflicts if faculty assigned (excluding current exam)
        if faculty_id:
            faculty_conflict = ExamSchedule.query.filter(
                ExamSchedule.id != exam_schedule_id,
                ExamSchedule.faculty_id == faculty_id,
                ExamSchedule.exam_date == exam_date,
                ExamSchedule.is_active == True,
                or_(
                    and_(ExamSchedule.start_time <= start_time, ExamSchedule.end_time > start_time),
                    and_(ExamSchedule.start_time < end_time, ExamSchedule.end_time >= end_time),
                    and_(ExamSchedule.start_time >= start_time, ExamSchedule.end_time <= end_time)
                )
            )
            
            if academic_year and semester:
                faculty_conflict = faculty_conflict.filter(
                    ExamSchedule.academic_year == academic_year,
                    ExamSchedule.semester == semester
                )
            
            if faculty_conflict.first():
                msg = 'Faculty conflict: This faculty member is already assigned to another exam at this time'
                if is_ajax:
                    return jsonify({'success': False, 'error': msg}), 409
                flash(msg, 'danger')
                return redirect(url_for('schedule.exam_edit_page', exam_id=exam_schedule_id))
        
        # Check for room conflicts if room assigned (excluding current exam)
        if room_id:
            room_conflict = ExamSchedule.query.filter(
                ExamSchedule.id != exam_schedule_id,
                ExamSchedule.room_id == room_id,
                ExamSchedule.exam_date == exam_date,
                ExamSchedule.is_active == True,
                or_(
                    and_(ExamSchedule.start_time <= start_time, ExamSchedule.end_time > start_time),
                    and_(ExamSchedule.start_time < end_time, ExamSchedule.end_time >= end_time),
                    and_(ExamSchedule.start_time >= start_time, ExamSchedule.end_time <= end_time)
                )
            )
            
            if academic_year and semester:
                room_conflict = room_conflict.filter(
                    ExamSchedule.academic_year == academic_year,
                    ExamSchedule.semester == semester
                )
            
            if room_conflict.first():
                msg = 'Room conflict: This room is already assigned to another exam at this time'
                if is_ajax:
                    return jsonify({'success': False, 'error': msg}), 409
                flash(msg, 'danger')
                return redirect(url_for('schedule.exam_edit_page', exam_id=exam_schedule_id))
        
        # Check proctor availability
        if faculty_id:
            availability_result = FacultyAvailability.check_faculty_available(
                faculty_id, exam_date, start_time, end_time
            )
            if availability_result['status'] == 'unavailable':
                faculty = Faculty.query.get(faculty_id)
                faculty_name = faculty.full_name if faculty else 'The selected faculty'
                reason = availability_result.get('reason')
                reason_text = f" Reason: {reason}" if reason else ""
                msg = f'Proctor unavailable: {faculty_name} has marked themselves as unavailable for this time slot.{reason_text}'
                if is_ajax:
                    return jsonify({'success': False, 'error': msg}), 409
                flash(msg, 'danger')
                return redirect(url_for('schedule.exam_edit_page', exam_id=exam_schedule_id))
            elif availability_result['status'] == 'not_in_schedule':
                faculty = Faculty.query.get(faculty_id)
                faculty_name = faculty.full_name if faculty else 'The selected faculty'
                warn_msg = f'Warning: {faculty_name} is not marked as available on this date/time. Exam updated anyway.'
                if is_ajax:
                    ajax_warnings.append(warn_msg)
                else:
                    flash(warn_msg, 'warning')
        
        # Check for lunch overlap and add warning (not blocking)
        if current_settings.exam_lunch_start and current_settings.exam_lunch_end:
            lunch_start = current_settings.exam_lunch_start
            lunch_end = current_settings.exam_lunch_end
            # Check overlap: exam starts before lunch ends AND exam ends after lunch starts
            start_mins = start_time.hour * 60 + start_time.minute
            end_mins = end_time.hour * 60 + end_time.minute
            lunch_start_mins = lunch_start.hour * 60 + lunch_start.minute
            lunch_end_mins = lunch_end.hour * 60 + lunch_end.minute
            
            if start_mins < lunch_end_mins and end_mins > lunch_start_mins:
                lunch_start_str = lunch_start.strftime('%I:%M %p').lstrip('0')
                lunch_end_str = lunch_end.strftime('%I:%M %p').lstrip('0')
                warn_msg = f'Note: This exam overlaps with the lunch break ({lunch_start_str} - {lunch_end_str})'
                if is_ajax:
                    ajax_warnings.append(warn_msg)
                else:
                    flash(warn_msg, 'warning')
        
        # Check exam duration limit
        exam_duration_limit = current_settings.exam_duration_limit if hasattr(current_settings, 'exam_duration_limit') else 120
        exam_duration = (end_time.hour * 60 + end_time.minute) - (start_time.hour * 60 + start_time.minute)
        if exam_duration > exam_duration_limit:
            hours = exam_duration_limit // 60
            mins = exam_duration_limit % 60
            limit_text = f'{hours} hour{"s" if hours > 1 else ""}' if mins == 0 else f'{hours}h {mins}m'
            warn_msg = f'Exam duration ({exam_duration} min) exceeds the maximum allowed ({limit_text})'
            if is_ajax:
                ajax_warnings.append(warn_msg)
            else:
                flash(warn_msg, 'warning')
        
        # Update exam schedule
        exam_schedule.section_id = section_id
        exam_schedule.subject_id = subject_id
        exam_schedule.faculty_id = faculty_id
        exam_schedule.room_id = room_id
        exam_schedule.exam_date = exam_date
        exam_schedule.start_time = start_time
        exam_schedule.end_time = end_time
        exam_schedule.semester = semester
        exam_schedule.academic_year = academic_year
        exam_schedule.exam_period = exam_period
        exam_schedule.schedule_type = schedule_type
        
        # Increment version for optimistic locking
        exam_schedule.version = (exam_schedule.version or 1) + 1
        
        # Release the edit lock
        exam_schedule.release_lock(current_user.id)
        
        # Log activity
        log_edit('exam_schedule', exam_schedule.id, f'{exam_schedule.subject.subject_code} - {exam_schedule.section.full_section_name}', {
            'exam_date': str(exam_date),
            'exam_period': exam_period
        })
        
        db.session.commit()
        
        # Broadcast exam schedule update to all connected users
        broadcast_schedule_change(exam_schedule, 'updated', 'exam')
        
        if is_ajax:
            return jsonify({
                'success': True,
                'message': 'Exam schedule updated successfully!',
                'warnings': ajax_warnings,
                'section_id': section_id,
                'exam_schedule_id': exam_schedule.id
            })
        flash('Exam schedule updated successfully!', 'success')
        
    except Exception as e:
        db.session.rollback()
        if is_ajax:
            return jsonify({'success': False, 'error': f'Error updating exam schedule: {str(e)}'}), 500
        flash(f'Error updating exam schedule: {str(e)}', 'danger')
    
    return redirect(url_for('schedule.exam_edit_page', exam_id=exam_schedule_id))


@exam_schedule_bp.route('/get/<int:exam_schedule_id>')
@login_required
def get_exam_schedule(exam_schedule_id):
    """Get exam schedule data for editing"""
    try:
        exam_schedule = ExamSchedule.query.get(exam_schedule_id)
        if not exam_schedule:
            return jsonify({'error': 'Exam schedule not found'}), 404
        
        return jsonify({
            'id': exam_schedule.id,
            'section_id': exam_schedule.section_id,
            'subject_id': exam_schedule.subject_id,
            'curriculum_id': exam_schedule.subject.semester.year_level.curriculum_id if exam_schedule.subject and exam_schedule.subject.semester and exam_schedule.subject.semester.year_level else None,
            'faculty_id': exam_schedule.faculty_id,
            'faculty_name': exam_schedule.faculty.full_name if exam_schedule.faculty else None,
            'room_id': exam_schedule.room_id,
            'room_number': exam_schedule.room.room_number if exam_schedule.room else None,
            'building_name': exam_schedule.room.building.building_name if exam_schedule.room and exam_schedule.room.building else None,
            'exam_date': exam_schedule.exam_date.strftime('%Y-%m-%d'),
            'start_time': exam_schedule.start_time.strftime('%H:%M'),
            'end_time': exam_schedule.end_time.strftime('%H:%M'),
            'semester': exam_schedule.semester,
            'academic_year': exam_schedule.academic_year,
            'exam_period': exam_schedule.exam_period,
            'schedule_type': exam_schedule.schedule_type or 'lecture',
            'version': exam_schedule.version if hasattr(exam_schedule, 'version') else None
        })
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@exam_schedule_bp.route('/section/<int:section_id>/exams')
@login_required
def get_section_exams(section_id):
    """Get all exam schedules for a section (for modal calendar)"""
    try:
        # Get current academic settings
        current_settings = AcademicSettings.query.filter_by(is_active=True).first()
        
        # Build query
        query = ExamSchedule.query.filter(
            ExamSchedule.section_id == section_id,
            ExamSchedule.is_active == True
        )
        
        # Filter by academic year and semester if set
        if current_settings:
            query = query.filter(
                ExamSchedule.academic_year == current_settings.academic_year,
                ExamSchedule.semester == current_settings.semester
            )
        
        # Order by date and time
        query = query.order_by(ExamSchedule.exam_date, ExamSchedule.start_time)
        
        exam_schedules = query.all()
        
        # Build response
        exams_data = []
        for exam in exam_schedules:
            exams_data.append({
                'id': exam.id,
                'section_id': exam.section_id,
                'subject_id': exam.subject_id,
                'subject_code': exam.subject.subject_code if exam.subject else None,
                'subject_name': exam.subject.course_description if exam.subject else None,
                'curriculum_id': exam.subject.semester.year_level.curriculum_id if exam.subject and exam.subject.semester and exam.subject.semester.year_level else None,
                'faculty_id': exam.faculty_id,
                'faculty_name': exam.faculty.full_name if exam.faculty else None,
                'room_id': exam.room_id,
                'room_number': exam.room.room_number if exam.room else None,
                'building_name': exam.room.building.building_name if exam.room and exam.room.building else None,
                'exam_date': exam.exam_date.strftime('%Y-%m-%d'),
                'start_time': exam.start_time.strftime('%H:%M'),
                'end_time': exam.end_time.strftime('%H:%M'),
                'academic_year': exam.academic_year,
                'semester': exam.semester,
                'exam_period': exam.exam_period,
                'schedule_type': exam.schedule_type or 'lecture',
                'version': exam.version if hasattr(exam, 'version') else None
            })
        
        # Get EXAM schedule hours from settings (minute-aware with legacy fallback)
        exam_start = _coerce_setting_time(getattr(current_settings, 'exam_start_time', None) if current_settings else None, 7)
        exam_end = _coerce_setting_time(getattr(current_settings, 'exam_end_time', None) if current_settings else None, 17)
        start_hour = exam_start.hour
        end_hour = exam_end.hour
        
        return jsonify({
            'exam_schedules': exams_data,
            'start_hour': start_hour,
            'end_hour': end_hour,
            'count': len(exams_data)
        })
        
    except Exception as e:
        print(f"[ERROR] get_section_exams: {str(e)}")
        return jsonify({'error': str(e), 'exam_schedules': []}), 500


@exam_schedule_bp.route('/delete', methods=['POST'])
@login_required
def delete():
    """Delete an exam schedule with concurrency check"""
    section_id = None  # Initialize to prevent UnboundLocalError
    try:
        # Frontend sends 'exam_schedule_id', not 'schedule_id'
        schedule_id = request.form.get('exam_schedule_id', type=int)
        exam_schedule = ExamSchedule.query.filter_by(id=schedule_id).with_for_update().first()
        if not exam_schedule:
            flash('Exam schedule not found', 'danger')
            return redirect(url_for('schedule.exam_view'))
        
        # Check if locked by another user
        if exam_schedule.is_locked_by_other(current_user.id):
            lock_info = exam_schedule.get_lock_info()
            flash(f'This exam schedule is being edited by {lock_info["locked_by_name"]}. Please try again later.', 'danger')
            return redirect(url_for('schedule.exam_view', section_id=exam_schedule.section_id))
        
        section_id = exam_schedule.section_id
        exam_info = f'{exam_schedule.subject.subject_code} - {exam_schedule.section.full_section_name}'
        
        # Store schedule data before deletion for broadcast
        schedule_copy = exam_schedule.to_dict()
        
        # Log activity before deletion
        log_delete('exam_schedule', exam_schedule.id, exam_info, {'exam_date': str(exam_schedule.exam_date)})
        
        db.session.delete(exam_schedule)
        db.session.commit()
        
        # Broadcast exam schedule deletion to all connected users
        # Create a simple object for broadcast since exam_schedule is deleted
        class DeletedSchedule:
            def __init__(self, data):
                self.id = data.get('id')
                self.academic_year = data.get('academic_year')
                self.semester = data.get('semester')
            def to_dict(self):
                return {'id': self.id, 'deleted': True}
        
        deleted_obj = DeletedSchedule(schedule_copy)
        broadcast_schedule_change(deleted_obj, 'deleted', 'exam')
        
        flash('Exam schedule deleted successfully!', 'success')
        
    except Exception as e:
        db.session.rollback()
        flash(f'Error deleting exam schedule: {str(e)}', 'danger')
    
    # Use section_id if available, otherwise redirect without it
    if section_id:
        return redirect(url_for('schedule.exam_view', section_id=section_id))
    else:
        return redirect(url_for('schedule.exam_view'))


@exam_schedule_bp.route('/delete-ajax', methods=['POST'])
@login_required
@csrf.exempt
def delete_ajax():
    """Delete an exam schedule via AJAX and return JSON response."""
    try:
        data = request.get_json()
        schedule_id = data.get('exam_schedule_id')

        if not schedule_id:
            return jsonify({'success': False, 'error': 'Exam schedule ID is required'}), 400

        exam_schedule = ExamSchedule.query.filter_by(id=schedule_id).with_for_update().first_or_404()
        section_id = exam_schedule.section_id

        # Check if locked by another user
        if exam_schedule.is_locked_by_other(current_user.id):
            lock_info = exam_schedule.get_lock_info()
            return jsonify({
                'success': False,
                'error': f'This exam schedule is being edited by {lock_info["locked_by_name"]}. Please try again later.'
            }), 409

        # Store data before deletion for broadcast
        schedule_copy = exam_schedule.to_dict()
        exam_info = f'{exam_schedule.subject.subject_code} - {exam_schedule.section.full_section_name}'

        # Log activity before deletion
        log_delete('exam_schedule', exam_schedule.id, exam_info, {'exam_date': str(exam_schedule.exam_date)})

        db.session.delete(exam_schedule)
        db.session.commit()

        # Broadcast deletion
        class DeletedSchedule:
            def __init__(self, d):
                self.id = d.get('id')
                self.academic_year = d.get('academic_year')
                self.semester = d.get('semester')
            def to_dict(self):
                return {'id': self.id, 'deleted': True}

        broadcast_schedule_change(DeletedSchedule(schedule_copy), 'deleted', 'exam')

        return jsonify({
            'success': True,
            'message': 'Exam schedule deleted successfully',
            'section_id': section_id
        })

    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'error': str(e)}), 500


@exam_schedule_bp.route('/batch-delete', methods=['POST'])
@login_required
@csrf.exempt
def batch_delete():
    """Delete multiple exam schedules at once via AJAX."""
    try:
        data = request.get_json()
        schedule_ids = data.get('schedule_ids', [])

        if not schedule_ids or not isinstance(schedule_ids, list):
            return jsonify({'success': False, 'error': 'No exam schedules selected'}), 400

        if len(schedule_ids) > 100:
            return jsonify({'success': False, 'error': 'Cannot delete more than 100 exam schedules at once'}), 400

        exams = ExamSchedule.query.filter(
            ExamSchedule.id.in_(schedule_ids),
            ExamSchedule.is_active == True
        ).all()

        if not exams:
            return jsonify({'success': False, 'error': 'No active exam schedules found for the given IDs'}), 404

        # Program access control for deans
        user_program_ids = current_user.get_program_ids()
        if user_program_ids is not None:
            exams = [e for e in exams if e.section and e.section.program_id in user_program_ids]

        # Check locks
        locked = [e for e in exams if e.is_locked_by_other(current_user.id)]
        if locked:
            return jsonify({
                'success': False,
                'error': f'{len(locked)} exam schedule(s) are locked by other users. Please try again later.'
            }), 409

        deleted_count = 0
        section_id = exams[0].section_id if exams else None
        broadcast_list = []
        for exam in exams:
            schedule_copy = exam.to_dict()
            exam_info = f'{exam.subject.subject_code} - {exam.section.full_section_name}' if exam.subject and exam.section else 'N/A'
            log_delete('exam_schedule', exam.id, exam_info, {'exam_date': str(exam.exam_date)})
            broadcast_list.append(schedule_copy)
            db.session.delete(exam)
            deleted_count += 1

        db.session.commit()

        # Broadcast deletions
        for sc in broadcast_list:
            class DeletedSchedule:
                def __init__(self, d):
                    self.id = d.get('id')
                    self.academic_year = d.get('academic_year')
                    self.semester = d.get('semester')
                def to_dict(self):
                    return {'id': self.id, 'deleted': True}
            broadcast_schedule_change(DeletedSchedule(sc), 'deleted', 'exam')

        return jsonify({
            'success': True,
            'message': f'Successfully deleted {deleted_count} exam schedule(s)',
            'deleted_count': deleted_count,
            'section_id': section_id
        })

    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'error': str(e)}), 500


@exam_schedule_bp.route('/ai-check-conflicts', methods=['POST'])
@login_required
@csrf.exempt  # Exempt CSRF for AJAX endpoints
def ai_check_exam_conflicts():
    """
    AI-powered conflict detection for exam schedules
    
    Uses new service layer architecture:
    - ConflictDetector: Fast pure Python conflict checking
    - RecommendationEngine: Smart suggestions
    - AISchedulerAssistant: Gemini AI explanations
    """
    from app.ai_scheduler import ai_scheduler
    from datetime import datetime as dt
    
    try:
        data = request.get_json()
        
        # Debug logging
        print(f"[AI CHECK EXAM] Received data: {data}")
        
        if not data:
            return jsonify({'error': 'No data received', 'ai_enabled': False}), 400
        
        # Parse exam schedule data
        section_id = data.get('section_id')
        subject_id = data.get('subject_id')
        faculty_id = data.get('faculty_id')
        room_id = data.get('room_id')
        exam_date_str = data.get('exam_date')
        start_time_str = data.get('start_time')
        end_time_str = data.get('end_time')
        exam_schedule_id = data.get('exam_schedule_id')  # For edit mode
        
        # Debug logging
        print(f"[AI CHECK EXAM] Parsed - section:{section_id} date:{exam_date_str} time:{start_time_str}-{end_time_str}")
        
        if not all([section_id, exam_date_str, start_time_str, end_time_str]):
            missing = []
            if not section_id: missing.append('section_id')
            if not exam_date_str: missing.append('exam_date')
            if not start_time_str: missing.append('start_time')
            if not end_time_str: missing.append('end_time')
            error_msg = f'Missing required fields: {", ".join(missing)}'
            print(f"[AI CHECK EXAM] Validation failed: {error_msg}")
            return jsonify({'error': error_msg, 'ai_enabled': False}), 400
        
        # Convert times and date
        start_time = dt.strptime(start_time_str, '%H:%M').time()
        end_time = dt.strptime(end_time_str, '%H:%M').time()
        exam_date = dt.strptime(exam_date_str, '%Y-%m-%d').date()
        
        # Get current academic settings
        current_settings = AcademicSettings.query.filter_by(is_active=True).first()
        
        # Check if exam times are outside configured hours (non-blocking warning)
        schedule_hours_warning = None
        if current_settings:
            exam_start_cfg = _coerce_setting_time(getattr(current_settings, 'exam_start_time', None), current_settings.exam_start_hour or 7)
            exam_end_cfg = _coerce_setting_time(getattr(current_settings, 'exam_end_time', None), current_settings.exam_end_hour or 17)

            if _minutes_of(start_time) < _minutes_of(exam_start_cfg):
                schedule_hours_warning = f'Start time is before configured exam hours ({_fmt_setting_time(exam_start_cfg)})'

            elif _minutes_of(end_time) > _minutes_of(exam_end_cfg):
                schedule_hours_warning = f'End time is after configured exam hours ({_fmt_setting_time(exam_end_cfg)})'
        
        # Get existing exam schedules for the same academic period
        existing_query = ExamSchedule.query.filter_by(is_active=True)
        
        if current_settings:
            existing_query = existing_query.filter_by(
                academic_year=current_settings.academic_year,
                semester=current_settings.semester,
                exam_period=current_settings.exam_period
            )
        
        existing_exams = existing_query.all()
        
        # Prepare exam schedule data for analysis
        exam_data = {
            'section_id': section_id,
            'subject_id': subject_id,
            'faculty_id': faculty_id,
            'room_id': room_id,
            'exam_date': exam_date,
            'start_time': start_time,
            'end_time': end_time,
            'schedule_type': data.get('schedule_type', 'lecture'),
            'is_exam': True
        }
        
        # Check if client wants full AI or offline-only conflict detection
        use_ai = data.get('use_ai', True)
        
        if use_ai:
            # Full AI analysis (conflict detection + recommendations + Gemini explanation)
            analysis = ai_scheduler.analyze_exam_conflicts(
                exam_data, 
                existing_exams,
                exclude_exam_id=int(exam_schedule_id) if exam_schedule_id else None
            )
        else:
            # Basic mode: rule-based conflict detection + recommendations (no Gemini AI)
            from app.services.conflict_detector import conflict_detector
            from app.services.recommendation_engine import recommendation_engine
            conflicts = conflict_detector.detect_exam_conflicts(
                exam_data,
                existing_exams,
                int(exam_schedule_id) if exam_schedule_id else None
            )
            recommendations = []
            offline_explanation = ''
            if conflicts:
                recommendations = recommendation_engine.generate_exam_recommendations(
                    exam_data, conflicts, existing_exams
                )
                offline_explanation = ai_scheduler._get_offline_explanation(conflicts, recommendations, is_exam=True)
            analysis = {
                'has_conflicts': len(conflicts) > 0,
                'conflicts': [c.to_dict() for c in conflicts],
                'recommendations': [r.to_dict() for r in recommendations],
                'ai_explanation': offline_explanation,
                'ai_enabled': False,
                'ai_fallback': False,
                'ai_fallback_reason': None
            }
        
        # Check faculty (proctor) availability (warning, not a hard conflict)
        faculty_availability_warning = None
        if faculty_id:
            # Get day of week from exam date
            day_names = ['Monday', 'Tuesday', 'Wednesday', 'Thursday', 'Friday', 'Saturday', 'Sunday']
            exam_day_of_week = day_names[exam_date.weekday()]
            
            availability_result = FacultyAvailability.check_faculty_available_by_day(
                faculty_id, exam_day_of_week, start_time, end_time
            )
            
            faculty = Faculty.query.get(faculty_id)
            faculty_name = faculty.full_name if faculty else 'Selected proctor'
            
            if availability_result['status'] == 'unavailable':
                # Hard conflict - faculty explicitly unavailable
                reason = availability_result.get('reason') or 'marked as unavailable'
                faculty_availability_warning = {
                    'type': 'error',
                    'message': f'{faculty_name} is {reason} for this time slot.',
                    'faculty_name': faculty_name,
                    'status': 'unavailable'
                }
            elif availability_result['status'] == 'not_in_schedule':
                # Soft warning - faculty has defined availability but not for this slot
                faculty_availability_warning = {
                    'type': 'warning',
                    'message': f'{faculty_name} is not marked as available on {exam_day_of_week} at this time.',
                    'faculty_name': faculty_name,
                    'status': 'not_in_schedule'
                }
            elif availability_result['status'] in ('available', 'preferred'):
                # Positive confirmation - proctor is available for this slot
                faculty_availability_warning = {
                    'type': 'success',
                    'message': f'{faculty_name} is available on {exam_day_of_week} at this time.',
                    'faculty_name': faculty_name,
                    'status': availability_result['status']
                }
        
        # Response already formatted by service layer
        response = {
            'ai_enabled': analysis.get('ai_enabled', False),
            'ai_fallback': analysis.get('ai_fallback', False),
            'ai_fallback_reason': analysis.get('ai_fallback_reason'),
            'ai_fallback_message': analysis.get('ai_fallback_reason') if analysis.get('ai_fallback', False) else '',
            'has_conflicts': analysis.get('has_conflicts', False),
            'conflicts': analysis.get('conflicts', []),
            'recommendations': analysis.get('recommendations', []),
            'ai_explanation': analysis.get('ai_explanation', ''),
            'schedule_hours_warning': schedule_hours_warning,
            'faculty_availability_warning': faculty_availability_warning
        }
        
        return jsonify(response)
        
    except Exception as e:
        print(f"AI check exam conflicts error: {str(e)}")
        import traceback
        traceback.print_exc()
        return jsonify({'error': str(e), 'ai_enabled': False}), 500


@exam_schedule_bp.route('/resolve-conflicts', methods=['POST'])
@login_required
@csrf.exempt
def resolve_exam_conflicts():
    """
    Generate a resolution plan for detected exam schedule conflicts.
    
    Uses ConflictResolver to find optimal form field changes that
    eliminate all conflicts. Returns a plan for user confirmation.
    """
    from app.services.conflict_resolver import conflict_resolver

    try:
        data = request.get_json()
        if not data:
            return jsonify({'error': 'No data received'}), 400

        section_id = data.get('section_id')
        subject_id = data.get('subject_id')
        faculty_id = data.get('faculty_id')
        room_id = data.get('room_id')
        exam_date_str = data.get('exam_date')
        start_time_str = data.get('start_time')
        end_time_str = data.get('end_time')
        exam_schedule_id = data.get('exam_schedule_id')
        conflicts = data.get('conflicts', [])

        if not all([section_id, exam_date_str, start_time_str, end_time_str]):
            return jsonify({'error': 'Missing required exam schedule fields'}), 400

        if not conflicts:
            return jsonify({'error': 'No conflicts to resolve'}), 400

        start_time = datetime.strptime(start_time_str, '%H:%M').time()
        end_time = datetime.strptime(end_time_str, '%H:%M').time()
        exam_date = datetime.strptime(exam_date_str, '%Y-%m-%d').date()

        current_settings = AcademicSettings.query.filter_by(is_active=True).first()
        existing_query = ExamSchedule.query.filter_by(is_active=True)
        if current_settings:
            existing_query = existing_query.filter_by(
                academic_year=current_settings.academic_year,
                semester=current_settings.semester,
                exam_period=current_settings.exam_period
            )
        existing_exams = existing_query.all()

        subject_obj = None
        if subject_id:
            subject_obj = Subject.query.get(subject_id)

        exam_data = {
            'section_id': section_id,
            'subject_id': subject_id,
            'faculty_id': faculty_id,
            'room_id': room_id,
            'exam_date': exam_date,
            'start_time': start_time,
            'end_time': end_time
        }

        exclude_id = int(exam_schedule_id) if exam_schedule_id else None

        plan = conflict_resolver.generate_exam_resolution_plan(
            exam_data=exam_data,
            conflicts=conflicts,
            existing_exams=existing_exams,
            subject=subject_obj,
            exclude_exam_id=exclude_id
        )

        return jsonify(plan)

    except Exception as e:
        print(f"Resolve exam conflicts error: {str(e)}")
        import traceback
        traceback.print_exc()
        return jsonify({'error': str(e)}), 500


@exam_schedule_bp.route('/batch-export/excel')
@login_required
def batch_export_excel():
    """Export all exam schedules for current academic period to Excel"""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    import io
    from flask import send_file
    
    try:
        # Get current academic settings
        current_settings = AcademicSettings.query.filter_by(is_active=True).first()
        if not current_settings:
            flash('No active academic settings found.', 'danger')
            return redirect(url_for('schedule.exam_view'))
        
        # Get user's program access
        user_program_ids = current_user.get_program_ids()
        
        # Build query for exam schedules
        query = ExamSchedule.query.filter_by(
            academic_year=current_settings.academic_year,
            semester=current_settings.semester,
            exam_period=current_settings.exam_period,
            is_active=True
        )
        
        # Filter by user's program access if Dean
        if user_program_ids is not None:
            query = query.join(Section).filter(Section.program_id.in_(user_program_ids))
        
        exam_schedules = query.order_by(ExamSchedule.exam_date, ExamSchedule.start_time).all()
        
        if not exam_schedules:
            flash('No exam schedules found for the current academic period.', 'info')
            return redirect(url_for('schedule.exam_view'))
        
        # Get exam period date range from settings
        exam_period_dates = ""
        if hasattr(current_settings, 'exam_period_start') and current_settings.exam_period_start:
            if hasattr(current_settings, 'exam_period_end') and current_settings.exam_period_end:
                exam_period_dates = f"({current_settings.exam_period_start.strftime('%B %d')} - {current_settings.exam_period_end.strftime('%B %d, %Y')})"
        
        # Create workbook
        wb = Workbook()
        ws = wb.active
        ws.title = 'Exam Schedules'
        
        # Add header
        ws['A1'] = 'EXAM SCHEDULES - BATCH EXPORT'
        ws['A1'].font = Font(bold=True, size=14)
        ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
        ws.merge_cells('A1:J1')
        
        ws['A2'] = f'{current_settings.semester} - {current_settings.academic_year}'
        ws['A2'].font = Font(bold=True, size=12)
        ws['A2'].alignment = Alignment(horizontal='center', vertical='center')
        ws.merge_cells('A2:J2')
        
        exam_period_text = f'Exam Period: {current_settings.exam_period}'
        if exam_period_dates:
            exam_period_text += f' {exam_period_dates}'
        ws['A3'] = exam_period_text
        ws['A3'].font = Font(bold=True, size=11)
        ws['A3'].alignment = Alignment(horizontal='center', vertical='center')
        ws.merge_cells('A3:J3')
        
        # Column headers
        headers = ['Section', 'Subject Code', 'Subject', 'Faculty', 'Room', 'Exam Date', 'Start Time', 'End Time', 'Program', 'Year Level']
        header_fill = PatternFill(start_color='2563eb', end_color='2563eb', fill_type='solid')
        header_font = Font(bold=True, color='FFFFFF', size=11)
        header_alignment = Alignment(horizontal='center', vertical='center')
        
        for col_idx, header in enumerate(headers, start=1):
            cell = ws.cell(row=5, column=col_idx, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment
        
        # Data rows
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        for row_idx, exam in enumerate(exam_schedules, start=6):
            ws.cell(row=row_idx, column=1, value=exam.section.section_name if exam.section else 'N/A')
            ws.cell(row=row_idx, column=2, value=exam.subject.subject_code if exam.subject else 'N/A')
            ws.cell(row=row_idx, column=3, value=exam.subject.course_description if exam.subject else 'N/A')
            ws.cell(row=row_idx, column=4, value=exam.faculty.full_name if exam.faculty else 'TBA')
            ws.cell(row=row_idx, column=5, value=exam.room.room_number if exam.room else 'TBA')
            ws.cell(row=row_idx, column=6, value=exam.exam_date.strftime('%B %d, %Y') if exam.exam_date else 'N/A')
            ws.cell(row=row_idx, column=7, value=exam.start_time.strftime('%I:%M %p') if exam.start_time else 'N/A')
            ws.cell(row=row_idx, column=8, value=exam.end_time.strftime('%I:%M %p') if exam.end_time else 'N/A')
            ws.cell(row=row_idx, column=9, value=exam.section.program.program_name if exam.section and exam.section.program else 'N/A')
            ws.cell(row=row_idx, column=10, value=str(exam.section.year_level) if exam.section else 'N/A')
            
            # Apply borders and alignment
            for col in range(1, 11):
                cell = ws.cell(row=row_idx, column=col)
                cell.border = thin_border
                cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # Set column widths
        ws.column_dimensions['A'].width = 15
        ws.column_dimensions['B'].width = 12
        ws.column_dimensions['C'].width = 35
        ws.column_dimensions['D'].width = 25
        ws.column_dimensions['E'].width = 12
        ws.column_dimensions['F'].width = 18
        ws.column_dimensions['G'].width = 12
        ws.column_dimensions['H'].width = 12
        ws.column_dimensions['I'].width = 25
        ws.column_dimensions['J'].width = 10
        
        # Save to BytesIO
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        filename = f'Exam_Schedules_Batch_{current_settings.semester}_{current_settings.academic_year.replace("/", "-")}.xlsx'
        
        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=filename
        )
        
    except Exception as e:
        print(f"Batch export error: {str(e)}")
        import traceback
        traceback.print_exc()
        flash(f'Error generating batch export: {str(e)}', 'danger')
        return redirect(url_for('schedule.exam_view'))


@exam_schedule_bp.route('/batch-export/pdf')
@login_required
def batch_export_pdf():
    """Export all exam schedules for current academic period to PDF"""
    from reportlab.lib.pagesizes import letter, landscape
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import inch
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib import colors as rl_colors
    from reportlab.lib.enums import TA_CENTER
    import io
    from flask import send_file
    
    try:
        # Get current academic settings
        current_settings = AcademicSettings.query.filter_by(is_active=True).first()
        if not current_settings:
            flash('No active academic settings found.', 'danger')
            return redirect(url_for('schedule.exam_view'))
        
        # Get user's program access
        user_program_ids = current_user.get_program_ids()
        
        # Build query for exam schedules
        query = ExamSchedule.query.filter_by(
            academic_year=current_settings.academic_year,
            semester=current_settings.semester,
            exam_period=current_settings.exam_period,
            is_active=True
        )
        
        # Filter by user's program access if Dean
        if user_program_ids is not None:
            query = query.join(Section).filter(Section.program_id.in_(user_program_ids))
        
        exam_schedules = query.order_by(ExamSchedule.exam_date, ExamSchedule.start_time).all()
        
        if not exam_schedules:
            flash('No exam schedules found for the current academic period.', 'info')
            return redirect(url_for('schedule.exam_view'))
        
        # Create PDF
        output = io.BytesIO()
        doc = SimpleDocTemplate(
            output,
            pagesize=landscape(letter),
            rightMargin=0.5*inch,
            leftMargin=0.5*inch,
            topMargin=0.5*inch,
            bottomMargin=0.5*inch
        )
        
        story = []
        styles = getSampleStyleSheet()
        
        # Custom styles
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Normal'],
            fontSize=14,
            alignment=TA_CENTER,
            fontName='Helvetica-Bold'
        )
        
        subtitle_style = ParagraphStyle(
            'CustomSubtitle',
            parent=styles['Normal'],
            fontSize=12,
            alignment=TA_CENTER,
            fontName='Helvetica-Bold'
        )
        
        # Get exam period date range from settings
        exam_period_dates = ""
        if hasattr(current_settings, 'exam_period_start') and current_settings.exam_period_start:
            if hasattr(current_settings, 'exam_period_end') and current_settings.exam_period_end:
                exam_period_dates = f"({current_settings.exam_period_start.strftime('%B %d')} - {current_settings.exam_period_end.strftime('%B %d, %Y')})"
        
        # Add title
        story.append(Paragraph('EXAM SCHEDULES - BATCH EXPORT', title_style))
        story.append(Paragraph(f'{current_settings.semester} - {current_settings.academic_year}', subtitle_style))
        exam_period_text = f'Exam Period: {current_settings.exam_period}'
        if exam_period_dates:
            exam_period_text += f' {exam_period_dates}'
        story.append(Paragraph(exam_period_text, subtitle_style))
        story.append(Spacer(1, 0.3*inch))
        
        # Prepare table data
        table_data = [['Section', 'Subject', 'Faculty', 'Room', 'Exam Date', 'Time']]
        
        for exam in exam_schedules:
            section_display = exam.section.section_name if exam.section else 'N/A'
            subject_display = f"{exam.subject.subject_code}\n{exam.subject.course_description[:30]}..." if exam.subject else 'N/A'
            faculty_display = exam.faculty.full_name if exam.faculty else 'TBA'
            room_display = exam.room.room_number if exam.room else 'TBA'
            date_display = exam.exam_date.strftime('%m/%d/%Y') if exam.exam_date else 'N/A'
            time_display = f"{exam.start_time.strftime('%I:%M %p')}-{exam.end_time.strftime('%I:%M %p')}" if exam.start_time and exam.end_time else 'N/A'
            
            table_data.append([
                section_display,
                subject_display,
                faculty_display,
                room_display,
                date_display,
                time_display
            ])
        
        # Create table
        col_widths = [1*inch, 2.5*inch, 1.8*inch, 1*inch, 1*inch, 1.5*inch]
        table = Table(table_data, colWidths=col_widths, repeatRows=1)
        
        # Style table
        table.setStyle(TableStyle([
            # Header
            ('BACKGROUND', (0, 0), (-1, 0), rl_colors.HexColor('#2563eb')),
            ('TEXTCOLOR', (0, 0), (-1, 0), rl_colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 10),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 8),
            
            # Data cells
            ('ALIGN', (0, 1), (-1, -1), 'CENTER'),
            ('VALIGN', (0, 1), (-1, -1), 'MIDDLE'),
            ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 1), (-1, -1), 8),
            
            # Grid
            ('GRID', (0, 0), (-1, -1), 0.5, rl_colors.grey),
            ('BOX', (0, 0), (-1, -1), 1, rl_colors.black),
            
            # Alternating rows
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [rl_colors.white, rl_colors.HexColor('#f9fafb')]),
        ]))
        
        story.append(table)
        
        # Build PDF
        doc.build(story)
        output.seek(0)
        
        filename = f'Exam_Schedules_Batch_{current_settings.semester}_{current_settings.academic_year.replace("/", "-")}.pdf'
        
        return send_file(
            output,
            mimetype='application/pdf',
            as_attachment=True,
            download_name=filename
        )
        
    except Exception as e:
        print(f"Batch PDF export error: {str(e)}")
        import traceback
        traceback.print_exc()
        flash(f'Error generating batch PDF export: {str(e)}', 'danger')
        return redirect(url_for('schedule.exam_view'))


@exam_schedule_bp.route('/export-for-posting/<int:section_id>')
@login_required
def export_for_posting(section_id):
    """Batch export - Export all exam schedules for the entire program in weekly calendar format"""
    from flask import send_file
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.drawing.image import Image as ExcelImage
    from openpyxl.worksheet.page import PageMargins
    from datetime import datetime, time as dt_time, timedelta
    from collections import defaultdict
    import io
    import os
    
    try:
        section = Section.query.get_or_404(section_id)
        
        # Check if section's program is archived
        if section.program and section.program.is_archived:
            flash('Cannot export archived section exam schedules.', 'error')
            return redirect(url_for('schedule.exam_view'))
        
        # Get current academic settings
        current_settings = AcademicSettings.query.filter_by(is_active=True).first()
        
        # Get all sections for this program
        program_id = section.program_id
        if not program_id:
            flash('Section must belong to a program for batch export', 'error')
            return redirect(url_for('schedule.exam_view', section_id=section_id))
        
        # Get ALL sections for this program
        all_sections = Section.query.filter_by(
            program_id=program_id
        ).order_by(Section.year_level, Section.section_name).all()
        
        # Create workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Exam Schedule"
        
        # Set page layout for Legal paper (8.5" x 14") - Landscape
        ws.page_setup.paperSize = ws.PAPERSIZE_LEGAL
        ws.page_setup.orientation = 'landscape'
        ws.page_setup.fitToPage = True
        ws.page_setup.fitToWidth = 1
        ws.page_setup.fitToHeight = 0  # Allow multiple pages vertically
        ws.page_margins = PageMargins(left=0.25, right=0.25, top=0.5, bottom=0.5)
        ws.print_options.horizontalCentered = True
        
        # Get program info
        program = section.program
        dept_code = program.program_code if program else ''
        dept_display_name = (program.department.department_name if (program and program.department) else (program.program_name if program else 'DEPT'))
        
        # === LOGOS AND HEADER ===
        base_dir = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
        images_dir = os.path.join(base_dir, 'static', 'images')
        
        # Set column A width wider to fit the logo without overlapping column B
        ws.column_dimensions['A'].width = 12  # Wider to contain logo
        
        # Left logo (Norzagaray College) - position at A1
        logo_left_path = os.path.join(images_dir, 'norzagaray-college-logo.png')
        if os.path.exists(logo_left_path):
            img_left = ExcelImage(logo_left_path)
            img_left.width = 75
            img_left.height = 75
            ws.add_image(img_left, 'A1')
        
        # Right logo (Bagong Pilipinas) - position at M1 (last column)
        logo_right_path = os.path.join(images_dir, 'bagong-pilipinas.png')
        if os.path.exists(logo_right_path):
            img_right = ExcelImage(logo_right_path)
            img_right.width = 80
            img_right.height = 80
            ws.add_image(img_right, 'M1')
        
        # Header text - LEFT ALIGNED at column B
        ws['B1'] = 'Republic of the Philippines'
        ws['B1'].font = Font(size=11)
        ws['B1'].alignment = Alignment(horizontal='left', vertical='center')
        
        ws['B2'] = 'Municipality of Norzagaray'
        ws['B2'].font = Font(size=11)
        ws['B2'].alignment = Alignment(horizontal='left', vertical='center')
        
        ws['B3'] = 'NORZAGARAY COLLEGE'
        ws['B3'].font = Font(bold=True, size=11)
        ws['B3'].alignment = Alignment(horizontal='left', vertical='center')
        
        ws['B4'] = dept_display_name
        ws['B4'].font = Font(bold=True, size=11)
        ws['B4'].alignment = Alignment(horizontal='left', vertical='center')
        
        # Set row heights for header area
        ws.row_dimensions[1].height = 18
        ws.row_dimensions[2].height = 15
        ws.row_dimensions[3].height = 15
        ws.row_dimensions[4].height = 15
        ws.row_dimensions[5].height = 8  # Spacer
        
        # Get exam period title
        if current_settings:
            exam_period_title = f"{current_settings.exam_period.upper()} EXAMINATION"
        else:
            exam_period_title = "EXAMINATION"
        
        # Row 6: PRELIMINARY EXAMINATION (centered across A-M)
        ws.merge_cells('A6:M6')
        cell = ws.cell(row=6, column=1, value=exam_period_title)
        cell.font = Font(bold=True, size=12)
        cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # Define styles
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        yellow_fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
        light_yellow_fill = PatternFill(start_color='FFFFCC', end_color='FFFFCC', fill_type='solid')
        header_fill = PatternFill(start_color='F5F5F5', end_color='F5F5F5', fill_type='solid')
        
        # Set column widths - Optimized for Legal paper (8.5" x 14") landscape
        # Note: Column A width already set above for logo
        ws.column_dimensions['B'].width = 10  # Monday Subject
        ws.column_dimensions['C'].width = 10  # Monday Proctor
        ws.column_dimensions['D'].width = 10  # Tuesday Subject
        ws.column_dimensions['E'].width = 10  # Tuesday Proctor
        ws.column_dimensions['F'].width = 10  # Wednesday Subject
        ws.column_dimensions['G'].width = 10  # Wednesday Proctor
        ws.column_dimensions['H'].width = 10  # Thursday Subject
        ws.column_dimensions['I'].width = 10  # Thursday Proctor
        ws.column_dimensions['J'].width = 10  # Friday Subject
        ws.column_dimensions['K'].width = 10  # Friday Proctor
        ws.column_dimensions['L'].width = 10  # Saturday Subject
        ws.column_dimensions['M'].width = 12  # Saturday Proctor - wider for right logo
        
        # Days of week mapping - each day has 2 columns (Subject at col, Proctor at col+1)
        days_order = AcademicSettings.get_active_operation_days()
        # Column mapping: dynamically compute start column for each day (2 columns per day starting at col 2)
        day_start_col = {day: 2 + idx * 2 for idx, day in enumerate(days_order)}
        
        # Get exam period date range from settings
        exam_period_start = None
        exam_period_end = None
        if current_settings:
            if hasattr(current_settings, 'exam_period_start') and current_settings.exam_period_start:
                exam_period_start = current_settings.exam_period_start
            if hasattr(current_settings, 'exam_period_end') and current_settings.exam_period_end:
                exam_period_end = current_settings.exam_period_end
        
        # Use configured exam period dates for the header if available
        if exam_period_start and exam_period_end:
            date_range_str = f"{exam_period_start.strftime('%B %d')}-{exam_period_end.strftime('%d, %Y')}"
            # Map the configured date range to days of the week
            configured_date_by_day = {}
            current_date = exam_period_start
            while current_date <= exam_period_end:
                day_name = current_date.strftime('%A')
                if day_name != 'Sunday':  # Skip Sundays
                    configured_date_by_day[day_name] = current_date
                current_date += timedelta(days=1)
        else:
            date_range_str = ""
            configured_date_by_day = {}
        
        # Start row for content (after header)
        current_row = 8
        
        # Loop through each section
        for sect in all_sections:
            # Query exam schedules for this specific section
            query = ExamSchedule.query.filter_by(section_id=sect.id, is_active=True)\
                .outerjoin(Faculty, ExamSchedule.faculty_id == Faculty.id)\
                .outerjoin(Room, ExamSchedule.room_id == Room.id)\
                .outerjoin(Building, Room.building_id == Building.id)\
                .filter(
                    or_(ExamSchedule.faculty_id == None,
                        and_(Faculty.is_active == True, Faculty.is_archived == False)),
                    or_(ExamSchedule.room_id == None,
                        and_(Room.is_available == True,
                            or_(Building.id == None,
                                and_(Building.is_active == True, Building.is_archived == False))))
                )
            if current_settings:
                query = query.filter(
                    ExamSchedule.academic_year == current_settings.academic_year,
                    ExamSchedule.semester == current_settings.semester,
                    ExamSchedule.exam_period == current_settings.exam_period
                )
            
            exam_schedules = query.order_by(
                ExamSchedule.exam_date,
                ExamSchedule.start_time
            ).all()
            
            # Skip sections with no exam schedules
            if not exam_schedules:
                continue
            
            # Use configured exam period dates if available, otherwise derive from actual exams
            if configured_date_by_day:
                date_by_day = configured_date_by_day.copy()
            else:
                # Find date range from actual exam schedules
                all_dates = sorted(set(exam.exam_date for exam in exam_schedules))
                date_by_day = {}
                for d in all_dates:
                    day_name = d.strftime('%A')
                    date_by_day[day_name] = d
            
            # Get section display name
            if hasattr(sect, 'full_section_name'):
                section_display = sect.full_section_name
            else:
                section_display = f"{dept_code}-{sect.year_level}{sect.section_name}"
            
            # Row 7: Date range (centered) - only set once for first section
            if current_row == 8:
                ws.merge_cells('A7:M7')
                cell = ws.cell(row=7, column=1, value=date_range_str)
                cell.font = Font(size=11)
                cell.alignment = Alignment(horizontal='center', vertical='center')
            
            # === SECTION HEADER (Yellow background) ===
            ws.merge_cells(f'A{current_row}:M{current_row}')
            cell = ws.cell(row=current_row, column=1, value=section_display)
            cell.font = Font(bold=True, size=11)
            cell.alignment = Alignment(horizontal='left', vertical='center')
            cell.fill = yellow_fill
            for col in range(1, 14):
                ws.cell(row=current_row, column=col).border = thin_border
            current_row += 1
            
            # === DAY ROW (merged cells for each day spanning Subject+Proctor) ===
            cell = ws.cell(row=current_row, column=1, value='Day')
            cell.font = Font(bold=True, size=9)
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = thin_border
            cell.fill = header_fill
            
            for day in days_order:
                col = day_start_col[day]
                ws.merge_cells(start_row=current_row, start_column=col, end_row=current_row, end_column=col+1)
                cell = ws.cell(row=current_row, column=col, value=day)
                cell.font = Font(bold=True, size=9)
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.border = thin_border
                ws.cell(row=current_row, column=col+1).border = thin_border
            current_row += 1
            
            # === DATE ROW (merged cells for each day) ===
            cell = ws.cell(row=current_row, column=1, value='Date')
            cell.font = Font(bold=True, size=9)
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = thin_border
            cell.fill = header_fill
            
            for day in days_order:
                col = day_start_col[day]
                ws.merge_cells(start_row=current_row, start_column=col, end_row=current_row, end_column=col+1)
                if day in date_by_day:
                    date_str = date_by_day[day].strftime('%B %d')
                else:
                    date_str = ''
                cell = ws.cell(row=current_row, column=col, value=date_str)
                cell.font = Font(size=9)
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.border = thin_border
                ws.cell(row=current_row, column=col+1).border = thin_border
            current_row += 1
            
            # === ROOM ROW (merged cells for each day) ===
            cell = ws.cell(row=current_row, column=1, value='Room')
            cell.font = Font(bold=True, size=9)
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = thin_border
            cell.fill = header_fill
            
            # Get rooms by day
            rooms_by_day = defaultdict(set)
            for exam in exam_schedules:
                day_name = exam.exam_date.strftime('%A')
                room_name = exam.room.room_number if exam.room else ''
                if room_name:
                    rooms_by_day[day_name].add(room_name)
            
            for day in days_order:
                col = day_start_col[day]
                ws.merge_cells(start_row=current_row, start_column=col, end_row=current_row, end_column=col+1)
                rooms = ', '.join(sorted(rooms_by_day[day])) if rooms_by_day[day] else ''
                cell = ws.cell(row=current_row, column=col, value=rooms)
                cell.font = Font(size=9)
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.border = thin_border
                ws.cell(row=current_row, column=col+1).border = thin_border
            current_row += 1
            
            # === TIME/SUBJECT/PROCTOR HEADER ROW ===
            cell = ws.cell(row=current_row, column=1, value='Time')
            cell.font = Font(bold=True, size=9)
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = thin_border
            cell.fill = header_fill
            
            # Subject and Proctor headers for each day
            for day in days_order:
                col = day_start_col[day]
                # Subject header
                cell = ws.cell(row=current_row, column=col, value='Subject')
                cell.font = Font(bold=True, size=8)
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.border = thin_border
                cell.fill = header_fill
                # Proctor header
                cell = ws.cell(row=current_row, column=col+1, value='Proctor')
                cell.font = Font(bold=True, size=8)
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.border = thin_border
                cell.fill = header_fill
            current_row += 1
            
            # === TIME SLOT ROWS ===
            # Group exams by time slot
            exams_by_time = defaultdict(lambda: defaultdict(list))
            all_time_slots = set()
            
            for exam in exam_schedules:
                start_str = exam.start_time.strftime('%I:%M').lstrip('0')
                end_str = exam.end_time.strftime('%I:%M').lstrip('0')
                time_key = f"{start_str}-{end_str}"
                day_name = exam.exam_date.strftime('%A')
                exams_by_time[time_key][day_name].append(exam)
                all_time_slots.add((exam.start_time, exam.end_time, time_key))
            
            # Sort time slots by start time
            sorted_time_slots = sorted(all_time_slots, key=lambda x: x[0])
            
            # Get lunch break times from settings
            if current_settings and current_settings.exam_lunch_start and current_settings.exam_lunch_end:
                lunch_end = current_settings.exam_lunch_end
                lunch_end_mins = lunch_end.hour * 60 + lunch_end.minute
            else:
                lunch_end_mins = 13 * 60
            
            # Track lunch break
            lunch_added = False
            
            for start_time, end_time, time_key in sorted_time_slots:
                # Check if we need to add LUNCH break
                if ((start_time.hour * 60 + start_time.minute) >= lunch_end_mins) and not lunch_added:
                    # Add LUNCH row (no fill - white background)
                    ws.merge_cells(f'A{current_row}:M{current_row}')
                    cell = ws.cell(row=current_row, column=1, value="LUNCH")
                    cell.font = Font(bold=True, size=9)
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                    for col in range(1, 14):
                        ws.cell(row=current_row, column=col).border = thin_border
                    current_row += 1
                    lunch_added = True
                
                # Time cell
                cell = ws.cell(row=current_row, column=1, value=time_key)
                cell.font = Font(size=9)
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.border = thin_border
                
                # For each day, show Subject and Proctor in separate columns
                for day in days_order:
                    col = day_start_col[day]
                    if day in exams_by_time[time_key]:
                        exams_for_slot = exams_by_time[time_key][day]
                        # Get subject(s) and proctor(s)
                        subjects = []
                        proctors = []
                        for exam in exams_for_slot:
                            subj = exam.subject.subject_code if exam.subject else ''
                            # Format proctor as "Mr./Ms. [Initial]. [Surname]"
                            if exam.faculty and exam.faculty.last_name:
                                salutation = "Ms." if exam.faculty.gender == "Female" else "Mr."
                                first_initial = exam.faculty.first_name[0].upper() if exam.faculty.first_name else ''
                                surname = exam.faculty.last_name
                                proctor = f"{salutation} {first_initial}. {surname}" if first_initial else f"{salutation} {surname}"
                            else:
                                proctor = ''
                            if subj:
                                subjects.append(subj)
                            if proctor:
                                proctors.append(proctor)
                        
                        # Subject cell
                        cell = ws.cell(row=current_row, column=col, value='\n'.join(subjects))
                        cell.font = Font(size=8)
                        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                        cell.border = thin_border
                        
                        # Proctor cell
                        cell = ws.cell(row=current_row, column=col+1, value='\n'.join(proctors))
                        cell.font = Font(size=8)
                        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                        cell.border = thin_border
                    else:
                        # Empty cells
                        cell = ws.cell(row=current_row, column=col, value='')
                        cell.border = thin_border
                        cell = ws.cell(row=current_row, column=col+1, value='')
                        cell.border = thin_border
                
                current_row += 1
            
            # Add spacing between sections
            current_row += 1
        
        # === SIGNATURE SECTION ===
        sig_start_row = current_row
        
        # Prepared by: (column A)
        ws.cell(row=sig_start_row, column=1, value='Prepared by:')
        ws.cell(row=sig_start_row, column=1).font = Font(size=10)
        
        # Get dean name from current user
        dean_name = current_user.full_name if current_user else 'Name of the Dean'
        dean_name = dean_name.upper()
        
        # Name (2 rows down, merged A-D for more space)
        ws.merge_cells(f'A{sig_start_row + 2}:D{sig_start_row + 2}')
        ws.cell(row=sig_start_row + 2, column=1, value=dean_name)
        ws.cell(row=sig_start_row + 2, column=1).font = Font(bold=True, size=10, underline='single')
        ws.cell(row=sig_start_row + 2, column=1).alignment = Alignment(horizontal='left')
        
        # Title (next row, merged A-D for more space to show full program name)
        ws.merge_cells(f'A{sig_start_row + 3}:D{sig_start_row + 3}')
        dean_title = f"Dean, {dept_display_name}"
        ws.cell(row=sig_start_row + 3, column=1, value=dean_title)
        ws.cell(row=sig_start_row + 3, column=1).font = Font(size=10)
        ws.cell(row=sig_start_row + 3, column=1).alignment = Alignment(horizontal='left')
        
        # Save to BytesIO
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        filename = f"{dept_code}_Exam_Schedule_Batch_Export.xlsx".replace(' ', '_')
        
        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=filename
        )
        
    except Exception as e:
        import traceback
        traceback.print_exc()
        flash(f'Error exporting exam schedule batch: {str(e)}', 'error')
        return redirect(url_for('schedule.exam_view', section_id=section_id))


@exam_schedule_bp.route('/export/<int:section_id>')
@login_required
def export_exam_schedule(section_id):
    """Export exam schedule to Excel - weekly grid format (same as batch export)"""
    from flask import send_file
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.drawing.image import Image as ExcelImage
    from openpyxl.worksheet.page import PageMargins
    from collections import defaultdict
    from datetime import timedelta
    import io
    import os
    
    try:
        section = Section.query.get_or_404(section_id)
        
        # Check if section's program is archived
        if section.program and section.program.is_archived:
            flash('Cannot export archived section exam schedules.', 'error')
            return redirect(url_for('schedule.exam_view'))
        
        # Get current academic settings
        current_settings = AcademicSettings.query.filter_by(is_active=True).first()
        
        # Query exam schedules for this section - exclude archived faculty and rooms
        query = ExamSchedule.query.filter_by(section_id=section_id, is_active=True)\
            .outerjoin(Faculty, ExamSchedule.faculty_id == Faculty.id)\
            .outerjoin(Room, ExamSchedule.room_id == Room.id)\
            .outerjoin(Building, Room.building_id == Building.id)\
            .filter(
                or_(
                    ExamSchedule.faculty_id == None,
                    and_(Faculty.is_active == True, Faculty.is_archived == False)
                ),
                or_(
                    ExamSchedule.room_id == None,
                    and_(
                        Room.is_available == True,
                        or_(
                            Building.id == None,
                            and_(Building.is_active == True, Building.is_archived == False)
                        )
                    )
                )
            )
        
        if current_settings:
            query = query.filter(
                ExamSchedule.academic_year == current_settings.academic_year,
                ExamSchedule.semester == current_settings.semester,
                ExamSchedule.exam_period == current_settings.exam_period
            )
        
        exam_schedules = query.order_by(ExamSchedule.exam_date, ExamSchedule.start_time).all()
        
        # Create workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Exam Schedule"
        
        # Set page layout for Legal paper (8.5" x 14") - Landscape
        ws.page_setup.paperSize = ws.PAPERSIZE_LEGAL
        ws.page_setup.orientation = 'landscape'
        ws.page_setup.fitToPage = True
        ws.page_setup.fitToWidth = 1
        ws.page_setup.fitToHeight = 0  # Allow multiple pages vertically
        ws.page_margins = PageMargins(left=0.25, right=0.25, top=0.5, bottom=0.5)
        ws.print_options.horizontalCentered = True
        
        # Get program info for logo and names
        program = section.program
        dept_logo_path = program.program_logo if program else None
        dept_code = program.program_code if program else ''
        dept_display_name = (program.department.department_name if (program and program.department) else (program.program_name if program else 'DEPT'))
        
        # === LOGOS AND HEADER (same as batch export) ===
        base_dir = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
        images_dir = os.path.join(base_dir, 'static', 'images')
        
        # Set column A width for logo
        ws.column_dimensions['A'].width = 12
        
        # Left logo (Norzagaray College) - position at A1
        logo_left_path = os.path.join(images_dir, 'norzagaray-college-logo.png')
        if os.path.exists(logo_left_path):
            img_left = ExcelImage(logo_left_path)
            img_left.width = 75
            img_left.height = 75
            ws.add_image(img_left, 'A1')
        
        # Right logo (Bagong Pilipinas) - position at M1 (to match batch export)
        logo_right_path = os.path.join(images_dir, 'bagong-pilipinas.png')
        if os.path.exists(logo_right_path):
            img_right = ExcelImage(logo_right_path)
            img_right.width = 80
            img_right.height = 80
            ws.add_image(img_right, 'M1')
        
        # Header text - LEFT ALIGNED at column B
        ws['B1'] = 'Republic of the Philippines'
        ws['B1'].font = Font(size=11)
        ws['B1'].alignment = Alignment(horizontal='left', vertical='center')
        
        ws['B2'] = 'Municipality of Norzagaray'
        ws['B2'].font = Font(size=11)
        ws['B2'].alignment = Alignment(horizontal='left', vertical='center')
        
        ws['B3'] = 'NORZAGARAY COLLEGE'
        ws['B3'].font = Font(bold=True, size=11)
        ws['B3'].alignment = Alignment(horizontal='left', vertical='center')
        
        ws['B4'] = dept_display_name
        ws['B4'].font = Font(bold=True, size=11)
        ws['B4'].alignment = Alignment(horizontal='left', vertical='center')
        
        # Set row heights for header area
        ws.row_dimensions[1].height = 18
        ws.row_dimensions[2].height = 15
        ws.row_dimensions[3].height = 15
        ws.row_dimensions[4].height = 15
        ws.row_dimensions[5].height = 8  # Spacer
        
        # Get exam period title
        if current_settings:
            exam_period_title = f"{current_settings.exam_period.upper()} EXAMINATION"
        else:
            exam_period_title = "EXAMINATION"
        
        # Row 6: EXAMINATION type (centered across A-M)
        ws.merge_cells('A6:M6')
        cell = ws.cell(row=6, column=1, value=exam_period_title)
        cell.font = Font(bold=True, size=12)
        cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # Define styles
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        yellow_fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
        header_fill = PatternFill(start_color='F5F5F5', end_color='F5F5F5', fill_type='solid')
        
        # Set column widths - same as batch export
        ws.column_dimensions['B'].width = 10  # Monday Subject
        ws.column_dimensions['C'].width = 10  # Monday Proctor
        ws.column_dimensions['D'].width = 10  # Tuesday Subject
        ws.column_dimensions['E'].width = 10  # Tuesday Proctor
        ws.column_dimensions['F'].width = 10  # Wednesday Subject
        ws.column_dimensions['G'].width = 10  # Wednesday Proctor
        ws.column_dimensions['H'].width = 10  # Thursday Subject
        ws.column_dimensions['I'].width = 10  # Thursday Proctor
        ws.column_dimensions['J'].width = 10  # Friday Subject
        ws.column_dimensions['K'].width = 10  # Friday Proctor
        ws.column_dimensions['L'].width = 10  # Saturday Subject
        ws.column_dimensions['M'].width = 12  # Saturday Proctor - wider for right logo
        
        # Days of week mapping - each day has 2 columns (Subject at col, Proctor at col+1)
        days_order = AcademicSettings.get_active_operation_days()
        day_start_col = {day: 2 + idx * 2 for idx, day in enumerate(days_order)}
        
        # Get exam period date range from settings
        exam_period_start = None
        exam_period_end = None
        if current_settings:
            if hasattr(current_settings, 'exam_period_start') and current_settings.exam_period_start:
                exam_period_start = current_settings.exam_period_start
            if hasattr(current_settings, 'exam_period_end') and current_settings.exam_period_end:
                exam_period_end = current_settings.exam_period_end
        
        # Use configured exam period dates for the header if available
        if exam_period_start and exam_period_end:
            date_range_str = f"{exam_period_start.strftime('%B %d')}-{exam_period_end.strftime('%d, %Y')}"
            # Map the configured date range to days of the week
            date_by_day = {}
            current_date = exam_period_start
            while current_date <= exam_period_end:
                day_name = current_date.strftime('%A')
                if day_name != 'Sunday':  # Skip Sundays
                    date_by_day[day_name] = current_date
                current_date += timedelta(days=1)
        else:
            # Find date range from actual exam schedules
            if exam_schedules:
                all_dates = sorted(set(exam.exam_date for exam in exam_schedules))
                date_by_day = {}
                for d in all_dates:
                    day_name = d.strftime('%A')
                    date_by_day[day_name] = d
                if all_dates:
                    date_range_str = f"{all_dates[0].strftime('%B %d')}-{all_dates[-1].strftime('%d, %Y')}"
                else:
                    date_range_str = ""
            else:
                date_range_str = ""
                date_by_day = {}
        
        # Row 7: Date range (centered across A-M)
        ws.merge_cells('A7:M7')
        cell = ws.cell(row=7, column=1, value=date_range_str)
        cell.font = Font(size=11)
        cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # Get section display name
        if hasattr(section, 'full_section_name'):
            section_display = section.full_section_name
        else:
            section_display = f"{dept_code}-{section.year_level}{section.section_name}"
        
        # Start row for content
        current_row = 8
        
        # === SECTION HEADER (Yellow background) ===
        ws.merge_cells(f'A{current_row}:M{current_row}')
        cell = ws.cell(row=current_row, column=1, value=section_display)
        cell.font = Font(bold=True, size=11)
        cell.alignment = Alignment(horizontal='left', vertical='center')
        cell.fill = yellow_fill
        for col in range(1, 14):
            ws.cell(row=current_row, column=col).border = thin_border
        current_row += 1
        
        # === DAY ROW (merged cells for each day spanning Subject+Proctor) ===
        cell = ws.cell(row=current_row, column=1, value='Day')
        cell.font = Font(bold=True, size=9)
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = thin_border
        cell.fill = header_fill
        
        for day in days_order:
            col = day_start_col[day]
            ws.merge_cells(start_row=current_row, start_column=col, end_row=current_row, end_column=col+1)
            cell = ws.cell(row=current_row, column=col, value=day)
            cell.font = Font(bold=True, size=9)
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = thin_border
            ws.cell(row=current_row, column=col+1).border = thin_border
        current_row += 1
        
        # === DATE ROW (merged cells for each day) ===
        cell = ws.cell(row=current_row, column=1, value='Date')
        cell.font = Font(bold=True, size=9)
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = thin_border
        cell.fill = header_fill
        
        for day in days_order:
            col = day_start_col[day]
            ws.merge_cells(start_row=current_row, start_column=col, end_row=current_row, end_column=col+1)
            if day in date_by_day:
                date_str = date_by_day[day].strftime('%B %d')
            else:
                date_str = ''
            cell = ws.cell(row=current_row, column=col, value=date_str)
            cell.font = Font(size=9)
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = thin_border
            ws.cell(row=current_row, column=col+1).border = thin_border
        current_row += 1
        
        # === ROOM ROW (merged cells for each day) ===
        cell = ws.cell(row=current_row, column=1, value='Room')
        cell.font = Font(bold=True, size=9)
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = thin_border
        cell.fill = header_fill
        
        # Get rooms by day
        rooms_by_day = defaultdict(set)
        for exam in exam_schedules:
            day_name = exam.exam_date.strftime('%A')
            room_name = exam.room.room_number if exam.room else ''
            if room_name:
                rooms_by_day[day_name].add(room_name)
        
        for day in days_order:
            col = day_start_col[day]
            ws.merge_cells(start_row=current_row, start_column=col, end_row=current_row, end_column=col+1)
            rooms = ', '.join(sorted(rooms_by_day[day])) if rooms_by_day[day] else ''
            cell = ws.cell(row=current_row, column=col, value=rooms)
            cell.font = Font(size=9)
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = thin_border
            ws.cell(row=current_row, column=col+1).border = thin_border
        current_row += 1
        
        # === TIME/SUBJECT/PROCTOR HEADER ROW ===
        cell = ws.cell(row=current_row, column=1, value='Time')
        cell.font = Font(bold=True, size=9)
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = thin_border
        cell.fill = header_fill
        
        # Subject and Proctor headers for each day
        for day in days_order:
            col = day_start_col[day]
            # Subject header
            cell = ws.cell(row=current_row, column=col, value='Subject')
            cell.font = Font(bold=True, size=8)
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = thin_border
            cell.fill = header_fill
            # Proctor header
            cell = ws.cell(row=current_row, column=col+1, value='Proctor')
            cell.font = Font(bold=True, size=8)
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = thin_border
            cell.fill = header_fill
        current_row += 1
        
        # === TIME SLOT ROWS ===
        # Group exams by time slot
        exams_by_time = defaultdict(lambda: defaultdict(list))
        all_time_slots = set()
        
        for exam in exam_schedules:
            start_str = exam.start_time.strftime('%I:%M').lstrip('0')
            end_str = exam.end_time.strftime('%I:%M').lstrip('0')
            time_key = f"{start_str}-{end_str}"
            day_name = exam.exam_date.strftime('%A')
            exams_by_time[time_key][day_name].append(exam)
            all_time_slots.add((exam.start_time, exam.end_time, time_key))
        
        # Sort time slots by start time
        sorted_time_slots = sorted(all_time_slots, key=lambda x: x[0])
        
        # Get lunch break times from settings
        if current_settings and current_settings.exam_lunch_start and current_settings.exam_lunch_end:
            lunch_end = current_settings.exam_lunch_end
            lunch_end_mins = lunch_end.hour * 60 + lunch_end.minute
        else:
            lunch_end_mins = 13 * 60
        
        # Track lunch break
        lunch_added = False
        
        for start_time, end_time, time_key in sorted_time_slots:
            # Check if we need to add LUNCH break
            if ((start_time.hour * 60 + start_time.minute) >= lunch_end_mins) and not lunch_added:
                # Add LUNCH row (no fill - white background)
                ws.merge_cells(f'A{current_row}:M{current_row}')
                cell = ws.cell(row=current_row, column=1, value="LUNCH")
                cell.font = Font(bold=True, size=9)
                cell.alignment = Alignment(horizontal='center', vertical='center')
                for col in range(1, 14):
                    ws.cell(row=current_row, column=col).border = thin_border
                current_row += 1
                lunch_added = True
            
            # Time cell
            cell = ws.cell(row=current_row, column=1, value=time_key)
            cell.font = Font(size=9)
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = thin_border
            
            # For each day, show Subject and Proctor in separate columns
            for day in days_order:
                col = day_start_col[day]
                if day in exams_by_time[time_key]:
                    exams_for_slot = exams_by_time[time_key][day]
                    # Get subject(s) and proctor(s)
                    subjects = []
                    proctors = []
                    for exam in exams_for_slot:
                        subj = exam.subject.subject_code if exam.subject else ''
                        # Format proctor as "Mr./Ms. [Initial]. [Surname]"
                        if exam.faculty and exam.faculty.last_name:
                            salutation = "Ms." if exam.faculty.gender == "Female" else "Mr."
                            first_initial = exam.faculty.first_name[0].upper() if exam.faculty.first_name else ''
                            surname = exam.faculty.last_name
                            proctor = f"{salutation} {first_initial}. {surname}" if first_initial else f"{salutation} {surname}"
                        else:
                            proctor = ''
                        if subj:
                            subjects.append(subj)
                        if proctor:
                            proctors.append(proctor)
                    
                    # Subject cell
                    cell = ws.cell(row=current_row, column=col, value='\n'.join(subjects))
                    cell.font = Font(size=8)
                    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                    cell.border = thin_border
                    
                    # Proctor cell
                    cell = ws.cell(row=current_row, column=col+1, value='\n'.join(proctors))
                    cell.font = Font(size=8)
                    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                    cell.border = thin_border
                else:
                    # Empty cells
                    cell = ws.cell(row=current_row, column=col, value='')
                    cell.border = thin_border
                    cell = ws.cell(row=current_row, column=col+1, value='')
                    cell.border = thin_border
            
            current_row += 1
        
        # === SIGNATURE SECTION ===
        sig_start_row = current_row + 1
        
        # Prepared by: (column A)
        ws.cell(row=sig_start_row, column=1, value='Prepared by:')
        ws.cell(row=sig_start_row, column=1).font = Font(size=10)
        
        # Get dean name from current user
        dean_name = current_user.full_name if current_user else 'Name of the Dean'
        dean_name = dean_name.upper()
        
        # Name (2 rows down, merged A-D for more space)
        ws.merge_cells(f'A{sig_start_row + 2}:D{sig_start_row + 2}')
        ws.cell(row=sig_start_row + 2, column=1, value=dean_name)
        ws.cell(row=sig_start_row + 2, column=1).font = Font(bold=True, size=10, underline='single')
        ws.cell(row=sig_start_row + 2, column=1).alignment = Alignment(horizontal='left')
        
        # Title (next row, merged A-D for more space to show full program name)
        ws.merge_cells(f'A{sig_start_row + 3}:D{sig_start_row + 3}')
        dean_title = f"Dean, {dept_display_name}"
        ws.cell(row=sig_start_row + 3, column=1, value=dean_title)
        ws.cell(row=sig_start_row + 3, column=1).font = Font(size=10)
        ws.cell(row=sig_start_row + 3, column=1).alignment = Alignment(horizontal='left')
        
        # Save to BytesIO
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        filename = f"{dept_code}_{section.year_level}{section.section_name}_Exam_Schedule.xlsx".replace(' ', '_')
        
        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=filename
        )
        
    except Exception as e:
        flash(f'Error exporting exam schedule: {str(e)}', 'error')
        return redirect(url_for('schedule.exam_view', section_id=section_id))


@exam_schedule_bp.route('/export/<int:section_id>/pdf')
@login_required
def export_exam_schedule_pdf(section_id):
    """Export exam schedule to PDF - weekly grid format"""
    try:
        section = Section.query.get_or_404(section_id)
        
        # Check if section's program is archived
        if section.program and section.program.is_archived:
            flash('Cannot export archived section exam schedules.', 'error')
            return redirect(url_for('schedule.exam_view'))
        
        current_settings = AcademicSettings.query.filter_by(is_active=True).first()
        
        # Query exam schedules - exclude archived faculty and rooms
        query = ExamSchedule.query.filter_by(section_id=section_id, is_active=True)\
            .outerjoin(Faculty, ExamSchedule.faculty_id == Faculty.id)\
            .outerjoin(Room, ExamSchedule.room_id == Room.id)\
            .outerjoin(Building, Room.building_id == Building.id)\
            .filter(
                or_(
                    ExamSchedule.faculty_id == None,
                    and_(Faculty.is_active == True, Faculty.is_archived == False)
                ),
                or_(
                    ExamSchedule.room_id == None,
                    and_(
                        Room.is_available == True,
                        or_(
                            Building.id == None,
                            and_(Building.is_active == True, Building.is_archived == False)
                        )
                    )
                )
            )
        
        if current_settings:
            query = query.filter(
                ExamSchedule.academic_year == current_settings.academic_year,
                ExamSchedule.semester == current_settings.semester,
                ExamSchedule.exam_period == current_settings.exam_period
            )
        exam_schedules = query.order_by(ExamSchedule.day_of_week, ExamSchedule.start_time).all()
        
        # Convert ExamSchedule to Schedule-like format for PDF function
        schedules = []
        for exam in exam_schedules:
            # Create a mock Schedule object with exam schedule attributes
            class MockSchedule:
                def __init__(self, exam):
                    self.day_of_week = exam.day_of_week
                    self.start_time = exam.start_time
                    self.end_time = exam.end_time
                    self.subject = exam.subject
                    self.faculty = exam.faculty
                    self.room = exam.room
                    self.schedule_type = 'EXAM'
            
            schedules.append(MockSchedule(exam))
        
        # Prepare metadata
        dept_name = section.program.program_name if section.program else 'DEPT'
        dept_code = section.program.program_code if section.program else ''
        
        # Import format_semester_text for proper semester formatting
        from app.services.export_service import format_semester_text
        
        if current_settings:
            semester_text = format_semester_text(current_settings.semester, current_settings.academic_year) + f" - {current_settings.exam_period.upper()}"
        else:
            semester_text = "EXAM SCHEDULE"
        
        section_display = section.full_section_name if hasattr(section, 'full_section_name') else f"{dept_code}-{section.year_level}{section.section_name}"
        
        # Import PDF creation function from schedule.py
        from app.routes.schedule import create_pdf_schedule
        
        # Create PDF
        output = create_pdf_schedule(
            schedules=schedules,
            title='EXAMINATION SCHEDULE',
            semester_text=semester_text,
            section_display=section_display,
            dept_name=dept_name.upper(),
            filename=f"{dept_code}_{section.year_level}{section.section_name}_Exam_Schedule.pdf"
        )
        
        filename = f"{dept_code}_{section.year_level}{section.section_name}_Exam_Schedule.pdf".replace(' ', '_')
        return send_file(
            output,
            mimetype='application/pdf',
            as_attachment=True,
            download_name=filename
        )
        
    except Exception as e:
        flash(f'Error exporting exam PDF schedule: {str(e)}', 'error')
        return redirect(url_for('schedule.exam_view', section_id=section_id))


@exam_schedule_bp.route('/cleanup-archived', methods=['POST'])
@login_required
@role_required('admin', 'super_admin')
def cleanup_archived():
    """Delete exam schedules that have archived sections, programs, faculty, or rooms"""
    try:
        # Get all active exam schedules
        all_exam_schedules = ExamSchedule.query.filter_by(is_active=True).all()
        
        deleted_count = 0
        deleted_details = []
        
        for exam in all_exam_schedules:
            if exam.has_archived_relationships():
                # Build detail string for logging
                reason_parts = []
                if exam.section and exam.section.program and exam.section.program.is_archived:
                    reason_parts.append(f"archived program: {exam.section.program.program_name}")
                if exam.faculty and exam.faculty.is_archived:
                    reason_parts.append(f"archived faculty: {exam.faculty.full_name}")
                if exam.room and not exam.room.is_available:
                    reason_parts.append(f"unavailable room: {exam.room.room_number}")
                if exam.room and exam.room.building and exam.room.building.is_archived:
                    reason_parts.append(f"archived building: {exam.room.building.building_name}")
                
                detail = f"{exam.subject.subject_code if exam.subject else 'N/A'} - {', '.join(reason_parts)}"
                deleted_details.append(detail)
                
                # Log deletion
                from app.utils.activity_logger import log_delete
                log_delete('exam_schedule', exam.id, f'{exam.subject.subject_code} - {exam.section.full_section_name}', {
                    'reason': 'Cleanup: ' + ', '.join(reason_parts),
                    'exam_date': str(exam.exam_date) if exam.exam_date else None
                })
                
                db.session.delete(exam)
                deleted_count += 1
        
        db.session.commit()
        
        if deleted_count > 0:
            flash(f'Successfully deleted {deleted_count} exam schedule(s) with archived relationships.', 'success')
            # Optionally log details
            print(f"[CLEANUP] Deleted exam schedules:\n" + "\n".join(deleted_details))
        else:
            flash('No exam schedules with archived relationships found.', 'info')
        
    except Exception as e:
        db.session.rollback()
        flash(f'Error cleaning up archived exam schedules: {str(e)}', 'danger')
        import traceback
        traceback.print_exc()
    
    return redirect(url_for('schedule.exam_view'))


# ============================================================================
# BATCH EXAM BUILDER ROUTES
# ============================================================================

@exam_schedule_bp.route('/batch-generate', methods=['POST'])
@login_required
@csrf.exempt
def batch_generate():
    """Generate batch exam preview for all unexamined subjects in a section."""
    from app.services.auto_scheduler import AutoScheduler

    try:
        data = request.get_json()
        section_id = data.get('section_id')
        curriculum_id = data.get('curriculum_id')
        preferred_building_id = data.get('preferred_building_id')

        if not section_id:
            return jsonify({'success': False, 'error': 'section_id required'}), 400

        section = Section.query.get_or_404(section_id)
        user_program_ids = current_user.get_program_ids()
        if user_program_ids is not None and section.program_id not in user_program_ids:
            return jsonify({'success': False, 'error': 'Access denied'}), 403

        scheduler = AutoScheduler()
        result = scheduler.generate_exam_batch_preview(
            section_id,
            curriculum_id=curriculum_id,
            preferred_building_id=preferred_building_id
        )
        return jsonify(result)

    except Exception as e:
        print(f"Exam batch generate error: {str(e)}")
        import traceback
        traceback.print_exc()
        return jsonify({'success': False, 'error': str(e)}), 500


@exam_schedule_bp.route('/batch-confirm', methods=['POST'])
@login_required
@csrf.exempt
def batch_confirm():
    """Confirm and save all batch exam schedules."""
    from app.services.auto_scheduler import AutoScheduler

    try:
        data = request.get_json()
        section_id = data.get('section_id')
        proposed = data.get('proposed', [])

        if not section_id or not proposed:
            return jsonify({'success': False, 'error': 'section_id and proposed required'}), 400

        section = Section.query.get_or_404(section_id)
        user_program_ids = current_user.get_program_ids()
        if user_program_ids is not None and section.program_id not in user_program_ids:
            return jsonify({'success': False, 'error': 'Access denied'}), 403

        scheduler = AutoScheduler()
        result = scheduler.confirm_exam_schedule(section_id, proposed, user_id=current_user.id)

        if result.get('success') and result.get('created', 0) > 0:
            try:
                # Fetch a newly-created exam to use as broadcast reference
                from app.models.settings import AcademicSettings as AS
                _as = AS.query.filter_by(is_active=True).first()
                if _as:
                    ref_exam = ExamSchedule.query.filter_by(
                        section_id=section_id,
                        is_active=True,
                        academic_year=_as.academic_year,
                        semester=_as.semester,
                        exam_period=_as.exam_period
                    ).first()
                    if ref_exam:
                        broadcast_schedule_change(ref_exam, 'batch_created', 'exam')
            except Exception:
                pass  # Don't let broadcast failure affect the response

        return jsonify(result)

    except Exception as e:
        print(f"Exam batch confirm error: {str(e)}")
        import traceback
        traceback.print_exc()
        return jsonify({'success': False, 'error': str(e)}), 500


@exam_schedule_bp.route('/batch-check-conflicts', methods=['POST'])
@login_required
@csrf.exempt
def batch_check_conflicts():
    """
    Check conflicts for ALL batch exam rows in a single request.
    Returns per-row conflict results including intra-batch detection.
    """
    from app.services.conflict_detector import conflict_detector
    from datetime import datetime as dt

    try:
        data = request.get_json()
        section_id = data.get('section_id')
        rows = data.get('rows', [])

        if not section_id or not rows:
            return jsonify({'success': False, 'error': 'section_id and rows required'}), 400

        current_settings = AcademicSettings.query.filter_by(is_active=True).first()
        exam_start_cfg = dt_time(7, 0)
        exam_end_cfg = dt_time(17, 0)
        exam_duration_limit = 120
        if current_settings:
            exam_start_cfg = _coerce_setting_time(getattr(current_settings, 'exam_start_time', None), current_settings.exam_start_hour or 7)
            exam_end_cfg = _coerce_setting_time(getattr(current_settings, 'exam_end_time', None), current_settings.exam_end_hour or 17)
            exam_duration_limit = getattr(current_settings, 'exam_duration_limit', 120) or 120

        existing_query = ExamSchedule.query.filter_by(is_active=True)
        if current_settings:
            existing_query = existing_query.filter_by(
                academic_year=current_settings.academic_year,
                semester=current_settings.semester,
                exam_period=current_settings.exam_period
            )
        existing_exams = existing_query.all()

        # Pre-load faculty names
        batch_faculty_ids = set()
        for r in rows:
            if r.get('faculty_id'):
                try:
                    batch_faculty_ids.add(int(r['faculty_id']))
                except (ValueError, TypeError):
                    pass
        faculty_name_map = {}
        if batch_faculty_ids:
            faculty_objs = Faculty.query.filter(Faculty.id.in_(batch_faculty_ids)).all()
            for f in faculty_objs:
                faculty_name_map[f.id] = f.full_name

        parsed_rows = []
        for i, row in enumerate(rows):
            start_str = row.get('start_time', '')
            end_str = row.get('end_time', '')
            date_str = row.get('exam_date', '')
            if not start_str or not end_str or not date_str:
                parsed_rows.append(None)
                continue
            try:
                st = dt.strptime(start_str, '%H:%M').time()
                et = dt.strptime(end_str, '%H:%M').time()
                ed = dt.strptime(date_str, '%Y-%m-%d').date()
            except ValueError:
                parsed_rows.append(None)
                continue

            parsed_rows.append({
                'index': i,
                'section_id': section_id,
                'subject_id': row.get('subject_id'),
                'faculty_id': int(row['faculty_id']) if row.get('faculty_id') else None,
                'room_id': int(row['room_id']) if row.get('room_id') else None,
                'exam_date': ed,
                'start_time': st,
                'end_time': et,
                'subject_code': row.get('subject_code', f'Row {i+1}'),
                'schedule_type': row.get('schedule_type', 'lecture'),
                'exam_schedule_id': int(row['exam_schedule_id']) if row.get('exam_schedule_id') else None,
            })

        results = []
        ok_count = 0
        conflict_count = 0
        warning_count = 0

        for i, parsed in enumerate(parsed_rows):
            if parsed is None:
                results.append({
                    'index': i,
                    'status': 'warning',
                    'conflicts': [{'type': 'time_invalid', 'severity': 'critical',
                                   'message': 'Missing or invalid date/time values'}]
                })
                warning_count += 1
                continue

            row_conflicts = []

            # 1) Check against existing DB exams
            exam_data = {
                'section_id': section_id,
                'subject_id': parsed['subject_id'],
                'faculty_id': parsed['faculty_id'],
                'room_id': parsed['room_id'],
                'exam_date': parsed['exam_date'],
                'start_time': parsed['start_time'],
                'end_time': parsed['end_time'],
                'schedule_type': parsed.get('schedule_type', 'lecture'),
            }
            db_conflicts = conflict_detector.detect_exam_conflicts(
                exam_data, existing_exams,
                exclude_exam_id=parsed.get('exam_schedule_id')
            )
            for c in db_conflicts:
                row_conflicts.append(c.to_dict())

            # 2) Intra-batch checks
            for j, other in enumerate(parsed_rows):
                if j == i or other is None:
                    continue

                if parsed['exam_date'] != other['exam_date']:
                    continue
                if parsed['start_time'] >= other['end_time'] or parsed['end_time'] <= other['start_time']:
                    continue

                time_disp = f"{other['start_time'].strftime('%I:%M %p')}-{other['end_time'].strftime('%I:%M %p')}"
                date_disp = other['exam_date'].strftime('%b %d')

                row_conflicts.append({
                    'type': 'section_batch',
                    'severity': 'critical',
                    'message': f"Overlaps with Row {j+1} ({other['subject_code']}) on {date_disp} {time_disp}",
                    'details': {'other_row': j+1}
                })

                if parsed['faculty_id'] and other['faculty_id'] and parsed['faculty_id'] == other['faculty_id']:
                    row_conflicts.append({
                        'type': 'faculty_batch',
                        'severity': 'high',
                        'message': f"Same proctor in Row {j+1} ({other['subject_code']}) at {time_disp}",
                        'details': {'other_row': j+1}
                    })

                if parsed['room_id'] and other['room_id'] and parsed['room_id'] == other['room_id']:
                    row_conflicts.append({
                        'type': 'room_batch',
                        'severity': 'high',
                        'message': f"Same room in Row {j+1} ({other['subject_code']}) at {time_disp}",
                        'details': {'other_row': j+1}
                    })

            # 3) Exam hours check
            if _minutes_of(parsed['start_time']) < _minutes_of(exam_start_cfg):
                row_conflicts.append({
                    'type': 'schedule_hours', 'severity': 'medium',
                    'message': f"Start time before exam hours ({exam_start_cfg.strftime('%H:%M')})", 'details': {}
                })
            if _minutes_of(parsed['end_time']) > _minutes_of(exam_end_cfg):
                row_conflicts.append({
                    'type': 'schedule_hours', 'severity': 'medium',
                    'message': f"End time after exam hours ({exam_end_cfg.strftime('%H:%M')})", 'details': {}
                })

            # 4) Duration limit
            duration = (parsed['end_time'].hour * 60 + parsed['end_time'].minute) - (parsed['start_time'].hour * 60 + parsed['start_time'].minute)
            if duration > exam_duration_limit:
                row_conflicts.append({
                    'type': 'duration_limit', 'severity': 'medium',
                    'message': f"Duration ({duration} min) exceeds limit ({exam_duration_limit} min)", 'details': {}
                })

            # 5) Proctor availability — already checked by conflict_detector in step 1
            # (detect_exam_conflicts → _check_proctor_availability handles both
            #  'unavailable' as HIGH and 'not_in_schedule' as MEDIUM)

            has_critical = any(c['severity'] in ('critical', 'high') for c in row_conflicts)
            has_warning = any(c['severity'] in ('medium', 'low') for c in row_conflicts)

            if has_critical:
                status = 'conflict'
                conflict_count += 1
            elif has_warning:
                status = 'warning'
                warning_count += 1
            else:
                status = 'ok'
                ok_count += 1

            results.append({'index': i, 'status': status, 'conflicts': row_conflicts})

        return jsonify({
            'success': True,
            'rows': results,
            'summary': {'total': len(rows), 'ok': ok_count, 'conflicts': conflict_count, 'warnings': warning_count}
        })

    except Exception as e:
        print(f"Exam batch check conflicts error: {str(e)}")
        import traceback
        traceback.print_exc()
        return jsonify({'success': False, 'error': str(e)}), 500


@exam_schedule_bp.route('/batch-available-rooms')
@login_required
def batch_available_rooms():
    """Get rooms available (no exam conflicts) at a specific date/time."""
    from app.services.auto_scheduler import AutoScheduler

    try:
        exam_date = request.args.get('exam_date')
        start_time = request.args.get('start_time')
        end_time = request.args.get('end_time')
        building_id = request.args.get('building_id', type=int)

        if not all([exam_date, start_time, end_time]):
            return jsonify({'rooms': [], 'error': 'exam_date, start_time, end_time required'}), 400

        scheduler = AutoScheduler()
        rooms = scheduler.get_available_exam_rooms(
            exam_date,
            start_time,
            end_time,
            preferred_building_id=building_id
        )
        return jsonify({'rooms': rooms})

    except Exception as e:
        print(f"Exam batch available rooms error: {str(e)}")
        return jsonify({'rooms': [], 'error': str(e)}), 500


@exam_schedule_bp.route('/batch-unscheduled-subjects/<int:section_id>')
@login_required
def batch_unscheduled_subjects(section_id):
    """Get subjects that don't yet have an exam for the current period."""
    from app.services.auto_scheduler import AutoScheduler

    try:
        section = Section.query.get_or_404(section_id)
        user_program_ids = current_user.get_program_ids()
        if user_program_ids is not None and section.program_id not in user_program_ids:
            return jsonify({'success': False, 'error': 'Access denied'}), 403

        scheduler = AutoScheduler()
        curriculum_id = request.args.get('curriculum_id', type=int)
        include_all = request.args.get('include_all', '').lower() in ('1', 'true')
        result = scheduler.get_unscheduled_exam_subjects(section_id, curriculum_id=curriculum_id,
                                                          include_all=include_all)
        return jsonify(result)

    except Exception as e:
        print(f"Exam batch unscheduled subjects error: {str(e)}")
        return jsonify({'success': False, 'error': str(e)}), 500
