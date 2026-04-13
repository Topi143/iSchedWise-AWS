import inspect
from types import SimpleNamespace

from flask import Flask
from openpyxl import Workbook

from app.routes import reports as reports_routes


MOJIBAKE_BULLET = '\u00e2\u20ac\u00a2'


def _unwrap(func):
    return inspect.unwrap(func)


def _sample_stats():
    return {
        'total_schedules': 6,
        'total_exam_schedules': 2,
        'lecture_count': 4,
        'lab_count': 2,
        'total_faculty': 5,
        'faculty_with_schedules': 4,
        'unassigned_faculty_count': 1,
        'overloaded_faculty_count': 1,
        'total_sections': 3,
        'sections_with_schedules': 2,
        'schedule_completion_rate': 67,
        'avg_room_utilization': 35,
        'total_rooms': 4,
        'rooms_in_use': 3,
        'unused_rooms_count': 1,
        'faculty_workloads': [],
        'room_utilizations': [],
        'room_utilization_by_building': {},
        'total_room_hours_used': 24,
        'schedule_by_day': {
            'Monday': 3,
            'Tuesday': 2,
            'Wednesday': 1,
        },
    }


def _collect_string_cells(ws):
    values = []
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
        for cell in row:
            if isinstance(cell.value, str):
                values.append(cell.value)
    return values


class _DummySettingsQuery:
    def filter_by(self, **_kwargs):
        return self

    def first(self):
        return SimpleNamespace(
            academic_year='2025-2026',
            semester='1st Semester',
            exam_period='Midterm',
        )


class _DummyAcademicSettings:
    query = _DummySettingsQuery()

    @staticmethod
    def get_active_operation_days():
        return ['Monday', 'Tuesday', 'Wednesday']


def test_format_export_bullet_uses_safe_ascii_prefix():
    assert reports_routes._format_export_bullet('sample insight') == '-  sample insight'


def test_excel_insight_builders_do_not_emit_mojibake(monkeypatch):
    stats = _sample_stats()

    monkeypatch.setattr(
        reports_routes,
        'create_reports_excel_header',
        lambda _ws, *_args, **_kwargs: 1,
    )
    monkeypatch.setattr(reports_routes, 'AcademicSettings', _DummyAcademicSettings)

    wb = Workbook()
    ws_summary = wb.active
    ws_summary.title = 'Summary'
    reports_routes.create_executive_summary_sheet(
        ws_summary,
        stats,
        '2025-2026',
        '1st Semester',
        'OFFICE OF THE REGISTRAR',
    )

    ws_weekly = wb.create_sheet('Weekly')
    reports_routes.create_enhanced_weekly_sheet(
        ws_weekly,
        stats,
        '2025-2026',
        '1st Semester',
        'OFFICE OF THE REGISTRAR',
    )

    values = _collect_string_cells(ws_summary) + _collect_string_cells(ws_weekly)

    assert any(v.startswith('-  ') for v in values)
    assert all(MOJIBAKE_BULLET not in v for v in values)


def test_pdf_export_insights_do_not_emit_mojibake(monkeypatch):
    app = Flask(__name__)

    captured_paragraphs = []

    class _DummyDoc:
        def __init__(self, *args, **kwargs):
            self.pagesize = kwargs.get('pagesize', (595, 842))
            self.leftMargin = kwargs.get('leftMargin', 36)
            self.rightMargin = kwargs.get('rightMargin', 36)

        def build(self, _elements, onFirstPage=None, onLaterPages=None):
            if onFirstPage:
                class _Canvas:
                    def saveState(self):
                        pass

                    def setFont(self, *_args, **_kwargs):
                        pass

                    def setFillColor(self, *_args, **_kwargs):
                        pass

                    def drawCentredString(self, *_args, **_kwargs):
                        pass

                    def drawString(self, *_args, **_kwargs):
                        pass

                    def drawRightString(self, *_args, **_kwargs):
                        pass

                    def restoreState(self):
                        pass

                    def getPageNumber(self):
                        return 1

                onFirstPage(_Canvas(), self)

    def _paragraph_stub(text, _style):
        captured_paragraphs.append(text)
        return text

    monkeypatch.setattr(reports_routes, 'SimpleDocTemplate', _DummyDoc)
    monkeypatch.setattr(reports_routes, 'Paragraph', _paragraph_stub)
    monkeypatch.setattr(reports_routes, '_create_report_header', lambda *_args, **_kwargs: [])
    monkeypatch.setattr(reports_routes, 'current_user', SimpleNamespace(get_program_ids=lambda: None))
    monkeypatch.setattr(reports_routes, 'calculate_statistics', lambda *_args, **_kwargs: _sample_stats())
    monkeypatch.setattr(reports_routes, 'AcademicSettings', _DummyAcademicSettings)

    route_func = _unwrap(reports_routes.export_pdf)

    with app.test_request_context('/reports/export/pdf'):
        response = route_func()

    assert response.status_code == 200
    assert any(text.startswith('-  ') for text in captured_paragraphs)
    assert all(MOJIBAKE_BULLET not in text for text in captured_paragraphs)
